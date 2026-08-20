import streamlit as st
import pandas as pd
import os
from src.core.excel_utils import df_to_excel_bytes, format_periodo

def render(empresa_seleccionada, empresa_path):
    if empresa_seleccionada == "🌐 [GLOBAL] Configuración General":
        st.warning("⚠️ Módulo de Sociedad Activa: Por favor, selecciona una empresa de trabajo específica (ej. Pacifico SpA) en la barra lateral izquierda para acceder a esta sección.")
        st.stop()
        
    from src.models.trial_balance_db import TrialBalanceDB
    TrialBalanceDB.initialize()
    
    st.title("⚙️ Ejecución y Generación de Reportes Financieros")
    
    available_periods = TrialBalanceDB.get_available_periods(empresa_seleccionada)
    
    if not available_periods:
        st.warning(f"⚠️ No hay periodos de Trial Balance cargados en la Base de Datos Histórica para {empresa_seleccionada}.")
    
    # Determinar las opciones para el selector
    lista_opciones = available_periods.copy()
    if not lista_opciones:
        if 'tb_df' in st.session_state:
            lista_opciones = ["Memoria Temporal (Sin guardar)"]
        else:
            lista_opciones = ["Sin historial"]

    col_p1, col_p2 = st.columns(2)
    with col_p1:
        periodo_actual = st.selectbox("Periodo Actual (Mes/Año a emitir)", lista_opciones, index=0, format_func=format_periodo)
    with col_p2:
        periodo_comp = st.selectbox("Periodo Comparativo", lista_opciones, index=1 if len(lista_opciones)>1 else 0, format_func=format_periodo)
        
    # Cargar dinámicamente desde base de datos a memoria estandar si hay datos (usando caché de sesión)
    db_cache_key = f"_ef_db_cache_{empresa_seleccionada}_{periodo_actual}_{periodo_comp}"
    if db_cache_key not in st.session_state or 'tb_df' not in st.session_state:
        if available_periods:
            st.session_state['tb_df'] = TrialBalanceDB.get_trial_balance(empresa_seleccionada, periodo_actual)
            st.session_state['tb_df_comp'] = TrialBalanceDB.get_trial_balance(empresa_seleccionada, periodo_comp)
            from src.models.pl_cubo_db import PlCuboDB
            st.session_state['pl_df'] = PlCuboDB.get_pl_cubo(empresa_seleccionada, periodo_actual)
            st.session_state['pl_df_comp'] = PlCuboDB.get_pl_cubo(empresa_seleccionada, periodo_comp)
            st.session_state[db_cache_key] = True
        else:
            if 'tb_df' in st.session_state:
                st.session_state['tb_df_comp'] = st.session_state['tb_df'].copy()
    
    tab_bal, tab_pl, tab_flujo, tab_patrimonio, tab_integral = st.tabs([
        "1️⃣ Balance Clasificado", 
        "2️⃣ E.R. Clasificados", 
        "3️⃣ Flujo Efectivo", 
        "4️⃣ Patrimonio", 
        "5️⃣ R. Integrales"
    ])
    
    with tab_bal:
        st.write("Generación del Estado de Situación Financiera Clasificado utilizando la plantilla base y reglas de mapeo.")
        
        with st.expander("👀 Ver/Descargar Plantilla Base"):
            template_bal_path = os.path.join(empresa_path, "Balance clasificado.xlsx")
            if os.path.exists(template_bal_path):
                with open(template_bal_path, "rb") as _f:
                    st.download_button(
                        label="📥 Descargar Plantilla Balance",
                        data=_f,
                        file_name="Plantilla_Balance_Clasificado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_plantilla_bal"
                    )
                st.info("💡 Descarga esta plantilla, modifícala si necesitas agregar/quitar filas o cambiar el formato, "
                        "y vuelve a subirla en la pestaña ‘Carga de Datos’. La hoja oculta \_CONFIG\_ dentro del archivo "
                        "contiene la configuración de columnas que usa el programa para leerla correctamente.")
            else:
                st.warning("⚠️ No se encontró la plantilla 'Balance clasificado.xlsx' en la carpeta de la empresa.")
        
        if 'tb_df' not in st.session_state or st.session_state['tb_df'] is None or 'map_balance_df' not in st.session_state:
            miss = []
            if 'tb_df' not in st.session_state: miss.append('tb_df (no está en memoria)')
            elif st.session_state['tb_df'] is None: miss.append('tb_df (es None)')
            if 'map_balance_df' not in st.session_state: miss.append('map_balance_df (Mapeo)')
            st.warning(f"⚠️ Debes importar un Trial Balance y haber configurado tu Mapeo de Balance previamente. (Falta: {', '.join(miss)})")
        else:
            unidad = st.radio("Unidad de Medida", ["M$ (Miles de pesos)", "Ch$ (Pesos)"], horizontal=True, key="um_bal")
            scale_factor = 1000.0 if "M$" in unidad else 1.0

            if st.button("🚀 Procesar Plantilla Balance Clasificado", type="primary", key="btn_run_balance"):
                with st.spinner("Mapeando cuentas, inyectando saldos (Actual y Comparativo) y calculando P&L implícito..."):
                    import time
                    start_time = time.time()
                    try:
                        from src.reporting.balance_generator import BalanceGenerator
                        
                        template_path = os.path.join(empresa_path, "Balance clasificado.xlsx")
                        if not os.path.exists(template_path):
                            st.error(f"❌ No se encontró la plantilla base '{template_path}' en la carpeta de la empresa activa.")
                        else:
                            tb_df = st.session_state['tb_df'].copy()
                            tb_df_comp = st.session_state.get('tb_df_comp')
                            map_balance_df = st.session_state['map_balance_df'].copy()
                            
                            engine = BalanceGenerator(template_path)
                            excel_output = engine.generate(tb_df, map_balance_df, scale_factor=scale_factor, tb_df_comp=tb_df_comp, periodo_actual_str=periodo_actual, periodo_comp_str=periodo_comp)
                            st.session_state['balance_excel_binary'] = excel_output.getvalue()
                            
                            # Read back for on-screen preview
                            excel_output.seek(0)
                            
                            import openpyxl
                            wb_check = openpyxl.load_workbook(excel_output, data_only=True)
                            ws_check = wb_check.active
                            
                            from src.core.excel_utils import detect_balance_columns, read_template_config, read_excel_preview
                            
                            cfg = read_template_config(wb_check)
                            if cfg is None:
                                name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx = detect_balance_columns(ws_check, wb_check)
                                cfg = {
                                    "name_col": name_col_idx,
                                    "nota_col": nota_col_idx if nota_col_idx else 0,
                                    "val_actual_col": val25_col_idx,
                                    "val_comp_col": val24_col_idx,
                                    "data_start_row": 5,
                                }
                            
                            col_actual = str(periodo_actual)
                            col_comp = str(periodo_comp)
                            if col_actual == col_comp:
                                col_comp = f"{col_comp} (Comp)"
                            
                            excel_output.seek(0)
                            preview_df = read_excel_preview(excel_output, cfg, col_actual, col_comp)
                            
                            if "Nota" in preview_df.columns:
                                def clean_nota(val):
                                    if pd.isna(val) or str(val).strip() == "" or str(val).lower() == "nan" or val == 0:
                                        return ""
                                    try:
                                        float_val = float(val)
                                        if float_val.is_integer():
                                            return str(int(float_val))
                                        return str(float_val)
                                    except:
                                        return str(val).strip()
                                preview_df["Nota"] = preview_df["Nota"].apply(clean_nota)
                            
                            if col_actual in preview_df.columns:
                                preview_df[col_actual] = pd.to_numeric(preview_df[col_actual], errors='coerce')
                            if col_comp in preview_df.columns:
                                preview_df[col_comp] = pd.to_numeric(preview_df[col_comp], errors='coerce')
                            
                            # Generate Word Version
                            from src.reporting.word_export import WordExportEngine
                            word_output = WordExportEngine.generate_classified_balance_word(
                                df=preview_df, 
                                title="Estado de Situación Financiera Clasificado",
                                unit=unidad
                            )
                            
                            st.session_state['preview_df'] = preview_df
                            st.session_state['balance_word_binary'] = word_output.getvalue() if hasattr(word_output, 'getvalue') else word_output
                            st.session_state['balance_col_actual'] = col_actual
                            st.session_state['balance_col_comp'] = col_comp
                            
                            elapsed_time = time.time() - start_time
                            st.success(f"✅ Balance Clasificado generado en {unidad}. La ecuación A-P+R=0 ha sido inyectada (Tiempo de ejecución: {elapsed_time:.2f} segundos).")
                    except Exception as e:
                        st.error(f"Error generando el reporte: {e}")

            if 'preview_df' in st.session_state and st.session_state.get('balance_excel_binary') is not None:
                preview_df = st.session_state['preview_df']
                col_actual = st.session_state.get('balance_col_actual', str(periodo_actual))
                col_comp = st.session_state.get('balance_col_comp', str(periodo_comp))
                
                # Apply global corporate styles
                from src.reporting.formatting import apply_corporate_style
                styled_df = apply_corporate_style(preview_df, excel_bytes=st.session_state.get('balance_excel_binary'))
                
                st.markdown(styled_df.to_html(index=False), unsafe_allow_html=True)
                st.write("")
                 
                # --- VALIDACIÓN DE ECUACIÓN CONTABLE ---
                row_activos = preview_df[preview_df['Clasificación'].astype(str).str.lower().str.strip() == "total activos"]
                row_pat_pas = preview_df[preview_df['Clasificación'].astype(str).str.lower().str.strip() == "total patrimonio y pasivos"]
                 
                if not row_activos.empty and not row_pat_pas.empty:
                     st.write("")
                     st.subheader("🔍 Validación de Ecuación Contable: Activos - (Pasivos + Patrimonio)")
                     
                     cols_verificar = [col_actual, col_comp]
                     cols_metrics = st.columns(len(cols_verificar))
                     
                     for idx, col in enumerate(cols_verificar):
                         try:
                             val_act = float(row_activos[col].values[0])
                         except:
                             val_act = 0.0
                             
                         try:
                             val_pp = float(row_pat_pas[col].values[0])
                         except:
                             val_pp = 0.0
                             
                         diff = abs(val_act) - abs(val_pp)
                         
                         with cols_metrics[idx]:
                             if abs(diff) < 1.0:
                                 st.success(f"**{col}**\n\nCuadrado (0)")
                             else:
                                 diff_fmt = f"{diff:,.0f}".replace(",", ".")
                                 st.error(f"**{col}**\n\nDescuadre: {diff_fmt}")
                 
                st.write("")
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    st.download_button(
                        label="📥 Descargar Balance Clasificado (Excel)",
                        data=st.session_state['balance_excel_binary'],
                        file_name="Balance_Clasificado_Procesado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        key="btn_dl_bal_excel"
                    )
                with col_btn2:
                    st.download_button(
                        label="📝 Descargar Balance Clasificado (Word)",
                        data=st.session_state.get('balance_word_binary', b''),
                        file_name="Balance_Clasificado_Procesado.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                        key="btn_dl_bal_word"
                    )

    with tab_pl:
        st.write("Generación del Estado de Resultados Clasificados por función y naturaleza.")
        
        with st.expander("👀 Ver/Descargar Plantilla Base"):
            template_er_path = os.path.join(empresa_path, "Estado de Resultados Clasificados.xlsx")
            if os.path.exists(template_er_path):
                with open(template_er_path, "rb") as file:
                    st.download_button(
                        label="📥 Descargar Plantilla E.R.",
                        data=file,
                        file_name="Plantilla_ER_Clasificados.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_plantilla_er"
                    )
                # Mostrar una vista previa de la plantilla
                try:
                    import re as _re
                    df_preview = pd.read_excel(template_er_path)
                    df_preview_display = df_preview.copy()
                    # Formatear todas las columnas que parezcan de año (no solo '2024')
                    for _col in df_preview_display.columns:
                        if _re.search(r'20\d{2}', str(_col)):
                            df_preview_display[_col] = df_preview_display[_col].apply(
                                lambda x: f"{int(x):,}".replace(",", ".") if pd.notna(x) and str(x).strip() not in ['', 'nan'] else x
                            )
                    st.dataframe(df_preview_display)
                except Exception as e:
                    st.error(f"No se pudo cargar la vista previa: {e}")
            else:
                st.warning("⚠️ El archivo base 'Estado de Resultados Clasificados.xlsx' no se encuentra en el directorio raíz.")
                
        unidad_medida = st.radio("Unidad de Medida", ["M$ (Miles de pesos)", "Ch$ (Pesos)"], horizontal=True, key="um_er")
        scale_factor = 1000.0 if unidad_medida.startswith("M$") else 1.0

        if st.button("🚀 Ejecutar E.R. Clasificados", type="primary"):
            missing_periods = []
            if 'pl_df' not in st.session_state or st.session_state['pl_df'] is None or (isinstance(st.session_state['pl_df'], pd.DataFrame) and st.session_state['pl_df'].empty):
                missing_periods.append(str(periodo_actual))
            if 'pl_df_comp' not in st.session_state or st.session_state['pl_df_comp'] is None or (isinstance(st.session_state['pl_df_comp'], pd.DataFrame) and st.session_state['pl_df_comp'].empty):
                if periodo_comp and str(periodo_comp) not in missing_periods:
                    missing_periods.append(str(periodo_comp))
                    
            if missing_periods:
                st.warning(f"⚠️ No se ha encontrado el P&L para los siguientes periodos seleccionados: {', '.join(missing_periods)}. Dirígete al módulo '1️⃣ Cargas de datos', selecciona el mes/año respectivo y sube su P&L asociado.")
            else:
                with st.spinner("Mapeando columnas y calculando sumatorias y subtotales del Estado de Resultados..."):
                    try:
                        import time
                        start_time = time.time()
                        import importlib
                        import src.reporting.er_generator
                        importlib.reload(src.reporting.er_generator)
                        from src.reporting.er_generator import ERGenerator
                        from src.reporting.excel_export import generate_excel_report
                        import io
                        
                        template_er_path = os.path.join(empresa_path, "Estado de Resultados Clasificados.xlsx")
                        if not os.path.exists(template_er_path):
                            st.error(f"❌ No se encontró la plantilla base '{template_er_path}' en la carpeta de la empresa.)")
                        else:
                            pl_df = st.session_state['pl_df'].copy()
                            
                            engine = ERGenerator(template_er_path)
                            # Guardamos ambos en session state para permitir las descargas
                            pl_df_comp = st.session_state.get('pl_df_comp')
                            excel_output, preview_df = engine.generate(pl_df, scale_factor=scale_factor, pl_df_comp=pl_df_comp, periodo_actual_str=periodo_actual, periodo_comp_str=periodo_comp)
                            st.session_state['er_preview_df'] = preview_df
                            st.session_state['er_excel_binary'] = excel_output.getvalue()
                            
                            elapsed_time = time.time() - start_time
                            st.success(f"✅ Estado de Resultados generado con éxito (Tiempo de ejecución: {elapsed_time:.2f} segundos).")
                    except Exception as e:
                        st.error(f"Error generando el reporte: {e}")

        if 'er_preview_df' in st.session_state:
            display_df = st.session_state['er_preview_df'].copy()
            
            # Detectar dinámicamente qué columnas son de año (periodos numéricos)
            # para no corromper columnas de texto como "Notas"
            import re as _re
            year_col_names = [
                col for col in display_df.columns
                if _re.search(r'20\d{2}', str(col))
                or (str(col).strip().isdigit() and col != display_df.columns[0])
            ]
            for col in year_col_names:
                display_df[col] = pd.to_numeric(display_df[col], errors='coerce')
            

            import sys
            import importlib
            import src.reporting.formatting
            importlib.reload(sys.modules['src.reporting.formatting'])
            from src.reporting.formatting import apply_corporate_style
            styled_df = apply_corporate_style(display_df, excel_bytes=st.session_state.get('er_excel_binary'))
            st.markdown(styled_df.to_html(index=False), unsafe_allow_html=True)
            st.write("")
            
            # --- VALIDACIÓN DE RESULTADO FINAL VS. CUBO P&L ---
            st.subheader("🔍 Validación de Resultado Final vs. Cubo P&L")
            
            df_valid_rows = display_df[display_df.iloc[:, 0].notna() & (display_df.iloc[:, 0].astype(str).str.strip() != "")]
            if not df_valid_rows.empty:
                target_names = [
                    "ganancias (perdida) del ejercicio",
                    "ganancias (pérdida) del ejercicio",
                    "ganancia (perdida) del ejercicio",
                    "ganancia (pérdida) del ejercicio",
                    "resultado del ejercicio",
                    "(perdida) procedente de operaciones continuadas",
                    "(pérdida) procedente de operaciones continuadas",
                    "perdida procedente de operaciones continuadas",
                    "pérdida procedente de operaciones continuadas",
                    "perdida",
                    "pérdida"
                ]
                last_row = df_valid_rows.iloc[-1]
                for name in target_names:
                    matched_rows = df_valid_rows[df_valid_rows.iloc[:, 0].astype(str).str.replace(r'\xa0', ' ', regex=True).str.strip().str.lower() == name]
                    if not matched_rows.empty:
                        last_row = matched_rows.iloc[0]
                        break
                
                # Detectar columnas de año dinámicamente (no por índice posicional)
                import re as _re
                year_cols_validation = [
                    col for col in display_df.columns
                    if _re.search(r'20\d{2}', str(col))
                    or (str(col).strip().isdigit() and col != display_df.columns[0])
                ]
                
                if not year_cols_validation:
                    st.warning("⚠️ No se detectaron columnas de periodo en el reporte para validar.")
                else:
                    col_actual_name = year_cols_validation[0]
                    val_actual_report = float(last_row[col_actual_name]) if pd.notna(last_row[col_actual_name]) else 0.0
                    val_actual_report_pesos = val_actual_report * scale_factor
                    
                    from src.models.pl_cubo_db import PlCuboDB
                    total_actual_cube = -1 * PlCuboDB.get_pl_cubo_total_sum(empresa_seleccionada, periodo_actual)
                    
                    cols_verificar = [(col_actual_name, val_actual_report_pesos, total_actual_cube)]
                    
                    if len(year_cols_validation) >= 2:
                        col_comp_name = year_cols_validation[1]
                        val_comp_report = float(last_row[col_comp_name]) if pd.notna(last_row[col_comp_name]) else 0.0
                        val_comp_report_pesos = val_comp_report * scale_factor
                        total_comp_cube = -1 * PlCuboDB.get_pl_cubo_total_sum(empresa_seleccionada, periodo_comp)
                        cols_verificar.append((col_comp_name, val_comp_report_pesos, total_comp_cube))
                    
                    cols_metrics = st.columns(len(cols_verificar))
                    for idx, (col_name, val_rep_pesos, val_cube_pesos) in enumerate(cols_verificar):
                        diff = abs(val_rep_pesos - val_cube_pesos)
                        
                        with cols_metrics[idx]:
                            if diff <= 999.0:
                                diff_fmt = f"{diff:,.0f}".replace(",", ".")
                                st.success(f"**{col_name}**\n\nCuadrado (Diferencia: ${diff_fmt})")
                            else:
                                diff_fmt = f"{diff:,.0f}".replace(",", ".")
                                rep_fmt = f"{val_rep_pesos:,.0f}".replace(",", ".")
                                cube_fmt = f"{val_cube_pesos:,.0f}".replace(",", ".")
                                st.error(
                                    f"**{col_name}**\n\n"
                                    f"❌ Descuadre: ${diff_fmt}\n\n"
                                    f"* Reporte: ${rep_fmt}\n"
                                    f"* Cubo P&L: ${cube_fmt}"
                                )
            st.write("")
            
            # Opciones de descarga (Excel / Word)
            st.subheader("Opciones de Exportación")
            col_ex1, col_ex2 = st.columns(2)
            
            with col_ex1:
                from src.reporting.excel_export import generate_excel_report
                
                # Recrear el binario para descarga
                final_xlsx = generate_excel_report(st.session_state['er_preview_df'], title="Estado de Resultados Clasificados", subtitle=f"Expresado en {unidad_medida}")
                st.download_button(
                    label="📥 Descargar E.R. en Excel",
                    data=final_xlsx,
                    file_name="Estado_Resultados_Clasificados.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="btn_dl_er_excel"
                )
                
            with col_ex2:
                try:
                    from src.reporting.word_export import generate_word_report
                    final_docx = generate_word_report(st.session_state['er_preview_df'], title="Estado de Resultados Clasificados", subtitle=f"Expresado en {unidad_medida}")
                    st.download_button(
                        label="📄 Descargar E.R. en Word",
                        data=final_docx,
                        file_name="Estado_Resultados_Clasificados.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                        key="btn_dl_er_word"
                    )
                except ImportError:
                    st.info("Módulo docx no instalado. Instala python-docx para habilitar exportación a Word.")
            
    with tab_flujo:
        st.write("Generación del Estado de Flujos de Efectivo.")
        
        with st.expander("👀 Ver/Descargar Plantilla Base"):
            _cf_names = ["Estado de Flujos de Efectivo.xlsx", "Estado de Flujos de Efectivo Indirecto.xlsx"]
            _found_any = False
            for _cf_name in _cf_names:
                _cf_path = os.path.join(empresa_path, _cf_name)
                if os.path.exists(_cf_path):
                    _found_any = True
                    with open(_cf_path, "rb") as _f:
                        st.download_button(
                            label=f"📥 Descargar {_cf_name}",
                            data=_f,
                            file_name=_cf_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_plantilla_cf_{_cf_name}"
                        )
            if not _found_any:
                st.warning("⚠️ No se encontraron plantillas de Flujo de Efectivo en la carpeta de la empresa.")
            else:
                st.info("💡 La hoja oculta \_CONFIG\_ en cada plantilla contiene la configuración de columnas. "
                        "Descarga, modifica y vuelve a subir para actualizar.")
        
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            metodo_flujo = st.radio("Método del Flujo de Efectivo", ["Directo", "Indirecto"], horizontal=True, key="metodo_cf")
        with col_m2:
            unidad_medida_cf = st.radio("Unidad de Medida", ["M$ (Miles de pesos)", "Ch$ (Pesos)"], horizontal=True, key="um_cf")
        
        scale_factor_cf = 1000.0 if unidad_medida_cf.startswith("M$") else 1.0

        # --- SECCIÓN DE AJUSTES DE DEPURACIÓN ---
        from src.models.database import SessionLocal
        from src.models.cash_flow_db import CashFlowAdjustment
        
        db = SessionLocal()
        try:
            existing_adjs = db.query(CashFlowAdjustment).filter_by(
                empresa=empresa_seleccionada,
                periodo=periodo_actual,
                es_consolidado=False
            ).all()
        finally:
            db.close()

        # Construir diccionarios de ajustes para usar
        adj_dict = {}
        for a in existing_adjs:
            li = a.linea_item.strip()
            if li not in adj_dict:
                adj_dict[li] = {'ingreso': 0.0, 'egreso': 0.0}
            adj_dict[li]['ingreso'] += float(a.ingreso_caja or 0.0)
            adj_dict[li]['egreso'] += float(a.egreso_caja or 0.0)

        with st.expander("🔮 Ajustes de Depuración de Flujo (Ingreso/Egreso Caja)"):
            template_cf_path = os.path.join(empresa_path, "Estado de Flujos de Efectivo.xlsx")
            lines_list = []
            if os.path.exists(template_cf_path):
                try:
                    import openpyxl
                    wb_t = openpyxl.load_workbook(template_cf_path, data_only=True)
                    ws_t = wb_t.active
                    for r in range(1, ws_t.max_row+1):
                        val = ws_t.cell(row=r, column=1).value
                        if val and isinstance(val, str) and len(val.strip()) > 3:
                            val_clean = val.strip()
                            if not any(x in val_clean.lower() for x in ["separados", "ejercicios", "flujo de efectivo", "actividades de", "nota"]):
                                lines_list.append(val_clean)
                except Exception as e:
                    pass
            if not lines_list:
                lines_list = ["Cobros procedentes de las ventas de bienes y prestación de servicios", 
                              "Pagos a proveedores por el suministro de bienes y servicios", 
                              "Pagos a y por cuenta de los empleados",
                              "Intereses pagados", "Intereses recibidos", "Impuestos a las ganancias reembolsados (pagados)",
                              "Compra de Propiedades, planta y equipo", "Compra de intangibles",
                              "Importes procedentes de préstamos de largo plazo", "Pagos de préstamos",
                              "Pagos de pasivos por arrendamientos financieros", "Otras entradas y (salidas) de dinero"]
            
            st.write("Registra reclasificaciones o depuraciones no monetarias sobre las líneas del flujo de efectivo.")
            
            col_l, col_i, col_e = st.columns([2, 1, 1])
            sel_line = col_l.selectbox("Línea de Flujo Destino", sorted(list(set(lines_list))), key="cf_adj_line")
            adj_ing = col_i.number_input("Ingreso Caja (+)", min_value=0.0, value=0.0, step=1000.0, key="cf_adj_ing")
            adj_egr = col_e.number_input("Egreso Caja (-)", min_value=0.0, value=0.0, step=1000.0, key="cf_adj_egr")
            adj_glosa = st.text_input("Glosa Explicativa", key="cf_adj_glosa")
            
            if st.button("➕ Registrar Ajuste de Flujo"):
                if not adj_glosa.strip():
                    st.error("Por favor, ingresa una glosa explicativa para el ajuste.")
                elif adj_ing == 0.0 and adj_egr == 0.0:
                    st.error("Por favor, ingresa un monto mayor a cero en Ingreso Caja o Egreso Caja.")
                else:
                    db = SessionLocal()
                    try:
                        new_adj = CashFlowAdjustment(
                            empresa=empresa_seleccionada,
                            periodo=periodo_actual,
                            glosa=adj_glosa.strip(),
                            linea_item=sel_line,
                            ingreso_caja=adj_ing,
                            egreso_caja=adj_egr,
                            es_consolidado=False
                        )
                        db.add(new_adj)
                        db.commit()
                        st.success("✅ Ajuste registrado con éxito.")
                        st.rerun()
                    except Exception as ex_db:
                        st.error(f"Error al guardar en base de datos: {ex_db}")
                    finally:
                        db.close()
                        
            if existing_adjs:
                st.write("---")
                st.write("**Ajustes Registrados para este Periodo:**")
                df_adj_disp = pd.DataFrame([{
                    "ID": a.id,
                    "Línea de Flujo": a.linea_item,
                    "Ingreso Caja": f"{int(round(a.ingreso_caja)):,}".replace(",", "."),
                    "Egreso Caja": f"{int(round(a.egreso_caja)):,}".replace(",", "."),
                    "Glosa": a.glosa
                } for a in existing_adjs])
                st.dataframe(df_adj_disp, use_container_width=True)
                
                col_del_1, col_del_2 = st.columns([3, 1])
                with col_del_1:
                    del_id = st.selectbox("Selecciona Ajuste para Eliminar", [a.id for a in existing_adjs], format_func=lambda x: next(f"ID {a.id} - {a.linea_item} ({a.glosa})" for a in existing_adjs if a.id == x))
                with col_del_2:
                    st.write("")
                    if st.button("🗑️ Eliminar Ajuste", use_container_width=True):
                        db = SessionLocal()
                        try:
                            db.query(CashFlowAdjustment).filter_by(id=del_id).delete()
                            db.commit()
                            st.success("✅ Ajuste eliminado.")
                            st.rerun()
                        except Exception as ex_db:
                            st.error(f"Error al eliminar: {ex_db}")
                        finally:
                            db.close()
            else:
                st.info("No hay ajustes de depuración registrados para este periodo.")

        if st.button("🚀 Ejecutar Flujo de Efectivo", type="primary"):
            if 'tb_df' not in st.session_state or 'map_balance_df' not in st.session_state:
                st.warning("⚠️ Debes importar un Trial Balance y haber configurado tu Mapeo de Balance antes de correr este reporte.")
            else:
                with st.spinner("Procesando balances detailed cuenta por cuenta y calculando variaciones..."):
                    try:
                        import time
                        start_time = time.time()
                        
                        import importlib
                        import src.reporting.cash_flow_generator
                        importlib.reload(src.reporting.cash_flow_generator)
                        from src.reporting.cash_flow_generator import CashFlowGenerator
                        cf_filename = "Estado de Flujos de Efectivo.xlsx" if metodo_flujo == "Directo" else "Estado de Flujos de Efectivo Indirecto.xlsx"
                        template_cf_path = os.path.join(empresa_path, cf_filename)
                        
                        if not os.path.exists(template_cf_path):
                            import shutil
                            src_template = os.path.join("templates", cf_filename)
                            if os.path.exists(src_template):
                                shutil.copy2(src_template, template_cf_path)
                                
                        if not os.path.exists(template_cf_path):
                            st.error(f"❌ Falta la plantilla '{template_cf_path}'.")
                        else:
                            # Cargar mapeos de balance y PL, con fallbacks desde el disco si no están en memoria
                            map_bal_copy = st.session_state.get('map_balance_df')
                            if map_bal_copy is None or (isinstance(map_bal_copy, pd.DataFrame) and map_bal_copy.empty):
                                map_bal_path = os.path.join(empresa_path, "map_balance.xlsx")
                                if os.path.exists(map_bal_path):
                                    map_bal_copy = pd.read_excel(map_bal_path)
                                    st.session_state['map_balance_df'] = map_bal_copy
                                else:
                                    map_bal_copy = pd.DataFrame()
                            else:
                                map_bal_copy = map_bal_copy.copy()

                            map_pl_copy = st.session_state.get('map_pl_df')
                            if map_pl_copy is None or (isinstance(map_pl_copy, pd.DataFrame) and map_pl_copy.empty):
                                map_pl_path = os.path.join(empresa_path, "map_pl.xlsx")
                                if os.path.exists(map_pl_path):
                                    map_pl_copy = pd.read_excel(map_pl_path)
                                    st.session_state['map_pl_df'] = map_pl_copy
                                else:
                                    map_pl_copy = None
                            else:
                                map_pl_copy = map_pl_copy.copy()
                            
                            cf_engine = CashFlowGenerator(template_cf_path)
                            ex_cf, matriz_audit = cf_engine.generate(
                                empresa=empresa_seleccionada,
                                periodo_actual_str=str(periodo_actual),
                                periodo_comp_str=str(periodo_comp),
                                map_balance_df=map_bal_copy,
                                map_pl_df=map_pl_copy,
                                method=metodo_flujo,
                                is_consolidado=False,
                                scale_factor=scale_factor_cf
                            )
                            
                            st.session_state['flujo_excel_binary'] = ex_cf.getvalue()
                            st.session_state['flujo_audit_data'] = matriz_audit
                            
                            # Mapear dinámicamente las columnas para Flujo de Efectivo
                            ex_cf.seek(0)
                            import openpyxl
                            import datetime
                            import re
                            wb_check = openpyxl.load_workbook(ex_cf, data_only=True)
                            ws_check = wb_check.active
                            
                            name_col_idx = 1
                            for col in range(1, 10):
                                for row in range(1, 15):
                                    val = ws_check.cell(row=row, column=col).value
                                    if val and str(val).strip().lower() in ["concepto", "descripcion", "detalle", "flujos", "origen/aplicacion"]:
                                        name_col_idx = col
                                        break
                                        
                            date_cols = []
                            for col in range(1, ws_check.max_column + 1):
                                if col == name_col_idx:
                                    continue
                                for row in range(1, 10):
                                    val = ws_check.cell(row=row, column=col).value
                                    if val is not None:
                                        is_date = (
                                            isinstance(val, (datetime.datetime, datetime.date)) or
                                            (isinstance(val, str) and re.search(r'20\d{2}', val))
                                        )
                                        if is_date:
                                            date_cols.append(col)
                                            break
                            date_cols = sorted(list(set(date_cols)))
                            
                            val25_col_idx = 3
                            val24_col_idx = 4
                            if len(date_cols) >= 2:
                                val25_col_idx = date_cols[0]
                                val24_col_idx = date_cols[1]
                                
                            nota_col_idx = None
                            for col in range(name_col_idx + 1, val24_col_idx):
                                if col != val25_col_idx:
                                    nota_col_idx = col
                                    break
                            if nota_col_idx is None:
                                nota_col_idx = val25_col_idx + 1 if val25_col_idx + 1 < val24_col_idx else name_col_idx + 1
                                
                            ex_cf.seek(0)
                            df_raw = pd.read_excel(ex_cf)
                            
                            col_actual = str(periodo_actual)
                            col_comp = str(periodo_comp)
                            if col_actual == col_comp:
                                col_comp = f"{col_comp} (Comp)"
                                
                            cols_to_keep = [name_col_idx - 1, nota_col_idx - 1, val25_col_idx - 1, val24_col_idx - 1]
                            preview_cf = df_raw.iloc[:, cols_to_keep].copy()
                            preview_cf.columns = ["Descripción", "Nota", col_actual, col_comp]
                            preview_cf = preview_cf.dropna(how='all', subset=["Descripción"])
                            
                            st.session_state['flujo_preview_df'] = preview_cf
                            elapsed_time = time.time() - start_time
                            st.success(f"✅ Flujo de Efectivo generado con éxito (Tiempo de ejecución: {elapsed_time:.2f} segundos).")
                    except Exception as e:
                        st.error(f"Error generando Flujo de Efectivo: {e}")
                        
        if 'flujo_preview_df' in st.session_state:
            tab_view1, tab_view2, tab_view3 = st.tabs(["📊 Vista de Reporte (Plantilla)", "📋 Hoja de Trabajo (Detalle Variaciones)", "🔍 Desglose por Cuenta Contable (Auditoría)"])
            
            with tab_view1:
                display_cf = st.session_state['flujo_preview_df'].copy()
                for col in display_cf.columns[1:]:
                    display_cf[col] = pd.to_numeric(display_cf[col], errors='coerce')
                
                import sys
                import importlib
                import src.reporting.formatting
                importlib.reload(sys.modules['src.reporting.formatting'])
                from src.reporting.formatting import apply_corporate_style
                styled_cf = apply_corporate_style(display_cf, excel_bytes=st.session_state.get('flujo_excel_binary'))
                st.markdown(styled_cf.to_html(index=False), unsafe_allow_html=True)
                
            with tab_view2:
                df_mat = pd.DataFrame(st.session_state['flujo_audit_data'])
                if not df_mat.empty:
                    df_gp = df_mat.groupby("Línea de Flujo Mapeada").agg({
                        "Variación Bruta": "sum",
                        "Ingreso Caja (Ajuste)": "first",
                        "Egreso Caja (Ajuste)": "first"
                    }).reset_index()
                    df_gp["Variación Depurada"] = df_gp["Variación Bruta"] + df_gp["Ingreso Caja (Ajuste)"] - df_gp["Egreso Caja (Ajuste)"]
                    df_gp.columns = ["Línea de Flujo Destino", "Variación Bruta", "Ingresos Caja (Depuración)", "Egresos Caja (Depuración)", "Variación Depurada"]
                    st.dataframe(df_gp, use_container_width=True)
                else:
                    st.info("No hay datos de auditoría de flujo disponibles.")
                    
            with tab_view3:
                df_mat = pd.DataFrame(st.session_state['flujo_audit_data'])
                if not df_mat.empty:
                    st.dataframe(df_mat, use_container_width=True)
                else:
                    st.info("No hay desglose detallado de cuentas contables disponible.")
            
            st.write("")
            col_cf1, col_cf2, col_cf3 = st.columns(3)
            with col_cf1:
                st.download_button(
                    label="📥 Descargar Flujo en Excel",
                    data=st.session_state['flujo_excel_binary'],
                    file_name="Estado_Flujo_Efectivo.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="btn_dl_cf_ex"
                )
            with col_cf2:
                if 'flujo_audit_data' in st.session_state and st.session_state['flujo_audit_data']:
                    df_matriz = pd.DataFrame(st.session_state['flujo_audit_data'])
                    df_matriz = df_matriz.sort_values(by="Línea de Flujo Mapeada")
                    matriz_bin = df_to_excel_bytes(df_matriz, 'Matriz_Auditoria_EFE')
                    st.download_button(
                        label="📥 Matriz Validación Flujo",
                        data=matriz_bin,
                        file_name="Matriz_Validacion_Flujo.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="secondary",
                        key="btn_dl_cf_matrix"
                    )
            with col_cf3:
                try:
                    from src.reporting.word_export import generate_word_report
                    final_docx_cf = generate_word_report(st.session_state['flujo_preview_df'], title="Estado de Flujos de Efectivo", subtitle=f"Expresado en {unidad_medida_cf}")
                    st.download_button(
                        label="📄 Descargar Flujo en Word",
                        data=final_docx_cf,
                        file_name="Estado_Flujo_Efectivo.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                        key="btn_dl_cf_word"
                    )
                except ImportError:
                    st.info("Módulo docx no instalado.")
            
    with tab_patrimonio:
        st.write("Generación del Estado de Cambios en el Patrimonio Neto.")
        
        with st.expander("👀 Ver/Descargar Plantilla Base"):
            _pat_path = os.path.join(empresa_path, "Estado de Cambios en el Patrimonio.xlsx")
            if os.path.exists(_pat_path):
                with open(_pat_path, "rb") as _f:
                    st.download_button(
                        label="📥 Descargar Plantilla Patrimonio",
                        data=_f,
                        file_name="Plantilla_Estado_Patrimonio.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_plantilla_pat"
                    )
                st.info("💡 La hoja oculta \_CONFIG\_ contiene la configuración de columnas. Descarga, modifica y sube de nuevo.")
            else:
                st.warning("⚠️ No se encontró la plantilla 'Estado de Cambios en el Patrimonio.xlsx'.")
        
        if st.button("🚀 Ejecutar Cambios en el Patrimonio", type="primary"):
            # Verificamos que se haya corrido el Balance y P&L primero para robarle los Bottom Lines matemáticos
            if 'preview_df' not in st.session_state or 'er_preview_df' not in st.session_state:
                st.warning("⚠️ Debes pulsar el botón de 'Procesar Balance' y 'Procesar E.R.' en esta misma pantalla antes de procesar el Patrimonio para que el sistema tenga cargado el resultado del ejercicio final.")
            else:
                with st.spinner("Realizando cuadratura de capital, reservas e inyectando resultados del ER..."):
                    try:
                        import sys
                        import time
                        start_time = time.time()
                        import importlib
                        import src.reporting.patrimonio_generator
                        importlib.reload(sys.modules['src.reporting.patrimonio_generator'])
                        from src.reporting.patrimonio_generator import PatrimonioGenerator
                        
                        import src.reporting.formatting
                        importlib.reload(sys.modules['src.reporting.formatting'])
                        template_pat_path = os.path.join(empresa_path, "Estado de Cambios en el Patrimonio.xlsx")
                        if not os.path.exists(template_pat_path):
                            st.error(f"❌ No se encontró la plantilla '{template_pat_path}'.")
                        else:
                            bal_df = st.session_state['preview_df'].copy()
                            pl_df = st.session_state['er_preview_df'].copy()
                            
                            pat_engine = PatrimonioGenerator(template_pat_path)
                            ex_pat = pat_engine.generate(bal_df, pl_df, periodo_actual_str=str(periodo_actual), periodo_comp_str=str(periodo_comp), empresa=empresa_seleccionada)
                            st.session_state['pat_excel_binary'] = ex_pat
                            
                            ex_pat.seek(0)
                            from src.core.excel_utils import detect_patrimonio_skiprows
                            pat_skip = detect_patrimonio_skiprows(ex_pat)
                            ex_pat.seek(0)
                            # Skip header space to grab the table rows
                            prev = pd.read_excel(ex_pat, skiprows=pat_skip)
                            from src.core.excel_utils import clean_preview_dataframe
                            prev = clean_preview_dataframe(prev)
                            first_col = prev.columns[0]
                            prev = prev.dropna(how='all', subset=[first_col])
                            st.session_state['pat_preview_df'] = prev
                            
                            elapsed_time = time.time() - start_time
                            st.success(f"✅ Estado de Patrimonio generado matemáticamente contra el P&L (Tiempo de ejecución: {elapsed_time:.2f} segundos).")
                    except Exception as e:
                        st.error(f"Error generando Patrimonio: {e}")
                        
        if 'pat_preview_df' in st.session_state:
            disp = st.session_state['pat_preview_df'].copy()
            for col in disp.columns[1:]:
                disp[col] = pd.to_numeric(disp[col], errors='coerce')
            

            import sys
            import importlib
            import src.reporting.formatting
            importlib.reload(sys.modules['src.reporting.formatting'])
            from src.reporting.formatting import apply_corporate_style
            styled_c = apply_corporate_style(disp, excel_bytes=st.session_state.get('pat_excel_binary'))
            st.markdown(styled_c.to_html(index=False), unsafe_allow_html=True)
            
            st.write("")
            col_pt1, col_pt2 = st.columns(2)
            with col_pt1:
                st.download_button(
                    label="📥 Descargar Patrimonio en Excel",
                    data=st.session_state['pat_excel_binary'],
                    file_name="Estado_Patrimonio_IFRS.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="btn_dl_pat_ex"
                )
            with col_pt2:
                try:
                    from src.reporting.word_export import generate_word_report
                    final_docx_pat = generate_word_report(st.session_state['pat_preview_df'], title="Estado de Cambios en el Patrimonio")
                    st.download_button(
                        label="📄 Descargar Patrimonio en Word",
                        data=final_docx_pat,
                        file_name="Estado_Patrimonio_IFRS.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                        key="btn_dl_pat_wd"
                    )
                except:
                    pass
            
    with tab_integral:
        st.write("Generación del Estado de Resultados Integrales (ORI).")
        
        with st.expander("👀 Ver/Descargar Plantilla Base"):
            _ori_path = os.path.join(empresa_path, "Estado de Resultados Integrales.xlsx")
            if os.path.exists(_ori_path):
                with open(_ori_path, "rb") as _f:
                    st.download_button(
                        label="📥 Descargar Plantilla ORI",
                        data=_f,
                        file_name="Plantilla_Estado_ORI.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_plantilla_ori"
                    )
                st.info("💡 La hoja oculta \_CONFIG\_ contiene la configuración de columnas. Descarga, modifica y sube de nuevo.")
            else:
                st.warning("⚠️ No se encontró la plantilla 'Estado de Resultados Integrales.xlsx'.")
        
        if st.button("🚀 Ejecutar Resultados Integrales", type="primary"):
            if 'er_preview_df' not in st.session_state:
                st.warning("⚠️ Debes pulsar el botón de 'Procesar E.R. Clasificados' primero para tener cargados sobre este entorno los resultados de las operaciones continuadas.")
            else:
                with st.spinner("Desplegando componentes de otro resultado integral..."):
                    try:
                        import sys
                        import time
                        start_time = time.time()
                        import src.reporting.ori_generator
                        importlib.reload(sys.modules['src.reporting.ori_generator'])
                        from src.reporting.ori_generator import OriGenerator
                        template_ori_path = os.path.join(empresa_path, "Estado de Resultados Integrales.xlsx")
                        if not os.path.exists(template_ori_path):
                            st.error(f"❌ No se encontró la plantilla '{template_ori_path}'.")
                        else:
                            pl_df = st.session_state['er_preview_df'].copy()
                            
                            ori_engine = OriGenerator(template_ori_path)
                            ex_ori = ori_engine.generate(
                                pl_df, 
                                periodo_actual_str=str(periodo_actual), 
                                periodo_comp_str=str(periodo_comp),
                                bal_preview_df=st.session_state.get('preview_df'),
                                empresa=empresa_seleccionada
                            )
                            st.session_state['ori_excel_binary'] = ex_ori
                            
                            # Mapear dinámicamente las columnas para Resultados Integrales
                            ex_ori.seek(0)
                            import openpyxl
                            import datetime
                            import re
                            wb_check = openpyxl.load_workbook(ex_ori, data_only=True)
                            ws_check = wb_check.active
                            
                            name_col_idx = 1
                            for col in range(1, 10):
                                for row in range(1, 15):
                                    val = ws_check.cell(row=row, column=col).value
                                    if val and str(val).strip().lower() in ["detalle", "concepto", "clasificacion", "descripcion"]:
                                        name_col_idx = col
                                        break
                                        
                            date_cols = []
                            for col in range(1, ws_check.max_column + 1):
                                if col == name_col_idx:
                                    continue
                                for row in range(1, 10):
                                    val = ws_check.cell(row=row, column=col).value
                                    if val is not None:
                                        is_date = (
                                            isinstance(val, (datetime.datetime, datetime.date)) or
                                            (isinstance(val, str) and re.search(r'20\d{2}', val))
                                        )
                                        if is_date:
                                            date_cols.append(col)
                                            break
                            date_cols = sorted(list(set(date_cols)))
                            
                            val25_col_idx = 2
                            val24_col_idx = 3
                            if len(date_cols) >= 2:
                                val25_col_idx = date_cols[0]
                                val24_col_idx = date_cols[1]
                                
                            ex_ori.seek(0)
                            from src.core.excel_utils import detect_general_skiprows
                            ori_skip = detect_general_skiprows(ex_ori)
                            ex_ori.seek(0)
                            df_raw = pd.read_excel(ex_ori, skiprows=ori_skip)
                            
                            col_actual = str(periodo_actual)
                            col_comp = str(periodo_comp)
                            if col_actual == col_comp:
                                col_comp = f"{col_comp} (Comp)"
                                
                            cols_to_keep = [name_col_idx - 1, val25_col_idx - 1, val24_col_idx - 1]
                            prev = df_raw.iloc[:, cols_to_keep].copy()
                            prev.columns = ["Detalle", col_actual, col_comp]
                            prev = prev.dropna(how='all', subset=["Detalle"])
                            st.session_state['ori_preview_df'] = prev
                            
                            elapsed_time = time.time() - start_time
                            st.success(f"✅ Estado de Resultados Integrales cargado con éxito (Tiempo de ejecución: {elapsed_time:.2f} segundos).")
                    except Exception as e:
                        st.error(f"Error generando ORI: {e}")
                        
        if 'ori_preview_df' in st.session_state:
            disp = st.session_state['ori_preview_df'].copy()
            for col in disp.columns[1:]:
                disp[col] = pd.to_numeric(disp[col], errors='coerce')
            

            import sys
            import importlib
            import src.reporting.formatting
            importlib.reload(sys.modules['src.reporting.formatting'])
            from src.reporting.formatting import apply_corporate_style
            styled_c = apply_corporate_style(disp, excel_bytes=st.session_state.get('ori_excel_binary'))
            st.markdown(styled_c.to_html(index=False), unsafe_allow_html=True)
            
            st.write("")
            col_pt1, col_pt2 = st.columns(2)
            with col_pt1:
                st.download_button(
                    label="📥 Descargar O.R.I. en Excel",
                    data=st.session_state['ori_excel_binary'],
                    file_name="Estado_Resultados_Integrales.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key="btn_dl_ori_ex"
                )
            with col_pt2:
                try:
                    from src.reporting.word_export import generate_word_report
                    final_docx_ori = generate_word_report(st.session_state['ori_preview_df'], title="Estado de Resultados Integrales")
                    st.download_button(
                        label="📄 Descargar O.R.I. en Word",
                        data=final_docx_ori,
                        file_name="Estado_Resultados_Integrales.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        type="primary",
                        key="btn_dl_ori_wd"
                    )
                except:
                    pass


