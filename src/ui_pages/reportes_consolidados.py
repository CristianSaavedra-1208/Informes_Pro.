import streamlit as st
import pandas as pd
import os
import datetime
import re
import openpyxl
from src.core.excel_utils import df_to_excel_bytes, format_periodo

def render(empresa_seleccionada, empresa_path):
    st.title("📄 Estados Financieros Consolidados")
    st.write("Emite reportes formales (PDF/Excel) inyectando la hoja de trabajo en tus plantillas base.")
    
    from src.models.database import SessionLocal
    from src.models.consolidacion import ConsolidationGroup
    from src.models.historical_data import HistoricalDataRecord
    from src.models.taxonomy_master import TaxonomyMasterRecord
    
    empresas_dir = os.path.join("data", "empresas")
    
    db = SessionLocal()
    grupos_disp3 = db.query(ConsolidationGroup).all()
    db.close()
    
    if not grupos_disp3:
        st.warning("Debes configurar un grupo de consolidación primero en el menú Operaciones -> Consolidación.")
    else:
        grupo_dict3 = {g.id: g.nombre_grupo for g in grupos_disp3}
        sel_g3 = st.selectbox("Seleccionar Grupo para Emisión", options=list(grupo_dict3.keys()), format_func=lambda x: grupo_dict3[x], key="sel_g3")
        
        grupo_name = grupo_dict3[sel_g3]
        grupo_folder = os.path.join(empresas_dir, f"[GRUPO] {grupo_name}")
        os.makedirs(grupo_folder, exist_ok=True)
        
        # (El estado de las plantillas se muestra al final de la página para no obstruir la operación)
        
        db = SessionLocal()
        per_recs3 = db.query(HistoricalDataRecord.periodo).distinct().all()
        db.close()
        periodos_hist3 = sorted([r[0] for r in per_recs3], reverse=True)
        if not periodos_hist3: periodos_hist3 = ["2026-12", "2025-12"]
        
        col_p_act, col_p_comp = st.columns(2)
        periodo_act_g = col_p_act.selectbox("Periodo Actual", periodos_hist3, key="per_act_g", format_func=format_periodo)
        periodo_comp_g = col_p_comp.selectbox("Periodo Comparativo (Opcional)", ["Ninguno"] + periodos_hist3, key="per_comp_g", format_func=format_periodo)
        
        report_type = st.radio("Tipo de Reporte a Generar", [
            "Balance Clasificado", 
            "Estado de Resultados", 
            "Estado de Flujos de Efectivo",
            "Estado de Cambios en el Patrimonio",
            "Estado de Resultados Integrales"
        ])
        
        metodo_flujo = "Directo"
        if report_type == "Estado de Flujos de Efectivo":
            metodo_flujo = st.radio("Método del Flujo de Efectivo", ["Directo", "Indirecto"], horizontal=True, key="metodo_cf_g")
            
        unidad = st.radio("Unidad de Medida", ["M$ (Miles de pesos)", "Ch$ (Pesos)"], horizontal=True)
        scale_factor = 1000.0 if "M$" in unidad else 1.0

        # --- SECCIÓN DE AJUSTES DE DEPURACIÓN DEL FLUJO CONSOLIDADO ---
        from src.models.database import SessionLocal
        from src.models.cash_flow_db import CashFlowAdjustment
        
        db = SessionLocal()
        try:
            existing_adjs = db.query(CashFlowAdjustment).filter_by(
                empresa=f"[GRUPO] {grupo_name}",
                periodo=periodo_act_g,
                es_consolidado=True
            ).all()
        finally:
            db.close()

        with st.expander("🔮 Ajustes de Depuración del Flujo Consolidado (Ingreso/Egreso Caja)"):
            template_cf_path = os.path.join(grupo_folder, "Estado de Flujos de Efectivo.xlsx")
            if not os.path.exists(template_cf_path) and os.path.exists("templates/Estado de Flujos de Efectivo.xlsx"):
                import shutil
                shutil.copy2("templates/Estado de Flujos de Efectivo.xlsx", template_cf_path)
                
            lines_list = []
            if os.path.exists(template_cf_path):
                try:
                    import openpyxl
                    wb_t = openpyxl.load_workbook(template_cf_path, data_only=True)
                    ws_t = wb_t.active
                    for r in range(1, ws_t.max_row+1):
                        val = ws_t.cell(row=r, column=1).value
                        if val and isinstance(val, str) and len(val.strip()) > 3:
                            val_clean = val.strip()
                            if not any(x in val_clean.lower() for x in ["separados", "ejercicios", "flujo de efectivo", "actividades de", "nota"]):
                                lines_list.append(val_clean)
                except:
                    pass
            if not lines_list:
                lines_list = ["Cobros procedentes de las ventas de bienes y prestación de servicios", 
                              "Pagos a proveedores por el suministro de bienes y servicios", 
                              "Pagos a y por cuenta de los empleados",
                              "Intereses pagados", "Intereses recibidos", "Impuestos a las ganancias reembolsados (pagados)",
                              "Compra de Propiedades, planta y equipo", "Compra de intangibles",
                              "Importes procedentes de préstamos de largo plazo", "Pagos de préstamos",
                              "Pagos de pasivos por arrendamientos financieros", "Otras entradas y (salidas) de dinero"]
            
            st.write("Registra reclasificaciones o depuraciones no monetarias en el EFE Consolidado.")
            
            col_l, col_i, col_e = st.columns([2, 1, 1])
            sel_line = col_l.selectbox("Línea de Flujo Destino", sorted(list(set(lines_list))), key="cf_adj_line_g")
            adj_ing = col_i.number_input("Ingreso Caja (+)", min_value=0.0, value=0.0, step=1000.0, key="cf_adj_ing_g")
            adj_egr = col_e.number_input("Egreso Caja (-)", min_value=0.0, value=0.0, step=1000.0, key="cf_adj_egr_g")
            adj_glosa = st.text_input("Glosa Explicativa", key="cf_adj_glosa_g")
            
            if st.button("➕ Registrar Ajuste de Flujo Consolidado"):
                if not adj_glosa.strip():
                    st.error("Por favor, ingresa una glosa explicativa para el ajuste.")
                elif adj_ing == 0.0 and adj_egr == 0.0:
                    st.error("Por favor, ingresa un monto mayor a cero en Ingreso Caja o Egreso Caja.")
                else:
                    db = SessionLocal()
                    try:
                        new_adj = CashFlowAdjustment(
                            empresa=f"[GRUPO] {grupo_name}",
                            periodo=periodo_act_g,
                            glosa=adj_glosa.strip(),
                            linea_item=sel_line,
                            ingreso_caja=adj_ing,
                            egreso_caja=adj_egr,
                            es_consolidado=True
                        )
                        db.add(new_adj)
                        db.commit()
                        st.success("✅ Ajuste registrado.")
                        st.rerun()
                    except Exception as ex_db:
                        st.error(f"Error al guardar: {ex_db}")
                    finally:
                        db.close()
                        
            if existing_adjs:
                st.write("---")
                st.write("**Ajustes Consolidados Registrados:**")
                df_adj_disp = pd.DataFrame([{
                    "ID": a.id,
                    "Línea de Flujo": a.linea_item,
                    "Ingreso Caja": f"{int(round(a.ingreso_caja)):,}".replace(",", "."),
                    "Egreso Caja": f"{int(round(a.egreso_caja)):,}".replace(",", "."),
                    "Glosa": a.glosa
                } for a in existing_adjs])
                st.dataframe(df_adj_disp, use_container_width=True)
                
                col_del_1, col_del_2 = st.columns([3, 1])
                with col_del_1:
                    del_id = st.selectbox("Selecciona Ajuste para Eliminar", [a.id for a in existing_adjs], format_func=lambda x: next(f"ID {a.id} - {a.linea_item} ({a.glosa})" for a in existing_adjs if a.id == x))
                with col_del_2:
                    st.write("")
                    if st.button("🗑️ Eliminar Ajuste Cons.", use_container_width=True):
                        db = SessionLocal()
                        try:
                            db.query(CashFlowAdjustment).filter_by(id=del_id).delete()
                            db.commit()
                            st.success("✅ Ajuste eliminado.")
                            st.rerun()
                        except Exception as ex_db:
                            st.error(f"Error al eliminar: {ex_db}")
                        finally:
                            db.close()
            else:
                st.info("No hay ajustes de depuración registrados para este periodo.")
        
        if st.button("🚀 Generar Reporte Consolidado", type="primary"):
            if report_type == "Balance Clasificado" and not os.path.exists(os.path.join(grupo_folder, "Balance clasificado.xlsx")):
                st.error("Sube la plantilla de Balance primero.")
            elif report_type == "Estado de Resultados" and not os.path.exists(os.path.join(grupo_folder, "Estado de Resultados Clasificados.xlsx")):
                st.error("Sube la plantilla de E.R. primero.")
            elif report_type == "Estado de Flujos de Efectivo" and not os.path.exists(os.path.join(grupo_folder, "Estado de Flujos de Efectivo.xlsx")) and not os.path.exists("templates/Estado de Flujos de Efectivo.xlsx"):
                st.error("Sube la plantilla de Flujo de Efectivo primero.")
            elif report_type == "Estado de Cambios en el Patrimonio" and not os.path.exists(os.path.join(grupo_folder, "Estado de Cambios en el Patrimonio.xlsx")):
                st.error("Sube la plantilla de Estado de Cambios en el Patrimonio primero.")
            elif report_type == "Estado de Resultados Integrales" and not os.path.exists(os.path.join(grupo_folder, "Estado de Resultados Integrales.xlsx")):
                st.error("Sube la plantilla de Estado de Resultados Integrales primero.")
            else:
                import time
                start_time = time.time()
                with st.spinner("Procesando hojas de trabajo y cruzando con plantillas..."):
                    import sys
                    import importlib
                    import src.core.consolidacion_engine
                    importlib.reload(src.core.consolidacion_engine)
                    from src.core.consolidacion_engine import generar_hoja_trabajo
                    df_hoja_act, msg_act = generar_hoja_trabajo(sel_g3, periodo_act_g)
                    df_hoja_comp = None
                    if periodo_comp_g != "Ninguno":
                        df_hoja_comp, msg_comp = generar_hoja_trabajo(sel_g3, periodo_comp_g)
                        
                    if df_hoja_act is not None:
                        def clean_str(s):
                            if pd.isna(s): return ""
                            return str(s).strip().lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
                            
                        # Slice and filter depending on the report type
                        if report_type == "Balance Clasificado":
                            bs_subtotals = {
                                "activos corrientes", "activos corrientes totales",
                                "activos no corrientes", "activos no corrientes totales",
                                "total activos", "patrimonio y pasivos",
                                "pasivos corrientes", "pasivo corrientes totales",
                                "pasivos no corrientes", "pasivo no corrientes totales",
                                "total pasivos", "patrimonio", "patrimonio total",
                                "total patrimonio y pasivos", "estado de resultados"
                            }
                            bs_subtotals_normalized = {clean_str(name) for name in bs_subtotals}
                            
                            # Slice up to 'Estado de Resultados' row
                            idx_er = df_hoja_act[df_hoja_act['Balance clasificado'] == "Estado de Resultados"].index
                            if not idx_er.empty:
                                df_hoja_act_sec = df_hoja_act.loc[:idx_er[0]-1]
                            else:
                                df_hoja_act_sec = df_hoja_act
                                
                            df_hoja_act_clean = df_hoja_act_sec[df_hoja_act_sec['Balance clasificado'].notna()]
                            df_hoja_act_clean = df_hoja_act_clean[df_hoja_act_clean['Balance clasificado'].str.strip() != ""]
                            df_hoja_act_clean = df_hoja_act_clean[~df_hoja_act_clean['Balance clasificado'].apply(clean_str).isin(bs_subtotals_normalized)]
                            
                            if df_hoja_comp is not None:
                                idx_er_comp = df_hoja_comp[df_hoja_comp['Balance clasificado'] == "Estado de Resultados"].index
                                if not idx_er_comp.empty:
                                    df_hoja_comp_sec = df_hoja_comp.loc[:idx_er_comp[0]-1]
                                else:
                                    df_hoja_comp_sec = df_hoja_comp
                                    
                                df_hoja_comp_clean = df_hoja_comp_sec[df_hoja_comp_sec['Balance clasificado'].notna()]
                                df_hoja_comp_clean = df_hoja_comp_clean[df_hoja_comp_clean['Balance clasificado'].str.strip() != ""]
                                df_hoja_comp_clean = df_hoja_comp_clean[~df_hoja_comp_clean['Balance clasificado'].apply(clean_str).isin(bs_subtotals_normalized)]
                            else:
                                df_hoja_comp_clean = None
                                
                        else:
                            # Estado de Resultados / other
                            pl_subtotals = {
                                "ganancia bruta", "resultado antes de impuestos",
                                "ganancias (perdida) del ejercicio", "ganancia (perdida) del ejercicio",
                                "estado de resultados", "otros rubros no clasificados"
                            }
                            pl_subtotals_normalized = {clean_str(name) for name in pl_subtotals}
                            
                            # Slice from 'Estado de Resultados' row onwards
                            idx_er = df_hoja_act[df_hoja_act['Balance clasificado'] == "Estado de Resultados"].index
                            if not idx_er.empty:
                                df_hoja_act_sec = df_hoja_act.loc[idx_er[0]+1:]
                            else:
                                df_hoja_act_sec = df_hoja_act
                                
                            df_hoja_act_clean = df_hoja_act_sec[df_hoja_act_sec['Balance clasificado'].notna()]
                            df_hoja_act_clean = df_hoja_act_clean[df_hoja_act_clean['Balance clasificado'].str.strip() != ""]
                            df_hoja_act_clean = df_hoja_act_clean[~df_hoja_act_clean['Balance clasificado'].apply(clean_str).isin(pl_subtotals_normalized)]
                            
                            if df_hoja_comp is not None:
                                idx_er_comp = df_hoja_comp[df_hoja_comp['Balance clasificado'] == "Estado de Resultados"].index
                                if not idx_er_comp.empty:
                                    df_hoja_comp_sec = df_hoja_comp.loc[idx_er_comp[0]+1:]
                                else:
                                    df_hoja_comp_sec = df_hoja_comp
                                    
                                df_hoja_comp_clean = df_hoja_comp_sec[df_hoja_comp_sec['Balance clasificado'].notna()]
                                df_hoja_comp_clean = df_hoja_comp_clean[df_hoja_comp_clean['Balance clasificado'].str.strip() != ""]
                                df_hoja_comp_clean = df_hoja_comp_clean[~df_hoja_comp_clean['Balance clasificado'].apply(clean_str).isin(pl_subtotals_normalized)]
                            else:
                                df_hoja_comp_clean = None

                        if report_type == "Balance Clasificado":
                            tb_df = pd.DataFrame({
                                'cuenta_id': df_hoja_act_clean['Balance clasificado'],
                                'saldo_final': df_hoja_act_clean['CONSOLIDADO']
                            })
                            tb_df_comp = None
                            if df_hoja_comp_clean is not None:
                                tb_df_comp = pd.DataFrame({
                                    'cuenta_id': df_hoja_comp_clean['Balance clasificado'],
                                    'saldo_final': df_hoja_comp_clean['CONSOLIDADO']
                                })
                                
                            # Exclude P&L accounts from tb_df and tb_df_comp to avoid double-counting in consolidated report
                            from src.models.database import SessionLocal
                            from src.models.historical_data import HistoricalDataRecord
                            from src.models.taxonomy_master import TaxonomyMasterRecord
                            from src.models.consolidacion import ConsolidationGroup
                            
                            db = SessionLocal()
                            try:
                                grupo = db.query(ConsolidationGroup).filter_by(id=sel_g3).first()
                                entidades_grupo = [grupo.empresa_matriz]
                                if grupo.filial_is_group:
                                    def get_subgroup_companies(sub_g_id):
                                        comps = []
                                        sub_g = db.query(ConsolidationGroup).filter_by(id=sub_g_id).first()
                                        if sub_g:
                                            comps.append(sub_g.empresa_matriz)
                                            if sub_g.filial_is_group:
                                                comps.extend(get_subgroup_companies(int(sub_g.empresa_filial)))
                                            else:
                                                comps.append(sub_g.empresa_filial)
                                        return comps
                                    entidades_grupo.extend(get_subgroup_companies(int(grupo.empresa_filial)))
                                else:
                                    entidades_grupo.append(grupo.empresa_filial)
                                    
                                pl_lines_db = db.query(HistoricalDataRecord.linea_item).filter(
                                    HistoricalDataRecord.reporte == 'P&L',
                                    HistoricalDataRecord.empresa.in_(entidades_grupo)
                                ).distinct().all()
                                
                                pl_tax_db = db.query(TaxonomyMasterRecord.nombre_linea_es).filter(
                                    TaxonomyMasterRecord.reporte_destino == 'P&L',
                                    TaxonomyMasterRecord.empresa.in_(entidades_grupo)
                                ).distinct().all()
                                
                                pl_lines_set = {r[0] for r in pl_lines_db}.union({r[0] for r in pl_tax_db})
                            finally:
                                db.close()
                                
                            def is_pl_line(li):
                                if li in pl_lines_set:
                                    return True
                                norm = str(li).lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
                                if any(x in norm for x in ["ingreso", "gasto", "costo", "egreso", "depreciacion", "amortizacion", "diferencia de cambio", "unidad de reajuste", "ganancia", "perdida"]):
                                    if not any(x in norm for x in ["diferido", "diferidos", "anticipado", "anticipados", "por pagar", "por cobrar", "acumulado", "acumulados", "acumulada", "acumuladas", "capital", "reserva", "ejercicio", "periodo"]):
                                        return True
                                return False
                                
                            tb_df = tb_df[~tb_df['cuenta_id'].apply(is_pl_line)]
                            if tb_df_comp is not None:
                                tb_df_comp = tb_df_comp[~tb_df_comp['cuenta_id'].apply(is_pl_line)]
                                
                            wb = openpyxl.load_workbook(os.path.join(grupo_folder, "Balance clasificado.xlsx"))
                            ws = wb.active
                            template_accounts = [str(ws.cell(row=r, column=2).value).strip() for r in range(1, ws.max_row+1) if ws.cell(row=r, column=2).value]
                            
                            dummy_map = pd.DataFrame({
                                'N° de Cuenta': tb_df['cuenta_id'],
                                'Clasificación balance': tb_df['cuenta_id']
                            })
                            dummy_map.loc[~dummy_map['N° de Cuenta'].isin(template_accounts), 'Clasificación balance'] = pd.NA
                            
                            import importlib
                            import src.reporting.balance_generator
                            importlib.reload(src.reporting.balance_generator)
                            from src.reporting.balance_generator import BalanceGenerator
                            gen = BalanceGenerator(os.path.join(grupo_folder, "Balance clasificado.xlsx"))
                            result_bytes = gen.generate(
                                tb_df=tb_df, 
                                map_balance_df=dummy_map, 
                                scale_factor=scale_factor,
                                tb_df_comp=tb_df_comp,
                                periodo_actual_str=periodo_act_g,
                                periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None
                            )
                            filename = f"Balance_Consolidado_{grupo_name}_{periodo_act_g}.xlsx"
                            
                        elif report_type == "Estado de Resultados":
                            pl_dict_act = {row['Balance clasificado']: [row['CONSOLIDADO']] for _, row in df_hoja_act_clean.iterrows()}
                            pl_df_wide = pd.DataFrame(pl_dict_act)
                            
                            pl_df_comp_wide = None
                            if df_hoja_comp_clean is not None:
                                pl_dict_comp = {row['Balance clasificado']: [row['CONSOLIDADO']] for _, row in df_hoja_comp_clean.iterrows()}
                                pl_df_comp_wide = pd.DataFrame(pl_dict_comp)
                            
                            import importlib
                            import src.reporting.er_generator
                            importlib.reload(src.reporting.er_generator)
                            from src.reporting.er_generator import ERGenerator
                            gen = ERGenerator(os.path.join(grupo_folder, "Estado de Resultados Clasificados.xlsx"))
                            result_bytes, _ = gen.generate(
                                pl_df=pl_df_wide, 
                                scale_factor=scale_factor,
                                pl_df_comp=pl_df_comp_wide,
                                periodo_actual_str=periodo_act_g,
                                periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None
                            )
                            filename = f"ER_Consolidado_{grupo_name}_{periodo_act_g}.xlsx"
                            
                        elif report_type == "Estado de Flujos de Efectivo":
                            cf_filename = "Estado de Flujos de Efectivo.xlsx" if metodo_flujo == "Directo" else "Estado de Flujos de Efectivo Indirecto.xlsx"
                            template_cf_path = os.path.join(grupo_folder, cf_filename)
                            if not os.path.exists(template_cf_path):
                                import shutil
                                src_template = os.path.join("templates", cf_filename)
                                if os.path.exists(src_template):
                                    shutil.copy2(src_template, template_cf_path)
                                    
                            db = SessionLocal()
                            try:
                                grupo = db.query(ConsolidationGroup).filter_by(id=sel_g3).first()
                                matriz_path = os.path.join(empresas_dir, grupo.empresa_matriz)
                            finally:
                                db.close()
                                
                            map_bal_path = os.path.join(matriz_path, "map_balance.xlsx")
                            map_pl_path = os.path.join(matriz_path, "map_pl.xlsx")
                            
                            map_balance_df = pd.read_excel(map_bal_path) if os.path.exists(map_bal_path) else None
                            map_pl_df = pd.read_excel(map_pl_path) if os.path.exists(map_pl_path) else None
                            
                            import importlib
                            import src.reporting.cash_flow_generator
                            importlib.reload(src.reporting.cash_flow_generator)
                            from src.reporting.cash_flow_generator import CashFlowGenerator
                            gen = CashFlowGenerator(template_cf_path)
                            result_bytes, matriz_audit = gen.generate(
                                empresa=f"[GRUPO] {grupo_name}",
                                periodo_actual_str=periodo_act_g,
                                periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None,
                                map_balance_df=map_balance_df,
                                map_pl_df=map_pl_df,
                                method=metodo_flujo,
                                is_consolidado=True,
                                consolidated_hoja_trabajo_df=df_hoja_act,
                                consolidated_hoja_trabajo_comp_df=df_hoja_comp,
                                scale_factor=scale_factor
                            )
                            filename = f"Flujo_Consolidado_{grupo_name}_{periodo_act_g}.xlsx"
                            st.session_state['flujo_audit_data_g'] = matriz_audit
                            
                        else: # Estado de Cambios en el Patrimonio o Estado de Resultados Integrales
                            # Generar Balance y P&L en memoria para pasarlos como argumentos
                            # 1. Balance en memoria
                            bs_subtotals = {
                                "activos corrientes", "activos corrientes totales",
                                "activos no corrientes", "activos no corrientes totales",
                                "total activos", "patrimonio y pasivos",
                                "pasivos corrientes", "pasivo corrientes totales",
                                "pasivos no corrientes", "pasivo no corrientes totales",
                                "total pasivos", "patrimonio", "patrimonio total",
                                "total patrimonio y pasivos", "estado de resultados"
                            }
                            bs_subtotals_normalized = {clean_str(name) for name in bs_subtotals}
                            
                            idx_er_b = df_hoja_act[df_hoja_act['Balance clasificado'] == "Estado de Resultados"].index
                            df_hoja_act_sec_b = df_hoja_act.loc[:idx_er_b[0]-1] if not idx_er_b.empty else df_hoja_act
                            df_hoja_act_clean_b = df_hoja_act_sec_b[df_hoja_act_sec_b['Balance clasificado'].notna() & (df_hoja_act_sec_b['Balance clasificado'].str.strip() != "")]
                            df_hoja_act_clean_b = df_hoja_act_clean_b[~df_hoja_act_clean_b['Balance clasificado'].apply(clean_str).isin(bs_subtotals_normalized)]
                            
                            tb_df_bal = pd.DataFrame({
                                'cuenta_id': df_hoja_act_clean_b['Balance clasificado'],
                                'saldo_final': df_hoja_act_clean_b['CONSOLIDADO']
                            })
                            tb_df_comp_bal = None
                            if df_hoja_comp is not None:
                                idx_er_comp_b = df_hoja_comp[df_hoja_comp['Balance clasificado'] == "Estado de Resultados"].index
                                df_hoja_comp_sec_b = df_hoja_comp.loc[:idx_er_comp_b[0]-1] if not idx_er_comp_b.empty else df_hoja_comp
                                df_hoja_comp_clean_b = df_hoja_comp_sec_b[df_hoja_comp_sec_b['Balance clasificado'].notna() & (df_hoja_comp_sec_b['Balance clasificado'].str.strip() != "")]
                                df_hoja_comp_clean_b = df_hoja_comp_clean_b[~df_hoja_comp_clean_b['Balance clasificado'].apply(clean_str).isin(bs_subtotals_normalized)]
                                tb_df_comp_bal = pd.DataFrame({
                                    'cuenta_id': df_hoja_comp_clean_b['Balance clasificado'],
                                    'saldo_final': df_hoja_comp_clean_b['CONSOLIDADO']
                                })
                                
                            db = SessionLocal()
                            try:
                                grupo = db.query(ConsolidationGroup).filter_by(id=sel_g3).first()
                                entidades_grupo = [grupo.empresa_matriz]
                                if grupo.filial_is_group:
                                    def get_subgroup_companies(sub_g_id):
                                        comps = []
                                        sub_g = db.query(ConsolidationGroup).filter_by(id=sub_g_id).first()
                                        if sub_g:
                                            comps.append(sub_g.empresa_matriz)
                                            if sub_g.filial_is_group:
                                                comps.extend(get_subgroup_companies(int(sub_g.empresa_filial)))
                                            else:
                                                comps.append(sub_g.empresa_filial)
                                        return comps
                                    entidades_grupo.extend(get_subgroup_companies(int(grupo.empresa_filial)))
                                else:
                                    entidades_grupo.append(grupo.empresa_filial)
                                    
                                pl_lines_db = db.query(HistoricalDataRecord.linea_item).filter(
                                    HistoricalDataRecord.reporte == 'P&L',
                                    HistoricalDataRecord.empresa.in_(entidades_grupo)
                                ).distinct().all()
                                
                                pl_tax_db = db.query(TaxonomyMasterRecord.nombre_linea_es).filter(
                                    TaxonomyMasterRecord.reporte_destino == 'P&L',
                                    TaxonomyMasterRecord.empresa.in_(entidades_grupo)
                                ).distinct().all()
                                
                                pl_lines_set = {r[0] for r in pl_lines_db}.union({r[0] for r in pl_tax_db})
                            finally:
                                db.close()
                                
                            def is_pl_line_local(li):
                                if li in pl_lines_set:
                                    return True
                                norm = str(li).lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
                                if any(x in norm for x in ["ingreso", "gasto", "costo", "egreso", "depreciacion", "amortizacion", "diferencia de cambio", "unidad de reajuste", "ganancia", "perdida"]):
                                    if not any(x in norm for x in ["diferido", "diferidos", "anticipado", "anticipados", "por pagar", "por cobrar", "acumulado", "acumulados", "acumulada", "acumuladas", "capital", "reserva", "ejercicio", "periodo"]):
                                        return True
                                return False
                                
                            tb_df_bal = tb_df_bal[~tb_df_bal['cuenta_id'].apply(is_pl_line_local)]
                            if tb_df_comp_bal is not None:
                                tb_df_comp_bal = tb_df_comp_bal[~tb_df_comp_bal['cuenta_id'].apply(is_pl_line_local)]
                                
                            wb_bal = openpyxl.load_workbook(os.path.join(grupo_folder, "Balance clasificado.xlsx"))
                            ws_bal = wb_bal.active
                            template_accounts = [str(ws_bal.cell(row=r, column=2).value).strip() for r in range(1, ws_bal.max_row+1) if ws_bal.cell(row=r, column=2).value]
                            
                            dummy_map = pd.DataFrame({
                                'N° de Cuenta': tb_df_bal['cuenta_id'],
                                'Clasificación balance': tb_df_bal['cuenta_id']
                            })
                            dummy_map.loc[~dummy_map['N° de Cuenta'].isin(template_accounts), 'Clasificación balance'] = pd.NA
                            
                            from src.reporting.balance_generator import BalanceGenerator
                            gen_bal = BalanceGenerator(os.path.join(grupo_folder, "Balance clasificado.xlsx"))
                            result_bytes_bal = gen_bal.generate(
                                tb_df=tb_df_bal, 
                                map_balance_df=dummy_map, 
                                scale_factor=scale_factor,
                                tb_df_comp=tb_df_comp_bal,
                                periodo_actual_str=periodo_act_g,
                                periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None
                            )
                            
                            col_actual_tmp = str(periodo_act_g)
                            col_comp_tmp = str(periodo_comp_g) if periodo_comp_g != "Ninguno" else "Comp"
                            if col_actual_tmp == col_comp_tmp:
                                col_comp_tmp = f"{col_comp_tmp} (Comp)"
                            
                            wb_check_bal = openpyxl.load_workbook(result_bytes_bal, data_only=True)
                            ws_check_bal = wb_check_bal.active
                            from src.core.excel_utils import detect_balance_columns, read_template_config, read_excel_preview
                            cfg_bal = read_template_config(wb_check_bal)
                            if cfg_bal is None:
                                name_col_idx_b, nota_col_idx_b, val25_col_idx_b, val24_col_idx_b = detect_balance_columns(ws_check_bal, wb_check_bal)
                                cfg_bal = {
                                    "name_col": name_col_idx_b,
                                    "nota_col": nota_col_idx_b if nota_col_idx_b else 0,
                                    "val_actual_col": val25_col_idx_b,
                                    "val_comp_col": val24_col_idx_b,
                                    "data_start_row": 5,
                                }
                            result_bytes_bal.seek(0)
                            bal_df = read_excel_preview(result_bytes_bal, cfg_bal, col_actual_tmp, col_comp_tmp)
                            if "Nota" in bal_df.columns:
                                bal_df = bal_df.drop(columns=["Nota"])
                            # Renombrar columna Clasificación para compatibilidad con Patrimonio/ORI
                            if "Clasificación" in bal_df.columns:
                                bal_df = bal_df.rename(columns={"Clasificación": "Clasificación"})
                            
                            # 2. P&L en memoria
                            pl_subtotals = {
                                "ganancia bruta", "resultado antes de impuestos",
                                "ganancias (perdida) del ejercicio", "ganancia (perdida) del ejercicio",
                                "estado de resultados", "otros rubros no clasificados"
                            }
                            pl_subtotals_normalized = {clean_str(name) for name in pl_subtotals}
                            
                            idx_er_p = df_hoja_act[df_hoja_act['Balance clasificado'] == "Estado de Resultados"].index
                            df_hoja_act_sec_pl = df_hoja_act.loc[idx_er_p[0]+1:] if not idx_er_p.empty else df_hoja_act
                            df_hoja_act_clean_pl = df_hoja_act_sec_pl[df_hoja_act_sec_pl['Balance clasificado'].notna() & (df_hoja_act_sec_pl['Balance clasificado'].str.strip() != "")]
                            df_hoja_act_clean_pl = df_hoja_act_clean_pl[~df_hoja_act_clean_pl['Balance clasificado'].apply(clean_str).isin(pl_subtotals_normalized)]
                            
                            pl_dict_act = {row['Balance clasificado']: [row['CONSOLIDADO']] for _, row in df_hoja_act_clean_pl.iterrows()}
                            pl_df_wide = pd.DataFrame(pl_dict_act)
                            
                            pl_df_comp_wide = None
                            if df_hoja_comp is not None:
                                idx_er_comp_p = df_hoja_comp[df_hoja_comp['Balance clasificado'] == "Estado de Resultados"].index
                                df_hoja_comp_sec_pl = df_hoja_comp.loc[idx_er_comp_p[0]+1:] if not idx_er_comp_p.empty else df_hoja_comp
                                df_hoja_comp_clean_pl = df_hoja_comp_sec_pl[df_hoja_comp_sec_pl['Balance clasificado'].notna() & (df_hoja_comp_sec_pl['Balance clasificado'].str.strip() != "")]
                                df_hoja_comp_clean_pl = df_hoja_comp_clean_pl[~df_hoja_comp_clean_pl['Balance clasificado'].apply(clean_str).isin(pl_subtotals_normalized)]
                                pl_dict_comp = {row['Balance clasificado']: [row['CONSOLIDADO']] for _, row in df_hoja_comp_clean_pl.iterrows()}
                                pl_df_comp_wide = pd.DataFrame(pl_dict_comp)
                                
                            from src.reporting.er_generator import ERGenerator
                            gen_er = ERGenerator(os.path.join(grupo_folder, "Estado de Resultados Clasificados.xlsx"))
                            result_bytes_er, pl_df = gen_er.generate(
                                pl_df=pl_df_wide, 
                                scale_factor=scale_factor,
                                pl_df_comp=pl_df_comp_wide,
                                periodo_actual_str=periodo_act_g,
                                periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None
                            )
                            # pl_df ya viene listo del ERGenerator como preview_df
                            
                            # 3. Generar reporte
                            if report_type == "Estado de Cambios en el Patrimonio":
                                import sys
                                import importlib
                                import src.reporting.patrimonio_generator
                                importlib.reload(sys.modules['src.reporting.patrimonio_generator'])
                                from src.reporting.patrimonio_generator import PatrimonioGenerator
                                import src.reporting.formatting
                                importlib.reload(sys.modules['src.reporting.formatting'])
                                gen = PatrimonioGenerator(os.path.join(grupo_folder, "Estado de Cambios en el Patrimonio.xlsx"))
                                result_bytes = gen.generate(
                                    bal_preview_df=bal_df, 
                                    pl_preview_df=pl_df, 
                                    periodo_actual_str=periodo_act_g, 
                                    periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None, 
                                    empresa=f"[GRUPO] {grupo_name}"
                                )
                                filename = f"Patrimonio_Consolidado_{grupo_name}_{periodo_act_g}.xlsx"
                            else:
                                import sys
                                import importlib
                                import src.reporting.ori_generator
                                importlib.reload(sys.modules['src.reporting.ori_generator'])
                                from src.reporting.ori_generator import OriGenerator
                                gen = OriGenerator(os.path.join(grupo_folder, "Estado de Resultados Integrales.xlsx"))
                                result_bytes = gen.generate(
                                    pl_preview_df=pl_df, 
                                    periodo_actual_str=periodo_act_g, 
                                    periodo_comp_str=periodo_comp_g if periodo_comp_g != "Ninguno" else None, 
                                    bal_preview_df=bal_df,
                                    empresa=f"[GRUPO] {grupo_name}"
                                )
                                filename = f"ORI_Consolidado_{grupo_name}_{periodo_act_g}.xlsx"
                            
                        # Mostrar Preview
                        result_bytes.seek(0)
                        preview_df = pd.read_excel(result_bytes)
                        
                        col_actual = str(periodo_act_g)
                        col_comp = str(periodo_comp_g) if periodo_comp_g != "Ninguno" else "Comp"
                        if col_actual == col_comp: col_comp = f"{col_comp} (Comp)"
                        
                        if report_type == "Balance Clasificado":
                            # Cargar con openpyxl para obtener _CONFIG_ o detectar columnas
                            result_bytes.seek(0)
                            wb_check = openpyxl.load_workbook(result_bytes, data_only=True)
                            ws_check = wb_check.active
                            
                            from src.core.excel_utils import detect_balance_columns, read_template_config, read_excel_preview
                            cfg_prev = read_template_config(wb_check)
                            if cfg_prev is None:
                                name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx = detect_balance_columns(ws_check, wb_check)
                                cfg_prev = {
                                    "name_col": name_col_idx,
                                    "nota_col": nota_col_idx if nota_col_idx else 0,
                                    "val_actual_col": val25_col_idx,
                                    "val_comp_col": val24_col_idx,
                                    "data_start_row": 5,
                                }
                            
                            result_bytes.seek(0)
                            preview_df = read_excel_preview(result_bytes, cfg_prev, col_actual, col_comp)
                            # Renombrar primera columna para consistencia
                            if "Clasificación" in preview_df.columns:
                                preview_df = preview_df.rename(columns={"Clasificación": "Balance clasificado"})
                        elif report_type == "Estado de Resultados":
                            # Detectar dinámicamente columnas de año (no por índice posicional)
                            # para soportar plantillas con columna Notas antes de los años
                            _year_cols_er = [
                                c for c in preview_df.columns
                                if re.search(r'20\d{2}', str(c))
                                or (str(c).strip().isdigit() and c != preview_df.columns[0])
                            ]
                            if _year_cols_er:
                                _er_cols = [preview_df.columns[0]] + _year_cols_er[:2]
                            else:
                                _er_cols = list(preview_df.columns[:3])
                            preview_df = preview_df[_er_cols].copy()
                            preview_df.columns = (["Clasificación"] + [col_actual, col_comp])[:len(_er_cols)]
                        elif report_type == "Estado de Flujos de Efectivo":
                            # Cargar con openpyxl para encontrar los índices de columnas dinámicamente para Flujo
                            wb_check = openpyxl.load_workbook(result_bytes, data_only=True)
                            ws_check = wb_check.active
                            
                            name_col_idx = 1
                            for col in range(1, 10):
                                for row in range(1, 15):
                                    val = ws_check.cell(row=row, column=col).value
                                    if val and str(val).strip().lower() in ["concepto", "descripcion", "detalle", "flujos", "origen/aplicacion"]:
                                        name_col_idx = col
                                        break
                                        
                            date_cols = []
                            for col in range(1, ws_check.max_column + 1):
                                if col == name_col_idx:
                                    continue
                                for row in range(1, 10):
                                    val = ws_check.cell(row=row, column=col).value
                                    if val is not None:
                                        is_date = (
                                            isinstance(val, (datetime.datetime, datetime.date)) or
                                            (isinstance(val, str) and re.search(r'20\d{2}', val))
                                        )
                                        if is_date:
                                            date_cols.append(col)
                                            break
                            date_cols = sorted(list(set(date_cols)))
                            
                            val25_col_idx = 3
                            val24_col_idx = 4
                            if len(date_cols) >= 2:
                                val25_col_idx = date_cols[0]
                                val24_col_idx = date_cols[1]
                                
                            nota_col_idx = None
                            for col in range(name_col_idx + 1, val24_col_idx):
                                if col != val25_col_idx:
                                    nota_col_idx = col
                                    break
                            if nota_col_idx is None:
                                nota_col_idx = val25_col_idx + 1 if val25_col_idx + 1 < val24_col_idx else name_col_idx + 1
                                
                            result_bytes.seek(0)
                            df_raw = pd.read_excel(result_bytes)
                            cols_to_keep = [name_col_idx - 1, val25_col_idx - 1, val24_col_idx - 1]
                            preview_df = df_raw.iloc[:, cols_to_keep].copy()
                            preview_df.columns = ["Descripción", col_actual, col_comp]
                        else: # Estado de Cambios en el Patrimonio o Estado de Resultados Integrales
                            result_bytes.seek(0)
                            if report_type == "Estado de Cambios en el Patrimonio":
                                from src.core.excel_utils import detect_patrimonio_skiprows
                                pat_skip = detect_patrimonio_skiprows(result_bytes)
                                result_bytes.seek(0)
                                preview_df = pd.read_excel(result_bytes, skiprows=pat_skip)
                            else:
                                from src.core.excel_utils import detect_general_skiprows
                                ori_skip = detect_general_skiprows(result_bytes)
                                result_bytes.seek(0)
                                preview_df = pd.read_excel(result_bytes, skiprows=ori_skip)
                                
                            from src.core.excel_utils import clean_preview_dataframe
                            preview_df = clean_preview_dataframe(preview_df)
                            first_col = preview_df.columns[0]
                            preview_df = preview_df.dropna(how='all', subset=[first_col]).reset_index(drop=True)
                            
                        # Generar versión Word
                        try:
                            if report_type == "Balance Clasificado":
                                from src.reporting.word_export import WordExportEngine
                                word_output = WordExportEngine.generate_classified_balance_word(
                                    df=preview_df,
                                    title=f"Estado de Situación Financiera Consolidado - {grupo_name}",
                                    unit=unidad
                                )
                                word_bytes_out = word_output.getvalue() if hasattr(word_output, 'getvalue') else word_output
                                word_filename = f"Balance_Consolidado_{grupo_name}_{periodo_act_g}.docx"
                            elif report_type == "Estado de Resultados":
                                from src.reporting.word_export import generate_word_report
                                word_bytes_out = generate_word_report(
                                    df=preview_df,
                                    title=f"Estado de Resultados Consolidado - {grupo_name}",
                                    subtitle=f"Periodo: {periodo_act_g} vs {periodo_comp_g} ({unidad})"
                                )
                                word_filename = f"ER_Consolidado_{grupo_name}_{periodo_act_g}.docx"
                            elif report_type == "Estado de Flujos de Efectivo":
                                from src.reporting.word_export import generate_word_report
                                word_bytes_out = generate_word_report(
                                    df=preview_df,
                                    title=f"Estado de Flujos de Efectivo Consolidado - {grupo_name}",
                                    subtitle=f"Periodo: {periodo_act_g} vs {periodo_comp_g} ({unidad})"
                                )
                                word_filename = f"Flujo_Consolidado_{grupo_name}_{periodo_act_g}.docx"
                            elif report_type == "Estado de Cambios en el Patrimonio":
                                from src.reporting.word_export import generate_word_report
                                word_bytes_out = generate_word_report(
                                    df=preview_df,
                                    title=f"Estado de Cambios en el Patrimonio Consolidado - {grupo_name}",
                                    subtitle=f"Periodo: {periodo_act_g} vs {periodo_comp_g} ({unidad})"
                                )
                                word_filename = f"Patrimonio_Consolidado_{grupo_name}_{periodo_act_g}.docx"
                            else: # Estado de Resultados Integrales
                                from src.reporting.word_export import generate_word_report
                                word_bytes_out = generate_word_report(
                                    df=preview_df,
                                    title=f"Estado de Resultados Integrales Consolidado - {grupo_name}",
                                    subtitle=f"Periodo: {periodo_act_g} vs {periodo_comp_g} ({unidad})"
                                )
                                word_filename = f"ORI_Consolidado_{grupo_name}_{periodo_act_g}.docx"
                        except Exception as ex_w:
                            word_bytes_out = None
                            word_filename = f"Reporte_Consolidado_{grupo_name}_{periodo_act_g}.docx"

                        st.session_state['consolidated_report_excel_binary'] = result_bytes.getvalue()
                        st.session_state['consolidated_report_word_binary'] = word_bytes_out
                        st.session_state['consolidated_report_preview_df'] = preview_df
                        st.session_state['consolidated_report_filename'] = filename
                        st.session_state['consolidated_report_word_filename'] = word_filename
                        st.session_state['consolidated_report_type'] = report_type
                        st.session_state['consolidated_report_col_actual'] = col_actual
                        st.session_state['consolidated_report_col_comp'] = col_comp
                        st.session_state['consolidated_report_df_hoja_act'] = df_hoja_act
                        st.session_state['consolidated_report_df_hoja_comp'] = df_hoja_comp
                        
                        elapsed_time = time.time() - start_time
                        st.success(f"✅ Reporte Generado (Tiempo de ejecución: {elapsed_time:.2f} segundos)")
                    else:
                        st.error(f"Error generando consolidación: {msg_act}")

        if 'consolidated_report_preview_df' in st.session_state and st.session_state.get('consolidated_report_excel_binary') is not None:
            preview_df = st.session_state['consolidated_report_preview_df']
            report_type = st.session_state.get('consolidated_report_type', report_type)
            col_actual = st.session_state.get('consolidated_report_col_actual', str(periodo_act_g))
            col_comp = st.session_state.get('consolidated_report_col_comp', str(periodo_comp_g))
            df_hoja_act = st.session_state.get('consolidated_report_df_hoja_act')
            df_hoja_comp = st.session_state.get('consolidated_report_df_hoja_comp')
            filename = st.session_state.get('consolidated_report_filename', 'Reporte_Consolidado.xlsx')
            word_filename = st.session_state.get('consolidated_report_word_filename', 'Reporte_Consolidado.docx')

            if report_type == "Estado de Flujos de Efectivo" and 'flujo_audit_data_g' in st.session_state:
                tab_view1, tab_view2 = st.tabs(["📊 Vista de Reporte (Plantilla)", "📋 Hoja de Trabajo (Detalle Variaciones)"])
                with tab_view1:
                    try:
                        import sys
                        import importlib
                        import src.reporting.formatting
                        importlib.reload(sys.modules['src.reporting.formatting'])
                        from src.reporting.formatting import apply_corporate_style
                        st.markdown(apply_corporate_style(preview_df, excel_bytes=st.session_state.get('consolidated_report_excel_binary')).to_html(index=False), unsafe_allow_html=True)
                    except Exception as e:
                        st.dataframe(preview_df, height=500)
                with tab_view2:
                    df_mat = pd.DataFrame(st.session_state['flujo_audit_data_g'])
                    if not df_mat.empty:
                        df_gp = df_mat.groupby("Línea de Flujo Mapeada").agg({
                            "Variación Bruta": "sum",
                            "Ingreso Caja (Ajuste)": "first",
                            "Egreso Caja (Ajuste)": "first"
                        }).reset_index()
                        df_gp["Variación Depurada"] = df_gp["Variación Bruta"] + df_gp["Ingreso Caja (Ajuste)"] - df_gp["Egreso Caja (Ajuste)"]
                        df_gp.columns = ["Línea de Flujo Destino", "Variación Bruta", "Ingresos Caja (Depuración)", "Egresos Caja (Depuración)", "Variación Depurada"]
                        st.dataframe(df_gp, use_container_width=True)
            else:
                try:
                    import sys
                    import importlib
                    import src.reporting.formatting
                    importlib.reload(sys.modules['src.reporting.formatting'])
                    from src.reporting.formatting import apply_corporate_style
                    st.markdown(apply_corporate_style(preview_df, excel_bytes=st.session_state.get('consolidated_report_excel_binary')).to_html(index=False), unsafe_allow_html=True)
                except Exception as e:
                    st.dataframe(preview_df, height=500)
                
            # --- VALIDACIÓN DE ECUACIÓN CONTABLE / RESULTADOS ---
            if report_type == "Balance Clasificado":
                row_activos = preview_df[preview_df['Balance clasificado'].astype(str).str.lower().str.strip() == "total activos"]
                row_pat_pas = preview_df[preview_df['Balance clasificado'].astype(str).str.lower().str.strip() == "total patrimonio y pasivos"]
                
                if not row_activos.empty and not row_pat_pas.empty:
                    st.write("")
                    st.subheader("🔍 Validación de Ecuación Contable: Activos - (Pasivos + Patrimonio)")
                    
                    cols_verificar = [col_actual, col_comp]
                    cols_metrics = st.columns(len(cols_verificar))
                    
                    for idx, col in enumerate(cols_verificar):
                        try:
                            val_act = float(row_activos[col].values[0])
                        except:
                            val_act = 0.0
                            
                        try:
                            val_pp = float(row_pat_pas[col].values[0])
                        except:
                            val_pp = 0.0
                            
                        diff = abs(val_act) - abs(val_pp)
                        
                        with cols_metrics[idx]:
                            if abs(diff) < 1.0:
                                st.success(f"**{col}**\n\nCuadrado (0)")
                            else:
                                diff_fmt = f"{diff:,.0f}".replace(",", ".")
                                st.error(f"**{col}**\n\nDescuadre: {diff_fmt}")
                    st.write("")
                    
            elif report_type == "Estado de Resultados":
                st.write("")
                st.subheader("🔍 Validación de Resultado Consolidado vs. Hoja de Trabajo")
                
                df_valid_rows = preview_df[preview_df.iloc[:, 0].notna() & (preview_df.iloc[:, 0].astype(str).str.strip() != "")]
                if not df_valid_rows.empty and df_hoja_act is not None:
                    target_names = [
                        "ganancias (perdida) del ejercicio",
                        "ganancias (pérdida) del ejercicio",
                        "ganancia (perdida) del ejercicio",
                        "ganancia (pérdida) del ejercicio",
                        "(perdida) procedente de operaciones continuadas",
                        "(pérdida) procedente de operaciones continuadas",
                        "perdida procedente de operaciones continuadas",
                        "pérdida procedente de operaciones continuadas",
                        "perdida",
                        "pérdida"
                    ]
                    net_income_row = df_valid_rows.iloc[-1]
                    for name in target_names:
                        matched_rows = df_valid_rows[df_valid_rows.iloc[:, 0].astype(str).str.replace(r'\xa0', ' ', regex=True).str.strip().str.lower() == name]
                        if not matched_rows.empty:
                            net_income_row = matched_rows.iloc[0]
                            break

                    val_actual_report = float(net_income_row[col_actual]) if pd.notna(net_income_row[col_actual]) else 0.0
                    val_actual_report_pesos = val_actual_report * scale_factor
                    
                    idx_er_act = df_hoja_act[df_hoja_act['Balance clasificado'] == "Estado de Resultados"].index
                    df_sec_act = df_hoja_act.loc[idx_er_act[0]+1:] if not idx_er_act.empty else df_hoja_act
                    
                    row_hoja_act = df_sec_act[df_sec_act['Balance clasificado'].astype(str).str.strip() == "Ganancias (Pérdida) del Ejercicio"]
                    val_actual_workpaper = float(row_hoja_act['CONSOLIDADO'].values[0]) if not row_hoja_act.empty else 0.0
                    
                    cols_verificar = [(col_actual, val_actual_report_pesos, val_actual_workpaper)]
                    
                    if col_comp in preview_df.columns and df_hoja_comp is not None:
                        val_comp_report = float(net_income_row[col_comp]) if pd.notna(net_income_row[col_comp]) else 0.0
                        val_comp_report_pesos = val_comp_report * scale_factor
                        
                        idx_er_comp = df_hoja_comp[df_hoja_comp['Balance clasificado'] == "Estado de Resultados"].index
                        df_sec_comp = df_hoja_comp.loc[idx_er_comp[0]+1:] if not idx_er_comp.empty else df_hoja_comp
                        
                        row_hoja_comp = df_sec_comp[df_sec_comp['Balance clasificado'].astype(str).str.strip() == "Ganancias (Pérdida) del Ejercicio"]
                        val_comp_workpaper = float(row_hoja_comp['CONSOLIDADO'].values[0]) if not row_hoja_comp.empty else 0.0
                        cols_verificar.append((col_comp, val_comp_report_pesos, val_comp_workpaper))
                        
                    cols_metrics = st.columns(len(cols_verificar))
                    for idx, (col_name, val_rep_pesos, val_wp_pesos) in enumerate(cols_verificar):
                        diff = abs(abs(val_rep_pesos) - abs(val_wp_pesos))
                        
                        with cols_metrics[idx]:
                            if diff <= 999.0:
                                diff_fmt = f"{diff:,.0f}".replace(",", ".")
                                st.success(f"**{col_name}**\n\nCuadrado (Diferencia: ${diff_fmt})")
                            else:
                                diff_fmt = f"{diff:,.0f}".replace(",", ".")
                                rep_fmt = f"{val_rep_pesos:,.0f}".replace(",", ".")
                                wp_fmt = f"{val_wp_pesos:,.0f}".replace(",", ".")
                                st.error(
                                    f"**{col_name}**\n\n"
                                    f"❌ Descuadre: ${diff_fmt}\n\n"
                                    f"* Reporte Cons.: ${rep_fmt}\n"
                                    f"* Hoja de Trabajo: ${wp_fmt}"
                                )
                st.write("")
                
            col_dn1, col_dn2 = st.columns(2)
            with col_dn1:
                st.download_button(
                    "📥 Descargar Reporte Final (Excel)", 
                    data=st.session_state['consolidated_report_excel_binary'], 
                    file_name=filename, 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                    type="primary",
                    use_container_width=True,
                    key="btn_dl_cons_excel"
                )
            with col_dn2:
                word_data_bytes = st.session_state.get('consolidated_report_word_binary')
                if word_data_bytes:
                    st.download_button(
                        "📝 Descargar Reporte Final (Word)", 
                        data=word_data_bytes, 
                        file_name=word_filename, 
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document", 
                        type="primary", 
                        use_container_width=True,
                        key="btn_dl_cons_word"
                    )
                else:
                    st.info("Exportación a Word no disponible para este reporte.")
                        
        st.write("")
        st.divider()
        st.subheader("⚙️ Diagnóstico de Plantillas del Grupo")
        st.info(f"📁 **Directorio del Grupo**: `{grupo_folder}`")
        
        col_status1, col_status2, col_status3, col_status4 = st.columns(4)
        with col_status1:
            st.write("**Plantilla Balance Clasificado**")
            if os.path.exists(os.path.join(grupo_folder, "Balance clasificado.xlsx")):
                st.success("✅ Plantilla Balance detectada")
            else:
                st.error("❌ Falta plantilla Balance")
                st.info("💡 Sube esta plantilla desde: *Operaciones ➔ Clasificación de Cuentas ➔ Gestión de Plantillas*.")
                
        with col_status2:
            st.write("**Plantilla Estado Resultados**")
            if os.path.exists(os.path.join(grupo_folder, "Estado de Resultados Clasificados.xlsx")):
                st.success("✅ Plantilla E.R. detectada")
            else:
                st.error("❌ Falta plantilla E.R.")
                st.info("💡 Sube esta plantilla desde: *Operaciones ➔ Clasificación de Cuentas ➔ Gestión de Plantillas*.")

        with col_status3:
            st.write("**Plantilla Estado de Cambios en el Patrimonio**")
            if os.path.exists(os.path.join(grupo_folder, "Estado de Cambios en el Patrimonio.xlsx")):
                st.success("✅ Plantilla Patrimonio detectada")
            else:
                st.error("❌ Falta plantilla Patrimonio")
                st.info("💡 Sube esta plantilla desde: *Operaciones ➔ Clasificación de Cuentas ➔ Gestión de Plantillas*.")

        with col_status4:
            st.write("**Plantilla Estado de Resultados Integrales**")
            if os.path.exists(os.path.join(grupo_folder, "Estado de Resultados Integrales.xlsx")):
                st.success("✅ Plantilla ORI detectada")
            else:
                st.error("❌ Falta plantilla ORI")
                st.info("💡 Sube esta plantilla desde: *Operaciones ➔ Clasificación de Cuentas ➔ Gestión de Plantillas*.")
