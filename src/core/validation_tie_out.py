import os
import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

from src.models.database import SessionLocal
from src.models.trial_balance_db import TrialBalanceDB
from src.models.pl_cubo_db import PlCuboDB
from src.models.consolidacion import ConsolidationGroup, ConsolidationJournalEntry
from src.core.excel_utils import read_excel_cached, df_to_excel_bytes
from src.core.sabana_builder import build_balance_sabana, build_pl_sabana, build_consolidated_balance_sabana
from src.core.consolidacion_engine import generar_hoja_trabajo, COLUMNAS_ORDENADAS

import re

RUBRO_NOTE_PATTERNS = [
    # Entidades relacionadas (prioridad)
    ("relacionada", 14, "Nota 14"),
    ("relacionadas", 14, "Nota 14"),
    ("inversion en empresas relacionadas", 14, "Nota 14"),

    # Pasivos por derechos de uso (prioridad sobre activo por derecho de uso)
    ("pasivos por derechos de uso", 16, "Nota 16"),
    ("pasivo por derechos de uso", 16, "Nota 16"),
    ("pasivos por derecho de uso", 16, "Nota 16"),
    ("pasivo por derecho de uso", 16, "Nota 16"),

    # Activos por derechos de uso
    ("activo por derechos de uso", 10, "Nota 10"),
    ("activos por derechos de uso", 10, "Nota 10"),
    ("derechos de uso", 10, "Nota 10"),
    ("derecho de uso", 10, "Nota 10"),

    # Activos Corrientes
    ("efectivo", 4, "Nota 04"),
    ("caja", 4, "Nota 04"),
    ("bancos", 4, "Nota 04"),
    
    ("otros activos no financieros", 5, "Nota 05"),
    
    ("deudores comerciales", 6, "Nota 06"),
    ("cuentas por cobrar", 6, "Nota 06"),
    
    ("inventario", 7, "Nota 07"),
    ("existencias", 7, "Nota 07"),
    
    # Activos No Corrientes
    ("intangible", 8, "Nota 08"),
    ("licencia", 8, "Nota 08"),
    
    ("propiedades, planta", 9, "Nota 09"),
    ("propiedades, plantas", 9, "Nota 09"),
    ("activo fijo", 9, "Nota 09"),
    
    ("plusvalia", 11, "Nota 11"),
    ("plusvalía", 11, "Nota 11"),
    ("goodwill", 11, "Nota 11"),
    
    ("impuestos corrientes", 12, "Nota 12"),
    ("impuestos, corrientes", 12, "Nota 12"),
    ("impuesto, corriente", 12, "Nota 12"),
    ("pasivo por impuestos, corrientes", 12, "Nota 12"),
    ("activo por impuestos, corrientes", 12, "Nota 12"),
    
    ("impuestos diferidos", 13, "Nota 13"),
    ("impuesto diferido", 13, "Nota 13"),

    # Pasivos
    ("otros pasivos financieros", 15, "Nota 15"),
    ("pasivos financieros", 15, "Nota 15"),
    ("instrumentos financieros", 15, "Nota 15"),
    
    ("cuentas comerciales y otras cuentas por pagar", 17, "Nota 17"),
    ("cuentas por pagar", 17, "Nota 17"),
    ("proveedores", 17, "Nota 17"),
    
    ("beneficios a los empleados", 18, "Nota 18"),
    ("vacaciones", 18, "Nota 18"),
    
    ("otras provisiones", 19, "Nota 19"),
    ("otros pasivos no financieros", 19, "Nota 19"),
    ("provisiones", 19, "Nota 19"),

    # Patrimonio
    ("capital emitido", 20, "Nota 20"),
    ("otras reservas", 20, "Nota 20"),
    ("resultados acumulados", 20, "Nota 20"),
    ("patrimonio", 20, "Nota 20"),

    # Resultados
    ("ingresos de actividades ordinarias", 21, "Nota 21"),
    ("costo de ventas", 21, "Nota 21"),
    ("infraestructura fibra", 21, "Nota 21"),
    ("uso fibra", 21, "Nota 21"),
    
    ("gastos de administración", 22, "Nota 22"),
    ("gastos de administracion", 22, "Nota 22"),
    ("depreciación operacional", 22, "Nota 22"),
    ("depreciacion operacional", 22, "Nota 22"),
    ("depreciación y amortizaciones", 22, "Nota 22"),
    ("depreciacion y amortizaciones", 22, "Nota 22"),
    
    ("diferencia de cambio", 23, "Nota 23"),
    ("diferencias de cambio", 23, "Nota 23"),
    ("unidades de reajuste", 23, "Nota 23"),
    
    ("ingresos financieros", 24, "Nota 24"),
    ("costos financieros", 24, "Nota 24"),
    
    ("otros egresos por función", 25, "Nota 25"),
    ("otros egresos por funcion", 25, "Nota 25"),
    ("otros ingresos por función", 25, "Nota 25"),
    ("otros ingresos por funcion", 25, "Nota 25"),
    ("impuestos a las ganancias", 25, "Nota 25"),

    ("segmentos", 26, "Nota 26")
]

def resolve_note_code(rubro: str, nota_asociada: str = ""):
    """Determina el número de orden y la etiqueta de la Nota para un rubro dado."""
    rubro_lower = str(rubro).lower().strip()
    nota_lower = str(nota_asociada).lower().strip()
    
    # 1. Buscar coincidencia en nota_asociada si trae "nota X" o "#NX"
    match_num = re.search(r'(?:nota|#n)\s*0*(\d+)', nota_lower)
    if match_num:
        num = int(match_num.group(1))
        return (num, f"Nota {num:02d}")
        
    # 2. Buscar por patrón en el texto del rubro
    for pattern, num, label in RUBRO_NOTE_PATTERNS:
        if pattern in rubro_lower:
            return (num, label)
            
    return (99, "Nota N/A")

class ValidationTieOutEngine:

    @staticmethod
    def obtener_matriz_tie_out(empresa_o_grupo: str, periodo: str, is_consolidated: bool = False):
        """
        Genera la Matriz de Validación Tie-Out (EEFF vs. Notas) y los Chequeos de Salud Financiera.
        Compara la suma real de las cuentas mapeadas a notas contra el total del rubro del EEFF.
        Ordena la tabla por el número de código de Nota y ubica 'N° Nota' como primera columna a la izquierda.
        """
        rows = []
        health_checks = {
            "is_valid": True,
            "total_descuadres": 0,
            "ec_patrimonial_diff": 0.0,
            "utilidad_ejercicio_diff": 0.0,
            "cuentas_huerfanas_count": 0,
            "tb_cuadre_diff": 0.0
        }

        # ---------------------------------------------------------------------
        # CASO 1: GRUPO CONSOLIDADO
        # ---------------------------------------------------------------------
        if is_consolidated:
            db = SessionLocal()
            grp = db.query(ConsolidationGroup).filter_by(nombre_grupo=empresa_o_grupo).first()
            db.close()

            if not grp:
                return pd.DataFrame(), health_checks

            df_hoja, msg = generar_hoja_trabajo(grp.id, periodo)
            if df_hoja is None or df_hoja.empty:
                return pd.DataFrame(), health_checks

            grupo_folder = os.path.join("data", "empresas", empresa_o_grupo)
            if not os.path.exists(grupo_folder):
                grupo_folder = os.path.join("data", "empresas", f"[GRUPO] {empresa_o_grupo}")
            map_bal_path = os.path.join(grupo_folder, "map_balance.xlsx")
            if not os.path.exists(map_bal_path): map_bal_path = "map_balance.xlsx"
            map_bal_df = read_excel_cached(map_bal_path, dtype=str) if os.path.exists(map_bal_path) else None

            # Sábana consolidada de Balance
            df_sab_bal = build_consolidated_balance_sabana(empresa_o_grupo, periodo, map_bal_df)

            idx_er = df_hoja[df_hoja['Balance clasificado'] == "Estado de Resultados"].index
            if not idx_er.empty:
                df_bal_sec = df_hoja.loc[:idx_er[0]-1]
                df_pl_sec = df_hoja.loc[idx_er[0]+1:]
            else:
                df_bal_sec = df_hoja
                df_pl_sec = pd.DataFrame()

            # --- A. BALANCE CONSOLIDADO ---
            if not df_bal_sec.empty:
                col_clasif_sab = next((c for c in df_sab_bal.columns if "clasifica" in str(c).lower() or "balance" in str(c).lower()), None) if not df_sab_bal.empty else None
                notas_cols_sab = [c for c in df_sab_bal.columns if "nota" in str(c).lower()] if not df_sab_bal.empty else []
                col_monto_sab = next((c for c in df_sab_bal.columns if "consolidado" in str(c).lower() or "saldo" in str(c).lower()), None) if not df_sab_bal.empty else None

                for _, r in df_bal_sec.iterrows():
                    rubro = str(r.get('Balance clasificado', '')).strip()
                    if not rubro or rubro.lower() == "nan" or rubro in ["Activos corrientes", "Activos corrientes totales", "Activos no corrientes", "Activos no corrientes totales", "TOTAL ACTIVOS", "Pasivos corrientes", "Pasivo corrientes totales", "Pasivos no corrientes", "Pasivo no corrientes totales", "TOTAL PASIVOS", "PATRIMONIO", "Patrimonio total", "TOTAL PATRIMONIO Y PASIVOS", "Estado de Resultados"]:
                        continue

                    val_eeff = float(r.get('CONSOLIDADO', 0.0)) if pd.notna(r.get('CONSOLIDADO')) else 0.0
                    nota_asociada = f"Nota {rubro}"
                    val_nota = 0.0

                    if not df_sab_bal.empty and col_clasif_sab and col_monto_sab:
                        sub_df = df_sab_bal[df_sab_bal[col_clasif_sab].astype(str).str.strip().str.lower() == rubro.lower()]
                        if not sub_df.empty:
                            if notas_cols_sab:
                                notas_set = []
                                for nc in notas_cols_sab:
                                    vals = [str(n).strip() for n in sub_df[nc].dropna().unique() if str(n).strip() and str(n).strip().lower() != "nan" and str(n).strip().lower() != rubro.lower()]
                                    notas_set.extend(vals)
                                notas_set = list(dict.fromkeys(notas_set))

                                if notas_set:
                                    nota_asociada = ", ".join(notas_set[:3])
                                    mask_nota = sub_df[notas_cols_sab].fillna('').astype(str).apply(
                                        lambda row: any(str(val).strip() != '' and str(val).strip().lower() != 'nan' and str(val).strip().lower() != rubro.lower() for val in row),
                                        axis=1
                                    )
                                    val_nota = float(pd.to_numeric(sub_df[mask_nota][col_monto_sab], errors='coerce').fillna(0.0).sum())
                                else:
                                    val_nota = val_eeff
                            else:
                                val_nota = val_eeff
                        else:
                            val_nota = 0.0
                    else:
                        val_nota = 0.0

                    diff = round(abs(val_eeff - val_nota), 2)
                    status = "✅ OK" if diff < 1.0 else "❌ DESCUADRE"

                    if status == "❌ DESCUADRE":
                        health_checks["total_descuadres"] += 1

                    rows.append({
                        "Reporte": "Balance (ESF)",
                        "Rubro Estado Financiero": rubro,
                        "Saldo EEFF ($)": val_eeff,
                        "Nota Asociada": nota_asociada,
                        "Suma Sábana ($)": val_nota,
                        "Suma Nota ($)": val_nota,
                        "Diferencia ($)": diff,
                        "Estado Sábana": status,
                        "Estado Nota Generada": status,
                        "Estado": status
                    })

            # --- B. P&L CONSOLIDADO ---
            if not df_pl_sec.empty:
                for _, r in df_pl_sec.iterrows():
                    rubro = str(r.get('Estado de Resultados clasificado', r.get('Balance clasificado', ''))).strip()
                    if not rubro or rubro.lower() == "nan":
                        continue

                    val_eeff = float(r.get('CONSOLIDADO', 0.0)) if pd.notna(r.get('CONSOLIDADO')) else 0.0
                    val_nota = val_eeff
                    nota_asociada = f"Nota {rubro}"

                    diff = round(abs(val_eeff - val_nota), 2)
                    status = "✅ OK" if diff < 1.0 else "❌ DESCUADRE"

                    if status == "❌ DESCUADRE":
                        health_checks["total_descuadres"] += 1

                    rows.append({
                        "Reporte": "Estado de Resultados (ERI)",
                        "Rubro Estado Financiero": rubro,
                        "Saldo EEFF ($)": val_eeff,
                        "Nota Asociada": nota_asociada,
                        "Suma Sábana ($)": val_nota,
                        "Suma Nota ($)": val_nota,
                        "Diferencia ($)": diff,
                        "Estado Sábana": status,
                        "Estado Nota Generada": status,
                        "Estado": status
                    })

        # ---------------------------------------------------------------------
        # CASO 2: EMPRESA INDIVIDUAL
        # ---------------------------------------------------------------------
        else:
            empresa_dir = os.path.join("data", "empresas", empresa_o_grupo)
            map_bal_path = os.path.join(empresa_dir, "map_balance.xlsx")
            if not os.path.exists(map_bal_path): map_bal_path = "map_balance.xlsx"
            map_pl_path = os.path.join(empresa_dir, "map_pl.xlsx")
            if not os.path.exists(map_pl_path): map_pl_path = "map_pl.xlsx"

            map_bal_df = pd.read_excel(map_bal_path, dtype=str) if os.path.exists(map_bal_path) else None
            map_pl_df = pd.read_excel(map_pl_path, dtype=str) if os.path.exists(map_pl_path) else None

            tb_df = TrialBalanceDB.get_trial_balance(empresa_o_grupo, periodo)
            pl_df = PlCuboDB.get_pl_cubo(empresa_o_grupo, periodo)

            sab_bal = build_balance_sabana(tb_df, map_bal_df) if tb_df is not None and not tb_df.empty else pd.DataFrame()

            # --- GENERAR Y EVALUAR PLANTILLA DE NOTAS PARA OBTENER SUMAS VISUALES REALES ---
            template_nota = os.path.join(empresa_dir, "Plantilla de notas_v1.xlsx")
            if not os.path.exists(template_nota):
                template_nota = "Plantilla de notas_v1.xlsx"

            wb_eval = None
            try:
                from src.ui_pages.informes_y_notas import load_all_entity_contexts, evaluate_formulas_in_workbook
                from src.reporting.note_generator import NoteGenerator
                from src.reporting.notes import NOTE_REGISTRY

                periodo_comp = "2025-12" if "2026" in str(periodo) else "2024-12"
                entity_contexts = load_all_entity_contexts(empresa_o_grupo, periodo, periodo_comp, map_bal_df, map_pl_df)
                all_sheets = []
                for code, info in NOTE_REGISTRY.items():
                    all_sheets.extend(info.get("sheets", []))
                all_sheets = list(dict.fromkeys(all_sheets))

                engine = NoteGenerator(template_nota)
                excel_nota_out = engine.generate(
                    sheet_names=all_sheets,
                    entity_contexts=entity_contexts,
                    active_entity_name=empresa_o_grupo,
                    is_consolidated=False,
                    scale_factor=1.0,
                    periodo_actual_str=periodo,
                    periodo_comp_str=periodo_comp,
                    map_balance_df=map_bal_df,
                    map_pl_df=map_pl_df
                )

                excel_nota_out.seek(0)
                excel_eval_out = evaluate_formulas_in_workbook(excel_nota_out)
                excel_eval_out.seek(0)
                wb_eval = openpyxl.load_workbook(excel_eval_out, data_only=True)
            except Exception as e:
                wb_eval = None

            def get_sheet_evaluated_total(sheet_name, target_keywords=None):
                if not wb_eval or sheet_name not in wb_eval.sheetnames:
                    return 0.0
                ws = wb_eval[sheet_name]
                for r in range(1, ws.max_row + 1):
                    lbl_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column+1, 4))]
                    lbl_str = " ".join([str(v).lower() for v in lbl_vals if v is not None])
                    if any(sub_kw in lbl_str for sub_kw in ["subtotal", "sub total", "sub-total", "sub_total"]):
                        continue
                    if any(kw in lbl_str for kw in ["total", "saldo final", "totales"]):
                        full_row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column+1, 15))]
                        full_row_str = " ".join([str(v).lower() for v in full_row_vals if v is not None])
                        if target_keywords:
                            if not any(tk in full_row_str for tk in target_keywords):
                                continue
                        nums = [float(v) for v in full_row_vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
                        if nums:
                            if len(nums) >= 6:
                                return nums[2]
                            return nums[-1] if len(nums) >= 3 else nums[0]
                return 0.0

            RUBRO_NOTE_LINK = {
                # Balance
                "Efectivo y efectivo equivalente": ("#N04", ["Efectivo"]),
                "Otros activos no financieros, corrientes": ("#N05", ["Otros activos no financieros, c"]),
                "Deudores comerciales y otras cuentas por cobrar, corrientes": ("#N06", ["Deudores"]),
                "Inventarios": ("#N07", ["Inventarios"]),
                "Activos intangibles distinto a la plusvalía": ("#N08", ["Intangibles"]),
                "Activos intangibles distinto a la plusvalia": ("#N08", ["Intangibles"]),
                "Propiedades, plantas y equipos": ("#N09", ["Activo Fijo"]),
                "Activo por derechos de uso": ("#N10", ["Activo por derechos de uso"]),
                "Plusvalia": ("#N11", ["Plusvalia"]),
                "Plusvalía": ("#N11", ["Plusvalia"]),
                "Activo por impuestos, corrientes": ("#N12", ["Impuestos corrientes"]),
                "Pasivo por impuestos, corrientes": ("#N12", ["Impuestos corrientes"]),
                "Activo por impuestos diferidos, no corrientes": ("#N13", ["Impuestos Diferidos"]),
                "Cuentas por cobrar a entidades relacionadas, no corrientes": ("#N14", ["Empresas relacionadas"]),
                "Cuentas por pagar entidades relacionadas, corrientes": ("#N14", ["Empresas relacionadas"]),
                "Cuentas por pagar entidades relacionadas, no corrientes": ("#N14", ["Empresas relacionadas"]),
                "Inversion en empresas relacionadas": ("#N14", ["Empresas relacionadas"]),
                "Otros pasivos financieros corrientes": ("#N15", ["Pasivos financieros "]),
                "Otros pasivos financieros, no corriente": ("#N15", ["Pasivos financieros "]),
                "Pasivos por derechos de uso, corrientes": ("#N16", ["Pasivos derechos de  uso"]),
                "Pasivos por derechos de uso, no corrientes": ("#N16", ["Pasivos derechos de  uso"]),
                "Cuentas comerciales y otras cuentas por pagar, corrientes": ("#N17", ["Cuentas por pagar"]),
                "Cuentas comerciales y otras cuentas por pagar, no corrientes": ("#N17", ["Cuentas por pagar"]),
                "Provisiones por beneficios a los empleados": ("#N18", ["Provisiones"]),
                "Otras provisiones, no corrientes": ("#N19", ["Otros pasivos no financieros"]),
                "Capital emitido": ("#N20", ["Patrimonio"]),
                "Otras reservas": ("#N20", ["Patrimonio"]),
                "Resultados acumulados": ("#N20", ["Patrimonio"]),

                # P&L
                "Ingresos de actividades ordinarias": ("#N21", ["Ingresos Ctos operacion"]),
                "Costo de ventas": ("#N21", ["Ingresos Ctos operacion"]),
                "Costos de uso fibra optica": ("#N21", ["Ingresos Ctos operacion"]),
                "Acceso a infraestructura fibra óptica": ("#N21", ["Ingresos Ctos operacion"]),
                "Depreciación operacional": ("#N21", ["Ingresos Ctos operacion"]),
                "Depreciación y amortizaciones": ("#N22", ["Gtos Adm"]),
                "Gastos de administración": ("#N22", ["Gtos Adm"]),
                "Diferencias de cambio": ("#N23", ["DC y Reajustes"]),
                "Resultados por unidades de reajuste": ("#N23", ["DC y Reajustes"]),
                "Costos financieros": ("#N24", ["Costos e ingresos Financieros"]),
                "Ingresos financieros": ("#N24", ["Costos e ingresos Financieros"]),
                "Otros egresos por función": ("#N25", ["Otros gastos por funcion"]),
                "Otros ingresos por función": ("#N25", ["Otros ingresos por funcion"]),
                "Resultado por impuestos a las ganancias": ("#N13", ["Impuestos Diferidos"]),
            }

            # --- A. BALANCE INDIVIDUAL ---
            if not sab_bal.empty:
                col_clasif = next((c for c in sab_bal.columns if "clasifica" in str(c).lower() or "balance" in str(c).lower()), None)
                if 'saldo_final' in sab_bal.columns:
                    col_monto = 'saldo_final'
                else:
                    col_monto = next((c for c in sab_bal.columns if "final" in str(c).lower() or "monto" in str(c).lower() or "saldo" in str(c).lower()), 'saldo_final')
                
                notas_cols = [c for c in sab_bal.columns if "nota" in str(c).lower()]
                unmapped_note_rubros = {"plusvalia", "plusvalía", "activo por impuestos diferidos, no corrientes", "capital emitido", "otras reservas"}

                if col_clasif and col_monto in sab_bal.columns:
                    rubros = sab_bal[col_clasif].dropna().astype(str).str.strip().unique()
                    for rubro in sorted(rubros):
                        if not rubro or rubro.lower() == "nan" or rubro in ["Activos corrientes", "Activos corrientes totales", "Activos no corrientes", "Activos no corrientes totales", "TOTAL ACTIVOS", "Pasivos corrientes", "Pasivo corrientes totales", "Pasivos no corrientes", "Pasivo no corrientes totales", "TOTAL PASIVOS", "PATRIMONIO", "Patrimonio total", "TOTAL PATRIMONIO Y PASIVOS", "Estado de Resultados"]:
                            continue

                        sub_df = sab_bal[sab_bal[col_clasif].astype(str).str.strip().str.lower() == rubro.lower()]
                        val_eeff = float(pd.to_numeric(sub_df[col_monto], errors='coerce').fillna(0.0).sum())

                        nota_asociada = f"Nota {rubro}"
                        val_sabana = 0.0

                        if rubro.lower().strip() in unmapped_note_rubros:
                            val_sabana = 0.0
                        elif notas_cols:
                            notas_set = []
                            for nc in notas_cols:
                                vals = [str(n).strip() for n in sub_df[nc].dropna().unique() if str(n).strip() and str(n).strip().lower() != "nan" and str(n).strip().lower() != rubro.lower()]
                                notas_set.extend(vals)
                            notas_set = list(dict.fromkeys(notas_set))

                            if notas_set:
                                nota_asociada = ", ".join(notas_set[:3])
                                mask_nota = sub_df[notas_cols].fillna('').astype(str).apply(
                                    lambda row: any(str(val).strip() != '' and str(val).strip().lower() != 'nan' and str(val).strip().lower() != rubro.lower() for val in row),
                                    axis=1
                                )
                                val_sabana = float(pd.to_numeric(sub_df[mask_nota][col_monto], errors='coerce').fillna(0.0).sum())
                            else:
                                val_sabana = val_eeff
                        else:
                            val_sabana = val_eeff

                        # Nivel 2: Nota Evaluada Visual
                        code, sheets = RUBRO_NOTE_LINK.get(rubro, (None, []))
                        val_nota_eval = sum([get_sheet_evaluated_total(sh) for sh in sheets]) if sheets and wb_eval else 0.0

                        diff_sabana = round(abs(val_eeff - val_sabana), 2)
                        diff_nota = round(min(abs(val_eeff - val_nota_eval), abs(abs(val_eeff) - abs(val_nota_eval))), 2)

                        status_sabana = "✅ OK" if diff_sabana < 1.0 else "❌ SIN MAPEO"
                        status_nota = "✅ OK" if diff_nota < 1000.0 else "❌ DESCUADRE EN NOTA EMITIDA"

                        if status_sabana != "✅ OK" or status_nota != "✅ OK":
                            health_checks["total_descuadres"] += 1

                        rows.append({
                            "Reporte": "Balance (ESF)",
                            "Rubro Estado Financiero": rubro,
                            "Saldo EEFF ($)": val_eeff,
                            "Nota Asociada": nota_asociada,
                            "Suma Sábana ($)": val_sabana,
                            "Suma Nota ($)": val_nota_eval,
                            "Diferencia ($)": diff_nota if diff_nota >= 1000.0 else diff_sabana,
                            "Estado Sábana": status_sabana,
                            "Estado Nota Generada": status_nota,
                            "Estado": status_nota if status_nota != "✅ OK" else status_sabana
                        })

            # --- B. P&L INDIVIDUAL ---
            if pl_df is not None and not pl_df.empty:
                exclude_cols = ['n° de cuenta', 'n de cuenta', 'cuenta', 'cuenta_id', 'nombre de la cuenta', 'nombre', 'descripcion', 'rubro']
                pl_rubros_cols = [c for c in pl_df.columns if not any(ex in str(c).lower() for ex in exclude_cols)]

                for col_r in pl_rubros_cols:
                    rubro_clean = str(col_r).replace('(Saldo)', '').strip()
                    if not rubro_clean or rubro_clean.lower() == "nan":
                        continue

                    series_vals = pd.to_numeric(pl_df[col_r], errors='coerce').fillna(0.0)
                    val_eeff = float(series_vals.sum())
                    val_sabana = val_eeff

                    code, sheets = RUBRO_NOTE_LINK.get(rubro_clean, (None, []))
                    PL_KEYWORDS = {
                        "Ingresos de actividades ordinarias": ["ingresos"],
                        "Costo de ventas": ["costos", "costo"],
                        "Costos de uso fibra optica": ["fibra"],
                        "Acceso a infraestructura fibra óptica": ["infraestructura", "fibra"],
                        "Depreciación operacional": ["depreciacion"],
                        "Depreciación y amortizaciones": ["amortizac"],
                        "Gastos de administración": ["administra"],
                        "Diferencias de cambio": ["cambio"],
                        "Resultados por unidades de reajuste": ["reajuste"],
                        "Costos financieros": ["costos financieros", "costos"],
                        "Ingresos financieros": ["ingresos financieros", "ingresos"],
                        "Otros egresos por función": ["gastos", "egresos"],
                        "Otros ingresos por función": ["ingresos"],
                        "Resultado por impuestos a las ganancias": ["impuesto"],
                    }
                    kw = PL_KEYWORDS.get(rubro_clean, None)
                    if kw is None:
                        if "egresos" in rubro_clean.lower() or "costo" in rubro_clean.lower(): kw = ["costos", "gastos", "egresos"]
                        elif "ingresos" in rubro_clean.lower(): kw = ["ingresos"]
                    val_nota_eval = sum([get_sheet_evaluated_total(sh, kw) for sh in sheets]) if sheets and wb_eval else 0.0

                    diff_sabana = round(abs(val_eeff - val_sabana), 2)
                    diff_nota = round(min(abs(val_eeff - val_nota_eval), abs(abs(val_eeff) - abs(val_nota_eval))), 2)

                    status_sabana = "✅ OK" if diff_sabana < 1.0 else "❌ SIN MAPEO"
                    status_nota = "✅ OK" if diff_nota < 1000.0 else "❌ DESCUADRE EN NOTA EMITIDA"

                    if status_sabana != "✅ OK" or status_nota != "✅ OK":
                        health_checks["total_descuadres"] += 1

                    rows.append({
                        "Reporte": "Estado de Resultados (ERI)",
                        "Rubro Estado Financiero": rubro_clean,
                        "Saldo EEFF ($)": val_eeff,
                        "Nota Asociada": f"Nota {rubro_clean}",
                        "Suma Sábana ($)": val_sabana,
                        "Suma Nota ($)": val_nota_eval,
                        "Diferencia ($)": diff_nota if diff_nota >= 1000.0 else diff_sabana,
                        "Estado Sábana": status_sabana,
                        "Estado Nota Generada": status_nota,
                        "Estado": status_nota if status_nota != "✅ OK" else status_sabana
                    })

        res_df = pd.DataFrame(rows)
        health_checks["is_valid"] = (health_checks["total_descuadres"] == 0)

        # Ordenar por Número de Código de Nota y poner 'N° Nota' como primera columna a la izquierda
        if not res_df.empty:
            note_info = [resolve_note_code(r['Rubro Estado Financiero'], r.get('Nota Asociada', '')) for _, r in res_df.iterrows()]
            res_df['_Sort_Order'] = [x[0] for x in note_info]
            res_df['N° Nota'] = [x[1] for x in note_info]

            # Excluir Nota 26 (Segmentos) en informes individuales (solo pertenece al consolidado)
            if not is_consolidated:
                res_df = res_df[res_df['_Sort_Order'] != 26].reset_index(drop=True)

            res_df = res_df.sort_values(by=['_Sort_Order', 'Reporte', 'Rubro Estado Financiero']).reset_index(drop=True)
            cols = ['N° Nota', 'Reporte', 'Rubro Estado Financiero', 'Saldo EEFF ($)', 'Nota Asociada', 'Suma Sábana ($)', 'Suma Nota ($)', 'Diferencia ($)', 'Estado Sábana', 'Estado Nota Generada', 'Estado']
            res_df = res_df[cols]

        return res_df, health_checks

        return res_df, health_checks

    @staticmethod
    def generar_excel_tie_out(df_matrix: pd.DataFrame, health_checks: dict, empresa_o_grupo: str, periodo: str) -> bytes:
        """
        Exporta la Matriz Tie-Out y el Informe de Salud Financiera a un libro Excel bellamente diseñado.
        """
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matriz Tie-Out"

        title_font = Font(name="Arial", size=14, bold=True, color="1F4E78")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

        ws.cell(row=1, column=1, value=f"MATRIZ DE VALIDACIÓN DE SALDOS Y TIE-OUT (EEFF vs. NOTAS)").font = title_font
        ws.cell(row=2, column=1, value=f"Entidad / Perímetro: {empresa_o_grupo} | Período: {periodo}").font = Font(name="Arial", size=11, italic=True)

        status_global = "✅ 100% CUADRADO Y AUDITADO" if health_checks["is_valid"] else f"❌ {health_checks['total_descuadres']} DESCUADRE(S) DETECTADOS"
        ws.cell(row=3, column=1, value=f"Estado Global: {status_global}").font = Font(name="Arial", size=11, bold=True, color="008000" if health_checks["is_valid"] else "C00000")

        start_row = 5
        headers = ["N° Nota", "Reporte", "Rubro Estado Financiero", "Saldo EEFF ($)", "Nota Asociada", "Suma Nota ($)", "Diferencia ($)", "Estado Cuadratura"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        current_row = start_row + 1
        if not df_matrix.empty:
            for _, r in df_matrix.iterrows():
                ws.cell(row=current_row, column=1, value=r.get("N° Nota", "")).alignment = center_align
                ws.cell(row=current_row, column=2, value=r.get("Reporte", "")).alignment = left_align
                ws.cell(row=current_row, column=3, value=r.get("Rubro Estado Financiero", "")).alignment = left_align

                c4 = ws.cell(row=current_row, column=4, value=r.get("Saldo EEFF ($)", 0.0))
                c4.number_format = "#,##0.00"
                c4.alignment = right_align

                ws.cell(row=current_row, column=5, value=r.get("Nota Asociada", "")).alignment = left_align

                c6 = ws.cell(row=current_row, column=6, value=r.get("Suma Nota ($)", 0.0))
                c6.number_format = "#,##0.00"
                c6.alignment = right_align

                c7 = ws.cell(row=current_row, column=7, value=r.get("Diferencia ($)", 0.0))
                c7.number_format = "#,##0.00"
                c7.alignment = right_align

                c8 = ws.cell(row=current_row, column=8, value=r.get("Estado", ""))
                c8.alignment = center_align
                c8.fill = ok_fill if r.get("Estado") == "✅ OK" else error_fill

                for c_i in range(1, 9):
                    ws.cell(row=current_row, column=c_i).border = thin_border

                current_row += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        wb.save(output)
        return output.getvalue()

