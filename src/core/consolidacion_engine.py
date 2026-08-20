import pandas as pd
from sqlalchemy import or_
from src.models.database import SessionLocal
from src.models.consolidacion import ConsolidationGroup, ConsolidationJournalEntry
from src.models.historical_data import HistoricalDataRecord
from src.models.taxonomy_master import TaxonomyMasterRecord

# Definición del orden estándar para las cuentas del Estado de Resultados (P&L) según requerimientos visuales
PL_ORDER_LIST = [
    "ingresos de arriendo fibra optica",
    "ingresos de actividades ordinarias",
    "costo de ventas",
    "acceso a infraestructura fibra optica",
    "costos de uso fibra optica",
    "depreciacion operacional",
    "depreciacion y amortizacion operacional", # Sinónimo de depreciación operacional
    "otros ingresos por funcion",
    "costos de distribucion",
    "gastos de administracion",
    "depreciacion y amortizaciones",           # Depreciación y amortizaciones del periodo
    "otros egresos por funcion",
    "resultado por inversion en empresas relacionadas",
    "ingresos financieros",
    "ingresos financieros con empresas relacionadas",
    "costos financieros",
    "diferencias de cambio",
    "resultado por unidad de reajuste",
    "resultados por unidades de reajuste", # Sinónimo
    "ganancia (perdida) por impuesto a las ganancias",
    "resultado por impuestos a las ganancias" # Sinónimo
]

PL_ORDER_MAP = {name: idx for idx, name in enumerate(PL_ORDER_LIST)}

def get_pl_sort_index(line_name: str) -> int:
    """
    Retorna el índice de ordenamiento de una cuenta de P&L en base a la lista ordenada.
    Realiza una normalización para evitar fallos por acentos, mayúsculas o espacios.
    """
    if not line_name:
        return 9999
    # Normalizar texto (minúsculas, sin acentos)
    norm = line_name.lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
    if norm in PL_ORDER_MAP:
        return PL_ORDER_MAP[norm]
        
    # Búsqueda parcial de respaldo
    for key, idx in PL_ORDER_MAP.items():
        if key in norm or norm in key:
            return idx
            
    return 9999

def generar_siguiente_codigo_asiento(grupo_id: int, periodo: str, db=None) -> str:
    """
    Genera el siguiente código único de comprobante contable mensual para un grupo y periodo.
    Ejemplo: AST-202605-001, AST-202605-002, etc.
    """
    import re
    is_local = False
    if db is None:
        db = SessionLocal()
        is_local = True
    try:
        entries = db.query(ConsolidationJournalEntry).filter(
            ConsolidationJournalEntry.grupo_id == grupo_id,
            ConsolidationJournalEntry.periodo == periodo
        ).all()

        max_seq = 0
        for a in entries:
            code = getattr(a, 'asiento_codigo', None)
            if code:
                match = re.search(r'AST-\d{6}-(\d+)', code)
                if match:
                    seq = int(match.group(1))
                    if seq > max_seq:
                        max_seq = seq

        next_seq = max_seq + 1
        periodo_clean = periodo.replace('-', '').strip()
        return f"AST-{periodo_clean}-{next_seq:03d}"
    finally:
        if is_local:
            db.close()

# Secuencia cronológica y visual estándar de columnas de consolidación
COLUMNAS_ORDENADAS = [
    "Elim inversión",
    "Elim Ctas IC",
    "reversa reclas Plusvalia",
    "PPA",
    "Amortizaciones",
    "Reclasificaciones",
    "Otras Eliminaciones"
]

COLUMNAS_ORDEN_MAP = {col: idx for idx, col in enumerate(COLUMNAS_ORDENADAS)}

def get_columna_index(col_name: str) -> int:
    return COLUMNAS_ORDEN_MAP.get(col_name, 999)

def is_pl_account(li: str) -> bool:
    if not li:
        return False
    norm = li.lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
    is_pl = False
    if "ingreso" in norm:
        if not any(x in norm for x in ["diferido", "diferidos", "anticipado", "anticipados", "por pagar"]):
            is_pl = True
    if any(x in norm for x in ["gasto", "costo", "egreso"]):
        if not any(x in norm for x in ["anticipado", "anticipados", "diferido", "diferidos", "por pagar"]):
            is_pl = True
    if "depreciacion" in norm or "amortizacion" in norm:
        if "acumulada" not in norm:
            is_pl = True
    if "diferencia de cambio" in norm or "diferencias de cambio" in norm:
        is_pl = True
    if "unidad de reajuste" in norm or "unidades de reajuste" in norm:
        is_pl = True
    if "resultado" in norm or "resultados" in norm:
        if not any(x in norm for x in ["acumulado", "acumulados"]):
            is_pl = True
    if "impuesto" in norm or "impuestos" in norm:
        if not any(x in norm for x in ["por cobrar", "por pagar", "diferido", "diferidos", "credito", "debito", "corriente", "corrientes"]):
            is_pl = True
    if "ganancia" in norm or "ganancias" in norm or "perdida" in norm or "perdidas" in norm:
        if not any(x in norm for x in ["acumulado", "acumulados", "acumulada", "acumuladas"]):
            is_pl = True
    return is_pl

def get_target_equity_line(keys_list) -> str:
    for kw in ["resultados acumulados", "ganancias acumuladas", "utilidades acumuladas"]:
        for li in keys_list:
            norm = li.lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')
            if kw in norm:
                return li
    return "Resultados acumulados"



def obtener_saldos_base(grupo_id: int, periodo: str, db=None):
    """
    Obtiene los saldos base de Matriz y Filial para el periodo dado,
    retornando un diccionario {linea_item: {'Matriz': float, 'Filial': float}}.
    """
    is_local_db = False
    if db is None:
        db = SessionLocal()
        is_local_db = True
    try:
        grupo = db.query(ConsolidationGroup).filter_by(id=grupo_id).first()
        if not grupo:
            return {}
            
        matriz_records = db.query(HistoricalDataRecord).filter_by(empresa=grupo.empresa_matriz, periodo=periodo).all()
        base_data = {}
        for r in matriz_records:
            base_data[r.linea_item] = {'Matriz': r.monto, 'Filial': 0.0}
            
        if grupo.filial_is_group:
            sub_grupo_id = int(grupo.empresa_filial)
            sub_df, _ = generar_hoja_trabajo(sub_grupo_id, periodo)
            if sub_df is not None:
                bs_subtotals = {
                    "Activos corrientes", "Activos corrientes totales",
                    "Activos no corrientes", "Activos no corrientes totales",
                    "Total activos", "Patrimonio y pasivos",
                    "Pasivos corrientes", "Pasivo corrientes totales",
                    "Pasivos no corrientes", "Pasivo no corrientes totales",
                    "Total pasivos", "Patrimonio", "Patrimonio total",
                    "Total patrimonio y pasivos"
                }
                pl_subtotals = {
                    "Ganancia bruta", "Resultado antes de impuestos",
                    "Ganancias (Pérdida) del Ejercicio", "Otros rubros no clasificados"
                }
                in_pl = False
                for index, row in sub_df.iterrows():
                    li = row['Balance clasificado']
                    if not li or pd.isna(li):
                        continue
                    li_str = str(li).strip()
                    if li_str == "Estado de Resultados":
                        in_pl = True
                        continue
                        
                    val = row['CONSOLIDADO']
                    if isinstance(val, str) or pd.isna(val):
                        continue
                        
                    if in_pl:
                        if li_str in pl_subtotals:
                            continue
                    else:
                        if li_str in bs_subtotals:
                            continue
                            
                    if li_str not in base_data:
                        base_data[li_str] = {'Matriz': 0.0, 'Filial': 0.0}
                    base_data[li_str]['Filial'] += float(val)
        else:
            filial_records = db.query(HistoricalDataRecord).filter_by(empresa=grupo.empresa_filial, periodo=periodo).all()
            for r in filial_records:
                if r.linea_item not in base_data:
                    base_data[r.linea_item] = {'Matriz': 0.0, 'Filial': 0.0}
                base_data[r.linea_item]['Filial'] += r.monto
                
        return base_data
    finally:
        if is_local_db:
            db.close()

def resolver_montos_asiento(grupo_id: int, periodo: str, lineas_asiento: list, db=None, columna_destino: str = None):
    """
    Resuelve los montos reales (Debe / Haber) para una lista de líneas de un asiento contable.
    Cada línea en `lineas_asiento` es un dict o un objeto con:
        - linea_item
        - debe
        - haber
        - elimina_saldo_total
    
    Retorna una lista de diccionarios, cada uno conteniendo:
        - linea_item
        - debe_calculado
        - haber_calculado
        - elimina_saldo_total
        - saldo_base
    """
    is_local_db = False
    if db is None:
        db = SessionLocal()
        is_local_db = True
    try:
        base_data = obtener_saldos_base(grupo_id, periodo, db)
        
        if columna_destino:
            # 1. Obtener asientos guardados que ocurran antes de la columna_destino cronológicamente
            asientos = db.query(ConsolidationJournalEntry).filter(
                ConsolidationJournalEntry.grupo_id == grupo_id,
                or_(
                    ConsolidationJournalEntry.periodo == periodo,
                    (ConsolidationJournalEntry.periodo < periodo) & (ConsolidationJournalEntry.es_recurrente == True)
                )
            ).all()

            # Ordenar columnas a procesar que sean estrictamente anteriores a columna_destino
            target_idx = get_columna_index(columna_destino)
            asientos_previos = [
                a for a in asientos 
                if get_columna_index(a.columna_ajuste) < target_idx
            ]

            # Si hay asientos previos, agrupamos y procesamos de forma similar a generar_hoja_trabajo
            if asientos_previos:
                from collections import defaultdict
                import copy
                
                col_to_vouchers = defaultdict(lambda: defaultdict(list))
                for a in asientos_previos:
                    code = getattr(a, 'asiento_codigo', None)
                    voucher_key = code if code else (a.periodo, a.glosa)
                    col_to_vouchers[a.columna_ajuste][voucher_key].append(a)
                
                todas_columnas_con_datos = set(a.columna_ajuste for a in asientos_previos)
                columnas_a_procesar = [col for col in COLUMNAS_ORDENADAS if col in todas_columnas_con_datos]
                for col in todas_columnas_con_datos:
                    if col not in columnas_a_procesar:
                        columnas_a_procesar.append(col)

                base_data_acumulado = copy.deepcopy(base_data)

                for col in columnas_a_procesar:
                    vouchers_in_col = col_to_vouchers[col]
                    for v_key, lines in vouchers_in_col.items():
                        netos_comprobante = {}
                        sum_total = 0.0
                        static_lines_zero = []
                        static_lines_other = []
                        dynamic_lines = []
                        dynamic_processed_lines = set()

                        for a in lines:
                            li = a.linea_item
                            if getattr(a, 'elimina_saldo_total', False):
                                if li in dynamic_processed_lines:
                                    neto = 0.0
                                else:
                                    monto_matriz = base_data_acumulado.get(li, {}).get('Matriz', 0.0)
                                    monto_filial = base_data_acumulado.get(li, {}).get('Filial', 0.0)
                                    neto = -(monto_matriz + monto_filial)
                                    dynamic_processed_lines.add(li)
                                netos_comprobante[a.id] = neto
                                sum_total += neto
                                dynamic_lines.append(a)
                            else:
                                debe_val = float(a.debe) if a.debe is not None else 0.0
                                haber_val = float(a.haber) if a.haber is not None else 0.0
                                neto = debe_val - haber_val
                                netos_comprobante[a.id] = neto
                                sum_total += neto
                                if debe_val == 0.0 and haber_val == 0.0:
                                    static_lines_zero.append(a)
                                else:
                                    static_lines_other.append(a)

                        if abs(sum_total) > 0.001:
                            if static_lines_zero:
                                balancer_entry = static_lines_zero[0]
                                netos_comprobante[balancer_entry.id] -= sum_total
                            elif static_lines_other:
                                balancer_entry = static_lines_other[0]
                                netos_comprobante[balancer_entry.id] -= sum_total
                            elif dynamic_lines:
                                balancer_entry = dynamic_lines[0]
                                netos_comprobante[balancer_entry.id] -= sum_total

                        for a in lines:
                            li = a.linea_item
                            neto = netos_comprobante.get(a.id, 0.0)
                            es_prior_year_static_pl = (a.periodo[:4] < periodo[:4]) and (not getattr(a, 'elimina_saldo_total', False)) and is_pl_account(li)
                            if es_prior_year_static_pl:
                                target_eq = get_target_equity_line(base_data_acumulado.keys())
                                if target_eq not in base_data_acumulado:
                                    base_data_acumulado[target_eq] = {'Matriz': 0.0, 'Filial': 0.0}
                                base_data_acumulado[target_eq]['Filial'] += neto
                            else:
                                if li not in base_data_acumulado:
                                    base_data_acumulado[li] = {'Matriz': 0.0, 'Filial': 0.0}
                                base_data_acumulado[li]['Filial'] += neto

                base_data = base_data_acumulado

        netos_comprobante = {} # index -> neto
        sum_total = 0.0
        static_lines_zero = [] # indices
        static_lines_other = [] # indices
        dynamic_lines = [] # indices
        dynamic_processed_lines = set()
        
        clean_lines = []
        for idx, l in enumerate(lineas_asiento):
            if hasattr(l, 'linea_item'):
                li = l.linea_item
                debe = float(l.debe or 0.0)
                haber = float(l.haber or 0.0)
                elimina = bool(l.elimina_saldo_total)
            else:
                li = l.get('linea_item')
                debe = float(l.get('debe', 0.0) or 0.0)
                haber = float(l.get('haber', 0.0) or 0.0)
                elimina = bool(l.get('elimina_saldo_total', False))
            
            clean_lines.append({
                'index': idx,
                'linea_item': li,
                'debe': debe,
                'haber': haber,
                'elimina_saldo_total': elimina
            })
            
        for item in clean_lines:
            idx = item['index']
            li = item['linea_item']
            elimina = item['elimina_saldo_total']
            
            if elimina:
                if li in dynamic_processed_lines:
                    neto = 0.0
                else:
                    monto_matriz = base_data.get(li, {}).get('Matriz', 0.0)
                    monto_filial = base_data.get(li, {}).get('Filial', 0.0)
                    neto = -(monto_matriz + monto_filial)
                    dynamic_processed_lines.add(li)
                netos_comprobante[idx] = neto
                sum_total += neto
                dynamic_lines.append(idx)
            else:
                debe_val = item['debe']
                haber_val = item['haber']
                neto = debe_val - haber_val
                netos_comprobante[idx] = neto
                sum_total += neto
                if debe_val == 0.0 and haber_val == 0.0:
                    static_lines_zero.append(idx)
                else:
                    static_lines_other.append(idx)
                    
        if abs(sum_total) > 0.001:
            if static_lines_zero:
                balancer_idx = static_lines_zero[0]
                netos_comprobante[balancer_idx] -= sum_total
            elif static_lines_other:
                balancer_idx = static_lines_other[0]
                netos_comprobante[balancer_idx] -= sum_total
            elif dynamic_lines:
                balancer_idx = dynamic_lines[0]
                netos_comprobante[balancer_idx] -= sum_total
                
        resolved_lines = []
        for item in clean_lines:
            idx = item['index']
            neto = netos_comprobante.get(idx, 0.0)
            li = item['linea_item']
            
            debe_calc = 0.0
            haber_calc = 0.0
            if neto >= 0.0:
                debe_calc = neto
            else:
                haber_calc = -neto
            
            orig_obj = lineas_asiento[idx]
            linea_nota_val = getattr(orig_obj, 'linea_nota', None) if hasattr(orig_obj, 'linea_nota') else orig_obj.get('linea_nota')

            m_val = base_data.get(li, {}).get('Matriz', 0.0)
            f_val = base_data.get(li, {}).get('Filial', 0.0)
            saldo_base = m_val + f_val
            
            resolved_lines.append({
                'linea_item': li,
                'linea_nota': linea_nota_val,
                'debe_calculado': debe_calc,
                'haber_calculado': haber_calc,
                'elimina_saldo_total': item['elimina_saldo_total'],
                'saldo_base': saldo_base
            })
            
        return resolved_lines
    finally:
        if is_local_db:
            db.close()

def generar_hoja_trabajo(grupo_id: int, periodo: str):
    """
    Genera el DataFrame consolidado cruzando los saldos base de Matriz y Filial,
    y aplicando los asientos formales de ajuste.
    Soporta sub-grupos recursivamente si la filial es otro grupo de consolidación.
    """
    db = SessionLocal()
    try:
        # 1. Obtener info del grupo
        grupo = db.query(ConsolidationGroup).filter_by(id=grupo_id).first()
        if not grupo:
            return None, "Grupo no encontrado."
            
        # Determinar nombre a mostrar de la filial
        filial_display_name = grupo.empresa_filial
        if grupo.filial_is_group:
            sub_grupo_id = int(grupo.empresa_filial)
            sub_grupo_obj = db.query(ConsolidationGroup).filter_by(id=sub_grupo_id).first()
            if sub_grupo_obj:
                filial_display_name = f"Consolidado {sub_grupo_obj.nombre_grupo}"
            
        # 2. Obtener saldos base cruzados
        base_data = obtener_saldos_base(grupo_id, periodo, db)
        if not base_data:
            return None, "No se encontraron saldos base (Matriz ni Filial) para este periodo."
            
        # Obtener todas las líneas de taxonomía definidas para las compañías de este grupo (incluyendo recursión)
        entidades_grupo = [grupo.empresa_matriz]
        if grupo.filial_is_group:
            def get_subgroup_companies(sub_g_id):
                comps = []
                sub_g = db.query(ConsolidationGroup).filter_by(id=sub_g_id).first()
                if sub_g:
                    comps.append(sub_g.empresa_matriz)
                    if sub_g.filial_is_group:
                        comps.extend(get_subgroup_companies(int(sub_g.empresa_filial)))
                    else:
                        comps.append(sub_g.empresa_filial)
                return comps
            entidades_grupo.extend(get_subgroup_companies(int(grupo.empresa_filial)))
        else:
            entidades_grupo.append(grupo.empresa_filial)
            
        tax_names = db.query(TaxonomyMasterRecord.nombre_linea_es).filter(
            TaxonomyMasterRecord.empresa.in_(entidades_grupo),
            TaxonomyMasterRecord.reporte_destino.in_(['Balance', 'P&L'])
        ).distinct().all()
        
        for t in tax_names:
            line_name = t[0]
            if line_name not in base_data:
                base_data[line_name] = {'Matriz': 0.0, 'Filial': 0.0}
            
        if not base_data:
            return None, "No se encontraron saldos base (Matriz ni Filial) para este periodo."
            
        # 3. Leer Asientos (del periodo + recurrentes de periodos anteriores)
        asientos = db.query(ConsolidationJournalEntry).filter(
            ConsolidationJournalEntry.grupo_id == grupo_id,
            or_(
                ConsolidationJournalEntry.periodo == periodo,
                (ConsolidationJournalEntry.periodo < periodo) & (ConsolidationJournalEntry.es_recurrente == True)
            )
        ).all()
        
        # Agrupar los asientos por columna y dentro de ella por comprobante (periodo, glosa) para poder balancear los dinámicos cronológicamente
        from collections import defaultdict
        col_to_vouchers = defaultdict(lambda: defaultdict(list))
        for a in asientos:
            code = getattr(a, 'asiento_codigo', None)
            voucher_key = code if code else (a.periodo, a.glosa)
            col_to_vouchers[a.columna_ajuste][voucher_key].append(a)
            
        # Determinar el orden en que se procesarán las columnas
        todas_columnas_con_datos = set(a.columna_ajuste for a in asientos)
        columnas_a_procesar = [col for col in COLUMNAS_ORDENADAS if col in todas_columnas_con_datos]
        # Agregar columnas no contempladas en el orden estándar al final
        for col in todas_columnas_con_datos:
            if col not in columnas_a_procesar:
                columnas_a_procesar.append(col)
                
        # Crear copia de base_data para ir acumulando los ajustes de forma cronológica
        import copy
        base_data_acumulado = copy.deepcopy(base_data)
        
        # Extraer columnas de ajuste dinámicamente y sumar Debe - Haber
        ajustes_data = {} # {linea_item: {columna_ajuste: neto}}
        columnas_ajuste = set()
        
        for col in columnas_a_procesar:
            columnas_ajuste.add(col)
            vouchers_in_col = col_to_vouchers[col]
            
            for v_key, lines in vouchers_in_col.items():
                netos_comprobante = {} # a.id -> neto
                sum_total = 0.0
                static_lines_zero = [] # lineas estaticas con debe=0 y haber=0
                static_lines_other = [] # otras lineas estaticas
                dynamic_lines = [] # lineas dinamicas
                dynamic_processed_lines = set()
                
                for a in lines:
                    li = a.linea_item
                    
                    if getattr(a, 'elimina_saldo_total', False):
                        if li in dynamic_processed_lines:
                            # Duplicado accidental en el borrador, ignorar para no duplicar el monto
                            neto = 0.0
                        else:
                            # Obtener saldos acumulados hasta el momento
                            monto_matriz = base_data_acumulado.get(li, {}).get('Matriz', 0.0)
                            monto_filial = base_data_acumulado.get(li, {}).get('Filial', 0.0)
                            neto = -(monto_matriz + monto_filial)
                            dynamic_processed_lines.add(li)
                        netos_comprobante[a.id] = neto
                        sum_total += neto
                        dynamic_lines.append(a)
                    else:
                        debe_val = float(a.debe) if a.debe is not None else 0.0
                        haber_val = float(a.haber) if a.haber is not None else 0.0
                        neto = debe_val - haber_val
                        netos_comprobante[a.id] = neto
                        sum_total += neto
                        if debe_val == 0.0 and haber_val == 0.0:
                            static_lines_zero.append(a)
                        else:
                            static_lines_other.append(a)
                
                # Si el comprobante está descuadrado (por ejemplo, porque tiene líneas dinámicas o diferencias IC)
                if abs(sum_total) > 0.001:
                    # Buscar la mejor línea para absorber la diferencia y cuadrar el asiento
                    if static_lines_zero:
                        balancer_entry = static_lines_zero[0]
                        netos_comprobante[balancer_entry.id] -= sum_total
                    elif static_lines_other:
                        balancer_entry = static_lines_other[0]
                        netos_comprobante[balancer_entry.id] -= sum_total
                    elif dynamic_lines:
                        # Si no hay estáticas, usamos la primera dinámica para absorber la diferencia y cuadrar la columna
                        balancer_entry = dynamic_lines[0]
                        netos_comprobante[balancer_entry.id] -= sum_total
                        
                # Aplicar los netos calculados y balanceados
                for a in lines:
                    li = a.linea_item
                    neto = netos_comprobante.get(a.id, 0.0)
                    
                    # 1. Registrar en la matriz de ajustes de la hoja
                    if li not in ajustes_data:
                        ajustes_data[li] = {}
                    ajustes_data[li][col] = ajustes_data[li].get(col, 0.0) + neto
                    
                    # 2. Acumular en base_data_acumulado para las siguientes columnas/comprobantes
                    es_prior_year_static_pl = (a.periodo[:4] < periodo[:4]) and (not getattr(a, 'elimina_saldo_total', False)) and is_pl_account(li)
                    if es_prior_year_static_pl:
                        target_eq = get_target_equity_line(base_data_acumulado.keys())
                        if target_eq not in base_data_acumulado:
                            base_data_acumulado[target_eq] = {'Matriz': 0.0, 'Filial': 0.0}
                        base_data_acumulado[target_eq]['Filial'] += neto
                    else:
                        if li not in base_data_acumulado:
                            base_data_acumulado[li] = {'Matriz': 0.0, 'Filial': 0.0}
                        base_data_acumulado[li]['Filial'] += neto
            
        # 4. Construir DataFrame final
        todas_lineas = set(base_data.keys()).union(set(ajustes_data.keys()))
        # Ordenar columnas de ajuste según la secuencia COLUMNAS_ORDENADAS
        columnas_ajuste_lista = [col for col in COLUMNAS_ORDENADAS if col in columnas_ajuste]
        # Agregar columnas no estándar
        for col in columnas_ajuste:
            if col not in columnas_ajuste_lista:
                columnas_ajuste_lista.append(col)
                
        if not columnas_ajuste_lista:
            columnas_ajuste_lista = ["Ajustes y Eliminaciones"]
        
        def _get_norm(x):
            return x.lower().strip().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')

        template_order = {}
        import os
        template_path = os.path.join("data", "empresas", f"[GRUPO] {grupo.nombre_grupo}", "Balance clasificado.xlsx")
        if os.path.exists(template_path):
            try:
                import openpyxl
                wb_t = openpyxl.load_workbook(template_path, data_only=True)
                ws_t = wb_t.active
                for r in range(1, ws_t.max_row + 1):
                    val = ws_t.cell(row=r, column=2).value
                    if val and isinstance(val, str):
                        norm_val = _get_norm(val)
                        if norm_val and norm_val not in template_order:
                            template_order[norm_val] = len(template_order)
            except:
                pass
        
        template_path_er = os.path.join("data", "empresas", f"[GRUPO] {grupo.nombre_grupo}", "Estado de Resultados Clasificados.xlsx")
        if os.path.exists(template_path_er):
            try:
                import openpyxl
                wb_t = openpyxl.load_workbook(template_path_er, data_only=True)
                ws_t = wb_t.active
                for r in range(1, ws_t.max_row + 1):
                    val_a = ws_t.cell(row=r, column=1).value
                    val_b = ws_t.cell(row=r, column=2).value
                    for val in [val_a, val_b]:
                        if val and isinstance(val, str):
                            norm_val = _get_norm(val)
                            if norm_val and norm_val not in template_order:
                                template_order[norm_val] = len(template_order)
            except:
                pass

        # Cargar todos los registros históricos de este grupo para asociar línea con reporte destino
        line_to_reporte = {}
        historicos_all = db.query(
            HistoricalDataRecord.linea_item, 
            HistoricalDataRecord.reporte
        ).filter(HistoricalDataRecord.empresa.in_(entidades_grupo)).distinct().all()
        for li_name, rep_name in historicos_all:
            line_to_reporte[_get_norm(li_name)] = rep_name

        # Cargar la taxonomía de este grupo para mapear nombres a reporte y obtener su id_reporte
        taxonomy_all = db.query(
            TaxonomyMasterRecord.nombre_linea_es,
            TaxonomyMasterRecord.reporte_destino,
            TaxonomyMasterRecord.id_reporte
        ).filter(
            TaxonomyMasterRecord.empresa.in_(entidades_grupo),
            TaxonomyMasterRecord.reporte_destino.in_(['Balance', 'P&L'])
        ).all()

        line_to_taxonomy = {}  # norm_name -> (reporte_destino, id_reporte)
        for nombre, destino, id_rep in taxonomy_all:
            norm_name = _get_norm(nombre)
            if norm_name not in line_to_taxonomy:
                line_to_taxonomy[norm_name] = (destino, id_rep)

        # Construir order_dict_full de forma dinámica ordenando los registros de taxonomía por su id_reporte
        def get_id_reporte_sort_key(id_rep):
            if not id_rep:
                return ()
            parts = id_rep.split('_')
            key_parts = []
            for part in parts:
                if part.isdigit():
                    key_parts.append((0, int(part)))
                else:
                    key_parts.append((1, part))
            return tuple(key_parts)

        # Ordenar registros de taxonomía
        recs_sorted = sorted(taxonomy_all, key=lambda r: get_id_reporte_sort_key(r.id_reporte))
        order_dict_full = {}
        for r in recs_sorted:
            norm_name = _get_norm(r.nombre_linea_es)
            if norm_name not in order_dict_full:
                order_dict_full[norm_name] = len(order_dict_full)

        def classify_single_item(li):
            norm = _get_norm(li)
            
            # Obtener tipo de reporte y id_reporte desde BD
            reporte_tipo = None
            id_rep = None
            
            if norm in line_to_reporte:
                reporte_tipo = line_to_reporte[norm]
            elif norm in line_to_taxonomy:
                reporte_tipo = line_to_taxonomy[norm][0]
                
            if norm in line_to_taxonomy:
                id_rep = line_to_taxonomy[norm][1]
                
            # Heurísticas fuertes para clasificar como P&L (Estado de Resultados)
            is_pl = False
            
            # 1. Si contiene "ingreso" o "ingresos"
            if "ingreso" in norm:
                # Excluir si es un pasivo de ingresos diferidos / anticipados o cuentas por pagar
                if not any(x in norm for x in ["diferido", "diferidos", "anticipado", "anticipados", "por pagar"]):
                    is_pl = True
                    
            # 2. Si contiene "gasto", "gastos", "costo", "costos", "egreso", "egresos"
            if any(x in norm for x in ["gasto", "costo", "egreso"]):
                # Excluir si es un activo o pasivo por gastos diferidos / anticipados / por pagar
                if not any(x in norm for x in ["anticipado", "anticipados", "diferido", "diferidos", "por pagar"]):
                    is_pl = True
                    
            # 3. Si contiene "depreciacion" o "amortizacion"
            if "depreciacion" in norm or "amortizacion" in norm:
                # Excluir si es depreciación acumulada (que va en Activo como menor valor)
                if "acumulada" not in norm:
                    is_pl = True
                    
            # 4. Si contiene "diferencia de cambio" o "diferencias de cambio"
            if "diferencia de cambio" in norm or "diferencias de cambio" in norm:
                is_pl = True
                
            # 5. Si contiene "unidad de reajuste" o "unidades de reajuste"
            if "unidad de reajuste" in norm or "unidades de reajuste" in norm:
                is_pl = True
                
            # 6. Si contiene "resultado" o "resultados"
            if "resultado" in norm or "resultados" in norm:
                # Excluir "resultados acumulados", "resultado acumulado" (que son patrimonio)
                if not any(x in norm for x in ["acumulado", "acumulados"]):
                    is_pl = True
                    
            # 7. Si contiene "impuesto" o "impuestos"
            if "impuesto" in norm or "impuestos" in norm:
                # Excluir cuentas de activos o pasivos por impuestos (corrientes, diferidos, por pagar, etc.)
                if not any(x in norm for x in ["por cobrar", "por pagar", "diferido", "diferidos", "credito", "debito", "corriente", "corrientes"]):
                    is_pl = True
                    
            # 8. Si contiene "ganancia", "ganancias", "perdida", "perdidas"
            if "ganancia" in norm or "ganancias" in norm or "perdida" in norm or "perdidas" in norm:
                # Excluir resultados acumulados o ganancias acumuladas
                if not any(x in norm for x in ["acumulado", "acumulados", "acumulada", "acumuladas", "ejercicio", "periodo"]):
                    is_pl = True
                    
            if is_pl or reporte_tipo == 'P&L':
                return "PL"
                
            # Es Balance: clasificar en PAT, AC, ANC, PC, PNC.
            # A. Patrimonio
            is_pat = False
            if id_rep and id_rep.startswith("PAT_"):
                is_pat = True
            elif any(kw in norm for kw in ["capital", "reserva", "aporte por enterar", "prima de emision", "acciones propias", "patrimonio", "controladora", "no controladora", "ejercicio", "periodo"]):
                is_pat = True
            elif "resultado" in norm and "acumulado" in norm:
                is_pat = True
            elif "ganancia" in norm and "acumulada" in norm:
                is_pat = True
            elif "perdida" in norm and "acumulada" in norm:
                is_pat = True
                
            if is_pat:
                return "PAT"
                
            # B. Determinar si es Pasivo
            is_pasivo = False
            if id_rep and id_rep.startswith("PAS_"):
                is_pasivo = True
            elif any(kw in norm for kw in ["pasivo", "cuenta por pagar", "cuentas por pagar", "provision", "obligacion", "beneficio"]):
                is_pasivo = True
                
            # C. Determinar si es Corriente o No Corriente
            is_non_current = False
            if any(kw in norm for kw in ["no corriente", "no corrientes"]):
                is_non_current = True
            elif any(kw in norm for kw in ["corriente", "corrientes"]):
                is_non_current = False
            else:
                # Fallback por palabras claves de activos/pasivos no corrientes
                if any(kw in norm for kw in ["diferido", "diferidos", "intangible", "plusvalia", "propiedades, planta", "propiedades, plantas", "propiedad de inversion", "propiedades de inversion", "inversion en", "inversiones contabilizadas", "derecho de uso", "derechos de uso"]):
                    is_non_current = True
                    
            if is_pasivo:
                return "PNC" if is_non_current else "PC"
            else:
                return "ANC" if is_non_current else "AC"

        filas = []
        
        def add_blank():
            filas.append({'Balance clasificado': ''})
            
        def add_title(title):
            filas.append({'Balance clasificado': title})
            
        def build_row(li):
            fila = {
                'Balance clasificado': li,
                f"{filial_display_name}": base_data.get(li, {}).get('Filial', 0.0),
                f"{grupo.empresa_matriz} Individual": base_data.get(li, {}).get('Matriz', 0.0)
            }
            total_ajustes = 0.0
            for col in columnas_ajuste_lista:
                val = ajustes_data.get(li, {}).get(col, 0.0)
                fila[col] = val
                total_ajustes += val
            fila['CONSOLIDADO'] = fila[f"{filial_display_name}"] + fila[f"{grupo.empresa_matriz} Individual"] + total_ajustes
            return fila
            
        def build_subtotal(title, rows_to_sum):
            sub_fila = {'Balance clasificado': title}
            cols_to_sum = [f"{filial_display_name}", f"{grupo.empresa_matriz} Individual"] + columnas_ajuste_lista + ['CONSOLIDADO']
            for col in cols_to_sum:
                sub_fila[col] = sum(r.get(col, 0.0) for r in rows_to_sum)
            return sub_fila

        # Identificar la línea de resultados acumulados en la taxonomía o crearla
        target_equity_line = None
        for kw in ["resultados acumulados", "ganancias acumuladas", "utilidades acumuladas"]:
            for li in todas_lineas:
                norm = _get_norm(li)
                if classify_single_item(li) == "PAT" and kw in norm:
                    target_equity_line = li
                    break
            if target_equity_line:
                break
        
        if not target_equity_line:
            for li in todas_lineas:
                if "resultados acumulados" in _get_norm(li):
                    target_equity_line = li
                    break
                    
        if not target_equity_line:
            target_equity_line = "Resultados acumulados"
            
        if target_equity_line not in base_data:
            base_data[target_equity_line] = {'Matriz': 0.0, 'Filial': 0.0}
        todas_lineas.add(target_equity_line)
        
        # 1. Redirección de Ajustes Recurrentes de Años Anteriores:
        # Si un ajuste recurrente viene de un año anterior, y su cuenta es de P&L,
        # su efecto acumulado debe ir a Patrimonio (Resultados Acumulados) del año actual
        # y NO contaminar las cuentas de P&L del año actual.
        for a in asientos:
            if a.periodo[:4] < periodo[:4]: # Año anterior
                li = a.linea_item
                col = a.columna_ajuste
                if classify_single_item(li) == "PL":
                    if getattr(a, 'elimina_saldo_total', False):
                        continue
                    # Calcular el neto de esta línea de ajuste
                    debe_val = float(a.debe) if a.debe is not None else 0.0
                    haber_val = float(a.haber) if a.haber is not None else 0.0
                    neto = debe_val - haber_val
                    
                    # Restar de la cuenta de P&L (donde se agregó en el loop inicial)
                    if li in ajustes_data and col in ajustes_data[li]:
                        ajustes_data[li][col] -= neto
                        if abs(ajustes_data[li][col]) < 0.001:
                            del ajustes_data[li][col]
                    
                    # Sumar a Resultados Acumulados
                    if target_equity_line not in ajustes_data:
                        ajustes_data[target_equity_line] = {}
                    ajustes_data[target_equity_line][col] = ajustes_data[target_equity_line].get(col, 0.0) + neto

        # Calcular el efecto neto en P&L para cada columna de ajuste y transferirlo a Patrimonio (Resultados Acumulados)
        for col in columnas_ajuste:
            pl_sum = 0.0
            for li in todas_lineas:
                if classify_single_item(li) == "PL":
                    pl_sum += ajustes_data.get(li, {}).get(col, 0.0)
            
            if pl_sum != 0.0:
                if target_equity_line not in ajustes_data:
                    ajustes_data[target_equity_line] = {}
                ajustes_data[target_equity_line][col] = ajustes_data[target_equity_line].get(col, 0.0) + pl_sum

        # Categorize input lines
        lines_ac = []
        lines_anc = []
        lines_pc = []
        lines_pnc = []
        lines_pat = []
        lines_pl = []
        lines_other = []
        
        for li in list(todas_lineas):
            cat = classify_single_item(li)
            if cat == "AC": lines_ac.append(li)
            elif cat == "ANC": lines_anc.append(li)
            elif cat == "PC": lines_pc.append(li)
            elif cat == "PNC": lines_pnc.append(li)
            elif cat == "PAT": lines_pat.append(li)
            elif cat == "PL": lines_pl.append(li)
            else: lines_other.append(li)
            
        # Helper to sort lines within categories based on standard keys
        def sort_cat(li_list):
            if template_order:
                return sorted(li_list, key=lambda x: (template_order.get(_get_norm(x), 9999), x))
            return sorted(li_list, key=lambda x: (order_dict_full.get(_get_norm(x), 9999), x))

        # Build Document
        # ACTIVOS
        add_title("Activos corrientes")
        rows_ac = [build_row(li) for li in sort_cat(lines_ac)]
        filas.extend(rows_ac)
        add_blank()
        sub_ac = build_subtotal("Activos corrientes totales", rows_ac)
        filas.append(sub_ac)
        add_blank()
        
        add_title("Activos no corrientes")
        rows_anc = [build_row(li) for li in sort_cat(lines_anc)]
        filas.extend(rows_anc)
        add_blank()
        sub_anc = build_subtotal("Activos no corrientes totales", rows_anc)
        filas.append(sub_anc)
        
        tot_activos = build_subtotal("Total activos", [sub_ac, sub_anc])
        filas.append(tot_activos)
        add_blank()
        
        # PASIVOS Y PATRIMONIO
        add_title("Patrimonio y pasivos")
        add_blank()
        
        add_title("Pasivos corrientes")
        rows_pc = [build_row(li) for li in sort_cat(lines_pc)]
        filas.extend(rows_pc)
        sub_pc = build_subtotal("Pasivo corrientes totales", rows_pc)
        filas.append(sub_pc)
        add_blank()
        
        add_title("Pasivos no corrientes")
        rows_pnc = [build_row(li) for li in sort_cat(lines_pnc)]
        filas.extend(rows_pnc)
        sub_pnc = build_subtotal("Pasivo no corrientes totales", rows_pnc)
        filas.append(sub_pnc)
        
        tot_pasivos = build_subtotal("Total pasivos", [sub_pc, sub_pnc])
        filas.append(tot_pasivos)
        add_blank()
        
        add_title("Patrimonio")
        rows_pat = [build_row(li) for li in sort_cat(lines_pat)]
        filas.extend(rows_pat)
        sub_pat = build_subtotal("Patrimonio total", rows_pat)
        filas.append(sub_pat)
        
        tot_pat_pas = build_subtotal("Total patrimonio y pasivos", [tot_pasivos, sub_pat])
        filas.append(tot_pat_pas)
        # P&L and Others
        add_blank()
        
        # Repetir cabecera para Estado de Resultados (separación visual clara similar al balance clasificado)
        er_header = {'Balance clasificado': 'Estado de Resultados'}
        cols_to_repeat = [f"{filial_display_name}", f"{grupo.empresa_matriz} Individual"] + columnas_ajuste_lista + ['CONSOLIDADO']
        for col in cols_to_repeat:
            er_header[col] = col
        filas.append(er_header)
        
        # Filtrar subtotales existentes para evitar duplicación
        subtotal_keywords = [
            "ganancia bruta", "ganancia antes de impuesto", "ganancia (perdida) antes de impuestos", 
            "resultado antes de impuestos", "ganancia (perdida)", "resultado final",
            "ganancia (perdida) de operaciones continuadas", "ganancias (perdida) del ejercicio",
            "ganancia (perdida) del ejercicio"
        ]
        clean_pl_lines = [li for li in lines_pl if _get_norm(li) not in subtotal_keywords]
        
        # Clasificar las cuentas del P&L en grupos para generar subtotales intermedios dinámicamente
        def is_group1(norm_name):
            if any(kw in norm_name for kw in ["ingreso de actividades ordinarias", "ingreso de arriendo", "ingresos de arriendo", "costo de ventas", "costos de ventas", "costo de venta", "costos de uso", "acceso a infraestructura", "acceso a fibra", "depreciacion operacional", "depreciación operacional", "depreciacion y amortizacion operacional", "depreciación y amortización operacional"]):
                return True
            if norm_name.startswith("ingreso") and not any(kw in norm_name for kw in ["financiero", "financieros", "otro", "otros"]):
                return True
            if "costo de" in norm_name and not any(kw in norm_name for kw in ["financiero", "financieros"]):
                return True
            return False

        def is_group3(norm_name):
            return "impuesto" in norm_name and any(kw in norm_name for kw in ["ganancia", "ganancias", "renta", "utilidad"])
        
        lines_g1 = []
        lines_g2 = []
        lines_g3 = []
        
        for li in clean_pl_lines:
            norm_li = _get_norm(li)
            if is_group1(norm_li):
                lines_g1.append(li)
            elif is_group3(norm_li):
                lines_g3.append(li)
            else:
                lines_g2.append(li)
                
        # Ordenar las líneas según el orden estándar visual y funcional
        def sort_pl_group(li_list):
            if template_order:
                return sorted(li_list, key=lambda x: (template_order.get(_get_norm(x), 9999), get_pl_sort_index(x)))
            return sorted(li_list, key=get_pl_sort_index)
            
        sorted_g1 = sort_pl_group(lines_g1)
        sorted_g2 = sort_pl_group(lines_g2)
        sorted_g3 = sort_pl_group(lines_g3)
        
        # Generar filas de datos
        rows_g1 = [build_row(li) for li in sorted_g1]
        rows_g2 = [build_row(li) for li in sorted_g2]
        rows_g3 = [build_row(li) for li in sorted_g3]
        
        # 1. Agregar Grupo 1 (Ingresos y Costos Operacionales)
        filas.extend(rows_g1)
        
        # 2. Agregar subtotal: Ganancia bruta
        if rows_g1:
            sub_gb = build_subtotal("Ganancia bruta", rows_g1)
            filas.append(sub_gb)
            
        # 3. Agregar Grupo 2 (Gastos de administración, financieros, etc.)
        filas.extend(rows_g2)
        
        # 4. Agregar subtotal: Resultado antes de impuestos (Grupo 1 + Grupo 2)
        if rows_g1 or rows_g2:
            sub_gai = build_subtotal("Resultado antes de impuestos", rows_g1 + rows_g2)
            filas.append(sub_gai)
            
        # 5. Agregar Grupo 3 (Impuestos)
        filas.extend(rows_g3)
        
        # 6. Agregar línea final: Ganancias (Pérdida) del Ejercicio (Grupo 1 + Grupo 2 + Group 3)
        if rows_g1 or rows_g2 or rows_g3:
            sub_ge = build_subtotal("Ganancias (Pérdida) del Ejercicio", rows_g1 + rows_g2 + rows_g3)
            filas.append(sub_ge)
        
        if lines_other:
            add_blank()
            add_title("Otros rubros no clasificados")
            rows_other = [build_row(li) for li in sorted(lines_other)]
            filas.extend(rows_other)
            
        df = pd.DataFrame(filas)
        if not df.empty:
            cols = ['Balance clasificado', f"{filial_display_name}", f"{grupo.empresa_matriz} Individual"] + columnas_ajuste_lista + ['CONSOLIDADO']
            df = df[cols]
        
        return df, "Generado correctamente"
    except Exception as e:
        return None, str(e)
    finally:
        db.close()
