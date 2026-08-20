import pandas as pd
import io

class ExcelExportEngine:
    """
    (Req 13) Genera Papeles de Trabajo (Working Papers) en Excel.
    Recrea la lógica de múltiples hojas: EEFF principal y desglose por Notas.
    """
    def __init__(self, mapped_df, rules_config):
        self.mapped_df = mapped_df
        self.rules_config = rules_config

    def generate_working_papers(self):
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Hoja Principal: EEFF
            eeff_data = []
            for line_item, accounts in self.rules_config.items():
                if "Total" in line_item:
                    continue # Salto los totales para simplificar de momento
                
                # Sumamos el saldo de las cuentas mapeadas
                subset = self.mapped_df[self.mapped_df['cuenta_id'].isin(accounts)]
                saldo = subset['saldo_final'].sum()
                eeff_data.append({"Rubro EEFF": line_item, "Saldo M$ (Actual)": saldo})
            
            eeff_df = pd.DataFrame(eeff_data)
            eeff_df.to_excel(writer, sheet_name="EEFF", index=False)
            
            # Obtener el objeto de la hoja para formateo básico
            workbook = writer.book
            worksheet = writer.sheets["EEFF"]
            
            # Ajuste de ancho de columnas
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 20
            
            # 2. Hojas de Notas (Working Papers Descendentes)
            for line_item, accounts in self.rules_config.items():
                if "Total" in line_item:
                    continue
                
                # Extraemos nombre limpio para la hoja
                sheet_name = line_item.split(":")[-1].strip()
                sheet_name = sheet_name[:31] # Límite de openpyxl
                
                subset = self.mapped_df[self.mapped_df['cuenta_id'].isin(accounts)]
                if not subset.empty:
                    nota_df = subset[['cuenta_id', 'descripcion', 'saldo_final']].rename(
                        columns={'cuenta_id': 'Código', 'descripcion': 'Detalle de la Nota', 'saldo_final': 'Saldo M$'}
                    )
                    nota_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    ws_nota = writer.sheets[sheet_name]
                    ws_nota.column_dimensions['A'].width = 15
                    ws_nota.column_dimensions['B'].width = 50
                    ws_nota.column_dimensions['C'].width = 20
                    
        return output.getvalue()

def generate_excel_report(df, title="Reporte Financiero", subtitle="Expresado en pesos"):
    """
    Función helper genérica para exportar cualquier DataFrame 
    a Excel con un estilo corporativo (azul/blanco).
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Reporte", index=False, startrow=4)
        worksheet = writer.sheets["Reporte"]
        
        # Inserción de Cabeceras
        worksheet['A1'] = "INFORMES PRO - IFRS16"
        worksheet['A2'] = title
        worksheet['A3'] = subtitle
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Estilos corporativos base
        blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=14, bold=True, color="1F4E78")
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
                           
        # Aplicar Título
        worksheet['A1'].font = bold_font
        worksheet['A2'].font = title_font
        worksheet['A3'].font = Font(italic=True, color="7F7F7F")
        
        # Formatear la tabla que empieza en fila 5
        # Ancho automático heurístico
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5)
            # Cap top width para descripciones muy largas
            if adjusted_width > 60: adjusted_width = 60
            if adjusted_width < 15: adjusted_width = 15
            worksheet.column_dimensions[column].width = adjusted_width

        # Pintar cabeceras de la tabla (Fila 5)
        for cell in worksheet[5]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Poner bordes al resto de la tabla
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                
                # Alineación de números
                if type(cell.value) in [int, float]:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = r'_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

    output.seek(0)
    return output.getvalue()
