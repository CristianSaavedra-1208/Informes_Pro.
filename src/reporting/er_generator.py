import pandas as pd
import io
import openpyxl
import unicodedata
import re
import datetime
from copy import copy
from io import BytesIO

def format_period_to_spanish_date(period_str):
    if not period_str:
        return ""
    try:
        import calendar
        parts = str(period_str).strip().split('-')
        if len(parts) >= 2:
            year = int(parts[0])
            month = int(parts[1])
            last_day = calendar.monthrange(year, month)[1]
            months_es = {
                1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
                5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
                9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
            }
            month_name = months_es[month]
            return f"{last_day} de {month_name} {year}"
    except Exception:
        pass
    return period_str

class ERGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

    def generate(self, pl_df, scale_factor=1.0, pl_df_comp=None, periodo_actual_str=None, periodo_comp_str=None):
        # 1. Cargar plantilla base con openpyxl para preservar todo el formato original
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active

        # 2. Detección dinámica de columnas en la plantilla
        # Encontrar columna de conceptos/clasificación
        clasif_col_idx = 1
        for col in range(1, 10):
            for row in range(1, 15):
                val = ws.cell(row=row, column=col).value
                if val and any(x in str(val).lower() for x in ["ingresos de actividades", "costo de ventas", "ganancia bruta", "resultado antes"]):
                    clasif_col_idx = col
                    break
            else:
                continue
            break
            
        # Encontrar columna de Notas
        nota_col_idx = None
        for col in range(1, 10):
            if col == clasif_col_idx:
                continue
            for row in range(1, 5):
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().lower() == "nota":
                    nota_col_idx = col
                    break

        # Detectar columnas de fecha (años o periodos)
        date_cols = []
        for col in range(1, ws.max_column + 1):
            if col == clasif_col_idx or col == nota_col_idx:
                continue
            for row in range(1, 5):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    is_date = (
                        isinstance(val, (datetime.datetime, datetime.date)) or
                        (isinstance(val, str) and re.search(r'20\d{2}', val)) or
                        (isinstance(val, str) and any(x in val.lower() for x in ["actual", "anterior", "auditado", "comparat", "comp"])) or
                        (isinstance(val, (int, float)) and 2000 <= val <= 2100)
                    )
                    if is_date:
                        date_cols.append(col)
                        break

        date_cols = sorted(list(set(date_cols)))

        # Ordenar por año de mayor a menor (ej. 2025 primero, luego 2024)
        def get_year_from_cell_val(col_idx):
            for row in range(1, 5):
                val = ws.cell(row=row, column=col_idx).value
                if val is not None:
                    match = re.search(r'20\d{2}', str(val))
                    if match: return int(match.group(0))
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        return val.year
                    if str(val).isdigit(): return int(val)
            return 0

        date_cols = sorted(date_cols, key=get_year_from_cell_val, reverse=True)

        val25_col_idx = date_cols[0] if len(date_cols) >= 1 else 3
        val24_col_idx = date_cols[1] if len(date_cols) >= 2 else 4

        # 3. Limpiar y actualizar los encabezados de fechas
        if periodo_actual_str:
            formatted_actual = format_period_to_spanish_date(periodo_actual_str)
            formatted_comp = format_period_to_spanish_date(periodo_comp_str) if periodo_comp_str else ""
            
            for row in range(1, 5):
                cell_25 = ws.cell(row=row, column=val25_col_idx)
                cell_24 = ws.cell(row=row, column=val24_col_idx)
                
                # Reemplazar si el valor es fecha o año
                if cell_25.value is not None:
                    is_date_or_year = (
                        isinstance(cell_25.value, (datetime.datetime, datetime.date)) or
                        (isinstance(cell_25.value, str) and ("20" in cell_25.value or "Actual" in cell_25.value)) or
                        (isinstance(cell_25.value, (int, float)) and 2000 <= cell_25.value <= 2100)
                    )
                    if is_date_or_year:
                        cell_25.value = formatted_actual
                        
                if cell_24.value is not None:
                    is_date_or_year_comp = (
                        isinstance(cell_24.value, (datetime.datetime, datetime.date)) or
                        (isinstance(cell_24.value, str) and ("20" in cell_24.value or "Anterior" in cell_24.value)) or
                        (isinstance(cell_24.value, (int, float)) and 2000 <= cell_24.value <= 2100)
                    )
                    if is_date_or_year_comp:
                        cell_24.value = formatted_comp if periodo_comp_str else ""

        # 4. Funciones de normalización de strings
        def sanitize(text):
            if pd.isna(text): return ""
            clean_str = str(text).strip().lower()
            clean_str = ''.join(c for c in unicodedata.normalize('NFD', clean_str) if unicodedata.category(c) != 'Mn')
            clean_str = re.sub(r'\s+', ' ', clean_str)
            return clean_str

        def extraer_saldos(df_target):
            res = {}
            if df_target is None or df_target.empty:
                return res
            ignorar_bases = {
                "n° de cuenta", "nombre de la cuenta", "cuenta", "nombre", "unnamed: 0",
                "saldo_inicial", "debitos", "creditos", "saldo_final",
                "id_reporte", "id_nota_asociada", "clasificacion balance", "clasificación balance",
                "nota 1", "nota 2", "clasificacion flujo efectivo", "clasificación flujo efectivo",
                "clasificacion", "clasificación"
            }
            for col in df_target.columns:
                col_clean = sanitize(col)
                if col_clean not in ignorar_bases:
                    temp_series = pd.to_numeric(df_target[col], errors='coerce').fillna(0)
                    res[col_clean] = (temp_series.sum() * -1) / scale_factor
            return res

        # Collect all template classifications to avoid synonym collisions
        template_classifications = set()
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=clasif_col_idx).value
            if val:
                template_classifications.add(sanitize(val))

        resultados_2025 = extraer_saldos(pl_df)
        resultados_comp = extraer_saldos(pl_df_comp)

        synonyms = {
            "depreciacion operacional": "depreciacion y amortizacion operacional",
            "costos de uso fibra optica": "acceso a infraestructura fibra optica",
            "resultado por unidad de reajuste": "resultados por unidades de reajuste",
            "resultados por unidad de reajuste": "resultado por unidad de reajuste",
            "resultados por unidades de reajuste": "resultado por unidad de reajuste",
            "ganancia (perdida) por impuesto a las ganancias": "resultado por impuestos a las ganancias",
            "resultado por impuestos a las ganancias": "ganancia (perdida) por impuesto a las ganancias",
            "ingresos financieros ic": "ingresos financieros",
            "ingresos financieros con empresas relacionadas": "ingresos financieros",
            "costos financieros ic": "costos financieros",
            "costos financieros con empresas relacionadas": "costos financieros"
        }

        def apply_synonyms(d, template_classifications):
            new_d = d.copy()
            for key, val in d.items():
                if key in synonyms:
                    syn_target = synonyms[key]
                    if key in template_classifications:
                        continue
                    elif syn_target in template_classifications:
                        new_d[syn_target] = new_d.get(syn_target, 0.0) + val
                        if key in new_d:
                            del new_d[key]
            return new_d

        resultados_2025 = apply_synonyms(resultados_2025, template_classifications)
        resultados_comp = apply_synonyms(resultados_comp, template_classifications)

        # 5. Inyección en Cascada (Waterfall) preservando formatos
        rolling_sum_25 = 0.0
        rolling_sum_24 = 0.0

        for row in range(1, ws.max_row + 1):
            cell_name = ws.cell(row=row, column=clasif_col_idx).value
            if cell_name and isinstance(cell_name, str):
                cell_name_clean = cell_name.strip()
                clasif_secreta = sanitize(cell_name)

                # A) Cuenta transaccional
                if (clasif_secreta in resultados_2025) or (resultados_comp and clasif_secreta in resultados_comp):
                    val_25 = resultados_2025.get(clasif_secreta, 0.0)
                    if pd.isna(val_25) or val_25 is None:
                        val_25 = 0.0
                    ws.cell(row=row, column=val25_col_idx).value = val_25
                    rolling_sum_25 += val_25

                    if resultados_comp:
                        val_24 = resultados_comp.get(clasif_secreta, 0.0)
                        if pd.isna(val_24) or val_24 is None:
                            val_24 = 0.0
                        ws.cell(row=row, column=val24_col_idx).value = val_24
                        rolling_sum_24 += val_24
                    else:
                        ws.cell(row=row, column=val24_col_idx).value = 0.0

                # B) Subtotal
                else:
                    def is_subtotal_row(label):
                        if not label:
                            return False
                        norm = sanitize(label)
                        subtotals = {
                            "ganancia bruta", 
                            "ganancia antes de impuesto",
                            "ganancia antes de impuestos",
                            "resultado antes de impuestos",
                            "resultado antes de impuesto",
                            "ganancias (perdida) del ejercicio",
                            "ganancia (perdida) del ejercicio",
                            "(perdida) procedente de operaciones continuadas",
                            "perdida procedente de operaciones continuadas",
                            "ganancia procedente de operaciones continuadas",
                            "ganancia (perdida) procedente de operaciones continuadas",
                            "perdida",
                            "ganancia",
                            "total"
                        }
                        if norm in subtotals:
                            return True
                        is_sub = any(word in norm for word in ["ganancia", "perdida", "antes de impuesto", "antes del impuesto"])
                        if is_sub:
                            if any(term in norm for term in ["inversion", "unidad de reajuste", "impuestos a las ganancias", "impuesto a las ganancias", "arriendo"]):
                                return False
                            return True
                        return False

                    es_subtotal = is_subtotal_row(cell_name)
                    if es_subtotal and clasif_secreta not in resultados_2025 and (not resultados_comp or clasif_secreta not in resultados_comp):
                        ws.cell(row=row, column=val25_col_idx).value = rolling_sum_25
                        if resultados_comp:
                            ws.cell(row=row, column=val24_col_idx).value = rolling_sum_24
                        else:
                            ws.cell(row=row, column=val24_col_idx).value = 0.0

        # Para compatibilidad con la UI de visualización que espera (output, template_df)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Evaluar fórmulas de Excel agregadas en la plantilla (ej. Resultado por acción)
        try:
            from src.ui_pages.informes_y_notas import evaluate_formulas_in_workbook
            output = evaluate_formulas_in_workbook(output)
        except Exception as e:
            output.seek(0)

        # Leer de vuelta en un DataFrame para previsualización en la UI
        try:
            template_df = pd.read_excel(BytesIO(output.getvalue()), header=0)
            def _normalize_col_name(col):
                s = str(col).strip()
                if re.match(r'^\d+\.0$', s):
                    s = s[:-2]
                return s
            template_df.columns = [_normalize_col_name(c) for c in template_df.columns]
        except Exception:
            template_df = pd.DataFrame()

        return output, template_df
