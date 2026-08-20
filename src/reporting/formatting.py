import pandas as pd

def apply_corporate_style(df, excel_bytes=None, sheet_name=None, col_indices=None):
    import openpyxl
    from io import BytesIO

    df = df.copy()
    if len(df.columns) > 0:
        df.iloc[:, 0] = df.iloc[:, 0].fillna("").astype(str).replace({"nan": "", "None": ""})

    for col in df.columns[1:]:
        if df[col].dtype == 'object':
            try:
                converted = pd.to_numeric(df[col], errors='ignore')
                df[col] = converted
            except:
                pass

    # Definir columnas de valores (numéricas) de forma robusta
    non_numeric_names = {'balance clasificado', 'nota', 'notas', 'descripcion', 'concepto', 'cuenta_id', 'glosa'}
    numeric_cols = [
        c for c in df.columns[1:]
        if str(c).lower().strip() not in non_numeric_names
    ]
    text_cols = [c for c in df.columns if c not in numeric_cols]
    
    def format_accounting(x):
        if pd.isna(x) or str(x).lower() == 'nan':
            return ""
        
        # If it's already a number
        if isinstance(x, (int, float)):
            if x == 0: return "-"
            if x < 0: return f"({abs(x):,.0f})".replace(',', '.')
            return f"{x:,.0f}".replace(',', '.')
            
        # Try to cast string numbers
        try:
            num = float(str(x).replace(',', ''))
            if pd.isna(num): return ""
            if num == 0: return "-"
            if num < 0: return f"({abs(num):,.0f})".replace(',', '.')
            return f"{num:,.0f}".replace(',', '.')
        except:
            # If it's text, date, or "M$"
            if "00:00:00" in str(x):
                return str(x).split(" ")[0] # Clean up dates
            return str(x)
        
    format_dict = {
        col: format_accounting
        for col in df.columns[1:]
    }
    
    styler = df.style.format(format_dict)
    
    table_styles = [
        {
            'selector': '',
            'props': [
                ('min-width', '100% !important'),
                ('width', 'auto !important'),
                ('table-layout', 'fixed !important'),
                ('border-collapse', 'collapse !important'),
                ('margin-left', '0px !important'),
                ('margin-right', 'auto !important')
            ]
        },
        {
            'selector': 'th',
            'props': [
                ('background-color', '#FFFFFF'),
                ('color', '#000000'),
                ('font-weight', 'bold'),
                ('font-family', 'Inter, sans-serif'),
                ('border', 'none'),
                ('padding', '8px'),
                ('border-top', '2px solid #000000'),
                ('border-bottom', '2px solid #000000')
            ]
        },
        {
            'selector': 'th:first-child, td:first-child',
            'props': [
                ('min-width', '250px !important'),
                ('width', 'auto !important'),
                ('white-space', 'normal !important')
            ]
        },
        {
            'selector': 'th:not(:first-child), td:not(:first-child)',
            'props': [
                ('white-space', 'nowrap !important'),
                ('width', '150px !important'),
                ('min-width', '150px !important'),
                ('max-width', '250px !important')
            ]
        },
        {
            'selector': 'td',
            'props': [
                ('font-family', 'Inter, sans-serif'),
                ('font-size', '0.85em'),
                ('border-left', 'none'),
                ('border-right', 'none'),
                ('border-top', 'none'),
                ('border-bottom', '1px solid #f0f0f0'),
                ('padding', '6px 10px'),
                ('background-clip', 'padding-box')
            ]
        },
        {
            'selector': 'tr:nth-child(even) td',
            'props': [('background-color', '#e3f0fe')]
        },
        {
            'selector': 'tr:nth-child(odd) td',
            'props': [('background-color', '#FFFFFF')]
        }
    ]
    
    styler = styler.set_table_styles(table_styles)
    styler = styler.hide(axis="index")
    
    for idx, col in enumerate(df.columns):
        align_val = 'left' if idx == 0 else 'right'
        styler = styler.set_properties(subset=[col], **{'text-align': align_val})

    dynamic_styled = False
    if excel_bytes is not None:
        try:
            if isinstance(excel_bytes, BytesIO):
                excel_bytes.seek(0)
                wb = openpyxl.load_workbook(excel_bytes, data_only=True)
            else:
                wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
                
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                
            def convert_border(border_side):
                if not border_side or border_side.style is None:
                    return None
                style = border_side.style
                if style == 'double':
                    css_style = 'double'
                    width = '3px'
                elif style in ('medium', 'mediumDashed', 'mediumDashDot', 'mediumDashDotDot'):
                    css_style = 'solid'
                    width = '2px'
                elif style == 'thick':
                    css_style = 'solid'
                    width = '3px'
                elif style in ('dashed', 'dashDot', 'dashDotDot'):
                    css_style = 'dashed'
                    width = '1px'
                elif style == 'dotted':
                    css_style = 'dotted'
                    width = '1px'
                else:
                    css_style = 'solid'
                    width = '1px'
                
                color_hex = '000000'
                if border_side.color and border_side.color.rgb:
                    c_rgb = str(border_side.color.rgb)
                    if len(c_rgb) == 8:
                        color_hex = c_rgb[2:]
                    elif len(c_rgb) == 6:
                        color_hex = c_rgb
                return f"{width} {css_style} #{color_hex}"

            df_cols = list(df.columns)
            ws_cols_mapping = {}
            if col_indices:
                ws_cols_mapping = {j: col_indices[j] for j in range(len(df_cols))}
            else:
                import datetime
                import re
                
                excel_name_col = 1
                for col in range(1, 11):
                    for row in range(1, 20):
                        val = ws.cell(row=row, column=col).value
                        if val and str(val).strip().lower() in ["concepto", "descripcion", "detalle", "flujos", "origen/aplicacion", "balance clasificado", "clasificacion", "activos"]:
                            excel_name_col = col
                            break
                    if excel_name_col > 1:
                        break
                        
                if excel_name_col == 1:
                    found = False
                    for col in range(1, 11):
                        for row in range(1, 25):
                            val = ws.cell(row=row, column=col).value
                            if val and isinstance(val, str) and len(val.strip()) > 3:
                                if not re.match(r'^\d+$', val.strip()):
                                    excel_name_col = col
                                    found = True
                                    break
                        if found:
                            break
                
                excel_date_cols = []
                for col in range(1, ws.max_column + 1):
                    for row in range(1, 15):
                        val = ws.cell(row=row, column=col).value
                        if val is not None:
                            is_date = (
                                isinstance(val, (datetime.datetime, datetime.date)) or
                                (isinstance(val, str) and re.search(r'20\d{2}', val))
                            )
                            if is_date:
                                excel_date_cols.append(col)
                                break
                excel_date_cols = sorted(list(set(excel_date_cols)))
                
                excel_nota_col = None
                for col in range(1, ws.max_column + 1):
                    for row in range(1, 15):
                        val = ws.cell(row=row, column=col).value
                        if val and str(val).strip().lower() == "nota":
                            excel_nota_col = col
                            break
                if excel_nota_col is None and excel_date_cols:
                    first_date_col = excel_date_cols[0]
                    if first_date_col > excel_name_col + 1:
                        excel_nota_col = excel_name_col + 1
                
                # Separar columnas del DataFrame en numéricas y no numéricas para alineación correcta
                non_numeric_df_cols = [0]
                for j in range(1, len(df_cols)):
                    col_name_lower = str(df_cols[j]).strip().lower()
                    if col_name_lower in ['nota', 'notas']:
                        non_numeric_df_cols.append(j)
                
                numeric_df_cols = [j for j in range(len(df_cols)) if j not in non_numeric_df_cols]

                # Mapear columnas no numéricas (Concepto, Moneda, Nota, etc.)
                excel_non_numeric_cols = [excel_name_col]
                if excel_nota_col is not None:
                    excel_non_numeric_cols.append(excel_nota_col)
                
                curr = excel_name_col + 1
                while len(excel_non_numeric_cols) < len(non_numeric_df_cols):
                    if curr not in excel_non_numeric_cols:
                        excel_non_numeric_cols.append(curr)
                    curr += 1
                
                for idx, j in enumerate(non_numeric_df_cols):
                    ws_cols_mapping[j] = excel_non_numeric_cols[idx]

                # Mapear columnas numéricas (Actual, Comparativa)
                excel_numeric_cols = list(excel_date_cols)
                curr = (excel_date_cols[-1] + 1) if excel_date_cols else (excel_name_col + 1)
                while len(excel_numeric_cols) < len(numeric_df_cols):
                    excel_numeric_cols.append(curr)
                    curr += 1
                
                for idx, j in enumerate(numeric_df_cols):
                    ws_cols_mapping[j] = excel_numeric_cols[idx]
            
            def clean_str(s):
                if s is None: return ""
                return str(s).replace("\xa0", " ").strip().lower()

            # Pre-construir mapa de búsqueda de filas en openpyxl
            row_map = {}
            for r in range(1, ws.max_row + 1):
                for col_idx in range(1, min(ws.max_column + 1, 6)):
                    val = ws.cell(row=r, column=col_idx).value
                    if val is not None:
                        s_val = clean_str(val)
                        if s_val and s_val not in row_map:
                            row_map[s_val] = r

            row_styles_matrix = pd.DataFrame('', index=df.index, columns=df.columns)
            row_top_borders = {}
            row_bottom_borders = {}
            
            for idx, df_row in df.iterrows():
                search_val = clean_str(df_row.iloc[0])
                found_row = row_map.get(search_val)
                
                if found_row is not None:
                    row_top_b = None
                    row_bottom_b = None
                    
                    for j, col_name in enumerate(df_cols):
                        excel_col = ws_cols_mapping.get(j, j + 1)
                        cell = ws.cell(row=found_row, column=excel_col)
                        
                        css_styles = []
                        
                        if cell.font:
                            if cell.font.bold:
                                css_styles.append("font-weight: bold")
                            else:
                                css_styles.append("font-weight: normal")
                            if cell.font.italic:
                                css_styles.append("font-style: italic")
                            if cell.font.size:
                                css_styles.append(f"font-size: {cell.font.size}pt")
                            if cell.font.name:
                                css_styles.append(f"font-family: '{cell.font.name}', sans-serif")
                        
                        if cell.alignment and cell.alignment.horizontal:
                            align = cell.alignment.horizontal
                            if align == 'center':
                                css_styles.append("text-align: center")
                            elif align == 'right':
                                css_styles.append("text-align: right")
                            elif align == 'left':
                                css_styles.append("text-align: left")
                                
                        if cell.fill and cell.fill.fill_type == 'solid' and cell.fill.fgColor:
                            c_rgb = str(cell.fill.fgColor.rgb)
                            if len(c_rgb) == 8 and c_rgb != '00000000' and not c_rgb.startswith('000000'):
                                css_styles.append(f"background-color: #{c_rgb[2:]}")
                            elif len(c_rgb) == 6:
                                css_styles.append(f"background-color: #{c_rgb}")
                        
                        top_b = None
                        bottom_b = None
                        if cell.border:
                            top_b = convert_border(cell.border.top)
                            bottom_b = convert_border(cell.border.bottom)
                        
                        if top_b:
                            css_styles.append(f"border-top: {top_b} !important")
                            if not row_top_b and top_b != 'none':
                                row_top_b = top_b
                        else:
                            css_styles.append("border-top: none")
                            
                        if bottom_b:
                            css_styles.append(f"border-bottom: {bottom_b} !important")
                            if not row_bottom_b and bottom_b not in ('none', '1px solid #f0f0f0'):
                                row_bottom_b = bottom_b
                        else:
                            css_styles.append("border-bottom: 1px solid #f0f0f0")
                        
                        if css_styles:
                            row_styles_matrix.loc[idx, col_name] = "; ".join(css_styles)
                            
                    if row_top_b:
                        row_top_borders[idx] = row_top_b
                    if row_bottom_b:
                        row_bottom_borders[idx] = row_bottom_b
            
            # Aplicar propagación de bordes horizontalmente en filas de totales
            for idx in row_styles_matrix.index:
                row_top_b = row_top_borders.get(idx)
                row_bottom_b = row_bottom_borders.get(idx)
                if row_top_b or row_bottom_b:
                    for col_name in df_cols:
                        style_str = row_styles_matrix.loc[idx, col_name]
                        if row_top_b:
                            style_str = re.sub(r'border-top:[^;]+', f'border-top: {row_top_b} !important', style_str)
                        if row_bottom_b:
                            style_str = re.sub(r'border-bottom:[^;]+', f'border-bottom: {row_bottom_b} !important', style_str)
                        row_styles_matrix.loc[idx, col_name] = style_str

            # Pasada de resolución de conflictos de colapso de bordes
            for i in range(len(df)):
                idx = df.index[i]
                row_top_b = row_top_borders.get(idx)
                row_bottom_b = row_bottom_borders.get(idx)
                if row_top_b and i > 0:
                    prev_idx = df.index[i - 1]
                    for col_name in df_cols:
                        row_styles_matrix.loc[prev_idx, col_name] += f"; border-bottom: {row_top_b} !important"
                if row_bottom_b and i < len(df) - 1:
                    next_idx = df.index[i + 1]
                    for col_name in df_cols:
                        row_styles_matrix.loc[next_idx, col_name] += f"; border-top: {row_bottom_b} !important"
            
            styler = styler.apply(lambda _: row_styles_matrix, axis=None)
            dynamic_styled = True
        except Exception as ex:
            import sys
            print(f"Error parsing Excel styles dynamically: {ex}", file=sys.stderr)
            dynamic_styled = False
            
    if not dynamic_styled:
        def highlight_totals(row):
            classif = str(row.iloc[0]).lower().strip()
            
            styles = [''] * len(row)
            if classif == "estado de resultados":
                for i in range(len(row)):
                    styles[i] = 'border-top: 2px solid #000000 !important; border-bottom: 2px solid #000000 !important; font-weight: bold; color: #000000; background-color: #FFFFFF;'
                return styles

            is_bottom_line = (
                "ganancias (pérdida) del ejercicio" in classif or
                "ganancia (pérdida) del ejercicio" in classif or
                "resultado final" in classif or
                classif == "ganancia (pérdida)"
            )
            
            if is_bottom_line:
                for i, col_name in enumerate(row.index):
                    if col_name in numeric_cols:
                        styles[i] = 'border-top: 1px solid #000000 !important; border-bottom: 3px double #000000 !important; font-weight: bold; color: #000000;'
                    elif i == 0:
                        styles[i] = 'font-weight: bold; color: #000000; border-top: 1px solid #000000 !important; border-bottom: 3px double #000000 !important;'
                return styles
                
            is_total = (
                "total" in classif or 
                "bruta" in classif or 
                "antes de" in classif or 
                "antes del" in classif or 
                "incremento" in classif or
                ("ganancia" in classif and "acumulada" not in classif) or
                ("pérdida" in classif and "acumulada" not in classif and "operacionales" not in classif and "operaciones continuadas" in classif) or
                "saldo final" in classif
            )
            
            if is_total:
                for i, col_name in enumerate(row.index):
                    if col_name in numeric_cols:
                        styles[i] = 'border-top: 1px solid #000000 !important; border-bottom: 1px solid #000000 !important; font-weight: bold; color: #000000; background-color: #f8f9fa;'
                    elif i == 0:
                        styles[i] = 'font-weight: bold; color: #000000; border-top: 1px solid #000000 !important; border-bottom: 1px solid #000000 !important; background-color: #f8f9fa;'
            return styles

        styler = styler.apply(highlight_totals, axis=1)
        
    return styler

