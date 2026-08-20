import pandas as pd
import openpyxl
from io import BytesIO
import unicodedata
import re
import datetime
import calendar
from copy import copy

def format_period_to_spanish_date(period_str):
    if not period_str:
        return ""
    try:
        parts = str(period_str).strip().split('-')
        if len(parts) >= 2:
            year = int(parts[0])
            month = int(parts[1])
            last_day = calendar.monthrange(year, month)[1]
            months_es = {
                1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
                5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
                9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
            }
            month_name = months_es[month]
            return f"{last_day} de {month_name} {year}"
    except Exception:
        pass
    return period_str

class BalanceGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

    def generate(self, tb_df, map_balance_df, scale_factor=1.0, tb_df_comp=None, periodo_actual_str=None, periodo_comp_str=None):
        """
        1. Cruza tb_df con map_balance_df para obtener la 'Clasificación balance' de cada cuenta.
        2. Suma los saldos por cada 'Clasificación balance'.
        3. Calcula el residuo de P&L (cuentas sin mapeo de balance).
        4. Inyecta los valores en el template de Excel.
        """
        # Ensure correct types
        tb_df['cuenta_id'] = tb_df['cuenta_id'].astype(str).str.strip()
        tb_df['saldo_final'] = pd.to_numeric(tb_df['saldo_final'], errors='coerce').fillna(0)
        
        map_balance_df['N° de Cuenta'] = map_balance_df['N° de Cuenta'].astype(str).str.strip()
        cls_col = next((c for c in map_balance_df.columns if "clasificaci" in c.lower() and "balance" in c.lower()), "Clasificación balance")
        if cls_col in map_balance_df.columns:
            map_balance_df[cls_col] = map_balance_df[cls_col].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        
        # Merge TB with Mapping using the sabana builder
        from src.core.sabana_builder import build_balance_sabana
        merged_df = build_balance_sabana(tb_df, map_balance_df)
        
        # Aggregate by Clasificación balance
        # Any row without a mapping will have NaN in 'Clasificación balance'
        mapped_mask = merged_df[cls_col].notna()
        
        agg_mapped = merged_df[mapped_mask].groupby(cls_col, as_index=False)['saldo_final'].sum()
        
        # Calculate P&L residual (A-P+R=0) -> everything unmapped belongs to P&L
        pl_residual = merged_df[~mapped_mask]['saldo_final'].sum()
        
        # Convert aggregation to dictionary for O(1) lookups
        sums_dict = dict(zip(agg_mapped[cls_col].str.strip(), agg_mapped['saldo_final']))
        
        sums_comp_dict = {}
        pl_residual_comp = 0.0
        if tb_df_comp is not None and not tb_df_comp.empty:
            merged_comp = build_balance_sabana(tb_df_comp, map_balance_df)
            mapped_mask_comp = merged_comp[cls_col].notna()
            agg_mapped_comp = merged_comp[mapped_mask_comp].groupby(cls_col, as_index=False)['saldo_final'].sum()
            pl_residual_comp = merged_comp[~mapped_mask_comp]['saldo_final'].sum()
            sums_comp_dict = dict(zip(agg_mapped_comp[cls_col].str.strip(), agg_mapped_comp['saldo_final']))

        
        # Load Template with openpyxl to preserve formatting
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Find column indices dynamically based on headers in the first few rows
        from src.core.excel_utils import detect_balance_columns
        name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx = detect_balance_columns(ws)
        
        # Relocate "Activos" header if needed
        r_activos = None
        r_corrientes = None
        for r in range(1, 15):
            val = ws.cell(row=r, column=name_col_idx).value
            if val and str(val).strip().lower() == "activos":
                r_activos = r
            if val and str(val).strip().lower() == "activos corrientes":
                r_corrientes = r

        if r_activos is not None and r_corrientes is not None:
            target_row = r_corrientes - 1
            if target_row > r_activos:
                orig_cell = ws.cell(row=r_activos, column=name_col_idx)
                target_cell = ws.cell(row=target_row, column=name_col_idx)
                target_cell.value = orig_cell.value
                if orig_cell.font:
                    target_cell.font = copy(orig_cell.font)
                orig_cell.value = None

        if periodo_actual_str:
            formatted_actual = format_period_to_spanish_date(periodo_actual_str)
            formatted_comp = format_period_to_spanish_date(periodo_comp_str) if periodo_comp_str else ""
            
            for r in range(1, 15):
                c25 = ws.cell(row=r, column=val25_col_idx)
                c24 = ws.cell(row=r, column=val24_col_idx)
                
                # Check for c25
                v25 = c25.value
                if v25 is not None:
                    is_date_or_year = (
                        isinstance(v25, (datetime.datetime, datetime.date)) or
                        (isinstance(v25, str) and ("20" in v25 or "Actual" in v25))
                    )
                    if is_date_or_year:
                        c25.value = formatted_actual
                        
                # Check for c24
                v24 = c24.value
                if v24 is not None:
                    is_date_or_year_comp = (
                        isinstance(v24, (datetime.datetime, datetime.date)) or
                        (isinstance(v24, str) and ("20" in v24 or "Anterior" in v24))
                    )
                    if is_date_or_year_comp:
                        c24.value = formatted_comp if periodo_comp_str else ""

        
        # Trackers for 2025 totals
        subtotal_corriente = 0.0
        subtotal_no_corriente = 0.0
        activos_totales = 0.0
        pasivos_totales = 0.0
        patrimonio_totales = 0.0
        
        # Trackers for 2024 totals
        subtotal_corriente_24 = 0.0
        subtotal_no_corriente_24 = 0.0
        activos_totales_24 = 0.0
        pasivos_totales_24 = 0.0
        patrimonio_totales_24 = 0.0
        
        def sanitize(text):
            if pd.isna(text): return ""
            clean_str = str(text).strip().lower()
            clean_str = ''.join(c for c in unicodedata.normalize('NFD', clean_str) if unicodedata.category(c) != 'Mn')
            clean_str = re.sub(r'\s+', ' ', clean_str)
            return clean_str
            
        def _is_total_activos_corrientes(s):
            return ('activo' in s or 'asset' in s) and 'corrien' in s and ('total' in s or 'subtotal' in s) and 'no corrien' not in s and 'nocorrien' not in s
        def _is_total_activos_no_corrientes(s):
            return ('activo' in s or 'asset' in s) and ('no corrien' in s or 'nocorrien' in s or 'largo' in s) and ('total' in s or 'subtotal' in s)
        def _is_total_activos(s):
            return ('total' in s or 'subtotal' in s) and ('activo' in s or 'asset' in s) and 'corrien' not in s and 'no corrien' not in s
        def _is_total_pasivos_corrientes(s):
            return ('pasivo' in s or 'liabilit' in s) and 'corrien' in s and ('total' in s or 'subtotal' in s) and 'no corrien' not in s and 'nocorrien' not in s
        def _is_total_pasivos_no_corrientes(s):
            return ('pasivo' in s or 'liabilit' in s) and ('no corrien' in s or 'nocorrien' in s or 'largo' in s) and ('total' in s or 'subtotal' in s)
        def _is_total_pasivos(s):
            return ('total' in s or 'subtotal' in s) and ('pasivo' in s or 'liabilit' in s) and 'corrien' not in s and 'no corrien' not in s and 'patrimonio' not in s
        def _is_total_patrimonio(s):
            return ('total' in s or 'subtotal' in s) and 'patrimonio' in s and 'pasivo' not in s
        def _is_total_pasivos_patrimonio(s):
            return ('total' in s or 'subtotal' in s) and 'patrimonio' in s and 'pasivo' in s
            
        current_section = "None"

        for row in range(1, ws.max_row + 1):
            cell_name = ws.cell(row=row, column=name_col_idx).value
            
            if cell_name and isinstance(cell_name, str):
                cell_name_clean = cell_name.strip()
                cell_lower = sanitize(cell_name)
                
                # Check for major sections to assign rolling sums
                if cell_lower == "activos corrientes":
                    current_section = "Activos Corrientes"
                elif cell_lower == "activos no corrientes":
                    current_section = "Activos No Corrientes"
                elif cell_lower == "pasivos corrientes":
                    current_section = "Pasivos Corrientes"
                elif cell_lower == "pasivos no corrientes":
                    current_section = "Pasivos No Corrientes"
                elif cell_lower == "patrimonio":
                    current_section = "Patrimonio"
                    
                # Load comparative either dynamically or statically
                val24 = 0.0
                if tb_df_comp is not None and not tb_df_comp.empty:
                    if cell_name_clean in sums_comp_dict and cell_name_clean not in ["Resultados acumulados", "Ganancias (pérdidas) acumuladas"]:
                        raw_val24 = sums_comp_dict[cell_name_clean]
                        if pd.isna(raw_val24) or raw_val24 is None:
                            raw_val24 = 0.0
                        val24 = (raw_val24 / scale_factor) * (-1 if current_section in ["Pasivos Corrientes", "Pasivos No Corrientes", "Patrimonio"] else 1)
                        ws.cell(row=row, column=val24_col_idx).value = val24
                    elif cell_name_clean in ["Resultados acumulados", "Ganancias (pérdidas) acumuladas"]:
                        mapped_bal_sum_comp = merged_comp[mapped_mask_comp]['saldo_final'].sum()
                        if pd.isna(mapped_bal_sum_comp) or abs(mapped_bal_sum_comp) < 10.0:
                            resid_comp = 0.0
                        else:
                            resid_comp = pl_residual_comp
                            if pd.isna(resid_comp):
                                resid_comp = 0.0
                        base_accum_comp = sums_comp_dict.get(cell_name_clean, 0)
                        if pd.isna(base_accum_comp) or base_accum_comp is None:
                            base_accum_comp = 0.0
                        final_resid_comp = (base_accum_comp + resid_comp) / scale_factor
                        val24 = final_resid_comp * (-1 if current_section in ["Pasivos Corrientes", "Pasivos No Corrientes", "Patrimonio"] else 1)
                        ws.cell(row=row, column=val24_col_idx).value = val24
                else:
                    val24_dirty = ws.cell(row=row, column=val24_col_idx).value
                    if val24_dirty is not None:
                        try:
                            clean_str_val = str(val24_dirty).replace(',', '').replace(' ', '').strip()
                            if clean_str_val: val24 = float(clean_str_val)
                        except (ValueError, TypeError): val24 = 0.0

                # Check if this row matches a mapped classification for 2025 (and isn't the special P&L residual row)
                if cell_name_clean in sums_dict and cell_name_clean not in ["Resultados acumulados", "Ganancias (pérdidas) acumuladas"]:
                    raw_val = sums_dict[cell_name_clean]
                    if pd.isna(raw_val) or raw_val is None:
                        raw_val = 0.0
                    val = raw_val / scale_factor
                    
                    # Store mathematically raw value internally, flip sign ONLY for visual output on details
                    visual_val = val
                    if current_section in ["Pasivos Corrientes", "Pasivos No Corrientes", "Patrimonio"]:
                        visual_val = val * -1
                        
                    ws.cell(row=row, column=val25_col_idx).value = visual_val
                    
                    # Accumulate for subtotals (2025) using VISUAL values so totals match what the user sees
                    if current_section == "Activos Corrientes":
                        subtotal_corriente += visual_val
                    elif current_section == "Activos No Corrientes":
                        subtotal_no_corriente += visual_val
                    elif current_section == "Pasivos Corrientes":
                        subtotal_corriente += visual_val
                    elif current_section == "Pasivos No Corrientes":
                        subtotal_no_corriente += visual_val
                    elif current_section == "Patrimonio":
                        patrimonio_totales += visual_val
                
                # Check if this is the target row for the P&L residual
                if cell_name_clean in ["Resultados acumulados", "Ganancias (pérdidas) acumuladas"]:
                    mapped_bal_sum = merged_df[mapped_mask]['saldo_final'].sum()
                    if pd.isna(mapped_bal_sum) or abs(mapped_bal_sum) < 10.0:
                        resid = 0.0
                    else:
                        resid = pl_residual / scale_factor
                        if pd.isna(resid):
                            resid = 0.0
                    current_val_raw = sums_dict.get(cell_name_clean, 0)
                    if pd.isna(current_val_raw) or current_val_raw is None:
                        current_val_raw = 0.0
                    current_val = current_val_raw / scale_factor
                    final_resid = current_val + resid
                    
                    visual_final_resid = final_resid
                    if current_section in ["Pasivos Corrientes", "Pasivos No Corrientes", "Patrimonio"]:
                        visual_final_resid = final_resid * -1
                        
                    ws.cell(row=row, column=val25_col_idx).value = visual_final_resid
                    
                    if current_section == "Activos Corrientes":
                        subtotal_corriente += visual_final_resid
                    elif current_section == "Activos No Corrientes":
                        subtotal_no_corriente += visual_final_resid
                    elif current_section == "Pasivos Corrientes":
                        subtotal_corriente += visual_final_resid
                    elif current_section == "Pasivos No Corrientes":
                        subtotal_no_corriente += visual_final_resid
                    elif current_section == "Patrimonio":
                        patrimonio_totales += visual_final_resid
                        
                # Independent accumulator for 2024 lines that just have pure static data
                # Since 2024 lines aren't driven by sums_dict matching, we accumulate them if they exist
                # But only if it's NOT a total row itself. How to know? The explicit if blocks below handle totals.
                if val24 != 0 and cell_lower not in ["activos corrientes totales", "total activos no corrientes", "total activos", "pasivos corrientes totales", "total pasivos no corrientes", "total pasivos", "patrimonio total", "total patrimonio y pasivos"] and cell_name_clean not in ["Resultados acumulados", "Ganancias (pérdidas) acumuladas"] and cell_name_clean not in sums_dict:
                    # 2024 is strictly read-only and un-inverted now based on user request.
                     pass

                
                # Lista de todas las categorias de totales para excluir de la acumulacion 2024
                _is_any_total = (
                    _is_total_activos_corrientes(cell_lower) or
                    _is_total_activos_no_corrientes(cell_lower) or
                    _is_total_activos(cell_lower) or
                    _is_total_pasivos_corrientes(cell_lower) or
                    _is_total_pasivos_no_corrientes(cell_lower) or
                    _is_total_pasivos(cell_lower) or
                    _is_total_patrimonio(cell_lower) or
                    _is_total_pasivos_patrimonio(cell_lower)
                )
                
                # We always accumulate whatever is sitting in val24 after processing
                # EXCEPT if this row is a total row, because then val24 will contain the Old Total, which we overwrite.
                if not _is_any_total:
                    if current_section == "Activos Corrientes":
                        subtotal_corriente_24 += val24
                    elif current_section == "Activos No Corrientes":
                        subtotal_no_corriente_24 += val24
                    elif current_section == "Pasivos Corrientes":
                        subtotal_corriente_24 += val24
                    elif current_section == "Pasivos No Corrientes":
                        subtotal_no_corriente_24 += val24
                    elif current_section == "Patrimonio":
                        patrimonio_totales_24 += val24
                
                # Inyeccion de totales/subtotales usando deteccion flexible (no exacta)
                if _is_total_activos_corrientes(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = subtotal_corriente
                    ws.cell(row=row, column=val24_col_idx).value = subtotal_corriente_24
                    activos_totales += subtotal_corriente
                    activos_totales_24 += subtotal_corriente_24
                    subtotal_corriente = 0.0  # reset 
                    subtotal_corriente_24 = 0.0
                elif _is_total_activos_no_corrientes(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = subtotal_no_corriente
                    ws.cell(row=row, column=val24_col_idx).value = subtotal_no_corriente_24
                    activos_totales += subtotal_no_corriente
                    activos_totales_24 += subtotal_no_corriente_24
                    subtotal_no_corriente = 0.0  # reset
                    subtotal_no_corriente_24 = 0.0
                elif _is_total_activos(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = activos_totales
                    ws.cell(row=row, column=val24_col_idx).value = activos_totales_24
                # The trackers now hold "Visual Sums" which represent true expected accounting arithmetic
                elif _is_total_pasivos_corrientes(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = subtotal_corriente
                    ws.cell(row=row, column=val24_col_idx).value = subtotal_corriente_24
                    pasivos_totales += subtotal_corriente
                    pasivos_totales_24 += subtotal_corriente_24
                    subtotal_corriente = 0.0
                    subtotal_corriente_24 = 0.0
                elif _is_total_pasivos_no_corrientes(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = subtotal_no_corriente
                    ws.cell(row=row, column=val24_col_idx).value = subtotal_no_corriente_24
                    pasivos_totales += subtotal_no_corriente
                    pasivos_totales_24 += subtotal_no_corriente_24
                    subtotal_no_corriente = 0.0
                    subtotal_no_corriente_24 = 0.0
                elif _is_total_pasivos(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = pasivos_totales
                    ws.cell(row=row, column=val24_col_idx).value = pasivos_totales_24
                elif _is_total_pasivos_patrimonio(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = (pasivos_totales + patrimonio_totales)
                    ws.cell(row=row, column=val24_col_idx).value = (pasivos_totales_24 + patrimonio_totales_24)
                elif _is_total_patrimonio(cell_lower):
                    ws.cell(row=row, column=val25_col_idx).value = patrimonio_totales
                    ws.cell(row=row, column=val24_col_idx).value = patrimonio_totales_24

                # Note: The template already has SUB-TOTAL and TOTAL formulas built-in natively.
                # Openpyxl preserves these formulas, so writing the base lines is enough.

        # Save to BytesIO for Streamlit download
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
