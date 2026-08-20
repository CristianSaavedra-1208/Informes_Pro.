import pandas as pd
import openpyxl
from io import BytesIO

class OriGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

    def generate(self, pl_preview_df=None, periodo_actual_str=None, periodo_comp_str=None, bal_preview_df=None, empresa=None):
        """
        Genera el formato de Resultados Integrales (ORI) inyectando la ganancia proveniente del ER y sumando líneas de Otros Resultados Integrales.
        pl_preview_df: DataFrame generado por ERGenerator (con 'Clasificación', col_actual, col_comp)
        bal_preview_df: DataFrame generado por BalanceGenerator (con 'Clasificación', col_actual, col_comp)
        """
        col_actual = periodo_actual_str if periodo_actual_str else '2025'
        col_comp = periodo_comp_str if periodo_comp_str else '2024'
        
        pl_25 = 0.0
        pl_24 = 0.0
        
        # Extraer Ganancia del Ejercicio (P&L Base) usando un extractor robusto
        if pl_preview_df is not None and not pl_preview_df.empty:
            cols = [str(c).strip() for c in pl_preview_df.columns]
            col_act_use = col_actual if col_actual in cols else (cols[1] if len(cols) > 1 else None)
            col_comp_use = col_comp if col_comp in cols else (cols[2] if len(cols) > 2 else None)
            
            for idx, row in pl_preview_df.iterrows():
                clasif = str(row.iloc[0]).lower().strip()
                if ('continuadas' in clasif or 
                    ('ganancia' in clasif and 'bruta' not in clasif and 'antes' not in clasif and 'otras' not in clasif and 'impuesto' not in clasif) or
                    ('resultado' in clasif and 'antes' not in clasif and 'bruto' not in clasif and 'impuesto' not in clasif)):
                    
                    if col_act_use:
                        v25 = row[col_act_use]
                        pl_25 = float(v25) if pd.notna(v25) and str(v25).strip() != '' else 0.0
                    if col_comp_use:
                        v24 = row[col_comp_use]
                        pl_24 = float(v24) if pd.notna(v24) and str(v24).strip() != '' else 0.0
                    break
        
        # ------------------ RESOLVER SALDOS DE OTRAS RESERVAS ------------------
        res_ext_actual = 0.0
        res_ext_comp = 0.0

        # Intentar extraer de bal_preview_df
        def get_bal_val(name, col):
            try:
                row = bal_preview_df[bal_preview_df['Clasificación'].astype(str).str.strip().str.lower() == name.lower()]
                if not row.empty:
                    v = row[col].iloc[0]
                    return float(v) if pd.notna(v) and str(v).strip() != '' else 0.0
            except:
                pass
            return None

        if bal_preview_df is not None and not bal_preview_df.empty:
            res_ext_actual = get_bal_val("Otras reservas", col_actual)
            res_ext_comp = get_bal_val("Otras reservas", col_comp)
            
        # Si alguno es None, cargamos de la DB
        # Obtener lista de empresas para consulta a base de datos
        companies_to_query = []
        if empresa:
            if empresa.startswith("[GRUPO]"):
                from src.models.database import SessionLocal
                from src.models.consolidacion import ConsolidationGroup
                grupo_name = empresa.replace("[GRUPO] ", "").strip()
                db_session = SessionLocal()
                try:
                    grupo_obj = db_session.query(ConsolidationGroup).filter_by(nombre_grupo=grupo_name).first()
                    if grupo_obj:
                        companies_to_query.append(grupo_obj.empresa_matriz)
                        if grupo_obj.filial_is_group:
                            def get_sub_companies(sub_g_id):
                                sub_g = db_session.query(ConsolidationGroup).filter_by(id=sub_g_id).first()
                                if sub_g:
                                    c = [sub_g.empresa_matriz]
                                    if sub_g.filial_is_group:
                                        c.extend(get_sub_companies(int(sub_g.empresa_filial)))
                                    else:
                                        c.append(sub_g.empresa_filial)
                                    return c
                                return []
                            companies_to_query.extend(get_sub_companies(int(grupo_obj.empresa_filial)))
                        else:
                            companies_to_query.append(grupo_obj.empresa_filial)
                except Exception as e:
                    print(f"Error resolving group in ORI: {e}")
                finally:
                    db_session.close()
            else:
                companies_to_query = [empresa]

        def get_reserves_from_db(period):
            if not companies_to_query or not period:
                return 0.0
            from src.models.database import SessionLocal
            from src.models.historical_data import HistoricalDataRecord
            db_session = SessionLocal()
            total_val = 0.0
            found_any = False
            try:
                if empresa and empresa.startswith("[GRUPO]"):
                    rec = db_session.query(HistoricalDataRecord).filter(
                        HistoricalDataRecord.empresa == empresa,
                        HistoricalDataRecord.periodo == period,
                        HistoricalDataRecord.reporte == 'Balance',
                        HistoricalDataRecord.linea_item.ilike('Otras reservas')
                    ).first()
                    if rec is not None:
                        return float(rec.monto)

                for co in companies_to_query:
                    rec = db_session.query(HistoricalDataRecord).filter(
                        HistoricalDataRecord.empresa == co,
                        HistoricalDataRecord.periodo == period,
                        HistoricalDataRecord.reporte == 'Balance',
                        HistoricalDataRecord.linea_item.ilike('Otras reservas')
                    ).first()
                    if rec is not None:
                        total_val += float(rec.monto)
                        found_any = True
            except Exception as e:
                print(f"Error querying reserves from DB in ORI for {period}: {e}")
            finally:
                db_session.close()
            return total_val if found_any else None

        if res_ext_actual is None:
            res_ext_actual = get_reserves_from_db(col_actual) or 0.0
        if res_ext_comp is None:
            res_ext_comp = get_reserves_from_db(col_comp) or 0.0

        # Buscar periodo anterior
        periodo_prev_str = None
        if col_comp and isinstance(col_comp, str) and '-' in col_comp:
            parts = col_comp.split('-')
            if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                y = int(parts[0])
                m = parts[1]
                periodo_prev_str = f"{y-1}-{m}"

        res_ext_prev = None
        if periodo_prev_str:
            res_ext_prev = get_reserves_from_db(periodo_prev_str)

        # Fallback if not found
        if res_ext_prev is None and col_comp:
            comp_year = col_comp.split('-')[0]
            from src.models.database import SessionLocal
            from src.models.historical_data import HistoricalDataRecord
            db_session = SessionLocal()
            try:
                earliest_rec = db_session.query(HistoricalDataRecord.periodo).filter(
                    HistoricalDataRecord.empresa.in_(companies_to_query),
                    HistoricalDataRecord.periodo.like(f"{comp_year}-%"),
                    HistoricalDataRecord.reporte == 'Balance'
                ).order_by(HistoricalDataRecord.periodo.asc()).first()
                if earliest_rec:
                    res_ext_prev = get_reserves_from_db(earliest_rec[0])
            except Exception as e:
                print(f"Error getting earliest reserve fallback in ORI: {e}")
            finally:
                db_session.close()

        if res_ext_prev is None:
            res_ext_prev = res_ext_comp

        # Calcular variaciones netas de reservas (Net ORI)
        net_ori_25 = res_ext_actual - res_ext_comp
        net_ori_24 = res_ext_comp - res_ext_prev

        # Calcular efectos brutos e impuestos (Tasa del 27%)
        # Neto = Bruto * 0.73  =>  Bruto = Neto / 0.73
        # Impuesto = -0.27 * Bruto
        gross_ori_25 = net_ori_25 / 0.73
        tax_ori_25 = -0.27 * gross_ori_25

        gross_ori_24 = net_ori_24 / 0.73
        tax_ori_24 = -0.27 * gross_ori_24

        total_integral_25 = pl_25 + net_ori_25
        total_integral_24 = pl_24 + net_ori_24
        
        # Inyectando a Excel Abierto
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Encontrar columna de conceptos/nombres y de valores de forma dinámica
        name_col_idx = 1
        for col in range(1, 5):
            for row in range(1, 15):
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().lower() in ["concepto", "descripcion", "detalle", "flujos", "origen/aplicacion", "clasificacion", "otros resultados"]:
                    name_col_idx = col
                    break
                    
        date_cols = []
        import datetime
        import re
        for col in range(name_col_idx + 1, ws.max_column + 1):
            for row in range(1, 10):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    val_str = str(val).strip().lower()
                    is_date = (
                        isinstance(val, (datetime.datetime, datetime.date)) or
                        re.search(r'20\d{2}', val_str) or
                        any(k in val_str for k in ["actual", "anterior", "comparativ", "auditado", "31-12", "31 de"])
                    )
                    if is_date:
                        date_cols.append(col)
                        break
        date_cols = sorted(list(set(date_cols)))
        
        val25_col_idx = date_cols[0] if len(date_cols) >= 1 else 2
        val24_col_idx = date_cols[1] if len(date_cols) >= 2 else 3
        
        # Row map based on our template script
        row_ganancia_base = 8
        row_coberturas = 12
        row_ori_pre_imp = 14
        row_imp_ori = 16
        row_sub_ori = 18
        row_total_integral = 19
        
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=name_col_idx).value
            if val and isinstance(val, str):
                val_clean = val.lower().strip()
                if "ganancia (perdida) del ejercicio" in val_clean or "utilidad (perdida) del ejercicio" in val_clean or "ganancia del ejercicio" in val_clean or "resultado del ejercicio" in val_clean:
                    row_ganancia_base = r
                elif "coberturas de flujo de efectivo" in val_clean or "coberturas" in val_clean:
                    row_coberturas = r
                elif "antes de impuestos" in val_clean and "otros" in val_clean:
                    row_ori_pre_imp = r
                elif "impuesto" in val_clean and ("otros" in val_clean or r > row_ori_pre_imp):
                    row_imp_ori = r
                elif "total de otros resultados integrales" in val_clean or ("neto de impuestos" in val_clean and "otros" in val_clean):
                    row_sub_ori = r
                elif "resultado integral total" in val_clean or "total resultado integral" in val_clean or "total de resultado integral" in val_clean:
                    row_total_integral = r
        
        def inyectar(r, v25, v24):
            ws.cell(row=r, column=val25_col_idx, value=v25)
            ws.cell(row=r, column=val24_col_idx, value=v24)
            
        inyectar(row_ganancia_base, pl_25, pl_24)
        inyectar(row_coberturas, gross_ori_25, gross_ori_24)
        inyectar(row_ori_pre_imp, gross_ori_25, gross_ori_24)
        inyectar(row_imp_ori, tax_ori_25, tax_ori_24)
        inyectar(row_sub_ori, net_ori_25, net_ori_24)
        inyectar(row_total_integral, total_integral_25, total_integral_24)
 
        if periodo_actual_str and periodo_comp_str:
            for r in range(1, 10):
                c25 = ws.cell(row=r, column=val25_col_idx)
                c24 = ws.cell(row=r, column=val24_col_idx)
                if c25.value and isinstance(c25.value, str) and ("20" in c25.value or "Actual" in c25.value):
                    c25.value = periodo_actual_str
                if c24.value and isinstance(c24.value, str) and ("20" in c24.value or "Anterior" in c24.value):
                    c24.value = periodo_comp_str
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
