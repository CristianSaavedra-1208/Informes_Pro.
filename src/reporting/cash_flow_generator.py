import pandas as pd
import openpyxl
from io import BytesIO
import numpy as np
from src.models.database import SessionLocal
from src.models.cash_flow_db import CashFlowAdjustment
from src.models.trial_balance_db import TrialBalanceDB

class CashFlowGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

    def generate(self, 
                 empresa, 
                 periodo_actual_str, 
                 periodo_comp_str, 
                 map_balance_df, 
                 map_pl_df=None, 
                 method="Directo", 
                 is_consolidado=False, 
                 consolidated_hoja_trabajo_df=None,
                 consolidated_hoja_trabajo_comp_df=None,
                 scale_factor=1.0):
        """
        Genera el Estado de Flujos de Efectivo (EFE) inyectando los saldos en la plantilla.
        Soporta cálculo a nivel de cuenta contable y aplicación de ajustes de depuración (Ingreso/Egreso Caja).
        """
        if not scale_factor:
            scale_factor = 1.0

        # 1. Obtener los balances detallados (Trial Balance) para el periodo actual y comparativo
        db = SessionLocal()
        try:
            # Cargar los ajustes de depuración del periodo actual
            adjustments = db.query(CashFlowAdjustment).filter(
                CashFlowAdjustment.empresa == empresa,
                CashFlowAdjustment.periodo == periodo_actual_str,
                CashFlowAdjustment.es_consolidado == is_consolidado
            ).all()
            
            # Cargar los ajustes de depuración del periodo comparativo
            adjustments_comp = db.query(CashFlowAdjustment).filter(
                CashFlowAdjustment.empresa == empresa,
                CashFlowAdjustment.periodo == periodo_comp_str,
                CashFlowAdjustment.es_consolidado == is_consolidado
            ).all()
        finally:
            db.close()

        # Construir diccionarios de ajustes
        adj_dict = {}
        for a in adjustments:
            li = a.linea_item.strip()
            if li not in adj_dict:
                adj_dict[li] = {'ingreso': 0.0, 'egreso': 0.0}
            adj_dict[li]['ingreso'] += float(a.ingreso_caja or 0.0)
            adj_dict[li]['egreso'] += float(a.egreso_caja or 0.0)

        adj_dict_comp = {}
        for a in adjustments_comp:
            li = a.linea_item.strip()
            if li not in adj_dict_comp:
                adj_dict_comp[li] = {'ingreso': 0.0, 'egreso': 0.0}
            adj_dict_comp[li]['ingreso'] += float(a.ingreso_caja or 0.0)
            adj_dict_comp[li]['egreso'] += float(a.egreso_caja or 0.0)

        # 2. Cargar mapeos de cuentas contables a líneas de flujo
        acc_info = {}
        map_cf_dict = {}

        # Mapeos desde map_balance_df
        if map_balance_df is not None:
            acc_col = next((c for c in map_balance_df.columns if 'cuenta' in c.lower() or 'cta' in c.lower()), None)
            cf_col = next((c for c in map_balance_df.columns if 'flujo' in c.lower() and 'efectivo' in c.lower()), None)
            id_rep_col = next((c for c in map_balance_df.columns if 'id_reporte' in c.lower() or 'd_reporte' in c.lower()), None)
            desc_col = next((c for c in map_balance_df.columns if 'nombre' in c.lower() or 'descripcion' in c.lower()), None)
            rub_col = next((c for c in map_balance_df.columns if 'clasificaci' in c.lower() and 'balance' in c.lower()), None)
            
            if acc_col and cf_col:
                for _, row in map_balance_df.iterrows():
                    acc_val = str(row[acc_col]).strip()
                    cf_val = str(row[cf_col]).strip()
                    id_rep = str(row[id_rep_col]).strip() if id_rep_col and pd.notna(row[id_rep_col]) else ""
                    desc = str(row[desc_col]).strip() if desc_col and pd.notna(row[desc_col]) else ""
                    rub = str(row[rub_col]).strip() if rub_col and pd.notna(row[rub_col]) else ""
                    
                    if acc_val and acc_val != 'nan':
                        section = 'A'
                        if acc_val.startswith('2'):
                            section = 'P'
                        elif acc_val.startswith('1'):
                            section = 'A'
                        elif id_rep.upper().startswith('PAS_') or id_rep.upper().startswith('PAT_'):
                            section = 'P'
                        elif id_rep.upper().startswith('ACT_'):
                            section = 'A'
                        else:
                            norm_rub = rub.lower()
                            if any(x in norm_rub for x in ["pasivo", "patrimonio", "capital", "reserva", "obligacion", "provision"]):
                                section = 'P'
                        
                        acc_info[acc_val] = {
                            'cf_line': cf_val if cf_val != 'nan' else '',
                            'section': section,
                            'description': desc,
                            'rubric': rub
                        }
                        # Map rubric name to cf_line (useful for consolidated rubrics)
                        if rub and cf_val and cf_val != 'nan':
                            map_cf_dict[rub] = cf_val

        # Mapeos desde map_pl_df
        if map_pl_df is not None:
            acc_col = next((c for c in map_pl_df.columns if 'cuenta' in c.lower() or 'cta' in c.lower()), None)
            cf_col = next((c for c in map_pl_df.columns if 'flujo' in c.lower() and 'efectivo' in c.lower()), None)
            desc_col = next((c for c in map_pl_df.columns if 'detalle' in c.lower() or 'descripcion' in c.lower() or 'nombre' in c.lower()), None)
            
            if acc_col:
                for _, row in map_pl_df.iterrows():
                    acc_val = str(row[acc_col]).strip()
                    cf_val = str(row[cf_col]).strip() if cf_col and pd.notna(row[cf_col]) else ""
                    desc = str(row[desc_col]).strip() if desc_col and pd.notna(row[desc_col]) else ""
                    
                    if acc_val and acc_val != 'nan':
                        rub = ""
                        for col in map_pl_df.columns:
                            if col not in [acc_col, cf_col, desc_col, 'Unnamed: 14'] and pd.notna(row[col]) and str(row[col]).strip() != '':
                                rub = col
                                break
                        acc_info[acc_val] = {
                            'cf_line': cf_val if cf_val != 'nan' else '',
                            'section': 'PL',
                            'description': desc,
                            'rubric': rub
                        }
                        if rub and cf_val and cf_val != 'nan' and cf_val != '':
                            map_cf_dict[rub] = cf_val

        # 3. Determinar los DataFrames a utilizar para el cálculo
        if is_consolidado:
            # En consolidación, tratamos las líneas/rubros de la hoja de trabajo consolidada como las "cuentas"
            # Actual
            df_act = pd.DataFrame()
            if consolidated_hoja_trabajo_df is not None:
                df_act_cleaned = consolidated_hoja_trabajo_df.copy()
                df_act_cleaned['CONSOLIDADO'] = pd.to_numeric(df_act_cleaned['CONSOLIDADO'], errors='coerce')
                df_act_cleaned = df_act_cleaned[df_act_cleaned['Balance clasificado'].notna() & (df_act_cleaned['Balance clasificado'].astype(str).str.strip() != "")]
                df_act_cleaned = df_act_cleaned[df_act_cleaned['CONSOLIDADO'].notna()]
                df_act = pd.DataFrame({
                    'cuenta_id': df_act_cleaned['Balance clasificado'],
                    'descripcion': df_act_cleaned['Balance clasificado'],
                    'saldo_final': df_act_cleaned['CONSOLIDADO']
                })
            
            # Comparativo
            df_comp = pd.DataFrame()
            if consolidated_hoja_trabajo_comp_df is not None:
                df_comp_cleaned = consolidated_hoja_trabajo_comp_df.copy()
                df_comp_cleaned['CONSOLIDADO'] = pd.to_numeric(df_comp_cleaned['CONSOLIDADO'], errors='coerce')
                df_comp_cleaned = df_comp_cleaned[df_comp_cleaned['Balance clasificado'].notna() & (df_comp_cleaned['Balance clasificado'].astype(str).str.strip() != "")]
                df_comp_cleaned = df_comp_cleaned[df_comp_cleaned['CONSOLIDADO'].notna()]
                df_comp = pd.DataFrame({
                    'cuenta_id': df_comp_cleaned['Balance clasificado'],
                    'descripcion': df_comp_cleaned['Balance clasificado'],
                    'saldo_final': df_comp_cleaned['CONSOLIDADO']
                })
            
            # Crear acc_info para los rubros consolidados usando map_cf_dict
            consolidated_acc_info = {}
            for _, row in df_act.iterrows():
                rub = str(row['cuenta_id']).strip()
                if rub:
                    # Determinar sección heurísticamente
                    section = 'A'
                    norm_rub = rub.lower()
                    if any(x in norm_rub for x in ["pasivo", "patrimonio", "capital", "reserva", "obligacion", "provision"]):
                        section = 'P'
                    elif "ingreso" in norm_rub or "costo" in norm_rub or "gasto" in norm_rub or "resultado" in norm_rub or "depreciacion" in norm_rub:
                        if not any(x in norm_rub for x in ["diferido", "diferidos", "anticipado", "anticipados", "acumulado", "acumulados"]):
                            section = 'PL'
                            
                    cf_line = map_cf_dict.get(rub, '')
                    consolidated_acc_info[rub] = {
                        'cf_line': cf_line,
                        'section': section,
                        'description': rub,
                        'rubric': rub
                    }
            acc_info = consolidated_acc_info
            
            # Para el consolidado, el periodo inicial para calcular el comparativo requiere un tercer periodo
            # que por ahora no cargamos recursivamente. Dejaremos el comparativo consolidado libre o calculado si se pasa
            df_prior_of_comp = None
        else:
            # Individual
            df_act = TrialBalanceDB.get_trial_balance(empresa, periodo_actual_str)
            df_comp = TrialBalanceDB.get_trial_balance(empresa, periodo_comp_str)
            
            # Buscar periodo previo al comparativo para poder calcular la variación comparativa (columna 4 del Excel)
            available_periods = TrialBalanceDB.get_available_periods(empresa)
            periodo_comp_prev = self._get_prior_period(periodo_comp_str, available_periods)
            df_prior_of_comp = TrialBalanceDB.get_trial_balance(empresa, periodo_comp_prev) if periodo_comp_prev else None

        # 4. Calcular variaciones para cada columna de datos
        # Actual Column (column 3)
        agrupacion_flujo, matriz_audit = self._compute_column_values(
            df_act, df_comp, acc_info, method=method
        )
        
        # Comparative Column (column 4)
        agrupacion_flujo_comp, matriz_audit_comp = self._compute_column_values(
            df_comp, df_prior_of_comp, acc_info, method=method
        )

        # 5. Obtener los saldos de Efectivo y Equivalentes para inyectar al final
        cash_prior, cash_actual = self._get_cash_balances(df_act, df_comp, map_balance_df, is_consolidado)
        cash_prior_comp, cash_actual_comp = self._get_cash_balances(df_comp, df_prior_of_comp, map_balance_df, is_consolidado)

        # 6. Combinar Variación Bruta con Ajustes de Depuración (Ingreso Caja / Egreso Caja)
        agrupacion_flujo_final = {}
        for k, v in agrupacion_flujo.items():
            adj = adj_dict.get(k, {'ingreso': 0.0, 'egreso': 0.0})
            agrupacion_flujo_final[k] = v + adj['ingreso'] - adj['egreso']

        agrupacion_flujo_comp_final = {}
        for k, v in agrupacion_flujo_comp.items():
            adj = adj_dict_comp.get(k, {'ingreso': 0.0, 'egreso': 0.0})
            agrupacion_flujo_comp_final[k] = v + adj['ingreso'] - adj['egreso']

        # 7. Actualizar la matriz de auditoría con los ajustes
        for item in matriz_audit:
            linea = item["Línea de Flujo Mapeada"]
            adj = adj_dict.get(linea, {'ingreso': 0.0, 'egreso': 0.0})
            # El ajuste de depuración se puede prorrratear o simplemente mostrar a nivel de desglose
            # Para la matriz cuenta por cuenta, los ajustes se muestran informativos al nivel de cuenta si corresponden,
            # pero dado que el ajuste se ingresa a nivel de línea, lo sumamos al total de la línea en el reporte.
            # Mostraremos el ajuste total de la línea en la previsualización del EFE.

        # 7.5. Calcular utilidad neta para el método indirecto
        pl_accounts = {acc for acc, info in acc_info.items() if info['section'] == 'PL'}
        net_income = 0.0
        if df_act is not None and not df_act.empty:
            net_income = -sum(float(row['saldo_final'] or 0.0) for _, row in df_act.iterrows() if str(row['cuenta_id']).strip() in pl_accounts)
            
        net_income_comp = 0.0
        if df_comp is not None and not df_comp.empty:
            net_income_comp = -sum(float(row['saldo_final'] or 0.0) for _, row in df_comp.iterrows() if str(row['cuenta_id']).strip() in pl_accounts)

        # 8. Inyectar saldos en la plantilla Excel
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        
        # Encontrar columna de conceptos/nombres y de valores de forma dinámica
        name_col_idx = 1
        for col in range(1, 5):
            for row in range(1, 15):
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().lower() in ["concepto", "descripcion", "detalle", "flujos", "origen/aplicacion", "actividades de"]:
                    name_col_idx = col
                    break
                    
        date_cols = []
        import datetime
        import re
        for col in range(name_col_idx + 1, ws.max_column + 1):
            for row in range(1, 10):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    val_str = str(val).strip().lower()
                    is_date = (
                        isinstance(val, (datetime.datetime, datetime.date)) or
                        re.search(r'20\d{2}', val_str) or
                        any(k in val_str for k in ["actual", "anterior", "comparativ", "auditado", "31-12", "31 de"])
                    )
                    if is_date:
                        date_cols.append(col)
                        break
        date_cols = sorted(list(set(date_cols)))
        
        val25_col_idx = date_cols[0] if len(date_cols) >= 1 else 3
        val24_col_idx = date_cols[1] if len(date_cols) >= 2 else 4
        
        # Preparar sinónimos normalizados para la utilidad del ejercicio
        import unicodedata
        def normalize_str(s):
            return ''.join(c for c in unicodedata.normalize('NFD', s.lower().strip()) if unicodedata.category(c) != 'Mn')
        
        utilidad_labels = [
            "utilidad (pérdida) del ejercicio", 
            "ganancia (pérdida) del ejercicio", 
            "resultado del ejercicio",
            "utilidad del ejercicio",
            "ganancia del ejercicio",
            "ganancia (pérdida) neta",
            "resultado neto del ejercicio",
            "resultado neto",
            "total profit",
            "resultado del periodo",
            "ganancia (perdida) del periodo"
        ]
        norm_utilidad_labels = {normalize_str(l) for l in utilidad_labels}

        # En la plantilla inyectar según columnas detectadas
        for row in range(1, ws.max_row + 1):
            cell_name = ws.cell(row=row, column=name_col_idx).value
            if cell_name and isinstance(cell_name, str):
                nm = cell_name.strip()
                nm_norm = normalize_str(nm)
                
                # Inyección del periodo actual (val25_col_idx)
                if nm in agrupacion_flujo_final:
                    ws.cell(row=row, column=val25_col_idx).value = agrupacion_flujo_final[nm] / scale_factor
                elif nm_norm in norm_utilidad_labels and method == "Indirecto":
                    ws.cell(row=row, column=val25_col_idx).value = net_income / scale_factor
                elif nm == "Saldo inicial de efectivo y equivalentes al efectivo":
                    ws.cell(row=row, column=val25_col_idx).value = cash_prior / scale_factor
                elif nm == "Saldo final de efectivo y equivalentes al efectivo":
                    ws.cell(row=row, column=val25_col_idx).value = cash_actual / scale_factor
                elif nm == "Incremento (decremento) neto en efectivo y equivalentes al efectivo":
                    ws.cell(row=row, column=val25_col_idx).value = (cash_actual - cash_prior) / scale_factor

                # Inyección del periodo comparativo (val24_col_idx)
                if nm in agrupacion_flujo_comp_final:
                    ws.cell(row=row, column=val24_col_idx).value = agrupacion_flujo_comp_final[nm] / scale_factor
                elif nm_norm in norm_utilidad_labels and method == "Indirecto":
                    ws.cell(row=row, column=val24_col_idx).value = net_income_comp / scale_factor
                elif nm == "Saldo inicial de efectivo y equivalentes al efectivo":
                    ws.cell(row=row, column=val24_col_idx).value = cash_prior_comp / scale_factor
                elif nm == "Saldo final de efectivo y equivalentes al efectivo":
                    ws.cell(row=row, column=val24_col_idx).value = cash_actual_comp / scale_factor
                elif nm == "Incremento (decremento) neto en efectivo y equivalentes al efectivo":
                    ws.cell(row=row, column=val24_col_idx).value = (cash_actual_comp - cash_prior_comp) / scale_factor

        # Reemplazar cabeceras de fechas en el excel si están especificadas
        if periodo_actual_str and periodo_comp_str:
            for r in range(1, 10):
                c25 = ws.cell(row=r, column=val25_col_idx)
                c24 = ws.cell(row=r, column=val24_col_idx)
                if c25.value and isinstance(c25.value, str) and ("20" in c25.value or "Actual" in c25.value or "31-12" in c25.value):
                    c25.value = self._format_period_to_spanish_date(periodo_actual_str)
                if c24.value and isinstance(c24.value, str) and ("20" in c24.value or "Anterior" in c24.value or "31-12" in c24.value):
                    c24.value = self._format_period_to_spanish_date(periodo_comp_str)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Enriquecer matriz_audit con campos de depuración para la salida
        for item in matriz_audit:
            linea = item["Línea de Flujo Mapeada"]
            item["Ingreso Caja (Ajuste)"] = (adj_dict.get(linea, {}).get('ingreso', 0.0) if linea else 0.0) / scale_factor
            item["Egreso Caja (Ajuste)"] = (adj_dict.get(linea, {}).get('egreso', 0.0) if linea else 0.0) / scale_factor
            
            # Escalar valores en la matriz para que coincidan con la previsualización y el Excel
            item["Saldo Anterior"] = (item["Saldo Anterior"] or 0.0) / scale_factor
            item["Saldo Actual"] = (item["Saldo Actual"] or 0.0) / scale_factor
            item["Variación Bruta"] = (item["Variación Bruta"] or 0.0) / scale_factor
            item["Variación Depurada"] = item["Variación Bruta"] + item["Ingreso Caja (Ajuste)"] - item["Egreso Caja (Ajuste)"]
            
        return output, matriz_audit

    def _compute_column_values(self, df_target, df_prior, acc_info, method="Directo"):
        """
        Calcula las variaciones brutas y genera la matriz de auditoría básica.
        """
        bal_target = {str(r['cuenta_id']).strip(): float(r['saldo_final']) for _, r in df_target.iterrows()} if df_target is not None else {}
        bal_prior = {str(r['cuenta_id']).strip(): float(r['saldo_final']) for _, r in df_prior.iterrows()} if df_prior is not None else {}
        
        all_accounts = set(bal_target.keys()).union(set(bal_prior.keys()))
        
        agrupacion_flujo = {}
        matriz_audit = []
        
        for acc in all_accounts:
            v_target = bal_target.get(acc, 0.0)
            v_prior = bal_prior.get(acc, 0.0)
            
            info = acc_info.get(acc, {'cf_line': '', 'section': 'A', 'description': acc, 'rubric': ''})
            cf_line = info['cf_line']
            section = info['section']
            desc = info['description']
            rubric = info['rubric']
            
            if not cf_line or cf_line.strip() == '':
                if method == "Indirecto" and section == 'PL':
                    cf_line = "PL_Default"
                else:
                    continue
                
            if method == "Indirecto":
                cf_line_lower = cf_line.lower()
                desc_lower = desc.lower()
                rubric_lower = rubric.lower() if rubric else ""
                
                if section == 'PL':
                    if "depreciaci" in cf_line_lower or "amortizaci" in cf_line_lower or "depreciaci" in rubric_lower or "amortizaci" in rubric_lower or "depreciaci" in desc_lower or "amortizaci" in desc_lower:
                        cf_line = "Depreciación y amortización"
                    elif "intereses pagados" in cf_line_lower or "costos financieros" in rubric_lower or "gastos financieros" in rubric_lower or "intereses" in desc_lower or "gastos financieros" in desc_lower:
                        cf_line = "Costos financieros"
                    elif "intereses recibidos" in cf_line_lower or "ingresos financieros" in rubric_lower or "ingresos financieros" in desc_lower or "intereses" in desc_lower:
                        cf_line = "Ingresos financieros"
                    elif "impuesto" in cf_line_lower or "impuesto" in rubric_lower or "gasto por impuesto" in rubric_lower or "impuesto" in desc_lower:
                        cf_line = "Gasto por impuesto a las ganancias"
                    elif "unidades de reajuste" in cf_line_lower or "diferencia de cambio" in rubric_lower or "unidades de reajuste" in rubric_lower or "variación de cambio" in rubric_lower or "diferencia de cambio" in cf_line_lower or "cambio" in desc_lower:
                        cf_line = "Diferencias de cambio no realizadas"
                    elif "provision" in desc_lower or "estimacion" in desc_lower or "deterioro" in desc_lower or "provisiones" in rubric_lower:
                        cf_line = "Provisiones y otros cargos no monetarios"
                    else:
                        continue
                else:
                    is_operational = any(x in cf_line_lower for x in [
                        "cobros", "pagos", "flujo operativo", "impuestos a las ganancias", 
                        "otras entradas y (salidas) de dinero", "eliminacion"
                    ])
                    is_investment_or_finance = any(x in cf_line_lower for x in [
                        "propiedades", "planta", "equipo", "intangibles", "prestamos", 
                        "préstamos", "arrendamientos", "acciones", "dividendos", "capital"
                    ])
                    
                    if is_operational and not is_investment_or_finance:
                        if section == 'A':
                            if "deudores" in rubric_lower or "clientes" in rubric_lower or "deudores" in desc_lower or "clientes" in desc_lower or "cuentas por cobrar" in rubric_lower or "cuentas por cobrar" in desc_lower:
                                cf_line = "Disminución (incremento) en deudores comerciales y otras cuentas por cobrar"
                            elif "inventario" in rubric_lower or "existencia" in rubric_lower or "inventario" in desc_lower or "existencia" in desc_lower or "mercaderia" in desc_lower:
                                cf_line = "Disminución (incremento) en inventarios"
                            else:
                                cf_line = "Disminución (incremento) en otros activos"
                        elif section == 'P':
                            if any(x in rubric_lower or x in desc_lower for x in ["empleados", "beneficios", "personal", "remuneraciones", "sueldos", "leyes sociales"]):
                                cf_line = "Incremento (disminución) en beneficios a los empleados"
                            elif any(x in rubric_lower or x in desc_lower for x in ["proveedores", "cuentas por pagar", "acreedores"]):
                                cf_line = "Incremento (disminución) en cuentas por pagar comerciales y otras cuentas por pagar"
                            else:
                                cf_line = "Incremento (disminución) en otros pasivos"
                
            # Regla de signos unificada
            if section == 'PL':
                if method == "Directo":
                    # Método Directo: el impacto de la cuenta de resultado es su saldo neto invertido
                    delta = -v_target
                else:
                    # Método Indirecto: el impacto es el saldo neto directo (para reincorporar gastos/ingresos no monetarios)
                    delta = v_target
            else:
                # Cuentas de Balance: delta = Inicial - Final (Debido a la convención de signo de contabilidad)
                delta = v_prior - v_target
                
            # Ignorar variaciones del efectivo de las secciones normales
            if "saldo inicial de efectivo" in cf_line.lower() or "saldo final de efectivo" in cf_line.lower() or "incremento (decremento) neto" in cf_line.lower():
                continue
                
            if cf_line not in agrupacion_flujo:
                agrupacion_flujo[cf_line] = 0.0
            agrupacion_flujo[cf_line] += delta
            
            matriz_audit.append({
                "Cuenta ID": acc,
                "Descripción": desc,
                "Rubro Mapeado": rubric,
                "Sección": "Activo" if section == 'A' else ("Pasivo/Patrimonio" if section == 'P' else "Estado de Resultados"),
                "Saldo Anterior": v_prior,
                "Saldo Actual": v_target,
                "Variación Bruta": delta,
                "Línea de Flujo Mapeada": cf_line
            })
            
        return agrupacion_flujo, matriz_audit

    def _get_cash_balances(self, df_target, df_prior, map_balance_df, is_consolidado=False):
        """
        Obtiene los saldos de caja sumando las cuentas de Efectivo y Equivalentes.
        """
        cash_accounts = set()
        
        if is_consolidado:
            # En consolidado el rubro se llama exactamente "Efectivo y equivalentes al efectivo" o similar
            cash_accounts.add("Efectivo y efectivo equivalente")
            cash_accounts.add("Efectivo y equivalentes al efectivo")
            cash_accounts.add("Efectivo y equivalentes de efectivo")
            if map_balance_df is not None:
                rub_col = next((c for c in map_balance_df.columns if 'clasificaci' in c.lower() and 'balance' in c.lower()), None)
                if rub_col:
                    for _, row in map_balance_df.iterrows():
                        rub_val = str(row[rub_col]).strip()
                        if "efectivo" in rub_val.lower() or "caja" in rub_val.lower():
                            cash_accounts.add(rub_val)
        else:
            if map_balance_df is not None:
                acc_col = next((c for c in map_balance_df.columns if 'cuenta' in c.lower() or 'cta' in c.lower()), None)
                rub_col = next((c for c in map_balance_df.columns if 'clasificaci' in c.lower() and 'balance' in c.lower()), None)
                id_rep_col = next((c for c in map_balance_df.columns if 'id_reporte' in c.lower() or 'd_reporte' in c.lower()), None)
                
                if acc_col:
                    for _, row in map_balance_df.iterrows():
                        acc_val = str(row[acc_col]).strip()
                        rub_val = str(row[rub_col]).lower().strip() if rub_col and pd.notna(row[rub_col]) else ""
                        id_rep = str(row[id_rep_col]).lower().strip() if id_rep_col and pd.notna(row[id_rep_col]) else ""
                        
                        if "efectivo" in rub_val or "caja" in rub_val or "efe_100" in id_rep:
                            cash_accounts.add(acc_val)
                            
        sum_target = 0.0
        sum_prior = 0.0
        
        if df_target is not None:
            for _, row in df_target.iterrows():
                if str(row['cuenta_id']).strip() in cash_accounts:
                    sum_target += float(row['saldo_final'] or 0.0)
                    
        if df_prior is not None:
            for _, row in df_prior.iterrows():
                if str(row['cuenta_id']).strip() in cash_accounts:
                    sum_prior += float(row['saldo_final'] or 0.0)
                    
        return sum_prior, sum_target

    def _get_prior_period(self, period_str, available_periods):
        try:
            parts = period_str.split('-')
            year = int(parts[0])
            month = int(parts[1])
            prior_yoy = f"{year-1}-{month:02d}"
            if prior_yoy in available_periods:
                return prior_yoy
            sorted_periods = sorted(available_periods)
            if period_str in sorted_periods:
                idx = sorted_periods.index(period_str)
                if idx > 0:
                    return sorted_periods[idx-1]
        except:
            pass
        return None

    def _format_period_to_spanish_date(self, period_str):
        if not period_str:
            return ""
        try:
            import calendar
            parts = str(period_str).strip().split('-')
            if len(parts) >= 2:
                year = int(parts[0])
                month = int(parts[1])
                last_day = calendar.monthrange(year, month)[1]
                months_es = {
                    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
                    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
                    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
                }
                return f"{last_day} de {months_es[month]} {year}"
        except:
            pass
        return period_str
