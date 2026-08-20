import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os

empresa = "Pacifico SpA"
path = f"data/empresas/{empresa}/Estado de Resultados Integrales.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active

ws.title = "Resultados Integrales"

# Titles
ws.cell(row=1, column=1, value="PACIFICO CABLE SpA.").font = Font(bold=True)
ws.cell(row=2, column=1, value="ESTADOS SEPARADOS DE RESULTADOS INTEGRALES").font = Font(bold=True)
ws.cell(row=3, column=1, value="Por los periodos correspondientes")

# All centered
for r in range(1, 4):
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

# Headers
c_act = ws.cell(row=5, column=2, value="Periodo Actual")
c_act.font = Font(underline='single')
c_act.alignment = Alignment(horizontal='center')

c_comp = ws.cell(row=5, column=3, value="Periodo Comparativo")
c_comp.font = Font(underline='single')
c_comp.alignment = Alignment(horizontal='center')

ws.cell(row=6, column=2, value="M$").alignment = Alignment(horizontal='center')
ws.cell(row=6, column=3, value="M$").alignment = Alignment(horizontal='center')

# Rows
rows_data = [
    (8, "(Pérdida) ganancia procedente de operaciones continuadas"),
    (10, "Componentes de otro resultado integral que se reclasificarán al resultado del ejercicio:"),
    (12, "Ganancias (pérdidas) por coberturas de flujos de efectivo, antes de impuestos"),
    (14, "Otro resultado integral, antes de impuestos"),
    (16, "Impuesto a las ganancias relativos a componentes de otro resultado integral"),
    (18, "Otro resultado integral"),
    (19, "Total resultado integral de operaciones continuadas")
]

for r, text in rows_data:
    cell = ws.cell(row=r, column=1, value=text)
    if "Total" in text:
        cell.alignment = Alignment(indent=2)
    elif "Componentes" not in text and "Pérdida" not in text:
        cell.alignment = Alignment(indent=1) # Indent sub-items

double_bottom = Border(top=Side(style='thin'), bottom=Side(style='double'))
thin_bottom = Border(bottom=Side(style='thin'))

# Format Total Row
for col in [2, 3]:
    ws.cell(row=18, column=col).border = thin_bottom
    ws.cell(row=19, column=col).border = double_bottom

ws.column_dimensions['A'].width = 75
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20

wb.save(path)
print(f"Plantilla ORI creada en {path}")
