import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def create_nota_template():
    # Rutas
    empresa_path = r"data/empresas/Pacifico SpA"
    os.makedirs(empresa_path, exist_ok=True)
    out_path = os.path.join(empresa_path, "Nota Efectivo y equivalentes.xlsx")
    
    os.makedirs("templates", exist_ok=True)
    template_path = os.path.join("templates", "Nota Efectivo y equivalentes.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Nota Efectivo"
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')
    
    # Titulo de la nota
    ws.merge_cells('A1:D1')
    ws['A1'] = "NOTAS A LOS ESTADOS FINANCIEROS"
    ws['A1'].font = bold_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A3:D3')
    ws['A3'] = "NOTA EFECTIVO Y EQUIVALENTES AL EFECTIVO"
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align
    
    # Encabezados (las columnas son A=1, B=2, C=3, D=4)
    ws['B5'] = "Detalle"
    ws['C5'] = "31-12-2025"
    ws['D5'] = "31-12-2024"
    ws['B5'].font = bold_font
    ws['C5'].font = bold_font
    ws['D5'].font = bold_font
    
    # Filas
    rows = [
        # Detalle, Monto 2025, Monto 2024
        ("", "", ""),
        ("Nota 1 - Por Naturaleza", "", ""),
        ("Efectivo en caja", "", ""),
        ("Fondo fijo", "", ""),
        ("Efectivo en bancos", "", ""),
        ("Deposito a plazo", "", ""),
        ("Fondos mutuos", "", ""),
        ("Total", "", ""),
        ("", "", ""),
        ("Nota 2 - Por Moneda", "", ""),
        ("Pesos chilenos", "", ""),
        # Agregar ambas opciones por temas de codificacion vistos en Excel
        ("Dólares estadounidenses", "", ""),
        ("Dlares estadounidenses", "", ""),
        ("Euros", "", ""),
        ("Total", "", ""),
    ]
    
    r_idx = 6
    for col_b, col_c, col_d in rows:
        ws.cell(row=r_idx, column=2, value=col_b)
        ws.cell(row=r_idx, column=3, value=col_c)
        ws.cell(row=r_idx, column=4, value=col_d)
        
        if "Nota" in col_b or "Total" in col_b:
            ws.cell(row=r_idx, column=2).font = bold_font
            
        r_idx += 1
        
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    wb.save(out_path)
    wb.save(template_path)
    print("Plantilla excel de Nota Efectivo generada en:", out_path)

if __name__ == "__main__":
    create_nota_template()
