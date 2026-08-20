import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def create_cf_template():
    # Caminos
    empresa_path = r"data/empresas/Pacifico SpA"
    os.makedirs(empresa_path, exist_ok=True)
    out_path = os.path.join(empresa_path, "Estado de Flujos de Efectivo.xlsx")
    
    # También actualizar el de templates
    os.makedirs("templates", exist_ok=True)
    template_path = os.path.join("templates", "Estado de Flujos de Efectivo.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo_Efectivo"
    
    # Estilos
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')
    
    top_border = Border(top=Side(style='thin'))
    double_bottom_border = Border(bottom=Side(style='double'), top=Side(style='thin'))
    
    # Títulos
    ws.merge_cells('A1:D1')
    ws['A1'] = "PACIFICO CABLE SpA."
    ws['A1'].font = bold_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A3:D3')
    ws['A3'] = "ESTADOS SEPARADOS DE FLUJOS DE EFECTIVO"
    ws['A3'].font = bold_font
    ws['A3'].alignment = center_align
    
    ws.merge_cells('A5:D5')
    ws['A5'] = "Por los ejercicios terminados al 31 de diciembre de 2025 y 2024"
    ws['A5'].alignment = center_align
    
    # Headers
    ws['B8'] = "Nota"
    ws['C8'] = "31-12-2025\nM$"
    ws['D8'] = "31-12-2024\nM$"
    
    ws['B8'].alignment = center_align
    ws['C8'].alignment = center_align
    ws['C8'].font = bold_font
    ws['D8'].alignment = center_align
    ws['D8'].font = bold_font
    
    rows = [
        ("Flujos de efectivo procedentes de (utilizados en) actividades de operación", "", "", "", True),
        ("Cobros procedentes de las ventas de bienes y prestación de servicios", "", "", "", False),
        ("Pagos a proveedores por el suministro de bienes y servicios", "", "", "", False),
        ("Pagos a y por cuenta de los empleados", "", "", "", False),
        ("Intereses pagados", "", "", "", False),
        ("Intereses recibidos", "", "", "", False),
        ("Impuestos a las ganancias reembolsados (pagados)", "", "", "", False),
        ("", "", "", "", False),
        ("Flujos de efectivo procedentes de actividades de operación", "", "", "", False), # subtotal
        ("", "", "", "", False),
        ("Flujos de efectivo procedentes de (utilizados en) actividades de inversión", "", "", "", True),
        ("Compra de Propiedades, planta y equipo", "", "", "", False),
        ("Compra de intangibles", "", "", "", False),
        ("", "", "", "", False),
        ("Flujos de efectivo utilizados en actividades de inversión", "", "", "", False), # subtotal
        ("", "", "", "", False),
        ("Flujos de efectivo procedentes de (utilizados en) actividades de financiación", "", "", "", True),
        ("Importes procedentes de préstamos de largo plazo", "", "", "", False),
        ("Pagos de préstamos", "", "", "", False),
        ("Pagos de pasivos por arrendamientos financieros", "", "", "", False),
        ("Otras entradas y (salidas) de dinero", "", "", "", False),
        ("", "", "", "", False),
        ("Flujos de efectivo procedentes de actividades de financiación", "", "", "", False), # subtotal
        ("", "", "", "", False),
        ("Efectos de la variación en la tasa de cambio sobre el efectivo y equivalentes al efectivo", "", "", "", False),
        ("", "", "", "", False),
        ("Incremento (decremento) neto en efectivo y equivalentes al efectivo", "", "", "", False),
        ("", "", "", "", False),
        ("Saldo inicial de efectivo y equivalentes al efectivo", "", "", "", False),
        ("", "", "", "", False),
        ("Saldo final de efectivo y equivalentes al efectivo", "", "", "", False),
    ]
    
    r_idx = 10
    for title, note, col1, col2, is_bold in rows:
        ws.cell(row=r_idx, column=1, value=title)
        ws.cell(row=r_idx, column=2, value=note)
        ws.cell(row=r_idx, column=3, value=col1)
        ws.cell(row=r_idx, column=4, value=col2)
        
        if is_bold:
            ws.cell(row=r_idx, column=1).font = bold_font
            
        r_idx += 1
        
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    wb.save(out_path)
    wb.save(template_path)
    print("Template Excel created successfully at", out_path)

if __name__ == "__main__":
    create_cf_template()
