import sys
import os
import openpyxl
import pandas as pd
import numpy as np

# Asegurar que el directorio del proyecto está en el PATH de Python
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src.models.database import SessionLocal
from src.core.consolidacion_engine import generar_hoja_trabajo
from src.reporting.balance_generator import BalanceGenerator
from src.reporting.er_generator import ERGenerator
from src.core.excel_utils import detect_balance_columns

def sanitize(text):
    if pd.isna(text): return ""
    import unicodedata
    import re
    clean_str = str(text).strip().lower()
    clean_str = ''.join(c for c in unicodedata.normalize('NFD', clean_str) if unicodedata.category(c) != 'Mn')
    clean_str = re.sub(r'\s+', ' ', clean_str)
    return clean_str

def run_integrity_tests():
    print("====================================================")
    print(" INICIANDO PRUEBAS DE INTEGRIDAD Y ASEGURAMIENTO    ")
    print("====================================================")

    # 1. Obtener la hoja de consolidación del Grupo 1 en 2026-05 y comp 2025-12
    print("\n[Paso 1] Generando hoja de trabajo consolidada...")
    df_hoja_act, msg_act = generar_hoja_trabajo(1, "2026-05")
    df_hoja_comp, msg_comp = generar_hoja_trabajo(1, "2025-12")

    if df_hoja_act is None:
        print("[ERROR]: No se pudo generar la hoja de trabajo de periodo actual.")
        return False
    print("[OK] Hoja de trabajo actual cargada con exito.")
    
    # Paths a plantillas
    grupo_folder = r"data\empresas\[GRUPO] Consolidado DB Terra Holdco"
    path_bal_tpl = os.path.join(grupo_folder, "Balance clasificado.xlsx")
    path_er_tpl = os.path.join(grupo_folder, "Estado de Resultados Clasificados.xlsx")

    scale_factor = 1000.0

    # ----------------------------------------------------
    # TEST DE INTEGRIDAD DEL BALANCE
    # ----------------------------------------------------
    print("\n[Paso 2] Ejecutando BalanceGenerator...")
    wb_bal_tpl = openpyxl.load_workbook(path_bal_tpl)
    ws_bal_tpl = wb_bal_tpl.active
    template_accounts_bal = [str(ws_bal_tpl.cell(row=r, column=2).value).strip() for r in range(1, ws_bal_tpl.max_row+1) if ws_bal_tpl.cell(row=r, column=2).value]

    def clean_str(s):
        if pd.isna(s): return ""
        return str(s).strip().lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')

    bs_subtotals = {
        "activos corrientes", "activos corrientes totales",
        "activos no corrientes", "activos no corrientes totales",
        "total activos", "patrimonio y pasivos",
        "pasivos corrientes", "pasivo corrientes totales",
        "pasivos no corrientes", "pasivo no corrientes totales",
        "total pasivos", "patrimonio", "patrimonio total",
        "total patrimonio y pasivos", "estado de resultados"
    }
    bs_subtotals_normalized = {clean_str(name) for name in bs_subtotals}

    idx_er = df_hoja_act[df_hoja_act['Balance clasificado'] == "Estado de Resultados"].index
    df_hoja_act_sec = df_hoja_act.loc[:idx_er[0]-1] if not idx_er.empty else df_hoja_act
    df_hoja_act_clean = df_hoja_act_sec[df_hoja_act_sec['Balance clasificado'].notna() & (df_hoja_act_sec['Balance clasificado'].str.strip() != "")]
    df_hoja_act_clean = df_hoja_act_clean[~df_hoja_act_clean['Balance clasificado'].apply(clean_str).isin(bs_subtotals_normalized)]

    tb_df_bal = pd.DataFrame({
        'cuenta_id': df_hoja_act_clean['Balance clasificado'],
        'saldo_final': df_hoja_act_clean['CONSOLIDADO']
    })

    tb_df_comp_bal = None
    if df_hoja_comp is not None:
        idx_er_comp = df_hoja_comp[df_hoja_comp['Balance clasificado'] == "Estado de Resultados"].index
        df_hoja_comp_sec = df_hoja_comp.loc[:idx_er_comp[0]-1] if not idx_er_comp.empty else df_hoja_comp
        df_hoja_comp_clean = df_hoja_comp_sec[df_hoja_comp_sec['Balance clasificado'].notna() & (df_hoja_comp_sec['Balance clasificado'].str.strip() != "")]
        df_hoja_comp_clean = df_hoja_comp_clean[~df_hoja_comp_clean['Balance clasificado'].apply(clean_str).isin(bs_subtotals_normalized)]
        tb_df_comp_bal = pd.DataFrame({
            'cuenta_id': df_hoja_comp_clean['Balance clasificado'],
            'saldo_final': df_hoja_comp_clean['CONSOLIDADO']
        })

    dummy_map_bal = pd.DataFrame({
        'N° de Cuenta': tb_df_bal['cuenta_id'],
        'Clasificación balance': tb_df_bal['cuenta_id']
    })
    dummy_map_bal.loc[~dummy_map_bal['N° de Cuenta'].isin(template_accounts_bal), 'Clasificación balance'] = pd.NA

    gen_bal = BalanceGenerator(path_bal_tpl)
    res_bal_bytes = gen_bal.generate(
        tb_df=tb_df_bal,
        map_balance_df=dummy_map_bal,
        scale_factor=scale_factor,
        tb_df_comp=tb_df_comp_bal,
        periodo_actual_str="2026-05",
        periodo_comp_str="2025-12"
    )

    wb_bal = openpyxl.load_workbook(res_bal_bytes)
    ws_bal = wb_bal.active
    name_col, nota_col, val25_col, val24_col = detect_balance_columns(ws_bal)

    print(f"   -> Columnas detectadas: concepto={name_col}, notas={nota_col}, actual={val25_col}, comp={val24_col}")

    # Verificar que no haya NaNs en columnas de valores y extraer totales
    totales = {}
    for col_idx, label in [(val25_col, "2026-05"), (val24_col, "2025-12")]:
        totales[label] = {}
        for r in range(1, ws_bal.max_row + 1):
            name = ws_bal.cell(row=r, column=name_col).value
            val = ws_bal.cell(row=r, column=col_idx).value
            
            if name and isinstance(name, str):
                name_clean = sanitize(name)
                
                # Chequear NaN
                if val is not None:
                    val_str = str(val).lower()
                    if "nan" in val_str or "none" in val_str or "#val" in val_str:
                        print(f"[ERROR]: Celda con valor invalido en fila {r} (columna {label}): {name} = {val}")
                        return False
                
                # Capturar totales clave
                if "total activos" in name_clean and "no corrientes" not in name_clean and "corrientes" not in name_clean:
                    totales[label]["Activos"] = float(val or 0.0)
                elif "total patrimonio y pasivos" in name_clean:
                    totales[label]["PasivosPatrimonio"] = float(val or 0.0)
                elif "patrimonio total" in name_clean:
                    totales[label]["Patrimonio"] = float(val or 0.0)
                elif "total pasivos" in name_clean and "no corrientes" not in name_clean and "corrientes" not in name_clean:
                    totales[label]["Pasivos"] = float(val or 0.0)
                elif "total activos no corrientes" in name_clean:
                    totales[label]["ActivosNoCorrientes"] = float(val or 0.0)
                elif "activos corrientes totales" in name_clean:
                    totales[label]["ActivosCorrientes"] = float(val or 0.0)

                # Validar preservacion de bordes en balance
                tpl_cell_border = ws_bal_tpl.cell(row=r, column=col_idx).border
                gen_cell_border = ws_bal.cell(row=r, column=col_idx).border
                if tpl_cell_border and (tpl_cell_border.top.style or tpl_cell_border.bottom.style):
                    if not gen_cell_border or gen_cell_border.top.style != tpl_cell_border.top.style or gen_cell_border.bottom.style != tpl_cell_border.bottom.style:
                        print(f"[ERROR]: Se perdio el borde del Balance en fila {r} ({label}): tpl={tpl_cell_border.top.style}/{tpl_cell_border.bottom.style}, gen={gen_cell_border.top.style}/{gen_cell_border.bottom.style}")
                        return False

    print("   -> Resultados de Totales en Balance:")
    for label, vals in totales.items():
        act = vals.get("Activos", 0.0)
        pas_pat = vals.get("PasivosPatrimonio", 0.0)
        act_c = vals.get("ActivosCorrientes", 0.0)
        act_nc = vals.get("ActivosNoCorrientes", 0.0)
        diff = abs(act - pas_pat)
        
        print(f"      [{label}] Activos={act:,.2f} (Corr={act_c:,.2f}, NoCorr={act_nc:,.2f}) | Pasivos+Pat={pas_pat:,.2f} | Diferencia={diff:,.2f}")
        
        # Validaciones críticas de negocio
        if act == 0.0 or pas_pat == 0.0:
            print(f"[ERROR]: Los totales de la columna {label} no pueden ser cero.")
            return False
        if act_nc == 0.0:
            print(f"[ERROR]: Total Activos No Corrientes es cero en {label}. Hay un problema en la clasificacion o acumulacion.")
            return False
        if diff > 10.0:
            print(f"[ERROR]: Descuadre matematico detectado en columna {label}. Activos != Pasivos + Patrimonio (Dif={diff:,.2f})")
            return False
        
    print("[OK] TEST BALANCE: PASO (Cuadre perfecto, bordes y cero NaNs).")

    # ----------------------------------------------------
    # TEST DE ESTADO DE RESULTADOS
    # ----------------------------------------------------
    print("\n[Paso 3] Ejecutando ERGenerator...")
    pl_dict_act = {row['Balance clasificado']: [row['CONSOLIDADO']] for _, row in df_hoja_act.iterrows() if pd.notna(row['Balance clasificado'])}
    pl_df_wide = pd.DataFrame(pl_dict_act)
    
    pl_df_comp_wide = None
    if df_hoja_comp is not None:
        pl_dict_comp = {row['Balance clasificado']: [row['CONSOLIDADO']] for _, row in df_hoja_comp.iterrows() if pd.notna(row['Balance clasificado'])}
        pl_df_comp_wide = pd.DataFrame(pl_dict_comp)

    wb_er_tpl = openpyxl.load_workbook(path_er_tpl)
    ws_er_tpl = wb_er_tpl.active

    gen_er = ERGenerator(path_er_tpl)
    res_er_bytes, _ = gen_er.generate(
        pl_df=pl_df_wide,
        scale_factor=scale_factor,
        pl_df_comp=pl_df_comp_wide,
        periodo_actual_str="2026-05",
        periodo_comp_str="2025-12"
    )

    wb_er = openpyxl.load_workbook(res_er_bytes)
    ws_er = wb_er.active

    # Detectar columnas ER
    # 1. Concepto
    clasif_col_idx = 1
    for col in range(1, 10):
        for row in range(1, 15):
            val = ws_er.cell(row=row, column=col).value
            if val and any(x in str(val).lower() for x in ["ingresos de actividades", "costo de ventas", "ganancia bruta", "resultado antes"]):
                clasif_col_idx = col
                break
        else:
            continue
        break
    
    # 2. Fechas
    date_cols_er = []
    for col in range(1, ws_er.max_column + 1):
        if col == clasif_col_idx:
            continue
        for row in range(1, 5):
            val = ws_er.cell(row=row, column=col).value
            if val and ("mayo" in str(val).lower() or "diciembre" in str(val).lower() or "2026" in str(val) or "2025" in str(val)):
                date_cols_er.append(col)
                break
    
    date_cols_er = sorted(list(set(date_cols_er)))
    val25_col_er = date_cols_er[0] if len(date_cols_er) >= 1 else 3
    val24_col_er = date_cols_er[1] if len(date_cols_er) >= 2 else 4

    print(f"   -> Columnas ER detectadas: concepto={clasif_col_idx}, actual={val25_col_er}, comp={val24_col_er}")

    # Verificar que no haya NaNs, que existan montos y que los subtotales conserven estilos
    verificado_ingresos = False
    verificado_borders = True
    
    for row in range(1, ws_er.max_row + 1):
        name = ws_er.cell(row=row, column=clasif_col_idx).value
        val_act = ws_er.cell(row=row, column=val25_col_er).value
        
        if name and isinstance(name, str):
            name_clean = sanitize(name)
            
            # Chequear NaN en valores
            for col_idx, label in [(val25_col_er, "Actual"), (val24_col_er, "Comp")]:
                val = ws_er.cell(row=row, column=col_idx).value
                if val is not None:
                    val_str = str(val).lower()
                    if "nan" in val_str or "none" in val_str or "#val" in val_str:
                        print(f"[ERROR]: Celda con valor invalido en ER fila {row} ({label}): {name} = {val}")
                        return False
            
            # Chequear que haya valores reales
            if "ingresos de actividades ordinarias" in name_clean:
                if val_act is not None and float(val_act) != 0.0:
                    verificado_ingresos = True
                    print(f"      Ingresos inyectados en ER: {float(val_act):,.2f}")
            
            # Validar preservacion de bordes en ER
            for col_idx, label in [(val25_col_er, "Actual"), (val24_col_er, "Comp")]:
                tpl_cell_border = ws_er_tpl.cell(row=row, column=col_idx).border
                gen_cell_border = ws_er.cell(row=row, column=col_idx).border
                if tpl_cell_border and (tpl_cell_border.top.style or tpl_cell_border.bottom.style):
                    if not gen_cell_border or gen_cell_border.top.style != tpl_cell_border.top.style or gen_cell_border.bottom.style != tpl_cell_border.bottom.style:
                        print(f"[ERROR]: Se perdió el borde en ER fila {row} ({label}): tpl={tpl_cell_border.top.style}/{tpl_cell_border.bottom.style}, gen={gen_cell_border.top.style}/{gen_cell_border.bottom.style}")
                        verificado_borders = False
                        return False

    if not verificado_ingresos:
        print("[ERROR]: No se inyectaron ingresos operacionales en el Estado de Resultados.")
        return False
    if not verificado_borders:
        print("[ERROR]: Se perdieron los bordes o estilos de subtotal en el Estado de Resultados.")
        return False

    print("[OK] TEST ESTADO DE RESULTADOS: PASO (Estilos, montos y cero NaNs validados).")

    print("\n====================================================")
    print("  ¡ASEGURAMIENTO DE INTEGRIDAD COMPLETADO CON EXITO! ")
    print("====================================================")
    return True

if __name__ == "__main__":
    success = run_integrity_tests()
    sys.exit(0 if success else 1)
