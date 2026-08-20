import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def update_efectivo(wb):
    ws = wb['Efectivo']
    print("Updating sheet: Efectivo")
    
    # Unmerge cells to prevent corruption when moving cells
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    
    # ----------------------------------------------------
    # TABLE 1 & 2: Shift values from C, E, F to E, F, G
    # and add column I as comparative (31.12.2024)
    # ----------------------------------------------------
    
    # Move headers
    ws['E3'] = 'Pacifico'
    ws['F3'] = 'Holdco'
    ws['G3'] = 'Consolidado'
    ws['I3'] = 'Consolidado'
    
    ws['E4'] = '31.12.2025'
    ws['F4'] = '31.12.2025'
    ws['G4'] = '31.12.2025'
    ws['I4'] = '31.12.2024'
    
    ws['E5'] = 'M$'
    ws['F5'] = 'M$'
    ws['G5'] = 'M$'
    ws['I5'] = 'M$'
    
    # Set borders and font styling for new headers
    header_font = Font(name='Calibri', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    
    for col in ['E', 'F', 'G', 'I']:
        for row in [3, 4, 5]:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = align_center
            
    # Shift rows 6 to 12 (Table 1)
    for r in range(6, 13):
        # Shift values: C -> E, E -> F, F -> G
        val_c = ws[f'C{r}'].value
        val_e = ws[f'E{r}'].value
        val_f = ws[f'F{r}'].value
        
        ws[f'E{r}'] = val_c
        ws[f'F{r}'] = val_e
        ws[f'G{r}'] = val_f
        ws[f'I{r}'] = None if r != 12 else 0 # validation row gets 0
        
        # Clear original columns C and D (value columns)
        ws[f'C{r}'] = None
        ws[f'D{r}'] = None
        
    # Re-apply formula for Table 1 Total (Row 10)
    ws['E10'] = '=SUM(E6:E9)'
    ws['F10'] = '=SUM(F6:F9)'
    ws['G10'] = '=SUM(G6:G9)'
    ws['I10'] = '=SUM(I6:I9)'
    
    # Re-apply validation formulas
    ws['E12'] = '=E10-E11'
    ws['F12'] = '=F10-F11'
    ws['G12'] = '=G10-G11'
    ws['I12'] = '=I10-I11'
    
    # Shift rows 14 to 21 (Table 2: Fondos Mutuos currency)
    ws['E14'] = '31.12.2025'
    ws['I14'] = '31.12.2024'
    ws['E15'] = 'M$'
    ws['F15'] = 'M$'
    ws['G15'] = 'M$'
    ws['I15'] = 'M$'
    
    for r in range(16, 22):
        val_c = ws[f'C{r}'].value
        val_e = ws[f'E{r}'].value
        val_f = ws[f'F{r}'].value
        
        ws[f'E{r}'] = val_c
        ws[f'F{r}'] = val_e
        ws[f'G{r}'] = val_f
        ws[f'I{r}'] = None if r != 21 else 0
        
        ws[f'C{r}'] = None
        ws[f'D{r}'] = None
        
    ws['E19'] = '=SUM(E16:E18)'
    ws['F19'] = '=SUM(F16:F18)'
    ws['G19'] = '=SUM(G16:G18)'
    ws['I19'] = '=SUM(I16:I18)'
    
    ws['E21'] = '=E19-E12'
    ws['F21'] = '=F19-F12'
    ws['G21'] = '=G19-G12'
    ws['I21'] = '=I19-I12'
    
    # ----------------------------------------------------
    # TABLE 3: Restructure Fondos Mutuos Bank Detail
    # ----------------------------------------------------
    
    # Set headers in Row 29 & 31 & 32
    ws['E29'] = 'Pacifico'
    ws['F29'] = 'Holdco'
    ws['G29'] = 'Consolidado'
    ws['I29'] = 'Consolidado'
    
    ws['E31'] = '31.12.2025'
    ws['I31'] = '31.12.2024'
    
    ws['E32'] = 'M$'
    ws['F32'] = 'M$'
    ws['G32'] = 'M$'
    ws['I32'] = 'M$'
    
    # Row 30 title update
    ws['B30'] = 'El detalle de los fondos mutuos al cierre del año 2025 y 2024 se presenta a continuación:'
    
    # Move comparative values from C8 (column H/Consolidado Comparative 2024) to Column I (9)
    # Note: In original Table 3, Pacifico comparative was F (6), Holdco was G (7), Consolidado was H (8)
    for r in range(33, 39):
        comp_val = ws.cell(row=r, column=8).value # Col H (8) is Consolidado 2024 comparative in original
        ws[f'I{r}'] = comp_val
        
        # Clear original Columns F(6), G(7), H(8) values for this table
        ws.cell(row=r, column=6).value = None
        ws.cell(row=r, column=7).value = None
        ws.cell(row=r, column=8).value = None
        
        # Also clear any other values in E to make sure it's ready for generator
        ws[f'E{r}'] = None
        ws[f'F{r}'] = None
        ws[f'G{r}'] = None
        
    # Re-apply formulas
    ws['E39'] = '=SUM(E33:E38)'
    ws['F39'] = '=SUM(F33:F38)'
    ws['G39'] = '=SUM(G33:G38)'
    ws['I39'] = '=SUM(I33:I38)'

    # Merge headers to match clean layout
    # Re-merge titles and entities
    ws.merge_cells('E3:E3')
    ws.merge_cells('F3:F3')
    ws.merge_cells('G3:G3')
    ws.merge_cells('I3:I3')
    ws.merge_cells('E29:E29')
    ws.merge_cells('F29:F29')
    ws.merge_cells('G29:G29')
    ws.merge_cells('I29:I29')

def update_deudores(wb):
    ws = wb['Deudores']
    print("Updating sheet: Deudores")
    
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    
    # ----------------------------------------------------
    # Table 1: Composition (Rows 1-17)
    # ----------------------------------------------------
    
    # Set headers
    ws['C2'] = 'PACIFICO'
    ws['D2'] = 'HOLDCO'
    ws['E2'] = 'CONSOLIDADO'
    ws['G2'] = 'CONSOLIDADO'
    
    ws['C3'] = '31.12.2025'
    ws['D3'] = '31.12.2025'
    ws['E3'] = '31.12.2025'
    ws['G3'] = '31.12.2024'
    
    ws['C4'] = 'M$'
    ws['D4'] = 'M$'
    ws['E4'] = 'M$'
    ws['G4'] = 'M$'
    
    header_font = Font(name='Calibri', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    
    for col in ['C', 'D', 'E', 'G']:
        for row in [2, 3, 4]:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = align_center
            
    # Set formulas in Totals
    ws['C14'] = '=SUM(C5:C13)'
    ws['D14'] = '=SUM(D5:D13)'
    ws['E14'] = '=SUM(E5:E13)'
    ws['G14'] = '=SUM(G5:G13)'
    
    # Validations
    ws['C17'] = '=C14-C17' # or clear them, generator handles validations
    
    # ----------------------------------------------------
    # Table 2: Vencimientos (Rows 20-35)
    # ----------------------------------------------------
    ws['C22'] = 'PACIFICO'
    ws['D22'] = 'HOLDCO'
    ws['E22'] = 'CONSOLIDADO'
    ws['G22'] = 'CONSOLIDADO'
    
    ws['C23'] = '31.12.2025'
    ws['D23'] = '31.12.2025'
    ws['E23'] = '31.12.2025'
    ws['G23'] = '31.12.2024'
    
    ws['C24'] = 'M$'
    ws['D24'] = 'M$'
    ws['E24'] = 'M$'
    ws['G24'] = 'M$'
    
    for col in ['C', 'D', 'E', 'G']:
        for row in [22, 23, 24]:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = align_center
            
    ws['C33'] = '=SUM(C26:C32)'
    ws['D33'] = '=SUM(D26:D32)'
    ws['E33'] = '=SUM(E26:E32)'
    ws['G33'] = '=SUM(G26:G32)'
    
    # ----------------------------------------------------
    # Table 3: Movimientos Deterioro (Rows 38-50)
    # ----------------------------------------------------
    ws['C40'] = 'PACIFICO'
    ws['D40'] = 'HOLDCO'
    ws['E40'] = 'CONSOLIDADO'
    ws['G40'] = 'CONSOLIDADO'
    
    ws['C41'] = '31.12.2025'
    ws['D41'] = '31.12.2025'
    ws['E41'] = '31.12.2025'
    ws['G41'] = '31.12.2024'
    
    ws['C42'] = 'M$'
    ws['D42'] = 'M$'
    ws['E42'] = 'M$'
    ws['G42'] = 'M$'
    
    for col in ['C', 'D', 'E', 'G']:
        for row in [40, 41, 42]:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = align_center
            
    ws['C48'] = '=SUM(C44:C47)'
    ws['D48'] = '=SUM(D44:D47)'
    ws['E48'] = '=SUM(E44:E47)'
    ws['G48'] = '=SUM(G44:G47)'

def update_pasivos_derechos_uso(wb):
    ws = wb['Pasivos derechos de  uso']
    print("Updating sheet: Pasivos derechos de  uso")
    
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    
    # Set years in row 3
    for col in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}3'] = '31.12.2025'
        
    ws['I3'] = '31.12.2024'
    ws['J3'] = '31.12.2024'
    
    # Set entities in row 4
    ws['C4'] = 'Pacifico'
    ws['E4'] = 'Holdco'
    ws['G4'] = 'Consolidado'
    ws['I4'] = 'Consolidado'
    
    # Set Conceptos in row 5
    ws['I5'] = 'Corriente'
    ws['J5'] = 'No corriente'
    
    # Set Units in row 6
    ws['I6'] = 'M$'
    ws['J6'] = 'M$'
    
    # Set fonts and styling
    header_font = Font(name='Calibri', size=11, bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        for row in [3, 4, 5, 6]:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = align_center
            
    # Set borders for new columns I and J
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in range(3, ws.max_row + 1):
        ws[f'I{r}'].border = thin_border
        ws[f'J{r}'].border = thin_border

if __name__ == "__main__":
    wb_file = "Plantilla de notas_v1.xlsx"
    wb = openpyxl.load_workbook(wb_file)
    
    update_efectivo(wb)
    update_deudores(wb)
    update_pasivos_derechos_uso(wb)
    
    wb.save(wb_file)
    print("Template updated successfully!")
