import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import unicodedata
import re
import datetime
import calendar

def clean_label(text):
    if text is None:
        return ""
    text_str = str(text).strip().lower()
    text_str = ''.join(c for c in unicodedata.normalize('NFD', text_str) if unicodedata.category(c) != 'Mn')
    return " ".join(text_str.split())

def determine_factor_for_accounts(accounts):
    factors = []
    for acc in accounts:
        acc_str = str(acc).strip()
        if acc_str.startswith('2') or acc_str.startswith('3101'):
            factors.append(-1)
        else:
            factors.append(1)
    # Si la mayoría o al menos la mitad de las cuentas son acreedoras (pasivo/patrimonio), el factor global es -1
    neg_count = sum(1 for f in factors if f == -1)
    if neg_count >= len(factors) / 2:
        return -1
    return 1

def check_extra_word_match(ew, acct_names):
    if not ew or ew in ['$', '%', 'm$', 'm', 'detalle', 'total']:
        return True
    
    acct_names_clean = [clean_label(n) for n in acct_names]
    ew_tokens = ew.split()
    
    keywords = ['bci', 'scotiabank', 'santander', 'estado', 'itau', 'chile', 'bice', 'security']
    for kw in keywords:
        if kw in ew_tokens:
            if not any(kw in name for name in acct_names_clean):
                return False
                
    return True

def detect_row_corriente_status(ws, row):
    # Escanear hacia atrás las celdas en Col 1 y 2
    for r in range(row, 0, -1):
        for c in [1, 2]:
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).strip().lower()
                if any(kw in val_str for kw in ["moneda", "vencimiento", "vencimientos"]):
                    return "unknown"
                if "no corriente" in val_str or "no-corriente" in val_str or "largo plazo" in val_str:
                    return "no_corriente"
                elif "corriente" in val_str:
                    return "corriente"
    return "unknown"

SHEET_TO_PL_RUBROS = {
    "Ingresos Ctos operacion": [
        "ingresos de actividades ordinarias",
        "costo de ventas",
        "costos de uso fibra optica",
        "depreciacion operacional"
    ],
    "Gtos Adm": [
        "gastos de administracion",
        "depreciacion y amortizaciones"
    ],
    "DC y Reajustes": [
        "diferencias de cambio",
        "resultados por unidades de reajuste"
    ],
    "Costos e ingresos Financieros": [
        "ingresos financieros",
        "costos financieros"
    ],
    "Otros ingresos por funcion": [
        "otros ingresos por funcion"
    ],
    "Otros gastos por funcion": [
        "otros egresos por funcion"
    ]
}

def get_filtered_value(period_ctx, matched_key, match_type, flow_key, row, ws):
    """
    Retorna el valor filtrado (corriente vs no corriente) para el matched_key.
    Si el estado de la fila es corriente o no corriente, se filtran las cuentas
    en el contexto según su Clasificación balance.
    """
    if match_type == 'pl':
        pl_ctx = period_ctx.get('pl', {})
        sheet_title = str(ws.title).strip()
        allowed_rubros = SHEET_TO_PL_RUBROS.get(sheet_title, [])
        
        # Si la pestaña no está en el mapa, buscar en todos los rubros como fallback
        if not allowed_rubros:
            for rubro_data in pl_ctx.values():
                if matched_key in rubro_data:
                    return float(rubro_data[matched_key].get(flow_key, 0.0))
            return 0.0
            
        # Buscar en los rubros específicos permitidos para esta pestaña
        for rubro in allowed_rubros:
            rubro_clean = ''.join(c for c in unicodedata.normalize('NFD', rubro.lower()) if unicodedata.category(c) != 'Mn')
            rubro_clean = re.sub(r'\s+', ' ', rubro_clean).strip()
            
            rubro_data = pl_ctx.get(rubro_clean)
            if rubro_data and matched_key in rubro_data:
                return float(rubro_data[matched_key].get(flow_key, 0.0))
        return 0.0

    group_data = period_ctx.get(match_type, {}).get(matched_key)
    if not group_data:
        return 0.0
        
    # Si no tiene acct_details (o es P&L), devolvemos el valor agregado directamente
    if 'acct_details' not in group_data:
        return float(group_data.get(flow_key, 0.0))
        
    status = detect_row_corriente_status(ws, row)
    if status == "unknown":
        return float(group_data.get(flow_key, 0.0))
        
    # Filtrar detalles de cuentas
    filtered_sum = 0.0
    for detail in group_data['acct_details']:
        clasif = detail['clasificacion_bal'].lower()
        is_nc = "no corriente" in clasif or "no-corriente" in clasif or "largo plazo" in clasif
        
        if status == "no_corriente" and is_nc:
            filtered_sum += detail.get(flow_key, 0.0)
        elif status == "corriente" and not is_nc:
            filtered_sum += detail.get(flow_key, 0.0)
            
    return filtered_sum


def format_date_dots(period_str):
    if not period_str or period_str == "Ninguno":
        return ""
    try:
        parts = period_str.strip().split('-')
        if len(parts) >= 2:
            year = parts[0]
            month = parts[1]
            last_day = str(calendar.monthrange(int(year), int(month))[1]).zfill(2)
            return f"{last_day}.{month}.{year}"
    except:
        pass
    return period_str

def format_period_to_spanish_date(period_str):
    if not period_str or period_str == "Ninguno":
        return ""
    try:
        parts = str(period_str).strip().split('-')
        if len(parts) >= 2:
            year = int(parts[0])
            month = int(parts[1])
            last_day = calendar.monthrange(year, month)[1]
            months_es = {
                1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
                5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
                9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
            }
            month_name = months_es[month]
            return f"{last_day} de {month_name} {year}"
    except Exception:
        pass
    return period_str

def build_entity_context(tb_df, map_balance_df, map_pl_df=None, empresa_name=None, periodo_str=None):
    """
    Agrupa los saldos de la TB para una entidad según las descripciones
    de mapeo de balance y P&L.
    """
    if tb_df is None or tb_df.empty:
        return {'nota1': {}, 'nota2': {}, 'pl': {}}
        
    mapping_cols = ['N° de Cuenta \n', 'N de Cuenta', 'N° de Cuenta', 'Cuenta', 'cuenta_id']
    tb_cuenta_col = next((c for c in tb_df.columns if c.strip() in [mc.strip() for mc in mapping_cols]), None)
    
    saldo_cols = ['Saldo DR/CR\n', 'Saldo DR/CR', 'Saldos', 'Saldo', 'saldo_final']
    saldo_col = next((c for c in tb_df.columns if c.strip() in [sc.strip() for sc in saldo_cols]), None)
    
    tb_nombre_cols = ['Nombre de la cuenta', 'Nombre cuenta', 'Nombre', 'descripcion', 'Detalle']
    tb_nombre_col = next((c for c in tb_df.columns if c.strip() in [nc.strip() for nc in tb_nombre_cols]), None)

    if not tb_cuenta_col or not saldo_col:
        return {'nota1': {}, 'nota2': {}, 'pl': {}}

    tb_df = tb_df.copy()
    tb_df[tb_cuenta_col] = tb_df[tb_cuenta_col].astype(str).str.strip()
    tb_df[saldo_col] = pd.to_numeric(tb_df[saldo_col], errors='coerce').fillna(0.0)

    sums_nota1 = {}
    sums_nota2 = {}
    pl_mappings = {}

    # Mapeo de Balance
    if map_balance_df is not None and not map_balance_df.empty:
        map_balance_copy = map_balance_df.copy()
        nota1_col = next((c for c in map_balance_copy.columns if "nota 1" in c.lower()), None)
        nota2_col = next((c for c in map_balance_copy.columns if "nota 2" in c.lower()), None)
        if nota1_col:
            map_balance_copy[nota1_col] = map_balance_copy[nota1_col].astype(str).str.strip()
        if nota2_col:
            map_balance_copy[nota2_col] = map_balance_copy[nota2_col].astype(str).str.strip()
            
        from src.core.sabana_builder import build_balance_sabana
        merged_df = build_balance_sabana(tb_df, map_balance_copy)
        
        map_cuenta_col = next((c for c in map_balance_copy.columns if c.strip() in [mc.strip() for mc in mapping_cols]), None)
        if map_cuenta_col:
            def get_grouped_sums(merged_df_obj, group_col):
                if not group_col or group_col not in merged_df_obj.columns:
                    return {}
                grouped = merged_df_obj.groupby(group_col)
                result = {}
                for name, gp in grouped:
                    records = gp.to_dict('records')
                    accts = [str(r.get(tb_cuenta_col, '')).strip() for r in records if tb_cuenta_col in r]
                    acct_names = [r.get(tb_nombre_col, '') for r in records if tb_nombre_col and tb_nombre_col in r]
                    
                    credit_abs = 0.0
                    debit_abs = 0.0
                    acct_details = []
                    
                    for r_val in records:
                        acc_id = str(r_val.get(tb_cuenta_col, '')).strip()
                        bal_raw = r_val.get(saldo_col, 0.0) if saldo_col in r_val else 0.0
                        bal = abs(float(bal_raw)) if pd.notna(bal_raw) else 0.0
                        if acc_id.startswith('2') or acc_id.startswith('3101'):
                            credit_abs += bal
                        else:
                            debit_abs += bal

                    if credit_abs == 0.0 and debit_abs == 0.0:
                        factor = determine_factor_for_accounts(accts)
                    else:
                        factor = -1 if credit_abs >= debit_abs else 1
                    
                    sum_val = float(gp[saldo_col].sum()) if saldo_col in gp.columns else 0.0
                    sum_ini = float(gp['saldo_inicial'].sum()) if 'saldo_inicial' in gp.columns else 0.0
                    sum_deb = float(gp['debitos'].sum()) if 'debitos' in gp.columns else 0.0
                    sum_cred = float(gp['creditos'].sum()) if 'creditos' in gp.columns else 0.0

                    for r_val in records:
                        acc_id = str(r_val.get(tb_cuenta_col, '')).strip()
                        val_raw = r_val.get(saldo_col, 0.0) if saldo_col in r_val else 0.0
                        val_num = float(val_raw) if pd.notna(val_raw) else 0.0
                        ini_num = float(r_val.get('saldo_inicial', 0.0)) if pd.notna(r_val.get('saldo_inicial')) else 0.0
                        deb_num = float(r_val.get('debitos', 0.0)) if pd.notna(r_val.get('debitos')) else 0.0
                        cred_num = float(r_val.get('creditos', 0.0)) if pd.notna(r_val.get('creditos')) else 0.0
                        
                        acct_details.append({
                            'cuenta_id': acc_id,
                            'nombre': r_val.get(tb_nombre_col, '') if tb_nombre_col else '',
                            'val': val_num * factor,
                            'inicial': ini_num * factor,
                            'debitos': deb_num * factor,
                            'creditos': cred_num * factor,
                            'clasificacion_bal': str(r_val.get('Clasificación balance', '') or '').strip()
                        })
                    
                    result[name] = {
                        'val': sum_val * factor,
                        'inicial': sum_ini * factor,
                        'debitos': sum_deb * factor,
                        'creditos': sum_cred * factor,
                        'accts': accts,
                        'acct_names': acct_names,
                        'acct_details': acct_details
                    }
                return result

            if nota1_col:
                sums_nota1 = get_grouped_sums(merged_df, nota1_col)
            if nota2_col:
                sums_nota2 = get_grouped_sums(merged_df, nota2_col)

    # Mapeo de P&L
    # Columnas a excluir del mapeo de notas en map_pl
    _PL_EXCLUDE_COLS = {
        'Cuenta', 'Detalle', 'Clasificación Flujo Efectivo', 'Clasificacion',
        'Clasificaci\u00f3n Flujo Efectivo'
    }
    if map_pl_df is not None and not map_pl_df.empty:
        from src.models.pl_cubo_db import PlCuboDB
        from src.core.sabana_builder import build_pl_sabana
        import unicodedata
        import re
        
        # Cargar pl_df combinando todas las empresas involucradas
        pl_df_list = []
        cos = [empresa_name] if isinstance(empresa_name, str) else empresa_name
        for co in cos:
            df_pl = PlCuboDB.get_pl_cubo(co, periodo_str) if periodo_str else None
            if df_pl is not None and not df_pl.empty:
                pl_df_list.append(df_pl)
                
        if pl_df_list:
            pl_df_comb = pd.concat(pl_df_list, ignore_index=True)
            pl_cuenta_col = next((c for c in pl_df_comb.columns if "cuenta" in str(c).lower() and "nombre" not in str(c).lower()), pl_df_comb.columns[0])
            pl_nombre_col = next((c for c in pl_df_comb.columns if "nombre" in str(c).lower()), pl_df_comb.columns[1] if len(pl_df_comb.columns) > 1 else None)
            
            ignorar_cols = [pl_cuenta_col]
            if pl_nombre_col:
                ignorar_cols.append(pl_nombre_col)
                
            agg_dict = {c: 'sum' for c in pl_df_comb.columns if c not in ignorar_cols}
            if pl_nombre_col:
                agg_dict[pl_nombre_col] = 'first'
                
            pl_df = pl_df_comb.groupby(pl_cuenta_col, as_index=False).agg(agg_dict)
        else:
            pl_df = None
            
        merged_pl = build_pl_sabana(None, map_pl_df, tb_df)
        if tb_df is not None and not tb_df.empty:
            tb_acct_set = set(tb_df[tb_cuenta_col].astype(str).str.strip())
            map_pl_cuenta_col = next((c for c in map_pl_df.columns if c.strip() in [mc.strip() for mc in mapping_cols]), None)
            acct_col_in_merged = tb_cuenta_col if tb_cuenta_col in merged_pl.columns else map_pl_cuenta_col
            if acct_col_in_merged and acct_col_in_merged in merged_pl.columns:
                merged_pl = merged_pl[merged_pl[acct_col_in_merged].astype(str).str.strip().isin(tb_acct_set)]
        
        # Pre-cargar datos de P&L Cubo si existen para cuadrar saldos al centavo
        pl_cubo_data = {}  # {col_sanitized: {acc_id: sum_value}}
        has_pl_cubo = pl_df is not None
        
        def pl_sanitize(text):
            if pd.isna(text): return ""
            clean_str = str(text).strip().lower()
            clean_str = ''.join(c for c in unicodedata.normalize('NFD', clean_str) if unicodedata.category(c) != 'Mn')
            clean_str = re.sub(r'\s+', ' ', clean_str)
            return clean_str
            
        if has_pl_cubo:
            ignorar_rubros = ["n° de cuenta", "nombre de la cuenta", "cuenta", "nombre", "unnamed: 0"]
            for col_pl in pl_df.columns:
                if str(col_pl).strip().lower() in ignorar_rubros:
                    continue
                col_san = pl_sanitize(col_pl)
                if col_san not in pl_cubo_data:
                    pl_cubo_data[col_san] = {}
                for _, r_val in pl_df.iterrows():
                    acc_id = str(r_val[pl_df.columns[0]]).strip()
                    val_str = r_val.get(col_pl)
                    if val_str is not None:
                        try:
                            pl_cubo_data[col_san][acc_id] = float(val_str)
                        except ValueError:
                            pass

        map_pl_cuenta_col = next((c for c in map_pl_df.columns if c.strip().lower() in [mc.strip().lower() for mc in mapping_cols]), None)
        if map_pl_cuenta_col:
            split_accounts = set()
            exclude_from_mapping = ['cuenta', 'detalle', 'nombre', 'descripcion', 'flujo', 'unnamed']
            mapping_cols_to_check = [
                c for c in map_pl_df.columns 
                if not any(x in c.lower() for x in exclude_from_mapping)
            ]
            for _, row_val in map_pl_df.iterrows():
                acc_id = str(row_val[map_pl_cuenta_col]).strip()
                non_empty_count = 0
                for c_col in mapping_cols_to_check:
                    val_cell = row_val.get(c_col)
                    if val_cell is not None and str(val_cell).strip().lower() not in ('nan', 'none', ''):
                        non_empty_count += 1
                if non_empty_count > 1:
                    split_accounts.add(acc_id)

            for col in map_pl_df.columns:
                if col in _PL_EXCLUDE_COLS or col.lower().startswith('unnamed'):
                    continue
                
                col_san = col.replace("\xa0", " ").strip().lower()
                clean_col = ''.join(c for c in unicodedata.normalize('NFD', col_san) if unicodedata.category(c) != 'Mn')
                clean_col = re.sub(r'\s+', ' ', clean_col)
                cubo_col_dict = pl_cubo_data.get(clean_col)
                
                if clean_col not in pl_mappings:
                    pl_mappings[clean_col] = {}
                
                grouped = merged_pl.groupby(col)
                for name, gp in grouped:
                    k_str = str(name).strip()
                    if k_str and k_str.lower() != 'nan':
                        ck = clean_label(k_str)
                        accts = gp[tb_cuenta_col].tolist()
                        acct_names = gp[tb_nombre_col].tolist() if tb_nombre_col else []
                        
                        credit_abs = 0.0
                        debit_abs = 0.0
                        for _, r_val in gp.iterrows():
                            acc_id = str(r_val[tb_cuenta_col]).strip()
                            bal = abs(float(r_val[saldo_col])) if (saldo_col in gp.columns and pd.notna(r_val[saldo_col])) else 0.0
                            if acc_id.startswith('2') or acc_id.startswith('3101'):
                                credit_abs += bal
                            else:
                                debit_abs += bal
                        
                        if credit_abs == 0.0 and debit_abs == 0.0:
                            factor = determine_factor_for_accounts(accts)
                        else:
                            factor = -1 if credit_abs >= debit_abs else 1
                        
                        if ck not in pl_mappings[clean_col]:
                            pl_mappings[clean_col][ck] = {
                                'val': 0.0, 'inicial': 0.0, 'debitos': 0.0, 'creditos': 0.0,
                                'accts': [], 'acct_names': []
                            }
                        
                        sum_val = 0.0
                        sum_ini = 0.0
                        sum_deb = 0.0
                        sum_cred = 0.0
                        visited_cubo_accs = set()
                        
                        for _, r_val in gp.iterrows():
                            acc_id = str(r_val[tb_cuenta_col]).strip()
                            
                            use_cubo = False
                            tb_val = float(r_val[saldo_col]) if (saldo_col in gp.columns and pd.notna(r_val[saldo_col])) else 0.0
                            tb_deb = float(r_val.get('debitos', 0.0)) if pd.notna(r_val.get('debitos', 0.0)) else 0.0
                            tb_cred = float(r_val.get('creditos', 0.0)) if pd.notna(r_val.get('creditos', 0.0)) else 0.0
                            tb_ini = float(r_val.get('saldo_inicial', 0.0)) if pd.notna(r_val.get('saldo_inicial', 0.0)) else 0.0
                            
                            has_tb_movement = (abs(tb_val) > 0.001 or abs(tb_deb) > 0.001 or abs(tb_cred) > 0.001 or abs(tb_ini) > 0.001)
                            is_split = (acc_id in split_accounts)
                            
                            if has_pl_cubo:
                                if acc_id not in visited_cubo_accs:
                                    cubo_val = cubo_col_dict.get(acc_id, 0.0) if cubo_col_dict else 0.0
                                    sum_val += cubo_val * factor
                                    sum_ini += cubo_val * factor
                                    visited_cubo_accs.add(acc_id)
                                use_cubo = True
                                    
                            if not use_cubo:
                                sum_val += tb_val * factor
                                sum_ini += tb_ini * factor
                                sum_deb += tb_deb * factor
                                sum_cred += tb_cred * factor
                                
                        pl_mappings[clean_col][ck]['val'] += sum_val
                        pl_mappings[clean_col][ck]['inicial'] += sum_ini
                        pl_mappings[clean_col][ck]['debitos'] += sum_deb
                        pl_mappings[clean_col][ck]['creditos'] += sum_cred
                        pl_mappings[clean_col][ck]['accts'].extend(accts)
                        pl_mappings[clean_col][ck]['acct_names'].extend(acct_names)

    # Limpiar y agregar etiquetas duplicadas después de normalizar espacios
    sums_nota1_clean = {}
    for k, v in sums_nota1.items():
        if k is not None:
            ck = clean_label(k)
            if ck not in sums_nota1_clean:
                sums_nota1_clean[ck] = {
                    'val': 0.0, 'inicial': 0.0, 'debitos': 0.0, 'creditos': 0.0,
                    'accts': [], 'acct_names': [], 'acct_details': []
                }
            sums_nota1_clean[ck]['val'] += v['val']
            sums_nota1_clean[ck]['inicial'] += v['inicial']
            sums_nota1_clean[ck]['debitos'] += v['debitos']
            sums_nota1_clean[ck]['creditos'] += v['creditos']
            sums_nota1_clean[ck]['accts'].extend(v['accts'])
            sums_nota1_clean[ck]['acct_names'].extend(v['acct_names'])
            sums_nota1_clean[ck]['acct_details'].extend(v['acct_details'])

    sums_nota2_clean = {}
    for k, v in sums_nota2.items():
        if k is not None:
            ck = clean_label(k)
            if ck not in sums_nota2_clean:
                sums_nota2_clean[ck] = {
                    'val': 0.0, 'inicial': 0.0, 'debitos': 0.0, 'creditos': 0.0,
                    'accts': [], 'acct_names': [], 'acct_details': []
                }
            sums_nota2_clean[ck]['val'] += v['val']
            sums_nota2_clean[ck]['inicial'] += v['inicial']
            sums_nota2_clean[ck]['debitos'] += v['debitos']
            sums_nota2_clean[ck]['creditos'] += v['creditos']
            sums_nota2_clean[ck]['accts'].extend(v['accts'])
            sums_nota2_clean[ck]['acct_names'].extend(v['acct_names'])
            sums_nota2_clean[ck]['acct_details'].extend(v['acct_details'])

    return {
        'nota1': sums_nota1_clean,
        'nota2': sums_nota2_clean,
        'pl': pl_mappings
    }


def fix_formulas_column_letters(ws, col_letter_map):
    import re
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                def replace_col(match):
                    c_let = match.group(1)
                    r_num = match.group(2)
                    new_c_let = col_letter_map.get(c_let, c_let)
                    return f"{new_c_let}{r_num}"
                new_val = re.sub(r'([A-Z]+)(\d+)', replace_col, val)
                cell.value = new_val

class NoteGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

    def generate(self,
                 sheet_names,
                 entity_contexts,
                 active_entity_name,
                 is_consolidated,
                 scale_factor=1.0,
                 periodo_actual_str=None,
                 periodo_comp_str=None,
                 map_balance_df=None,
                 map_pl_df=None):
        """
        Genera un Excel de notas filtrando y llenando las pestañas indicadas 
        según el contexto consolidado o individual.
        """
        if not sheet_names:
            raise ValueError("No sheet names specified for note generation. The master template must contain at least one sheet for this note.")

        # Determine if this note is a results note
        from src.reporting.notes import NOTE_REGISTRY
        is_results_note = False
        clean_sheet_names = [s.strip().lower() for s in sheet_names]
        for code, info in NOTE_REGISTRY.items():
            reg_sheets = [rs.strip().lower() for rs in info.get('sheets', [])]
            if any(s in clean_sheet_names for s in reg_sheets):
                if info.get('category') == 'resultados':
                    is_results_note = True
                    break

        wb = load_workbook(self.template_path)
        
        # 1. Eliminar hojas no relacionadas (preservando DB_DATA)
        for name in list(wb.sheetnames):
            if name not in sheet_names and name != "DB_DATA":
                wb.remove(wb[name])

        # 1.5 Volcar base de datos plana a DB_DATA
        try:
            from src.models.trial_balance_db import TrialBalanceDB
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            companies = []
            if is_consolidated:
                from src.models.database import SessionLocal
                from src.models.consolidacion import ConsolidationGroup
                grupo_name = active_entity_name.replace("[GRUPO] ", "").strip()
                db = SessionLocal()
                try:
                    grupo_obj = db.query(ConsolidationGroup).filter_by(nombre_grupo=grupo_name).first()
                    if grupo_obj:
                        companies.append(grupo_obj.empresa_matriz)
                        if grupo_obj.filial_is_group:
                            def get_sub_companies(sub_g_id):
                                sub_g = db.query(ConsolidationGroup).filter_by(id=sub_g_id).first()
                                if sub_g:
                                    c = [sub_g.empresa_matriz]
                                    if sub_g.filial_is_group:
                                        c.extend(get_sub_companies(int(sub_g.empresa_filial)))
                                    else:
                                        c.append(sub_g.empresa_filial)
                                    return c
                                return []
                            companies.extend(get_sub_companies(int(grupo_obj.empresa_filial)))
                        else:
                            companies.append(grupo_obj.empresa_filial)
                except Exception as e:
                    print(f"Error al obtener filiales para DB_DATA: {e}")
                finally:
                    db.close()
            else:
                companies = [active_entity_name]
                
            if not companies:
                companies = [active_entity_name]
                
            if map_balance_df is None:
                try:
                    map_balance_df = pd.read_excel(self.template_path, sheet_name="Mapeo Balance")
                except:
                    pass
            if map_pl_df is None:
                try:
                    map_pl_df = pd.read_excel(self.template_path, sheet_name="Mapeo Ctas P&L Cubo")
                except:
                    pass
                    
            rows = []
            periods = [periodo_actual_str]
            if periodo_comp_str and periodo_comp_str != "Ninguno":
                periods.append(periodo_comp_str)
                
            map_bal_clean = None
            if map_balance_df is not None and not map_balance_df.empty:
                map_bal_clean = map_balance_df.copy()
                mc_cols = ['N° de Cuenta \n', 'N de Cuenta', 'N° de Cuenta', 'Cuenta', 'cuenta_id']
                mc_col = next((c for c in map_bal_clean.columns if c.strip() in [x.strip() for x in mc_cols]), None)
                if mc_col:
                    map_bal_clean[mc_col] = map_bal_clean[mc_col].astype(str).str.strip()
                    map_bal_clean.rename(columns={mc_col: 'cuenta_key'}, inplace=True)
                    
            map_pl_clean = None
            if map_pl_df is not None and not map_pl_df.empty:
                map_pl_clean = map_pl_df.copy()
                mc_cols = ['N° de Cuenta \n', 'N de Cuenta', 'N° de Cuenta', 'Cuenta', 'cuenta_id']
                mc_col = next((c for c in map_pl_clean.columns if c.strip() in [x.strip() for x in mc_cols]), None)
                if mc_col:
                    map_pl_clean[mc_col] = map_pl_clean[mc_col].astype(str).str.strip()
                    map_pl_clean.rename(columns={mc_col: 'cuenta_key'}, inplace=True)
            
            for co in companies:
                for per in periods:
                    tb = TrialBalanceDB.get_trial_balance(co, per)
                    if tb is not None and not tb.empty:
                        tb_cols = ['N° de Cuenta \n', 'N de Cuenta', 'N° de Cuenta', 'Cuenta', 'cuenta_id']
                        tb_cuenta_col = next((c for c in tb.columns if c.strip() in [x.strip() for x in tb_cols]), None)
                        
                        tb_nombre_cols = ['Nombre de la cuenta', 'Nombre cuenta', 'Nombre', 'descripcion', 'Detalle']
                        tb_nombre_col = next((c for c in tb.columns if c.strip() in [x.strip() for x in tb_nombre_cols]), None)
                        
                        saldo_cols = ['Saldo DR/CR\n', 'Saldo DR/CR', 'Saldos', 'Saldo', 'saldo_final']
                        saldo_col = next((c for c in tb.columns if c.strip() in [sc.strip() for sc in saldo_cols]), None)
                        
                        if tb_cuenta_col and saldo_col:
                            tb = tb.copy()
                            tb[tb_cuenta_col] = tb[tb_cuenta_col].astype(str).str.strip()
                            
                            if map_bal_clean is not None and 'cuenta_key' in map_bal_clean.columns:
                                tb = pd.merge(tb, map_bal_clean, left_on=tb_cuenta_col, right_on='cuenta_key', how='left')
                                
                            if map_pl_clean is not None and 'cuenta_key' in map_pl_clean.columns:
                                tb = pd.merge(tb, map_pl_clean, left_on=tb_cuenta_col, right_on='cuenta_key', how='left')
                                
                            for _, r_val in tb.iterrows():
                                c_id = r_val[tb_cuenta_col]
                                c_name = r_val[tb_nombre_col] if tb_nombre_col in r_val else ""
                                s_ini = r_val.get('saldo_inicial', 0.0)
                                deb = r_val.get('debitos', 0.0)
                                cred = r_val.get('creditos', 0.0)
                                s_fin = r_val.get(saldo_col, 0.0)
                                
                                clas_bal = r_val.get('Clasificación balance', '')
                                n1 = r_val.get('nota 1', '')
                                n2 = r_val.get('nota 2', '')
                                clas_pl = r_val.get('Clasificacion estado de resultados', '')
                                
                                rows.append({
                                    'empresa': co,
                                    'periodo': per,
                                    'tipo_periodo': 'actual' if per == periodo_actual_str else 'comparativo',
                                    'cuenta_id': c_id,
                                    'nombre_cuenta': c_name,
                                    'saldo_inicial': s_ini,
                                    'debitos': deb,
                                    'creditos': cred,
                                    'saldo_final': s_fin,
                                    'rubro_balance': clas_bal,
                                    'nota_1_linea': n1,
                                    'nota_2_linea': n2,
                                    'clasificacion_pl': clas_pl
                                })
            
            if rows:
                df_db_data = pd.DataFrame(rows)
                if "DB_DATA" in wb.sheetnames:
                    wb.remove(wb["DB_DATA"])
                ws_db = wb.create_sheet("DB_DATA")
                    
                for r_idx, row_data in enumerate(dataframe_to_rows(df_db_data, index=False, header=True)):
                    ws_db.append(row_data)
        except Exception as e:
            print(f"Error generando hoja DB_DATA: {e}")

        # Caché de celdas combinadas y lista de empresas para O(1) lookups y cero I/O redundante
        import os, datetime as _dt
        empresas_dir = os.path.join("data", "empresas")
        real_companies = [d for d in os.listdir(empresas_dir) if os.path.isdir(os.path.join(empresas_dir, d))] if os.path.exists(empresas_dir) else []
        
        _ws_merged_cache = {}

        def get_merged_cell_value(ws_obj, row, col):
            if ws_obj not in _ws_merged_cache:
                m_map = {}
                for merged_range in ws_obj.merged_cells.ranges:
                    top_left_val = ws_obj.cell(row=merged_range.min_row, column=merged_range.min_col).value
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        for c in range(merged_range.min_col, merged_range.max_col + 1):
                            m_map[(r, c)] = top_left_val
                _ws_merged_cache[ws_obj] = m_map
            m_map = _ws_merged_cache[ws_obj]
            if (row, col) in m_map:
                return m_map[(row, col)]
            return ws_obj.cell(row=row, column=col).value

        def analyze_columns(ws):
            col_mapping = {} # col_idx -> {'entity': str, 'year': int, 'period_type': str}
            
            for c in range(1, ws.max_column + 1):

                col_entity = None
                col_year = None
                col_period_type = None
                
                # Buscar en filas 1 a 12
                for r in range(1, 13):
                    val = get_merged_cell_value(ws, r, c)
                    if val is None:
                        continue

                    # --- Manejar datetime de Excel ---
                    if isinstance(val, (_dt.datetime, _dt.date)):
                        col_year = val.year
                        # Primera columna de fecha = actual, siguiente = comp
                        # Se resuelve abajo por posición
                        continue

                    val_lower = str(val).strip().lower()
                    
                    # Buscar coincidencia dinámica con alguna empresa real
                    for co in real_companies:
                        co_clean = co.replace("SpA", "").replace("Parent", "").strip().lower()
                        if len(co_clean) > 2 and co_clean in val_lower:
                            col_entity = co
                            break
                            
                    if not col_entity:
                        if 'consolidado' in val_lower or 'grupo' in val_lower:
                            col_entity = 'Consolidado'
                        elif 'pacifico' in val_lower:
                            col_entity = 'Pacifico SpA'
                        elif 'holdco' in val_lower or 'holco' in val_lower:
                            col_entity = 'DB Holdco Terra SpA'
                    
                    # Buscar año en texto
                    if col_year is None:
                        match = re.search(r'(202\d)', val_lower)
                        if match:
                            col_year = int(match.group(1))
                        
                    # Usar límites de palabra o regex más estrictos para evitar falsos positivos con 'composición', 'compañía', etc.
                    if re.search(r'\bcomparativ[ao]s?\b|\bcomp\b|\bcomparativa\b', val_lower):
                        col_period_type = 'comp'
                    elif re.search(r'\bactual\b', val_lower):
                        col_period_type = 'actual'
                        
                # --- Validar si es columna descriptiva de texto (no de datos) ---
                # Si en las cabeceras hay palabras como 'detalle', 'nombre', 'concepto', etc., no es columna de datos.
                is_label_col = False
                for r in range(1, 13):
                    val_h = get_merged_cell_value(ws, r, c)
                    if val_h is not None and not isinstance(val_h, (_dt.datetime, _dt.date)):
                        val_h_clean = clean_label(val_h)
                        # Usar límites de palabra \b para evitar falsos positivos como 'rut' dentro de 'bruto'
                        if re.search(r'\b(detalle|nombre|concepto|cuenta|descripcion|nombre entidad|rut|codigo|especificacion|glosa|item|tipo de relacion|relacion)\b', val_h_clean):
                            is_label_col = True
                            break

                if is_label_col:
                    continue

                if col_entity or col_year or col_period_type:
                    # Si hay año pero no entidad, asignar la entidad activa
                    if not col_entity:
                        col_entity = active_entity_name
                    col_mapping[c] = {
                        'entity': col_entity,
                        'year': col_year,
                        'period_type': col_period_type
                    }

            # --- Fallback para plantillas individuales sin etiqueta de empresa ---
            # Si no se detectó ninguna columna con datos, buscar columnas que tienen
            # encabezados como "M$", "$", "%" o fechas y asignarlas a la entidad activa
            if not col_mapping:
                numeric_header_cols = []
                for c in range(1, ws.max_column + 1):
                    for r in range(1, 13):
                        val = get_merged_cell_value(ws, r, c)
                        if val is None:
                            continue
                        if isinstance(val, (_dt.datetime, _dt.date)):
                            numeric_header_cols.append((c, val.year))
                            break
                        val_str = str(val).strip().lower()
                        if val_str in ['m$', '$', '%', 'm', 'miles'] or re.search(r'202\d', val_str):
                            yr_match = re.search(r'(202\d)', val_str)
                            numeric_header_cols.append((c, int(yr_match.group(1)) if yr_match else None))
                            break

                for idx, (c, yr) in enumerate(numeric_header_cols):
                    period_type = 'actual' if idx == 0 else 'comp'
                    col_mapping[c] = {
                        'entity': active_entity_name,
                        'year': yr,
                        'period_type': period_type
                    }

            return col_mapping


        def find_best_context(col_entity, entity_contexts):
            if not col_entity:
                return None
            col_ent_lower = col_entity.lower()
            for k in entity_contexts.keys():
                k_lower = k.lower()
                if col_ent_lower in k_lower or k_lower in col_ent_lower:
                    return entity_contexts[k]
            if 'consolidado' in col_ent_lower:
                for k in entity_contexts.keys():
                    if 'grupo' in k.lower() or 'consolidado' in k.lower():
                        return entity_contexts[k]
            return None

        # 2. Formatear fechas para reemplazo en cabeceras
        formatted_actual = format_period_to_spanish_date(periodo_actual_str) if periodo_actual_str else ""
        formatted_comp = format_period_to_spanish_date(periodo_comp_str) if periodo_comp_str and periodo_comp_str != "Ninguno" else ""
        
        actual_yr = periodo_actual_str.split('-')[0] if periodo_actual_str else "2025"
        comp_yr = periodo_comp_str.split('-')[0] if periodo_comp_str and periodo_comp_str != "Ninguno" else "2024"

        # 3. Procesar cada hoja seleccionada
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            
            col_mapping = analyze_columns(ws)
            
            # Detectar todos los años (202x) en toda la hoja para tener la lista de años de la plantilla
            all_years = set()
            for r_y in range(1, ws.max_row + 1):
                for c_y in range(1, ws.max_column + 1):
                    val_y = ws.cell(row=r_y, column=c_y).value
                    if val_y is not None:
                        if isinstance(val_y, (datetime.datetime, datetime.date)):
                            all_years.add(val_y.year)
                        else:
                            match_y = re.search(r'\b(202\d)\b', str(val_y))
                            if match_y:
                                all_years.add(int(match_y.group(1)))
            years_found = sorted(list(all_years))
            template_year_to_period_type = {}
            if len(years_found) >= 2:
                template_year_to_period_type[years_found[-1]] = 'actual'
                template_year_to_period_type[years_found[-2]] = 'comp'
            elif len(years_found) == 1:
                template_year_to_period_type[years_found[0]] = 'actual'

            def detect_block_year(ws_obj, current_row, col_idx):
                for r_b in range(current_row - 1, 0, -1):
                    is_boundary_b = False
                    for c_lbl in range(1, min(5, ws_obj.max_column + 1)):
                        cell_lbl = ws_obj.cell(row=r_b, column=c_lbl).value
                        if cell_lbl is not None:
                            cell_lbl_clean = clean_label(cell_lbl)
                            if ("total" in cell_lbl_clean and "sub" not in cell_lbl_clean) or "transaccion" in cell_lbl_clean or "las principales" in cell_lbl_clean or "la composicion" in cell_lbl_clean:
                                is_boundary_b = True
                                break
                    
                    val_found = None
                    for c_b in range(col_idx, 0, -1):
                        val_b = get_merged_cell_value(ws_obj, r_b, c_b)
                        if val_b is not None:
                            if c_b < col_idx and c_b not in col_mapping:
                                break
                            val_found = val_b
                            break
                    
                    if val_found is not None:
                        if isinstance(val_found, (datetime.datetime, datetime.date)):
                            return val_found.year
                        val_str_b = str(val_found).strip()
                        match_b = re.search(r'\b(202\d|199\d)\b', val_str_b)
                        if match_b:
                            return int(match_b.group(1))
                    if is_boundary_b:
                        break
                return None

            # Reemplazar cabeceras de fechas en filas 1-8
            # Construir fecha del periodo actual en varios formatos para comparación
            _actual_dot = format_date_dots(periodo_actual_str)   # ej. "30.06.2026"
            _comp_dot   = format_date_dots(periodo_comp_str) if periodo_comp_str and periodo_comp_str != 'Ninguno' else ''
            _actual_yr  = periodo_actual_str.split('-')[0] if periodo_actual_str else actual_yr
            _actual_mo  = periodo_actual_str.split('-')[1] if periodo_actual_str and '-' in periodo_actual_str else ''
            _comp_yr    = periodo_comp_str.split('-')[0] if periodo_comp_str and '-' in (periodo_comp_str or '') else comp_yr
            _comp_mo    = periodo_comp_str.split('-')[1] if periodo_comp_str and '-' in (periodo_comp_str or '') else ''

            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    if val and isinstance(val, str):
                        val_str = val
                        val_lower_trimmed = val.strip().lower()
                        if val_lower_trimmed in ["fecha  actual", "fecha actual", "fecha_actual", "actual"]:
                            val_str = _actual_dot
                        elif val_lower_trimmed in ["fecha  comparativa", "fecha comparativa", "fecha_comparativa", "comparativa", "comparativo"]:
                            val_str = _comp_dot
                        else:
                            # Reemplazo dinámico según el tipo de periodo de la columna mapeada
                            col_info = col_mapping.get(c)
                            if col_info:
                                col_per_type = col_info.get('period_type') or template_year_to_period_type.get(col_info['year'])
                                target_dot = _actual_dot if col_per_type == 'actual' else _comp_dot
                                
                                # 1. Detectar y reemplazar fechas completas (dd.mm.yyyy)
                                date_pattern = re.search(r'\d{1,2}[.\-,]\d{1,2}[.\-,]\d{4}', val)
                                if date_pattern:
                                    val_str = re.sub(r'\d{1,2}[.\-,]\d{1,2}[.\-,]\d{4}', target_dot, val)
                                else:
                                    # 2. Detectar y reemplazar años solos (202x)
                                    year_pattern = re.search(r'\b(202\d)\b', val)
                                    if year_pattern:
                                        target_yr = target_dot.split('.')[-1] if target_dot else year_pattern.group(1)
                                        val_str = re.sub(r'\b202\d\b', target_yr, val)
                        ws.cell(row=r, column=c).value = val_str

            # Si es reporte individual, buscar, renombrar y remapear la columna comparativa consolidada
            if not is_consolidated:
                comp_col_idx = None
                for c, info in col_mapping.items():
                    period_type = info.get('period_type') or (template_year_to_period_type.get(info['year']) if info['year'] is not None else None)
                    if period_type == 'comp' and info['entity'] == 'Consolidado':
                        comp_col_idx = c
                        break
                if comp_col_idx is not None:
                    # Actualizar mapeo dinámico
                    col_mapping[comp_col_idx]['entity'] = active_entity_name
                    # Renombrar visualmente en las cabeceras
                    for r in range(1, ws.max_row + 1):
                        val = ws.cell(row=r, column=comp_col_idx).value
                        if val is not None and isinstance(val, str):
                            val_lower = val.strip().lower()
                            if "consolidado" in val_lower:
                                ws.cell(row=r, column=comp_col_idx).value = val.replace("Consolidado", active_entity_name).replace("CONSOLIDADO", active_entity_name)

            # 3.5 Limpiar columnas activas en filas de datos para evitar números obsoletos de la plantilla
            active_cols = [c for c, info in col_mapping.items() if info.get('entity') is not None or info.get('year') is not None or info.get('period_type') is not None]
            
            # Conjunto de palabras clave que identifican cabeceras o filas no-datos
            HEADER_KEYWORDS = {
                'detalle', 'm$', '$', '%', 'neto', 'bruto', 
                'concepto', 'nombre', 'total', 'validacion', 'cuadro', 'fecha', 'actual', 
                'comparativa', 'unidades', 'codigo', 'cuenta', 'nombre cuenta', 'detalle de'
            }
            if not is_results_note:
                HEADER_KEYWORDS.add('amortizacion')
                HEADER_KEYWORDS.add('depreciacion')
            
            for row in range(1, ws.max_row + 1):
                # Evitar limpiar filas de cabecera (1 a 4 siempre protegidas)
                if row < 5:
                    continue
                
                # Identificar el primer texto no vacío de la fila
                first_text = None
                for c_idx in range(1, min(5, ws.max_column + 1)):
                    val = ws.cell(row=row, column=c_idx).value
                    if val is not None and str(val).strip() != "":
                        first_text = clean_label(val)
                        break
                
                # Si la fila es cabecera, total o validación, no se limpian sus columnas de datos
                has_year_in_row = False
                for c_idx in range(1, ws.max_column + 1):
                    val_c = ws.cell(row=row, column=c_idx).value
                    if val_c is not None:
                        if isinstance(val_c, (datetime.datetime, datetime.date)):
                            has_year_in_row = True
                            break
                        if re.search(r'202\d', str(val_c)):
                            has_year_in_row = True
                            break

                # Contar celdas de texto en las columnas activas de datos para identificar cabeceras apiladas
                string_cells_count = 0
                for c_temp in active_cols:
                    val_temp = ws.cell(row=row, column=c_temp).value
                    if val_temp is not None and isinstance(val_temp, str) and not val_temp.startswith('='):
                        try:
                            float(val_temp)
                        except ValueError:
                            string_cells_count += 1

                if first_text is None or any(kw in first_text for kw in HEADER_KEYWORDS) or has_year_in_row or string_cells_count > 1:
                    continue


                # Si es una fila de datos legítima, limpiar los valores obsoletos
                for c in active_cols:
                    cell = ws.cell(row=row, column=c)
                    if cell.value is not None:
                        val_str = str(cell.value).strip()
                        if not val_str.startswith('='):
                            cell.value = None

            last_total_row = 0
            running_sums = {}

            # Rellenar filas de datos
            for row in range(1, ws.max_row + 1):
                # Validar fila de Total
                is_total_row = False
                total_col = None
                for c_idx in range(1, min(5, ws.max_column + 1)):
                    cell_val = ws.cell(row=row, column=c_idx).value
                    if cell_val is not None and isinstance(cell_val, str):
                        cell_clean = cell_val.strip().lower()
                        if cell_clean == "total" or cell_clean.startswith("total ") or cell_clean.startswith("total:"):
                            is_total_row = True
                            total_col = c_idx
                            break
                            
                if is_total_row:
                    for col_idx in range(1, ws.max_column + 1):
                        if col_idx > (total_col or 2) and col_idx in col_mapping:
                            # Calcular la suma leyendo las celdas en el rango (last_total_row + 1) a (row - 1)
                            sum_val = 0.0
                            for r in range(last_total_row + 1, row):
                                val_cell = ws.cell(row=r, column=col_idx).value
                                if val_cell is not None:
                                    try:
                                        sum_val += float(val_cell)
                                    except:
                                        pass
                            
                            cell = ws.cell(row=row, column=col_idx)
                            # Conservar la fórmula en la celda si ya existe en la plantilla, de lo contrario escribir la suma
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = sum_val
                                cell.number_format = '#,##0;(#,##0);"-"'
                    last_total_row = row
                    running_sums = {}
                    continue

                # Si la fila tiene textos de cabecera como 'M$', '$', '%' en alguna de sus columnas de datos, no la procesamos como datos
                is_header_row = False
                for c_idx in range(1, ws.max_column + 1):
                    val_c = ws.cell(row=row, column=c_idx).value
                    if val_c is not None and isinstance(val_c, str):
                        val_c_clean = val_c.strip().lower()
                        if val_c_clean in ['m$', '$', '%', 'miles']:
                            is_header_row = True
                            break
                if is_header_row:
                    continue

                # Buscar etiqueta descriptiva
                desc_col = None
                matched_key = None
                match_type = None

                # 1. Encontrar la columna del primer texto sustancial (la columna de etiqueta)
                for c_idx in range(1, min(6, ws.max_column + 1)):
                    cell_val = ws.cell(row=row, column=c_idx).value
                    if cell_val is not None and isinstance(cell_val, str):
                        cell_clean = clean_label(cell_val)
                        # Contar celdas de texto en las columnas de datos para identificar cabeceras apiladas
                        string_cells_count = 0
                        for c_temp in col_mapping.keys():
                            val_temp = ws.cell(row=row, column=c_temp).value
                            if val_temp is not None and isinstance(val_temp, str) and not val_temp.startswith('='):
                                try:
                                    float(val_temp)
                                except ValueError:
                                    string_cells_count += 1
                        
                        is_header = False
                        if cell_val.isupper():
                            is_header = True
                        elif cell_clean in HEADER_KEYWORDS or any(cell_clean.startswith(kw + ' ') for kw in HEADER_KEYWORDS) or string_cells_count > 1:
                            is_header = True
                        if is_header or not cell_clean or len(cell_clean) <= 2:
                            continue
                        desc_col = c_idx
                        break

                # 2. Si encontramos la columna de etiqueta, verificar si tiene correspondencia en los datos
                if desc_col is not None:
                    cell_val = ws.cell(row=row, column=desc_col).value
                    cell_clean = clean_label(cell_val)
                    
                    if is_results_note:
                        PL_ALIASES = {
                            'depreciacion operacional': 'depreciacion activos fijos de operacion',
                            'amortizacion derechos de uso': 'amortizacion por derechos de uso',
                            'costos arriendos ic': 'costos arriendos intercompania',
                            'depreciacion en gtos de adm': 'depreciacion del ejercicio',
                            'amortizacion intangibles': 'amortizacion de intangibles'
                        }
                        if cell_clean in PL_ALIASES:
                            cell_clean = PL_ALIASES[cell_clean]
                    
                    sample_context = None
                    for k in [active_entity_name, 'Consolidado'] + list(entity_contexts.keys()):
                        if k in entity_contexts:
                            sample_context = entity_contexts[k]['actual']
                            break
                    
                    if sample_context:
                        # 2.1 Búsqueda exacta
                        if is_results_note:
                            sheet_title = str(ws.title).strip()
                            allowed_rubros = SHEET_TO_PL_RUBROS.get(sheet_title, [])
                            if not allowed_rubros:
                                for rubro, rubro_data in sample_context.get('pl', {}).items():
                                    if cell_clean in rubro_data:
                                        matched_key = cell_clean
                                        match_type = 'pl'
                                        break
                            else:
                                for rubro in allowed_rubros:
                                    rubro_clean = ''.join(c for c in unicodedata.normalize('NFD', rubro.lower()) if unicodedata.category(c) != 'Mn')
                                    rubro_clean = re.sub(r'\s+', ' ', rubro_clean).strip()
                                    rubro_data = sample_context.get('pl', {}).get(rubro_clean, {})
                                    if cell_clean in rubro_data:
                                        matched_key = cell_clean
                                        match_type = 'pl'
                                        break
                        else:
                            if cell_clean in sample_context.get('nota1', {}):
                                matched_key = cell_clean
                                match_type = 'nota1'
                            elif cell_clean in sample_context.get('nota2', {}):
                                matched_key = cell_clean
                                match_type = 'nota2'
                            else:
                                sheet_title = str(ws.title).strip()
                                allowed_rubros = SHEET_TO_PL_RUBROS.get(sheet_title, [])
                                if not allowed_rubros:
                                    for rubro, rubro_data in sample_context.get('pl', {}).items():
                                        if cell_clean in rubro_data:
                                            matched_key = cell_clean
                                            match_type = 'pl'
                                            break
                                else:
                                    for rubro in allowed_rubros:
                                        rubro_clean = ''.join(c for c in unicodedata.normalize('NFD', rubro.lower()) if unicodedata.category(c) != 'Mn')
                                        rubro_clean = re.sub(r'\s+', ' ', rubro_clean).strip()
                                        rubro_data = sample_context.get('pl', {}).get(rubro_clean, {})
                                        if cell_clean in rubro_data:
                                            matched_key = cell_clean
                                            match_type = 'pl'
                                            break
                        
                        # 2.2 Búsqueda tolerante por prefijo/sufijo para cortes menores (máx 5 de diferencia)
                        if matched_key is None:
                            ctx_types = ['pl'] if is_results_note else ['nota1', 'nota2', 'pl']
                            for ctx_type in ctx_types:
                                if ctx_type == 'pl':
                                    sheet_title = str(ws.title).strip()
                                    allowed_rubros = SHEET_TO_PL_RUBROS.get(sheet_title, [])
                                    rubros_to_inspect = []
                                    if not allowed_rubros:
                                        rubros_to_inspect = list(sample_context.get('pl', {}).keys())
                                    else:
                                        for rubro in allowed_rubros:
                                            rubro_clean = ''.join(c for c in unicodedata.normalize('NFD', rubro.lower()) if unicodedata.category(c) != 'Mn')
                                            rubro_clean = re.sub(r'\s+', ' ', rubro_clean).strip()
                                            rubros_to_inspect.append(rubro_clean)
                                            
                                    for rubro in rubros_to_inspect:
                                        rubro_data = sample_context.get('pl', {}).get(rubro, {})
                                        for k_ctx in rubro_data.keys():
                                            if (k_ctx.startswith(cell_clean) or cell_clean.startswith(k_ctx)) and abs(len(k_ctx) - len(cell_clean)) <= 5:
                                                matched_key = k_ctx
                                                match_type = 'pl'
                                                break
                                        if matched_key is not None:
                                            break
                                else:
                                    for k_ctx in sample_context.get(ctx_type, {}):
                                        if (k_ctx.startswith(cell_clean) or cell_clean.startswith(k_ctx)) and abs(len(k_ctx) - len(cell_clean)) <= 5:
                                            matched_key = k_ctx
                                            match_type = ctx_type
                                            break
                                if matched_key is not None:
                                    break

                if desc_col is not None:
                    extra_words = []
                    for c_idx in range(1, ws.max_column + 1):
                        if c_idx != desc_col:
                            val_temp = ws.cell(row=row, column=c_idx).value
                            if val_temp is not None and isinstance(val_temp, str):
                                extra_words.append(clean_label(val_temp))
                                
                    # Obtener nombres de cuentas del contexto activo para validación de palabras extra
                    active_ctx = entity_contexts.get(active_entity_name, {}).get('actual', {})
                    acct_names = []
                    if active_ctx:
                        if match_type == 'nota1' and matched_key in active_ctx.get('nota1', {}):
                            acct_names = active_ctx['nota1'][matched_key].get('acct_names', [])
                        elif match_type == 'nota2' and matched_key in active_ctx.get('nota2', {}):
                            acct_names = active_ctx['nota2'][matched_key].get('acct_names', [])
                        elif match_type == 'pl':
                            sheet_title = str(ws.title).strip()
                            allowed_rubros = SHEET_TO_PL_RUBROS.get(sheet_title, [])
                            found_data = None
                            if not allowed_rubros:
                                for rubro_data in active_ctx.get('pl', {}).values():
                                    if matched_key in rubro_data:
                                        found_data = rubro_data[matched_key]
                                        break
                            else:
                                for rubro in allowed_rubros:
                                    rubro_clean = ''.join(c for c in unicodedata.normalize('NFD', rubro.lower()) if unicodedata.category(c) != 'Mn')
                                    rubro_clean = re.sub(r'\s+', ' ', rubro_clean).strip()
                                    rubro_data = active_ctx.get('pl', {}).get(rubro_clean, {})
                                    if matched_key in rubro_data:
                                        found_data = rubro_data[matched_key]
                                        break
                            if found_data:
                                acct_names = found_data.get('acct_names', [])

                    skip_row = False
                    for ew in extra_words:
                        if not check_extra_word_match(ew, acct_names):
                            skip_row = True
                            break
                    if skip_row:
                        continue

                    # Rellenar primero columnas individuales
                    for c in range(desc_col + 1, ws.max_column + 1):
                        col_info = col_mapping.get(c)
                        if not col_info or col_info['entity'] == 'Consolidado':
                            continue
                            
                        entity_ctx = find_best_context(col_info['entity'], entity_contexts)
                        if not entity_ctx:
                            continue
                            
                        # Intentar detectar el año del bloque de filas actual
                        block_yr = None if is_results_note else detect_block_year(ws, row, c)
                        if block_yr and block_yr in template_year_to_period_type:
                            period_type = template_year_to_period_type[block_yr]
                        else:
                            period_type = col_info.get('period_type') or template_year_to_period_type.get(col_info['year'], 'actual')
                        period_ctx = entity_ctx.get(period_type)
                        # Si no hay contexto comparativo (período Ninguno), usar el actual como fallback
                        if not period_ctx and period_type == 'comp':
                            period_ctx = entity_ctx.get('actual')
                        if not period_ctx:
                            continue
                            
                        # --- Validar si la columna pertenece al bloque de tabla activo de esta fila ---
                        # Escaneamos hacia atrás en la columna actual. Si cruzamos una celda con "Total" o títulos en las columnas
                        # descriptivas de la izquierda, o celdas de corte, antes de encontrar un encabezado de fecha en esta columna,
                        # significa que esta columna no es un campo de datos para la subtabla de la fila actual (tablas apiladas).
                        has_header_in_block = False
                        if is_results_note:
                            has_header_in_block = True
                        else:
                            for r_idx in range(row - 1, 0, -1):
                                # Comprobar si cruzamos el límite de la subtabla
                                is_boundary = False
                                for c_lbl in range(1, min(5, ws.max_column + 1)):
                                    cell_lbl = ws.cell(row=r_idx, column=c_lbl).value
                                    if cell_lbl is not None:
                                        cell_lbl_clean = clean_label(cell_lbl)
                                        if ("total" in cell_lbl_clean and "sub" not in cell_lbl_clean) or "transaccion" in cell_lbl_clean or "las principales" in cell_lbl_clean or "la composicion" in cell_lbl_clean:
                                            is_boundary = True
                                            break
                                if is_boundary:
                                    break
                                
                                val_h = get_merged_cell_value(ws, r_idx, c)
                                if val_h is not None:
                                    if isinstance(val_h, (datetime.datetime, datetime.date)) or re.search(r'(202\d)', str(val_h)):
                                        has_header_in_block = True
                                        break
                        
                        if not has_header_in_block:
                            continue
                            
                        col_headers = []
                        for r_idx in range(1, row):
                            h_val = get_merged_cell_value(ws, r_idx, c)
                            if h_val is not None:
                                col_headers.append(str(h_val).strip())
                                
                        is_gross = any("bruto" in h.lower() or "costo" in h.lower() for h in col_headers)
                        is_depr = any("depreciaci" in h.lower() or "acumulada" in h.lower() or "amortizaci" in h.lower() for h in col_headers)
                        is_net = any("neto" in h.lower() for h in col_headers)
                        
                        # Detectar el flujo según el texto de cabecera
                        is_inicial = any("inicial" in h.lower() or "apertura" in h.lower() or "comienzo" in h.lower() for h in col_headers)
                        is_debits = any("debit" in h.lower() or "debe" in h.lower() or "cargo" in h.lower() or "adicion" in h.lower() or "aumento" in h.lower() for h in col_headers)
                        is_credits = any("credit" in h.lower() or "haber" in h.lower() or "abono" in h.lower() or "retiro" in h.lower() or "disminuc" in h.lower() for h in col_headers)
                        
                        flow_key = 'val'
                        if is_inicial:
                            flow_key = 'inicial'
                        elif is_debits:
                            flow_key = 'debitos'
                        elif is_credits:
                            flow_key = 'creditos'
                            
                        val_to_write = None
                        if is_gross:
                            val_to_write = get_filtered_value(period_ctx, matched_key, 'nota1', flow_key, row, ws) / scale_factor
                        elif is_depr:
                            val_to_write = get_filtered_value(period_ctx, matched_key, 'nota2', flow_key, row, ws) / scale_factor
                        elif is_net:
                            val_to_write = (get_filtered_value(period_ctx, matched_key, 'nota1', flow_key, row, ws) + 
                                            get_filtered_value(period_ctx, matched_key, 'nota2', flow_key, row, ws)) / scale_factor
                        else:
                            if match_type == 'nota1':
                                val_to_write = get_filtered_value(period_ctx, matched_key, 'nota1', flow_key, row, ws) / scale_factor
                                # Fallback a nota2 si el resultado da cero
                                if val_to_write == 0.0:
                                    val_to_write = get_filtered_value(period_ctx, matched_key, 'nota2', flow_key, row, ws) / scale_factor
                            elif match_type == 'nota2':
                                val_to_write = get_filtered_value(period_ctx, matched_key, 'nota2', flow_key, row, ws) / scale_factor
                                # Fallback a nota1 si el resultado da cero
                                if val_to_write == 0.0:
                                    val_to_write = get_filtered_value(period_ctx, matched_key, 'nota1', flow_key, row, ws) / scale_factor
                            elif match_type == 'pl':
                                val_to_write = get_filtered_value(period_ctx, matched_key, 'pl', flow_key, row, ws) / scale_factor
                                
                        if val_to_write is not None:
                            cell = ws.cell(row=row, column=c)
                            if cell.__class__.__name__ == 'MergedCell':
                                continue
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = val_to_write
                                cell.number_format = '#,##0;(#,##0);"-"'
                            running_sums[c] = running_sums.get(c, 0.0) + val_to_write

                    # Rellenar columnas de Consolidado por suma de las individuales o fallback
                    for c in range(desc_col + 1, ws.max_column + 1):
                        col_info = col_mapping.get(c)
                        if not col_info or col_info['entity'] != 'Consolidado':
                            continue
                            
                        block_yr = detect_block_year(ws, row, c)
                        if block_yr and block_yr in template_year_to_period_type:
                            period_type = template_year_to_period_type[block_yr]
                        else:
                            period_type = col_info.get('period_type') or template_year_to_period_type.get(col_info['year'], 'actual')
                        
                        # Buscar columnas individuales en esta hoja que tengan el mismo periodo
                        indiv_cols = [col_idx for col_idx, info in col_mapping.items()                                       if info['entity'] not in ['Consolidado', None] and (info.get('period_type') or template_year_to_period_type.get(info['year'], 'actual')) == period_type]
                                      
                        if indiv_cols:
                            sum_val = 0.0
                            for col_idx in indiv_cols:
                                val_cell = ws.cell(row=row, column=col_idx).value
                                if val_cell is not None:
                                    try:
                                        sum_val += float(val_cell)
                                    except:
                                        pass
                            cell = ws.cell(row=row, column=c)
                            if cell.__class__.__name__ == 'MergedCell':
                                continue
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = sum_val
                                cell.number_format = '#,##0;(#,##0);"-"'
                            running_sums[c] = running_sums.get(c, 0.0) + sum_val
                        else:
                            # Fallback: usar contexto consolidado/activo directo
                            entity_ctx = find_best_context(active_entity_name, entity_contexts)
                            if entity_ctx:
                                period_ctx = entity_ctx.get(period_type)
                                if period_ctx:
                                    col_headers = []
                                    for r_idx in range(1, row):
                                        h_val = ws.cell(row=r_idx, column=c).value
                                        if h_val is not None:
                                            col_headers.append(str(h_val).strip())
                                            
                                    is_gross = any("bruto" in h.lower() or "costo" in h.lower() for h in col_headers)
                                    is_depr = any("depreciaci" in h.lower() or "acumulada" in h.lower() or "amortizaci" in h.lower() for h in col_headers)
                                    is_net = any("neto" in h.lower() for h in col_headers)
                                    
                                    val_to_write = None
                                    if is_gross:
                                        val_to_write = period_ctx['nota1'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                    elif is_depr:
                                        val_to_write = period_ctx['nota2'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                    elif is_net:
                                        val_to_write = (period_ctx['nota1'].get(matched_key, {}).get('val', 0.0) + period_ctx['nota2'].get(matched_key, {}).get('val', 0.0)) / scale_factor
                                    else:
                                        if match_type == 'nota1':
                                            val_to_write = period_ctx['nota1'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                            # Fallback a nota2 si el resultado da cero
                                            if val_to_write == 0.0:
                                                val_to_write = period_ctx['nota2'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                        elif match_type == 'nota2':
                                            val_to_write = period_ctx['nota2'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                            # Fallback a nota1 si el resultado da cero
                                            if val_to_write == 0.0:
                                                val_to_write = period_ctx['nota1'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                        elif match_type == 'pl':
                                            val_to_write = period_ctx['pl'].get(matched_key, {}).get('val', 0.0) / scale_factor
                                            
                                    if val_to_write is not None:
                                        cell = ws.cell(row=row, column=c)
                                        if cell.__class__.__name__ == 'MergedCell':
                                            continue
                                        if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                            cell.value = val_to_write
                                            cell.number_format = '#,##0;(#,##0);"-"'
                                        running_sums[c] = running_sums.get(c, 0.0) + val_to_write
            # --- Regla especial para la nota de Deudores (Cuadro #N06.3 - Movimiento de Incobrables) ---
            if sheet_name == 'Deudores':
                r_ini = None
                r_per = None
                r_rev = None
                r_cas = None
                r_fin = None
                
                # Buscamos a partir de la fila 35 para enfocarnos en el cuadro #N06.3
                for r in range(35, min(ws.max_row + 1, 60)):
                    val_c = ws.cell(row=r, column=2).value
                    if val_c is not None and isinstance(val_c, str):
                        clean_lbl = clean_label(val_c)
                        if 'saldo inicial' in clean_lbl:
                            r_ini = r
                        elif 'perdida reconocida por deterioro' in clean_lbl or 'perdida reconocida' in clean_lbl:
                            r_per = r
                        elif 'reverso de deterioro' in clean_lbl:
                            r_rev = r
                        elif 'castigos' in clean_lbl:
                            r_cas = r
                        elif 'saldo final' in clean_lbl:
                            r_fin = r
                
                if r_ini is not None and r_fin is not None:
                    c_act = None
                    c_comp = None
                    for col_idx, info in col_mapping.items():
                        if info['entity'] == active_entity_name:
                            p_type = info.get('period_type') or template_year_to_period_type.get(info['year'])
                            if p_type == 'actual':
                                c_act = col_idx
                            elif p_type == 'comp':
                                c_comp = col_idx
                                
                    entity_ctx = find_best_context(active_entity_name, entity_contexts)
                    val_act = 0.0
                    val_comp = 0.0
                    if entity_ctx:
                        ctx_act = entity_ctx.get('actual', {})
                        ctx_cmp = entity_ctx.get('comp', {})
                        
                        for k_ctx in ctx_act.get('nota1', {}):
                            if 'estimacion de incobrables' in k_ctx or 'estimacion de incobr' in k_ctx:
                                val_act = abs(ctx_act['nota1'][k_ctx].get('val', 0.0)) / scale_factor
                                break
                        for k_ctx in ctx_cmp.get('nota1', {}):
                            if 'estimacion de incobrables' in k_ctx or 'estimacion de incobr' in k_ctx:
                                val_comp = abs(ctx_cmp['nota1'][k_ctx].get('val', 0.0)) / scale_factor
                                break
                                
                    if c_act is not None:
                        saldo_fin_act = val_act
                        saldo_ini_act = val_comp
                        
                        loss_act = max(0.0, saldo_fin_act - saldo_ini_act)
                        reverso_act = max(0.0, saldo_ini_act - saldo_fin_act)
                        castigos_act = 0.0
                        
                        ws.cell(row=r_ini, column=c_act).value = saldo_ini_act
                        ws.cell(row=r_ini, column=c_act).number_format = '#,##0;(#,##0);"-"'
                        
                        if r_per is not None:
                            ws.cell(row=r_per, column=c_act).value = loss_act
                            ws.cell(row=r_per, column=c_act).number_format = '#,##0;(#,##0);"-"'
                        if r_rev is not None:
                            ws.cell(row=r_rev, column=c_act).value = reverso_act
                            ws.cell(row=r_rev, column=c_act).number_format = '#,##0;(#,##0);"-"'
                        if r_cas is not None:
                            ws.cell(row=r_cas, column=c_act).value = castigos_act
                            ws.cell(row=r_cas, column=c_act).number_format = '#,##0;(#,##0);"-"'
                            
                        ws.cell(row=r_fin, column=c_act).value = saldo_fin_act
                        ws.cell(row=r_fin, column=c_act).number_format = '#,##0;(#,##0);"-"'
                        
                    if c_comp is not None:
                        saldo_fin_comp = val_comp
                        saldo_ini_comp = 0.0
                        
                        loss_comp = max(0.0, saldo_fin_comp - saldo_ini_comp)
                        reverso_comp = max(0.0, saldo_ini_comp - saldo_fin_comp)
                        castigos_comp = 0.0
                        
                        ws.cell(row=r_ini, column=c_comp).value = saldo_ini_comp
                        ws.cell(row=r_ini, column=c_comp).number_format = '#,##0;(#,##0);"-"'
                        
                        if r_per is not None:
                            ws.cell(row=r_per, column=c_comp).value = loss_comp
                            ws.cell(row=r_per, column=c_comp).number_format = '#,##0;(#,##0);"-"'
                        if r_rev is not None:
                            ws.cell(row=r_rev, column=c_comp).value = reverso_comp
                            ws.cell(row=r_rev, column=c_comp).number_format = '#,##0;(#,##0);"-"'
                        if r_cas is not None:
                            ws.cell(row=r_cas, column=c_comp).value = castigos_comp
                            ws.cell(row=r_cas, column=c_comp).number_format = '#,##0;(#,##0);"-"'
                            
                        ws.cell(row=r_fin, column=c_comp).value = saldo_fin_comp
                        ws.cell(row=r_fin, column=c_comp).number_format = '#,##0;(#,##0);"-"'

            # --- Regla especial para la nota de Intangibles (Cuadro #N08.3 - Movimiento de Intangibles) ---
            if sheet_name == 'Intangibles':
                r_ini = None
                r_adi = None
                r_amo = None
                r_fin = None
                r_hdr = None
                
                # Buscar las filas en el cuadro #N08.3
                for r in range(25, min(ws.max_row + 1, 45)):
                    val_c = ws.cell(row=r, column=3).value
                    if val_c is not None and isinstance(val_c, str):
                        clean_lbl = clean_label(val_c)
                        if 'inicio del periodo' in clean_lbl or 'inicial' in clean_lbl:
                            r_ini = r
                        elif 'adiciones' in clean_lbl:
                            r_adi = r
                        elif 'amortizacion' in clean_lbl:
                            r_amo = r
                        elif 'final del periodo' in clean_lbl or 'saldo final' in clean_lbl:
                            r_fin = r
                    
                    if r_hdr is None:
                        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                        if any(val and isinstance(val, str) and clean_label(val) in ['licencias', 'concesiones', 'licencia'] for val in row_vals):
                            r_hdr = r

                if r_hdr is not None:
                    entity_ctx = find_best_context(active_entity_name, entity_contexts)
                    ctx_act = entity_ctx.get('actual', {}) if entity_ctx else None
                    ctx_cmp = entity_ctx.get('comp', {}) if entity_ctx else None
                    
                    def find_ctx_key(clean_hdr, ctx):
                        if not ctx:
                            return None
                        for k in ctx.get('nota1', {}).keys():
                            if k == clean_hdr or k.startswith(clean_hdr) or clean_hdr.startswith(k):
                                return k
                        for k in ctx.get('nota2', {}).keys():
                            if k == clean_hdr or k.startswith(clean_hdr) or clean_hdr.startswith(k):
                                return k
                        return None

                    # Rellenar cada columna que tenga una clase de intangibles
                    for c in range(1, ws.max_column + 1):
                        hdr = ws.cell(row=r_hdr, column=c).value
                        if not hdr or 'total' in str(hdr).lower() or 'detalle' in str(hdr).lower() or 'm$' in str(hdr).lower():
                            continue
                        clean_hdr = clean_label(hdr)
                        if len(clean_hdr) <= 2:
                            continue
                            
                        # Buscar la clave correspondiente en los contextos
                        key_act = find_ctx_key(clean_hdr, ctx_act)
                        key_cmp = find_ctx_key(clean_hdr, ctx_cmp)
                        
                        key_cmp = key_cmp or key_act
                        key_act = key_act or key_cmp
                        
                        if not key_act:
                            continue
                            
                        # Valores Iniciales (Diciembre 2025)
                        gross_ini = get_filtered_value(ctx_cmp, key_cmp, 'nota1', 'val', r_ini or r_hdr, ws) if ctx_cmp else 0.0
                        depr_ini = get_filtered_value(ctx_cmp, key_cmp, 'nota2', 'val', r_ini or r_hdr, ws) if ctx_cmp else 0.0
                        net_ini = (gross_ini + depr_ini) / scale_factor
                        
                        # Valores Finales (Mayo 2026 / Junio 2026)
                        gross_fin = get_filtered_value(ctx_act, key_act, 'nota1', 'val', r_fin or r_hdr, ws) if ctx_act else 0.0
                        depr_fin = get_filtered_value(ctx_act, key_act, 'nota2', 'val', r_fin or r_hdr, ws) if ctx_act else 0.0
                        net_fin = (gross_fin + depr_fin) / scale_factor
                        
                        # Adiciones: max(0, gross_fin - gross_ini)
                        adiciones = max(0.0, (gross_fin - gross_ini) / scale_factor)
                        
                        # Amortización del ejercicio: net_fin - net_ini - adiciones
                        amortizacion = net_fin - net_ini - adiciones
                        
                        # Escribir en las celdas
                        if r_ini is not None:
                            cell = ws.cell(row=r_ini, column=c)
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = net_ini
                                cell.number_format = '#,##0;(#,##0);"-"'
                        if r_adi is not None:
                            cell = ws.cell(row=r_adi, column=c)
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = adiciones
                                cell.number_format = '#,##0;(#,##0);"-"'
                        if r_amo is not None:
                            cell = ws.cell(row=r_amo, column=c)
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = amortizacion
                                cell.number_format = '#,##0;(#,##0);"-"'
                        if r_fin is not None:
                            cell = ws.cell(row=r_fin, column=c)
                            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                                cell.value = net_fin
                                cell.number_format = '#,##0;(#,##0);"-"'

            # 4. Filtrar columnas si es reporte individual
            from openpyxl.utils import get_column_letter
            col_letter_map = {get_column_letter(c): get_column_letter(c) for c in range(1, ws.max_column + 1)}
            
            if not is_consolidated:
                # Solo filtrar si hay columnas que pertenecen a otra empresa que no sea la activa ni Consolidado
                has_other_entities = any(
                    info['entity'] is not None and info['entity'] != active_entity_name and info['entity'] != 'Consolidado'
                    for info in col_mapping.values()
                )
                
                if has_other_entities:
                    columns_to_keep = {1, 2}
                    for c in range(3, ws.max_column + 1):
                        info = col_mapping.get(c)
                        # Mantener si pertenece a la empresa activa, o si no se tiene mapeado de entidad
                        if info and (info['entity'] == active_entity_name or info['entity'] is None):
                            columns_to_keep.add(c)
                            
                    # Construir col_letter_map con los índices nuevos resultantes
                    col_letter_map = {}
                    for new_idx, orig_idx in enumerate(sorted(list(columns_to_keep)), 1):
                        orig_letter = get_column_letter(orig_idx)
                        new_letter = get_column_letter(new_idx)
                        col_letter_map[orig_letter] = new_letter
                        
                    for c in range(ws.max_column, 1, -1):
                        if c not in columns_to_keep:
                            ws.delete_cols(c)

            # Corregir letras de columna en las fórmulas después de reubicar/filtrar columnas
            fix_formulas_column_letters(ws, col_letter_map)

        # 5. Guardar libro resultante
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
