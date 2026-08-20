import pandas as pd
import io
import os
import streamlit as st

@st.cache_data(show_spinner=False)
def _cached_read_excel(path, mtime, engine=None, **kwargs):
    if engine:
        return pd.read_excel(path, engine=engine, **kwargs)
    return pd.read_excel(path, **kwargs)

def read_excel_cached(path, engine=None, **kwargs):
    if os.path.exists(path):
        mtime = os.path.getmtime(path)
    else:
        mtime = 0
    return _cached_read_excel(path, mtime, engine=engine, **kwargs)


def propagate_global_file(file_name: str, empresas_dir: str = None, source_empresa: str = "Pacifico SpA"):
    """
    Propaga un archivo desde la empresa fuente global hacia todas las empresas activas.
    """
    import shutil
    if empresas_dir is None:
        empresas_dir = os.path.join("data", "empresas")
    
    source_path = os.path.join(empresas_dir, source_empresa, file_name)
    if not os.path.exists(source_path):
        return
        
    if os.path.exists(empresas_dir):
        for co in os.listdir(empresas_dir):
            co_path = os.path.join(empresas_dir, co)
            if os.path.isdir(co_path) and co != source_empresa:
                try:
                    shutil.copy2(source_path, os.path.join(co_path, file_name))
                except Exception:
                    pass



def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Hoja1") -> bytes:
    """
    Motor central para convertir un DataFrame de Pandas a bytes de un archivo Excel descargable.
    Previene re-importaciones repetitivas de 'io' y centraliza cualquier futura estilización de reportes.
    """
    # Prevención de error "IndexError: At least one sheet must be visible" de openpyxl
    if df is None or (df.empty and len(df.columns) == 0):
        df = pd.DataFrame({"Aviso": ["No hay datos o la tabla está vacía"]})
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()


def sort_accounts(df: pd.DataFrame, account_col: str, tipo_col: str = None) -> pd.DataFrame:
    """
    Ordena un DataFrame de cuentas en base a su número correlativo y opcionalmente por tipo.
    Limpia caracteres no numéricos para ordenar numéricamente pero conserva el formato original.
    """
    if df is None or df.empty:
        return df
    
    df_sorted = df.copy()
    
    # Resolver la columna de cuenta real de forma robusta
    actual_col = account_col
    if account_col not in df_sorted.columns:
        for c in df_sorted.columns:
            c_lower = c.lower().strip()
            acc_lower = account_col.lower().strip()
            if acc_lower in c_lower or c_lower in acc_lower:
                actual_col = c
                break
    
    if actual_col not in df_sorted.columns:
        return df_sorted
        
    # Limpiar caracteres no numéricos y ordenar
    clean_series = df_sorted[actual_col].astype(str).str.replace(r'\D', '', regex=True)
    clean_series = clean_series.replace('', '999999999999')
    sort_key = pd.to_numeric(clean_series, errors='coerce').fillna(999999999999)
    df_sorted['_sort_key_num'] = sort_key
    
    actual_tipo_col = None
    if tipo_col:
        if tipo_col in df_sorted.columns:
            actual_tipo_col = tipo_col
        else:
            for c in df_sorted.columns:
                if tipo_col.lower().strip() in c.lower().strip():
                    actual_tipo_col = c
                    break
                    
    if actual_tipo_col:
        # Ordena por Tipo (ej. Balance antes de Resultado) y luego por cuenta correlativa
        df_sorted = df_sorted.sort_values(by=[actual_tipo_col, '_sort_key_num'], ascending=[True, True])
    else:
        df_sorted = df_sorted.sort_values(by=['_sort_key_num'], ascending=True)
        
    df_sorted = df_sorted.drop(columns=['_sort_key_num'], errors='ignore')
    return df_sorted


def heal_mapping_fields(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sana/rellena automáticamente columnas como 'ID_Reporte', 'ID_Nota_Asociada' 
    y 'Clasificación Flujo Efectivo' en el DataFrame de mapeo de Balance 
    basándose en la 'Clasificación balance' de otras filas del mismo DataFrame.
    """
    if df is None or df.empty:
        return df
        
    df_healed = df.copy()
    
    # 1. Encontrar las columnas relevantes de forma robusta
    cls_col = next((c for c in df_healed.columns if "clasificaci" in c.lower() and "balance" in c.lower()), None)
    id_rep_col = next((c for c in df_healed.columns if "id_reporte" in c.lower() or "d_reporte" in c.lower()), None)
    id_nota_col = next((c for c in df_healed.columns if "id_nota" in c.lower() or "nota_asociada" in c.lower() or c.lower().strip() == "nota_as"), None)
    cf_col = next((c for c in df_healed.columns if "flujo" in c.lower() and "efectivo" in c.lower()), None)
    
    if not cls_col:
        return df_healed
        
    # 2. Construir diccionarios de referencia a partir de las filas que sí tienen datos
    # Filtramos filas donde Clasificación balance no sea nula/vacía
    df_valid = df_healed[df_healed[cls_col].notna() & (df_healed[cls_col].astype(str).str.strip() != "")]
    
    # 2. Construir diccionarios de referencia vectorizados
    rep_map = {}
    nota_map = {}
    cf_map = {}
    
    if id_rep_col:
        s_cls = df_valid[cls_col].astype(str).str.strip()
        s_val = df_valid[id_rep_col].astype(str).str.strip()
        mask_v = df_valid[id_rep_col].notna() & (s_val != "") & (s_val.str.lower() != "nan")
        rep_map = dict(zip(s_cls[mask_v], s_val[mask_v]))
            
    if id_nota_col:
        s_cls = df_valid[cls_col].astype(str).str.strip()
        s_val = df_valid[id_nota_col].astype(str).str.strip()
        mask_v = df_valid[id_nota_col].notna() & (s_val != "") & (s_val.str.lower() != "nan")
        nota_map = dict(zip(s_cls[mask_v], s_val[mask_v]))
            
    if cf_col:
        s_cls = df_valid[cls_col].astype(str).str.strip()
        s_val = df_valid[cf_col].astype(str).str.strip()
        mask_v = df_valid[cf_col].notna() & (s_val != "") & (s_val.str.lower() != "nan")
        cf_map = dict(zip(s_cls[mask_v], s_val[mask_v]))
            
    # 3. Rellenar los valores nulos o vacíos de forma vectorizada
    cls_series = df_healed[cls_col].astype(str).str.strip()

    def fill_column_vectorized(col_name, mapping):
        if not col_name or not mapping:
            return
        s_target = df_healed[col_name].astype(str).str.strip()
        mask_empty = df_healed[col_name].isna() | (s_target == "") | (s_target.str.lower() == "nan")
        mapped_vals = cls_series[mask_empty].map(mapping)
        mask_filled = mapped_vals.notna()
        if mask_filled.any():
            df_healed.loc[mapped_vals[mask_filled].index, col_name] = mapped_vals[mask_filled]

    fill_column_vectorized(id_rep_col, rep_map)
    fill_column_vectorized(id_nota_col, nota_map)
    fill_column_vectorized(cf_col, cf_map)

    return df_healed


def format_periodo(periodo_str: str) -> str:
    """
    Formatea un periodo tipo '2025-06' a '2025-Jun' (formato abreviado).
    Conserva textos que no cumplan el formato 'YYYY-MM'.
    """
    if not periodo_str or not isinstance(periodo_str, str) or len(periodo_str) != 7 or '-' not in periodo_str:
        return str(periodo_str) if periodo_str is not None else ""
    
    parts = periodo_str.split('-')
    if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
        return periodo_str
        
    year, month = parts
    months_map = {
        '01': 'Ene', '02': 'Feb', '03': 'Mar', '04': 'Abr',
        '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Ago',
        '09': 'Sept', '10': 'Oct', '11': 'Nov', '12': 'Dic'
    }
    return f"{year}-{months_map.get(month, month)}"


def extract_named_ranges_from_excel(excel_bytes_io, note_code):
    """
    Busca en el libro Excel rangos nombrados que contengan el note_code (ej. 'N04').
    Retorna una lista de tuplas (nombre_rango, df) con los contenidos de dichos rangos.
    Si no encuentra ninguno, retorna una lista vacía.
    """
    import openpyxl
    excel_bytes_io.seek(0)
    wb = openpyxl.load_workbook(excel_bytes_io, data_only=True)
    
    # Normalizar note_code (ej: '#N04' -> 'n04')
    code_clean = str(note_code).replace("#", "").strip().lower()
    
    matching_ranges = []
    # Buscar todos los defined names
    for dn in wb.defined_names.values():
        dn_name_lower = dn.name.lower()
        if code_clean in dn_name_lower:
            matching_ranges.append(dn.name)
            
    # Ordenar por nombre para asegurar un orden predecible (ej: Nota_N04_Tabla1, Nota_N04_Tabla2)
    matching_ranges.sort()
    
    results = []
    for r_name in matching_ranges:
        try:
            dn = wb.defined_names.get(r_name)
            destinations = list(dn.destinations)
            if not destinations:
                continue
            sheet_name, cell_range = destinations[0]
            
            # Limpiar nombre de hoja (a veces viene con comillas simples si tiene espacios)
            sheet_name = sheet_name.replace("'", "")
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            
            # Obtener celdas del rango
            cells = ws[cell_range]
            
            # Si es una sola celda, ws[cell_range] retorna la celda directamente
            if not isinstance(cells, tuple):
                cells = ((cells,),)
                
            # Convertir celdas a matriz de valores
            data = []
            for row in cells:
                row_vals = [cell.value for cell in row]
                data.append(row_vals)
                
            df = pd.DataFrame(data)
            results.append((r_name, df))
        except Exception as ex:
            print(f"Error extrayendo rango nombrado {r_name}: {ex}")
            
    return results


def read_template_config(wb):
    """
    Lee la hoja oculta '_CONFIG_' de la plantilla Excel y retorna un diccionario
    con los metadatos de layout (name_col, nota_col, val_actual_col, val_comp_col,
    data_start_row, template_type).
    Retorna None si la hoja no existe o está vacía.
    """
    if wb is None or "_CONFIG_" not in wb.sheetnames:
        return None
    ws_cfg = wb["_CONFIG_"]
    cfg = {}
    for row in ws_cfg.iter_rows(values_only=True):
        if row and row[0] is not None and len(row) >= 2 and row[1] is not None:
            key = str(row[0]).strip()
            val = row[1]
            if key:
                # Convertir a int si es numérico
                try:
                    cfg[key] = int(val)
                except (ValueError, TypeError):
                    cfg[key] = val if val != "" else None
    return cfg if cfg else None


def write_template_config(wb, cfg: dict):
    """
    Escribe (o sobreescribe) la hoja '_CONFIG_' en el workbook con los metadatos
    de layout dados en el diccionario cfg.
    La hoja se crea oculta para no interferir con la visualización del reporte.
    Retorna el workbook modificado.
    """
    # Eliminar hoja previa si existe
    if "_CONFIG_" in wb.sheetnames:
        del wb["_CONFIG_"]
    ws_cfg = wb.create_sheet("_CONFIG_")
    ws_cfg.sheet_state = "hidden"
    # Escribir cada par clave/valor en una fila
    for key, val in cfg.items():
        ws_cfg.append([key, val])
    return wb


def read_excel_preview(excel_io, cfg: dict, col_actual: str = "Actual", col_comp: str = "Comp") -> pd.DataFrame:
    """
    Lee el Excel generado usando los índices del diccionario cfg (proveniente de _CONFIG_
    o de detect_balance_columns). Usa header=None para ser completamente inmune a cambios
    de layout en la plantilla. Retorna un DataFrame listo para preview con columnas nombradas.
    
    cfg debe contener:
      - name_col       : índice 1-based de la columna de descripción/clasificación
      - val_actual_col : índice 1-based de la columna de valor actual
      - val_comp_col   : índice 1-based de la columna de valor comparativo
      - nota_col       : índice 1-based de la columna de Nota (o None/0 si no existe)
      - data_start_row : número de fila 1-based donde empiezan los datos (después del header)
    """
    excel_io.seek(0)
    df_raw = pd.read_excel(excel_io, header=None, sheet_name=0)

    # Convertir índices 1-based → 0-based
    name_col    = cfg.get("name_col", 2) - 1
    val_actual  = cfg.get("val_actual_col", 3) - 1
    val_comp    = cfg.get("val_comp_col", 5) - 1
    nota_col_1b = cfg.get("nota_col", None)
    data_start  = cfg.get("data_start_row", 5) - 1  # 1-based → 0-based slice

    # Recortar a partir de la fila de datos
    df_data = df_raw.iloc[data_start:].reset_index(drop=True)

    # Construir mapeo índice → nombre de columna (solo columnas que existen)
    cols_map = {}
    n_cols = df_data.shape[1]
    if 0 <= name_col < n_cols:
        cols_map[name_col] = "Clasificación"
    if 0 <= val_actual < n_cols:
        cols_map[val_actual] = col_actual
    if 0 <= val_comp < n_cols:
        cols_map[val_comp] = col_comp
    if nota_col_1b and nota_col_1b > 0:
        nc0 = nota_col_1b - 1
        if 0 <= nc0 < n_cols:
            cols_map[nc0] = "Nota"

    # Seleccionar y renombrar columnas en orden
    ordered_idx  = sorted(cols_map.keys())
    result = df_data[ordered_idx].copy()
    result.columns = [cols_map[i] for i in ordered_idx]

    # Eliminar filas completamente vacías en la columna de clasificación
    if "Clasificación" in result.columns:
        result = result[
            result["Clasificación"].notna() &
            (result["Clasificación"].astype(str).str.strip() != "") &
            (result["Clasificación"].astype(str).str.strip().str.lower() != "nan")
        ].reset_index(drop=True)

    return result


def detect_balance_columns(ws, wb=None):
    """
    Detecta de forma dinámica y robusta las columnas del Balance Clasificado.
    Si se pasa el workbook `wb` y este contiene una hoja '_CONFIG_', usa esa
    información directamente (más confiable). En caso contrario, hace la detección
    automática por escaneo de celdas.
    
    Retorna: (name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx) — todos 1-based.
    """
    # Intentar leer desde _CONFIG_ primero
    if wb is not None:
        cfg = read_template_config(wb)
        if cfg:
            name_col_idx  = cfg.get("name_col", 2)
            nota_col_idx  = cfg.get("nota_col", None) or None
            val25_col_idx = cfg.get("val_actual_col", 3)
            val24_col_idx = cfg.get("val_comp_col", 5)
            # Asegurar que nota_col_idx=0 se trate como None
            if nota_col_idx == 0:
                nota_col_idx = None
            return name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx

    import datetime
    import re

    # 1. Buscar columna de descripción (donde aparece "Activos" o "Clasificación")
    name_col_idx = 2
    for col in range(1, 10):
        for row in range(1, 15):
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().lower() in ["activos", "clasificación", "clasificacion", "concepto", "detalle"]:
                name_col_idx = col
                break

    # 2. Buscar si existe una columna de Nota real
    nota_col_idx = None
    for col in range(1, 10):
        for row in range(1, 10):
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().lower() == "nota":
                nota_col_idx = col
                break

    # 3. Escanear columnas para detectar fechas
    date_cols = []
    for col in range(1, ws.max_column + 1):
        if col == name_col_idx or col == nota_col_idx:
            continue
        for row in range(1, 6):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                is_date = (
                    isinstance(val, (datetime.datetime, datetime.date)) or
                    (isinstance(val, str) and re.search(r'20\d{2}', val)) or
                    (isinstance(val, str) and any(x in val.lower() for x in ["actual", "anterior", "auditado", "comparat", "comp"]))
                )
                if is_date:
                    date_cols.append(col)
                    break
    date_cols = sorted(list(set(date_cols)))

    # 4. Asignar índices de periodos según el layout detectado
    if nota_col_idx is not None:
        # Layout con Nota (ej: Holdco) -> B=Name, C=Nota, D=Actual, F=Comp
        val25_col_idx = date_cols[0] if len(date_cols) >= 1 else 4
        val24_col_idx = date_cols[1] if len(date_cols) >= 2 else 6
    else:
        # Layout sin Nota (ej: Pacifico) -> B=Name, C=Actual, E=Comp
        val25_col_idx = date_cols[0] if len(date_cols) >= 1 else 3
        val24_col_idx = date_cols[1] if len(date_cols) >= 2 else 5

    return name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx


def detect_patrimonio_skiprows(excel_io):
    """
    Detecta de forma dinámica la cantidad de filas a omitir (skiprows) 
    para el Estado de Cambios en el Patrimonio.
    """
    try:
        import pandas as pd
        excel_io.seek(0)
        df = pd.read_excel(excel_io, header=None)
        for idx, row in df.iterrows():
            if idx >= 15:
                break
            row_vals = [str(v).lower().strip() for v in row.values if pd.notna(v)]
            has_capital = any("capital" in v for v in row_vals)
            has_res_gan = any("reserva" in v or "acumulad" in v or "ganancia" in v or "patrimonio" in v for v in row_vals)
            if has_capital and has_res_gan:
                return idx
    except Exception as e:
        print(f"Error detecting patrimonio skiprows: {e}")
    return 4  # Default fallback


def detect_general_skiprows(excel_io):
    """
    Detecta de forma dinámica la cantidad de filas a omitir (skiprows) 
    para un reporte general (como ORI/Estado de Resultados Integrales) 
    buscando palabras claves de cabeceras de columnas conceptuales.
    """
    try:
        import pandas as pd
        excel_io.seek(0)
        df = pd.read_excel(excel_io, header=None)
        for idx, row in df.iterrows():
            if idx >= 15:
                break
            row_vals = [str(v).lower().strip() for v in row.values if pd.notna(v)]
            if any(v in ["detalle", "concepto", "clasificacion", "clasificación", "descripcion", "descripción"] for v in row_vals):
                return idx
    except Exception as e:
        print(f"Error detecting general skiprows: {e}")
    return 4  # Default fallback


def clean_preview_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Limpia un DataFrame de previsualización:
    1. Elimina columnas 'Unnamed:' excepto la primera.
    2. Renombra la primera columna a '' si comienza con 'Unnamed'.
    3. Limpia los nombres de las columnas reemplazando espacios no separables con espacios estándar.
    """
    if df is None or df.empty:
        return df
        
    df = df.copy()
    
    # 1. Limpiar espacios normales e internos (como no-breaking spaces \xa0) en cabeceras
    df.columns = [str(c).replace('\xa0', ' ').replace('\u200b', '').strip() for c in df.columns]
    
    # 2. Identificar y eliminar columnas Unnamed (excepto la primera que contiene descripciones)
    cols_to_drop = []
    for i, col in enumerate(df.columns):
        if i > 0 and str(col).lower().startswith("unnamed:"):
            cols_to_drop.append(col)
            
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
        
    # 3. Renombrar la primera columna si es Unnamed
    if len(df.columns) > 0 and str(df.columns[0]).lower().startswith("unnamed"):
        df = df.rename(columns={df.columns[0]: ""})
        
    return df





