import streamlit as st
import pandas as pd
import os
import datetime
from src.core.excel_utils import df_to_excel_bytes, format_periodo, read_excel_cached

from src.reporting.notes import NOTE_REGISTRY

# Mapeo de notas dinámico reconstruido desde NOTE_REGISTRY
sheet_name_map = {code: info['sheets'] for code, info in NOTE_REGISTRY.items()}

# Agrupación de notas dinámica agrupada por categorías del estado financiero
notes_by_category = {
    "activos_corrientes": [],
    "activos_no_corrientes": [],
    "pasivos_corrientes": [],
    "pasivos_no_corrientes": [],
    "patrimonio": [],
    "resultados": []
}

for code, info in NOTE_REGISTRY.items():
    cat = info['category']
    if cat in notes_by_category:
        notes_by_category[cat].append((code, f"[{code}] {info['title']}"))

def load_all_entity_contexts(active_entity, periodo_actual, periodo_comp, map_balance_df, map_pl_df):
    from src.models.database import SessionLocal
    from src.models.consolidacion import ConsolidationGroup
    from src.models.trial_balance_db import TrialBalanceDB
    from src.reporting.note_generator import build_entity_context
    
    contexts = {}
    is_consolidated = active_entity.startswith("[GRUPO]")
    
    # 1. Determinar empresas involucradas
    companies = []
    if is_consolidated:
        grupo_name = active_entity.replace("[GRUPO] ", "").strip()
        db = SessionLocal()
        try:
            grupo_obj = db.query(ConsolidationGroup).filter_by(nombre_grupo=grupo_name).first()
            if grupo_obj:
                companies.append(grupo_obj.empresa_matriz)
                if grupo_obj.filial_is_group:
                    def get_sub_companies(sub_g_id):
                        sub_g = db.query(ConsolidationGroup).filter_by(id=sub_g_id).first()
                        if sub_g:
                            c = [sub_g.empresa_matriz]
                            if sub_g.filial_is_group:
                                c.extend(get_sub_companies(int(sub_g.empresa_filial)))
                            else:
                                c.append(sub_g.empresa_filial)
                            return c
                        return []
                    companies.extend(get_sub_companies(int(grupo_obj.empresa_filial)))
                else:
                    companies.append(grupo_obj.empresa_filial)
        finally:
            db.close()
    else:
        companies = [active_entity]
        
    # 2. Cargar contextos de datos para cada empresa individual
    # Guardamos los TB en caché para reutilizarlos en el paso 3 (caso consolidado)
    # y evitar consultar la BD dos veces por empresa/período.
    _tb_cache_act  = {}  # { empresa: DataFrame | None }
    _tb_cache_comp = {}  # { empresa: DataFrame | None }

    for co in companies:
        tb_act  = TrialBalanceDB.get_trial_balance(co, periodo_actual)
        tb_comp = TrialBalanceDB.get_trial_balance(co, periodo_comp) if periodo_comp != "Ninguno" else None
        _tb_cache_act[co]  = tb_act
        _tb_cache_comp[co] = tb_comp

        ctx_act  = build_entity_context(tb_act,  map_balance_df, map_pl_df, empresa_name=co, periodo_str=periodo_actual)
        ctx_comp = build_entity_context(tb_comp, map_balance_df, map_pl_df, empresa_name=co, periodo_str=periodo_comp) if tb_comp is not None else {'nota1': {}, 'nota2': {}, 'pl': {}}

        contexts[co] = {
            'actual': ctx_act,
            'comp':   ctx_comp
        }

    # 3. Cargar contexto acumulado para el Consolidado (soporte de fallback)
    # Reutiliza los DataFrames ya cargados en el paso 2 — sin nuevas consultas a la BD.
    if is_consolidated:
        tb_act_list  = [df for df in _tb_cache_act.values()  if df is not None and not df.empty]
        tb_comp_list = [df for df in _tb_cache_comp.values() if df is not None and not df.empty]

        tb_act_comb  = pd.concat(tb_act_list,  ignore_index=True) if tb_act_list  else None
        tb_comp_comb = pd.concat(tb_comp_list, ignore_index=True) if tb_comp_list else None

        ctx_act_comb  = build_entity_context(tb_act_comb,  map_balance_df, map_pl_df, empresa_name=companies, periodo_str=periodo_actual)
        ctx_comp_comb = build_entity_context(tb_comp_comb, map_balance_df, map_pl_df, empresa_name=companies, periodo_str=periodo_comp) if tb_comp_comb is not None else {'nota1': {}, 'nota2': {}, 'pl': {}}

        contexts[active_entity] = {
            'actual': ctx_act_comb,
            'comp':   ctx_comp_comb
        }

    return contexts


import datetime

def evaluate_openpyxl_formula(formula, ws, col_idx, row_idx=None, visited=None):
    import re
    from openpyxl.utils import column_index_from_string
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
        
    if visited is None:
        visited = set()
        
    if row_idx is not None:
        coord = (row_idx, col_idx)
        if coord in visited:
            return 0.0
        visited.add(coord)
        
    formula = formula.strip().upper()
    
    # Pattern 1: =SUM(D36:G36) o =SUM(C5:C6)
    sum_match = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', formula)
    if sum_match:
        col1_let = sum_match.group(1)
        row1_num = int(sum_match.group(2))
        col2_let = sum_match.group(3)
        row2_num = int(sum_match.group(4))
        
        from openpyxl.utils import column_index_from_string
        col1_idx = column_index_from_string(col1_let)
        col2_idx = column_index_from_string(col2_let)
        
        total = 0.0
        
        if col1_let == col2_let:
            # Suma vertical (misma columna)
            for r in range(row1_num, row2_num + 1):
                if r <= ws.max_row:
                    val = ws.cell(row=r, column=col1_idx).value
                    if val is not None:
                        if isinstance(val, str) and val.startswith('='):
                            val = evaluate_openpyxl_formula(val, ws, col1_idx, r, visited)
                        try:
                            total += float(val)
                        except ValueError:
                            pass
        elif row1_num == row2_num:
            # Suma horizontal (misma fila)
            r = row1_num
            for c in range(col1_idx, col2_idx + 1):
                if c <= ws.max_column:
                    val = ws.cell(row=r, column=c).value
                    if val is not None:
                        if isinstance(val, str) and val.startswith('='):
                            val = evaluate_openpyxl_formula(val, ws, c, r, visited)
                        try:
                            total += float(val)
                        except ValueError:
                            pass
        else:
            # Suma bidimensional (caja/rango)
            for r in range(row1_num, row2_num + 1):
                for c in range(col1_idx, col2_idx + 1):
                    if r <= ws.max_row and c <= ws.max_column:
                        val = ws.cell(row=r, column=c).value
                        if val is not None:
                            if isinstance(val, str) and val.startswith('='):
                                val = evaluate_openpyxl_formula(val, ws, c, r, visited)
                            try:
                                total += float(val)
                            except ValueError:
                                pass
                                
        if row_idx is not None:
            visited.discard((row_idx, col_idx))
        return total
        
    # Pattern 2: =X12-X34 or similar arithmetic expression
    tokens = re.findall(r'([A-Z]+)(\d+)', formula)
    if tokens:
        expr = formula[1:] # remove '='
        sorted_tokens = sorted(tokens, key=lambda t: len(t[0] + t[1]), reverse=True)
        
        for col_letter, row_str in sorted_tokens:
            row_idx_ref = int(row_str)
            if row_idx_ref <= ws.max_row:
                c_idx = column_index_from_string(col_letter)
                val = ws.cell(row=row_idx_ref, column=c_idx).value
                if isinstance(val, str) and val.startswith('='):
                    val = evaluate_openpyxl_formula(val, ws, c_idx, row_idx_ref, visited)
                val_float = 0.0
                if val is not None:
                    try:
                        val_float = float(val)
                    except ValueError:
                        pass
                token_str = f"{col_letter}{row_str}"
                expr = expr.replace(token_str, str(val_float))
            
        expr = expr.replace('+-', '-').replace('--', '+')
        if re.match(r'^[0-9. +\-*/()]+$', expr):
            try:
                ret_val = eval(expr)
                if row_idx is not None:
                    visited.discard((row_idx, col_idx))
                return ret_val
            except:
                pass
                
    if row_idx is not None:
        visited.discard((row_idx, col_idx))
    return None

def evaluate_formulas_in_workbook(excel_bytes_in):
    import openpyxl
    from io import BytesIO
    excel_bytes_in.seek(0)
    wb = openpyxl.load_workbook(excel_bytes_in, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    visited = set()
                    evaluated = evaluate_openpyxl_formula(val, ws, col, row, visited)
                    if evaluated is not None:
                        cell.value = evaluated
    
    excel_bytes_out = BytesIO()
    wb.save(excel_bytes_out)
    excel_bytes_out.seek(0)
    return excel_bytes_out

def classify_row(row):
    cells = [val for val in row if pd.notna(val) and str(val).strip() != ""]
    if len(cells) == 0:
        return "empty", None
    
    # Check if it contains any cell starting with "validacion" (case-insensitive)
    is_validation = False
    for cell in row:
        if pd.notna(cell) and str(cell).strip().lower().startswith("validacion"):
            is_validation = True
            break
            
    if is_validation:
        return "validation", cells
        
    if len(cells) == 1:
        return "text", cells[0]
        
    return "table", cells

def is_header_start_row(row):
    # Contar celdas no vacías para omitir filas de detalle con fechas
    non_empty_cells = [val for val in row if pd.notna(val) and str(val).strip() != ""]
    if len(non_empty_cells) > 3:
        if any(isinstance(val, datetime.datetime) for val in row):
            return False
        row_strs = [str(val).strip().lower() for val in row if pd.notna(val)]
        if any(any(k in s for k in ["clp", "us$", "uf", "préstamo", "leasing", "tasa", "banco", "prestamo"]) for s in row_strs):
            return False

    has_date = False
    has_entity = False
    has_numeric_value = False
    
    for val in row:
        if pd.isna(val) or val == "":
            continue
        if isinstance(val, (int, float)):
            if val > 2000 and val < 2100:
                has_date = True
            else:
                has_numeric_value = True
        elif isinstance(val, datetime.datetime):
            has_date = True
        else:
            val_str = str(val).strip().lower()
            if any(k in val_str for k in ["31.12.", "31.03.", "30.06.", "30.09.", "31 de ", "30 de "]):
                has_date = True
            if any(k in val_str for k in ["pacifico", "holdco", "consolidado", "matriz", "filial", "empresa"]):
                has_entity = True
                
    return (has_date or has_entity) and not has_numeric_value

def split_sheet_into_elements(df):
    df = df.dropna(how='all', axis=1)
    
    elements = []
    current_chunk = []
    has_data = False
    after_total = False
    
    for idx, row in df.iterrows():
        row_vals = list(row.values)
        row_type, cell_val = classify_row(row_vals)
        
        is_end_of_table = False
        if row_type == "validation":
            is_end_of_table = True
        elif row_vals:
            non_empty_cells = [v for v in row_vals if pd.notna(v) and str(v).strip() != ""]
            if non_empty_cells:
                label_str = str(non_empty_cells[0]).strip().lower()
                if (
                    label_str == "total" or 
                    label_str.startswith("total") or 
                    label_str.startswith("sub-total") or 
                    label_str.startswith("subtotal") or 
                    "totales" in label_str
                ):
                    is_end_of_table = True
                    
        if is_end_of_table:
            after_total = True
            
        if row_type == "empty":
            if has_data and after_total:
                # Look ahead to see if the next non-empty row is validation
                is_followed_by_validation = False
                for j in range(idx + 1, len(df)):
                    next_row_vals = list(df.iloc[j].values)
                    next_type, _ = classify_row(next_row_vals)
                    if next_type == "empty":
                        continue
                    if next_type == "validation":
                        is_followed_by_validation = True
                    break
                
                if is_followed_by_validation:
                    current_chunk.append(row_vals)
                else:
                    elements.append(("table", pd.DataFrame(current_chunk)))
                    current_chunk = []
                    has_data = False
                    after_total = False
            else:
                current_chunk.append(row_vals)
        elif row_type == "text":
            if has_data and after_total:
                elements.append(("table", pd.DataFrame(current_chunk)))
                current_chunk = []
                has_data = False
                after_total = False
                elements.append(("text", cell_val))
            else:
                current_chunk.append(row_vals)
        elif is_header_start_row(row_vals):
            if has_data:
                elements.append(("table", pd.DataFrame(current_chunk)))
                current_chunk = []
                has_data = False
                after_total = False
            current_chunk.append(row_vals)
        else:
            has_nums = False
            for val in row_vals:
                if pd.notna(val) and isinstance(val, (int, float)) and not (val > 2000 and val < 2100):
                    has_nums = True
                    break
            if has_nums:
                has_data = True
            current_chunk.append(row_vals)
            
    if current_chunk:
        elements.append(("table", pd.DataFrame(current_chunk)))
        
    return elements

def render_note_section(notes_list, key_prefix, scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp):
    options = [n[1] for n in notes_list]
    selected_label = st.selectbox(
        "Seleccione la nota a generar:",
        options,
        key=f"{key_prefix}_select"
    )
    
    # Buscar nota seleccionada
    selected_id = None
    for n_id, n_label in notes_list:
        if n_label == selected_label:
            selected_id = n_id
            break
            
    if selected_id:
        st.write("---")
        
        # Validar si la nota está en la plantilla y tiene pestañas asignadas
        if selected_id not in sheet_name_map or not sheet_name_map[selected_id]:
            st.warning(f"⚠️ La nota '{selected_label}' no tiene pestañas Excel configuradas en la plantilla (es de carácter informativo o no se encuentra implementada aún).")
            return
            
        col_b1, col_b2 = st.columns([2, 2])
        with col_b1:
            run_note = st.button("🚀 Ejecutar y Visualizar Nota", type="primary", use_container_width=True, key=f"{key_prefix}_run")
        with col_b2:
            st.empty()
            
        if run_note:
            template_nota = "Plantilla de notas_v1.xlsx"
            if not os.path.exists(template_nota):
                st.error("❌ No se encontró la plantilla maestra global 'Plantilla de notas_v1.xlsx' en la raíz.")
            else:
                st.info(f"Procesando y mapeando saldos para {selected_label}...")
                
                # Cargar mapeos con caché (evita re-lectura si el archivo no cambió)
                map_bal_local = os.path.join(empresa_path, "map_balance.xlsx")
                map_pl_local  = os.path.join(empresa_path, "map_pl.xlsx")
                map_balance_df = read_excel_cached(map_bal_local, dtype=str) if os.path.exists(map_bal_local) else None
                map_pl_df      = read_excel_cached(map_pl_local,  dtype=str) if os.path.exists(map_pl_local)  else None

                if map_balance_df is None:
                    st.error("❌ No se pudo cargar el maestro de Mapeo Balance.")
                    st.stop()

                from src.reporting.note_generator import NoteGenerator
                
                try:
                    # Cargar contextos para todas las entidades involucradas
                    with st.spinner("Cargando saldos y compilando contextos de datos..."):
                        entity_contexts = load_all_entity_contexts(
                            active_entity=empresa_seleccionada,
                            periodo_actual=periodo_actual,
                            periodo_comp=periodo_comp,
                            map_balance_df=map_balance_df,
                            map_pl_df=map_pl_df
                        )
                        
                    is_consolidated = empresa_seleccionada.startswith("[GRUPO]")
                    target_sheets = sheet_name_map[selected_id]
                    
                    engine = NoteGenerator(template_nota)
                    excel_nota_out = engine.generate(
                        sheet_names=target_sheets,
                        entity_contexts=entity_contexts,
                        active_entity_name=empresa_seleccionada,
                        is_consolidated=is_consolidated,
                        scale_factor=scale_factor_nota,
                        periodo_actual_str=periodo_actual,
                        periodo_comp_str=periodo_comp,
                        map_balance_df=map_balance_df,
                        map_pl_df=map_pl_df
                    )
                    
                    st.success(f"✅ Nota generada exitosamente en {unidad_nota}.")
                    
                    # Vista previa de la nota con sub-tablas visualmente separadas
                    excel_nota_out.seek(0)
                    excel_eval_out = evaluate_formulas_in_workbook(excel_nota_out)
                    preview_nota_df = pd.read_excel(excel_eval_out, sheet_name=target_sheets[0], header=None)
                    
                    # --- SEMÁFORO FINANCIERO DE CUADRATURA (TIE-OUT) ---
                    try:
                        from src.models.database import SessionLocal
                        from src.models.historical_data import HistoricalDataRecord
                        
                        NOTE_TO_RUBRO_MAP = {
                             "#N04": [("Balance", "Efectivo y efectivo equivalente")],
                             "#N06": [("Balance", "Deudores comerciales y otras cuentas por cobrar, corrientes")],
                             "#N08": [("Balance", "Activos intangibles distinto a la plusvalía"), ("Balance", "Activos intangibles distinto a la plusvalia")],
                             "#N14": [("Balance", "Cuentas por pagar entidades relacionadas, corrientes"), 
                                      ("Balance", "Cuentas por pagar entidades relacionadas, no corrientes")],
                             "#N17": [("Balance", "Cuentas comerciales y otras cuentas por pagar, corrientes"),
                                      ("Balance", "Cuentas comerciales y otras cuentas por pagar, no corrientes")],
                             "#N18": [("Balance", "Provisiones por beneficios a los empleados")],
                             "#N23": [("P&L", "Diferencias de cambio"), ("P&L", "Diferencia de cambio")]
                        }
                        
                        rubros = NOTE_TO_RUBRO_MAP.get(selected_id)
                        if rubros:
                            db = SessionLocal()
                            db_total = 0.0
                            for rep, lin in rubros:
                                r_recs = db.query(HistoricalDataRecord).filter_by(
                                    empresa=empresa_seleccionada,
                                    periodo=periodo_actual,
                                    reporte=rep,
                                    linea_item=lin
                                ).all()
                                for r_rec in r_recs:
                                    db_total += abs(float(r_rec.monto))
                            db.close()
                            
                            # Buscar la fila de "Total" en el preview_nota_df
                            total_row_idx = None
                            # 1. Buscar fila de gran total específica
                            for idx, row_series in preview_nota_df.iterrows():
                                row_str = " ".join([str(val).strip().lower() for val in row_series.values if pd.notna(val)])
                                if "total" in row_str and "sub" not in row_str:
                                    if "corriente" in row_str or "general" in row_str or "administrac" in row_str or "operac" in row_str:
                                        total_row_idx = idx
                                        break
                                        
                            # 2. Si no se encuentra, tomar la primera fila con 'total'
                            if total_row_idx is None:
                                for idx, row_series in preview_nota_df.iterrows():
                                    row_str = " ".join([str(val).strip().lower() for val in row_series.values if pd.notna(val)])
                                    if "total" in row_str and "sub" not in row_str:
                                        total_row_idx = idx
                                        break
                                    
                            if total_row_idx is not None:
                                total_row = preview_nota_df.iloc[total_row_idx]
                                numeric_cols = []
                                
                                import datetime
                                import re
                                for col_idx, val in enumerate(total_row.values):
                                    if pd.notna(val) and isinstance(val, (int, float)):
                                        # Verificar si la columna es de datos (tiene fecha/año en cabecera)
                                        is_date_col = False
                                        for r in range(0, total_row_idx):
                                            cell_val = preview_nota_df.iloc[r, col_idx]
                                            # Si está vacío, buscar a la izquierda en la misma fila para celdas combinadas
                                            if pd.isna(cell_val) or str(cell_val).strip() == "":
                                                for c in range(col_idx - 1, -1, -1):
                                                    left_val = preview_nota_df.iloc[r, c]
                                                    if pd.notna(left_val) and str(left_val).strip() != "":
                                                        cell_val = left_val
                                                        break
                                            
                                            if pd.notna(cell_val) and str(cell_val).strip() != "":
                                                if isinstance(cell_val, (datetime.datetime, pd.Timestamp)):
                                                    is_date_col = True
                                                    break
                                                if len(str(cell_val)) <= 25 and re.search(r'\b(202\d|199\d)\b', str(cell_val)):
                                                    is_date_col = True
                                                    break
                                        if is_date_col:
                                            numeric_cols.append((col_idx, float(val)))
                                        
                                excel_total = None
                                if numeric_cols:
                                    if len(numeric_cols) == 1:
                                        excel_total = abs(numeric_cols[0][1]) * scale_factor_nota
                                    else:
                                        neto_col_idx = None
                                        for col_idx, val in numeric_cols:
                                            for r in range(0, total_row_idx):
                                                cell_val = preview_nota_df.iloc[r, col_idx]
                                                if pd.notna(cell_val) and "neto" in str(cell_val).strip().lower():
                                                    neto_col_idx = col_idx
                                                    break
                                            if neto_col_idx is not None:
                                                break
                                        
                                        if neto_col_idx is not None:
                                            for col_idx, val in numeric_cols:
                                                if col_idx == neto_col_idx:
                                                    excel_total = abs(val) * scale_factor_nota
                                                    break
                                        else:
                                            excel_total = abs(numeric_cols[0][1]) * scale_factor_nota
                                            
                                if excel_total is not None:
                                    diff = abs(excel_total - db_total)
                                    
                                    # Tolerancia flexible de $1,000 debido a redondeos de escala
                                    if diff < 1000.0:
                                        st.success(f"⚖️ **[Tie-Out OK]** El total de la Nota (${excel_total:,.0f}) coincide perfectamente con el Estado Financiero Principal (${db_total:,.0f}).")
                                    else:
                                        st.warning(f"⚖️ **[Tie-Out Diferencia]** Se detectó un descuadre de **${diff:,.0f}** entre el total de la Nota (${excel_total:,.0f}) y los Estados Financieros (${db_total:,.0f}).")
                        else:
                            st.info("ℹ️ Esta nota no tiene rubros principales en el balance asociados para validación automática.")
                    except Exception as tie_err:
                        st.info(f"ℹ️ Validación de Tie-Out omitida temporalmente: {tie_err}")
                    
                    # Intentar extraer usando rangos nombrados
                    from src.core.excel_utils import extract_named_ranges_from_excel
                    named_ranges = extract_named_ranges_from_excel(excel_nota_out, selected_id)
                    
                    elements = []
                    if named_ranges:
                        for r_name, df_chunk in named_ranges:
                            elements.append(("table", df_chunk, target_sheets[0]))
                    else:
                        for sheet_name in target_sheets:
                            sheet_df = pd.read_excel(excel_eval_out, sheet_name=sheet_name, header=None)
                            if len(target_sheets) > 1:
                                elements.append(("text", f"Pestaña: {sheet_name.strip()}", sheet_name))
                            sheet_elements = split_sheet_into_elements(sheet_df)
                            for el_type, el_val in sheet_elements:
                                elements.append((el_type, el_val, sheet_name))
                    
                    # Mostrar las sub-tablas visualmente separadas y estilizadas
                    from src.reporting.formatting import apply_corporate_style
                    
                    table_counter = 0
                    for item in elements:
                        if len(item) == 3:
                            el_type, el_val, sh_name = item
                        else:
                            el_type, el_val = item
                            sh_name = target_sheets[0]
                            
                        if el_type == "text":
                            st.markdown(f"##### 📋 {el_val}")
                        else:
                            chunk_df = el_val.dropna(how='all', axis=0).reset_index(drop=True)
                            if chunk_df.empty:
                                continue
                            
                            table_counter += 1
                            sub_code = f"{selected_id}.{table_counter}"
                            st.markdown(f"**📍 Cuadro `{sub_code}`**")
                            
                            # Renombrar columnas a espacios únicos
                            new_cols = []
                            for idx in range(len(chunk_df.columns)):
                                new_cols.append(" " * (idx + 1))
                            chunk_df.columns = new_cols
                            chunk_df = chunk_df.fillna("")
                            
                            # Aplicar estilo corporativo
                            styled_df = apply_corporate_style(chunk_df, excel_bytes=excel_eval_out, sheet_name=sh_name)
                            
                            # Convertir a HTML y ocultar cabecera thead
                            html_str = styled_df.to_html(index=False)
                            html_str = html_str.replace("<thead>", '<thead style="display:none">')
                            
                            # Envolver en un contenedor con scroll horizontal para tablas anchas
                            scrollable_html = f'<div style="overflow-x: auto; width: 100%;">{html_str}</div>'
                            st.markdown(scrollable_html, unsafe_allow_html=True)
                            st.write("") # Espaciador
                    
                    # Generación en Word usando el nuevo exportador de tablas
                    from src.reporting.word_export import WordExportEngine
                    elements_for_word = [(item[0], item[1]) for item in elements]
                    word_nota_out = WordExportEngine.generate_notes_word(
                        elements=elements_for_word,
                        title=selected_label,
                        unit=unidad_nota,
                        note_code=selected_id
                    )
                    
                    excel_nota_out.seek(0)
                    st.write("")
                    c1, c2 = st.columns(2)
                    with c1:
                        st.download_button(
                            "📥 Descargar Nota en Excel",
                            data=excel_nota_out,
                            file_name=f"{selected_label.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                            use_container_width=True,
                            key=f"{key_prefix}_dl_ex"
                        )
                    with c2:
                        st.download_button(
                            "📝 Descargar Nota en Word",
                            data=word_nota_out,
                            file_name=f"{selected_label.replace(' ', '_')}.docx",
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                            type="primary",
                            use_container_width=True,
                            key=f"{key_prefix}_dl_wd"
                        )
                except Exception as e:
                    st.error(f"❌ Error al procesar nota: {e}")
                    st.exception(e)

def render(empresa_seleccionada, empresa_path):
    st.title("📑 Informes y Notas a los Estados Financieros")
    
    if empresa_seleccionada == "🌐 [GLOBAL] Configuración General":
        st.info("🌐 **Modo Global Activo**: Desde esta sección puedes administrar la **Plantilla Maestra Global** de Notas (`Plantilla de notas_v1.xlsx`). Para ejecutar y visualizar notas contables de una empresa específica, selecciónala en la barra lateral izquierda.")
        
        with st.expander("⚙️ Administrar Plantilla Maestra Global de Notas", expanded=True):
            st.write("Descarga la plantilla Excel global actual de notas, realiza modificaciones en su estructura, textos o mapeos, y vuelve a subirla para que aplique a todas las empresas.")
            
            template_nota = "Plantilla de notas_v1.xlsx"
            if os.path.exists(template_nota):
                with open(template_nota, "rb") as file:
                    st.download_button(
                        label="📥 Descargar Plantilla Maestra Global Actual",
                        data=file,
                        file_name="Plantilla_de_notas_v1.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_global_master_template_notes"
                    )
            else:
                st.warning("⚠️ No se encontró la plantilla maestra global 'Plantilla de notas_v1.xlsx' en la raíz.")

            uploaded_template = st.file_uploader(
                "Subir Nueva Plantilla Maestra Global (Reemplazar en la raíz):",
                type=["xlsx"],
                key="global_notes_template_uploader"
            )
            
            if uploaded_template is not None:
                if st.button("💾 Guardar Plantilla Maestra Global", type="primary", use_container_width=True):
                    with open(template_nota, "wb") as f:
                        f.write(uploaded_template.getbuffer())
                    st.session_state['success_msg'] = "✅ ¡Plantilla maestra global subida y actualizada con éxito en la raíz del proyecto!"
                    st.rerun()
        st.stop()
        
    with st.expander("🌐 Información de Plantilla Maestra Global"):
        st.write("Todas las notas utilizan la **Plantilla Maestra Global** (`Plantilla de notas_v1.xlsx`). Para modificar mapeos o la estructura base, selecciona **🌐 [GLOBAL] Configuración General** en el menú lateral.")
        template_nota = "Plantilla de notas_v1.xlsx"
        if os.path.exists(template_nota):
            with open(template_nota, "rb") as file:
                st.download_button(
                    label="📥 Descargar Plantilla Maestra Global",
                    data=file,
                    file_name="Plantilla_de_notas_v1.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_master_template_notes_company"
                )

    
    from src.models.database import SessionLocal
    from src.models.historical_data import HistoricalDataRecord
    
    # Obtener períodos disponibles en el histórico
    db = SessionLocal()
    per_recs = db.query(HistoricalDataRecord.periodo).distinct().all()
    db.close()
    periodos_hist = sorted([r[0] for r in per_recs], reverse=True)
    if not periodos_hist:
        periodos_hist = ["2026-03", "2025-12", "2024-12"]
        
    # Parámetros globales en columnas
    col_p1, col_p2, col_esc = st.columns([1.5, 1.5, 2])
    with col_p1:
        periodo_actual = st.selectbox("Periodo Actual:", periodos_hist, index=0, format_func=format_periodo)
    with col_p2:
        periodo_comp = st.selectbox("Periodo Comparativo (Opcional):", ["Ninguno"] + periodos_hist, index=1 if len(periodos_hist) > 1 else 0, format_func=format_periodo)
    with col_esc:
        unidad_nota = st.selectbox(
            "Unidad de medida en reportes:",
            ["Miles de pesos (M$)", "Unidades ($)", "Millones de pesos (MM$)"],
            index=0
        )
        
        if "Miles" in unidad_nota:
            scale_factor_nota = 1000.0
        elif "Millones" in unidad_nota:
            scale_factor_nota = 1000000.0
        else:
            scale_factor_nota = 1.0
            
    st.write("")
    
    # Crear pestañas para cada rubro
    tabs = st.tabs([
        "💰 Activos Corrientes",
        "🏢 Activos No Corrientes",
        "💳 Pasivos Corrientes",
        "🛡️ Pasivos No Corrientes",
        "📊 Patrimonio",
        "📈 Resultados",
        "📂 Informes Corporativos"
    ])
    
    # Tab 1: Activos Corrientes
    with tabs[0]:
        st.subheader("Notas de Activos Corrientes")
        render_note_section(notes_by_category["activos_corrientes"], "act_corr", scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp)
        
    # Tab 2: Activos No Corrientes
    with tabs[1]:
        st.subheader("Notas de Activos No Corrientes")
        render_note_section(notes_by_category["activos_no_corrientes"], "act_no_corr", scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp)
        
    # Tab 3: Pasivos Corrientes
    with tabs[2]:
        st.subheader("Notas de Pasivos Corrientes")
        render_note_section(notes_by_category["pasivos_corrientes"], "pas_corr", scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp)
        
    # Tab 4: Pasivos No Corrientes
    with tabs[3]:
        st.subheader("Notas de Pasivos No Corrientes")
        render_note_section(notes_by_category["pasivos_no_corrientes"], "pas_no_corr", scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp)
        
    # Tab 5: Patrimonio
    with tabs[4]:
        st.subheader("Notas de Patrimonio")
        render_note_section(notes_by_category["patrimonio"], "patrimonio", scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp)
        
    # Tab 6: Resultados
    with tabs[5]:
        st.subheader("Notas de Resultados (P&L)")
        render_note_section(notes_by_category["resultados"], "resultados", scale_factor_nota, unidad_nota, empresa_path, empresa_seleccionada, periodo_actual, periodo_comp)
        
    # Tab 7: Informes
    with tabs[6]:
        st.subheader("Informes Adicionales")
        st.write("Generación de paquetes de gerencia corporativos. (PDF, Word).")
        st.info("Funcionalidad en desarrollo.")
