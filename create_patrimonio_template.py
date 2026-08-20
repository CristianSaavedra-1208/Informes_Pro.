import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os

empresa = "Pacifico SpA"
path = f"data/empresas/{empresa}/Estado de Cambios en el Patrimonio.xlsx"
os.makedirs(os.path.dirname(path), exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active

ws.title = "Estado Cambios Patrimonio"

# Titles
ws.cell(row=1, column=1, value="PACIFICO CABLE SpA.").font = Font(bold=True)
ws.cell(row=2, column=1, value="ESTADOS SEPARADOS DE CAMBIOS EN EL PATRIMONIO").font = Font(bold=True)
ws.cell(row=3, column=1, value="Por los ejercicios terminados al 31 de diciembre de 2025 y 2024")

# All centered
for r in range(1, 4):
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

headers = ["Capital emitido", "Ganancias (pérdidas) acumuladas", "Otras reservas", "Patrimonio Total"]
col_letters = ['B', 'C', 'D', 'E']

# Row 5 headers
for i, h in enumerate(headers):
    c = ws.cell(row=5, column=i+2, value=h)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

# Row 6 (M$)
for i in range(4):
    c = ws.cell(row=6, column=i+2, value="M$")
    c.alignment = Alignment(horizontal='center')

thin_bottom = Border(bottom=Side(style='thin'))
thin_top_bottom = Border(top=Side(style='thin'), bottom=Side(style='thin'))

for i in range(4):
    ws.cell(row=6, column=i+2).border = thin_bottom

ws.cell(row=7, column=1, value="Saldo inicial ejercicio actual (fecha inicial)")
ws.cell(row=8, column=1, value="Cambios en el patrimonio:")
ws.cell(row=9, column=1, value="Ganancia (pérdida)")
ws.cell(row=10, column=1, value="Reclasificación reservas periodos anteriores")
ws.cell(row=11, column=1, value="Total resultados integrales")
ws.cell(row=12, column=1, value="Saldo final ejercicio actual (fecha final)")

# Borders for totals
for i in range(4):
    ws.cell(row=11, column=i+2).border = thin_top_bottom
    ws.cell(row=12, column=i+2).border = thin_top_bottom

# Add spacing row to prevent UI from merging tables
ws.cell(row=14, column=1, value=" ")

# 2024 Block
ws.cell(row=15, column=1, value="Saldo inicial ejercicio anterior (fecha inicial)")
ws.cell(row=16, column=1, value="Cambios en el patrimonio:")
ws.cell(row=17, column=1, value="Ganancia (pérdida)")
ws.cell(row=18, column=1, value="Reclasificación reservas periodos anteriores")
ws.cell(row=19, column=1, value="Total resultados integrales")
ws.cell(row=20, column=1, value="Saldo final ejercicio anterior (fecha final)")

for i in range(4):
    ws.cell(row=19, column=i+2).border = thin_top_bottom
    ws.cell(row=20, column=i+2).border = thin_top_bottom

ws.column_dimensions['A'].width = 45
for L in col_letters:
    ws.column_dimensions[L].width = 20

wb.save(path)
print(f"Plantilla creada en {path}")
