"""
migrate_templates.py
====================
Migra todas las plantillas Excel existentes en el proyecto agregándoles la hoja
oculta '_CONFIG_' con metadatos de layout.

Este script es idempotente: si una plantilla ya tiene '_CONFIG_', la omite.
Se puede ejecutar manualmente o se llama automáticamente al iniciar la app.

Uso manual:
    python migrate_templates.py

Uso programático:
    from migrate_templates import run_migration_if_needed
    run_migration_if_needed()
"""

import os
import sys
import shutil
import openpyxl

# Asegurar que el path del proyecto esté disponible
_PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from src.core.excel_utils import detect_balance_columns, write_template_config

# ─────────────────────────────────────────────────────────────────────────────
# Configuraciones conocidas: si el nombre de archivo coincide, se usa
# esta config directa en lugar de la detección automática.
# Sirve como override para casos donde la detección automática puede fallar.
# ─────────────────────────────────────────────────────────────────────────────
KNOWN_CONFIGS = {
    # Balance con Nota (ej. Holdco: col B=name, C=nota, D=actual, F=comp)
    # Se detecta automáticamente vía detect_balance_columns
    # Balance sin Nota (ej. Pacifico: col B=name, C=actual, E=comp)
    # Se detecta automáticamente
}

# Archivos de plantilla conocidos y su tipo
TEMPLATE_TYPES = {
    "Balance clasificado.xlsx": "balance",
    "Estado de Resultados Clasificados.xlsx": "er",
    "Estado de Flujos de Efectivo.xlsx": "flujo",
    "Estado de Flujos de Efectivo Indirecto.xlsx": "flujo_indirecto",
    "Estado de Resultados Integrales.xlsx": "ori",
    "Estado de Cambios en el Patrimonio.xlsx": "patrimonio",
}


def _detect_data_start_row(ws, name_col_idx: int) -> int:
    """Detecta la primera fila que contiene datos reales (después del encabezado)."""
    # Busca la primera fila donde la columna name_col tiene texto real
    # que no sea un título de la empresa (muchos caracteres en mayúscula)
    # sino una categoría contable conocida
    trigger_words = [
        "activos corrientes", "activos", "ingresos", "flujos", "saldo",
        "capital", "clasificacion", "clasificación", "concepto",
        "efectivo", "pasivos corrientes"
    ]
    for row in range(1, min(20, ws.max_row + 1)):
        val = ws.cell(row=row, column=name_col_idx).value
        if val and isinstance(val, str):
            val_lower = val.strip().lower()
            if any(tw in val_lower for tw in trigger_words):
                return row
    return 5  # fallback


def _migrate_single_template(path: str, template_type: str) -> bool:
    """
    Agrega la hoja '_CONFIG_' a una plantilla Excel.
    Retorna True si se migró, False si ya tenía _CONFIG_ o si ocurrió un error.
    """
    try:
        wb = openpyxl.load_workbook(path)

        # Ya migrada → saltar
        if "_CONFIG_" in wb.sheetnames:
            return False

        ws = wb.active

        # Detectar columnas automáticamente
        name_col_idx, nota_col_idx, val25_col_idx, val24_col_idx = detect_balance_columns(ws)
        data_start_row = _detect_data_start_row(ws, name_col_idx)

        cfg = {
            "template_type":  template_type,
            "name_col":       name_col_idx,
            "nota_col":       nota_col_idx if nota_col_idx is not None else 0,
            "val_actual_col": val25_col_idx,
            "val_comp_col":   val24_col_idx,
            "data_start_row": data_start_row,
            "version":        1,
        }

        write_template_config(wb, cfg)

        # Respaldar original antes de sobreescribir
        bak_path = path + ".bak"
        if not os.path.exists(bak_path):
            shutil.copy2(path, bak_path)

        wb.save(path)
        return True

    except Exception as e:
        print(f"  [WARN] Error migrando '{path}': {e}")
        return False


def run_migration_if_needed(verbose: bool = True) -> dict:
    """
    Recorre todas las carpetas de empresas y la carpeta templates/ del proyecto,
    y agrega '_CONFIG_' a cada plantilla que aún no la tenga.
    
    Retorna un dict con estadísticas: {migrated: int, skipped: int, errors: int}
    """
    base = _PROJECT_ROOT
    stats = {"migrated": 0, "skipped": 0, "errors": 0}

    # Rutas a escanear
    scan_paths = [
        os.path.join(base, "templates"),
    ]
    empresas_root = os.path.join(base, "data", "empresas")
    if os.path.isdir(empresas_root):
        for empresa in os.listdir(empresas_root):
            empresa_path = os.path.join(empresas_root, empresa)
            if os.path.isdir(empresa_path):
                scan_paths.append(empresa_path)

    if verbose:
        print("=" * 60)
        print("  MIGRACION DE PLANTILLAS -> Agregando hoja _CONFIG_")
        print("=" * 60)

    for folder in scan_paths:
        if not os.path.isdir(folder):
            continue
        folder_label = os.path.relpath(folder, base)
        for filename, ttype in TEMPLATE_TYPES.items():
            fpath = os.path.join(folder, filename)
            if not os.path.exists(fpath):
                continue
            migrated = _migrate_single_template(fpath, ttype)
            if migrated:
                stats["migrated"] += 1
                if verbose:
                    print(f"  [OK]  Migrada: {folder_label}/{filename}")
            else:
                # Verificar si ya tenía _CONFIG_ (normal) o hubo error (silencioso)
                try:
                    wb = openpyxl.load_workbook(fpath)
                    if "_CONFIG_" in wb.sheetnames:
                        stats["skipped"] += 1
                        if verbose:
                            print(f"  [--]  Ya tiene _CONFIG_: {folder_label}/{filename}")
                    else:
                        stats["errors"] += 1
                except Exception:
                    stats["errors"] += 1

    if verbose:
        print("-" * 60)
        print(f"  Migradas: {stats['migrated']} | Omitidas: {stats['skipped']} | Errores: {stats['errors']}")
        print("=" * 60)


    return stats


if __name__ == "__main__":
    run_migration_if_needed(verbose=True)
